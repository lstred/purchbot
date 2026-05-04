from __future__ import annotations

from datetime import date, timedelta
from pathlib import Path
import os
import json
import calendar
import re
from time import perf_counter

import pandas as pd
import numpy as np
import streamlit as st
import plotly.express as px
from io import BytesIO

# PDF generation
try:
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
        PageBreak, Image, KeepTogether
    )
    from reportlab.graphics.shapes import Drawing, Rect, String
    from reportlab.graphics.charts.lineplots import LinePlot
    from reportlab.graphics.charts.barcharts import VerticalBarChart
    from reportlab.graphics import renderPDF
    import matplotlib
    matplotlib.use('Agg')  # Non-interactive backend
    import matplotlib.pyplot as plt
    _HAS_REPORTLAB = True
except Exception:
    _HAS_REPORTLAB = False

from app.config import get_config
from app.data import loaders
from app.services.metrics_service import compute_dashboard_data, MetricsFilters
from app.data.seasonality_store import load_seasonality, save_for_cost_center
from app.data.stockturn_store import load_targets, save_target, get_target_for_cc
from app.data.launch_store import load_launch_dates

# Global configuration
config = get_config()

# User preferences persistence under %APPDATA%/PurchaseOrderBot/history
def _prefs_path() -> Path:
    base = Path(os.environ.get("APPDATA", "")).expanduser()
    root = (base if str(base).strip() else Path.home()) / "PurchaseOrderBot" / "history"
    root.mkdir(parents=True, exist_ok=True)
    return root / "user_prefs.json"

def _load_prefs() -> dict:
    try:
        p = _prefs_path()
        if p.exists():
            return json.loads(p.read_text(encoding="utf-8")) or {}
    except Exception:
        pass
    return {}

def _save_prefs(prefs: dict) -> None:
    try:
        p = _prefs_path()
        p.write_text(json.dumps(prefs, indent=2), encoding="utf-8")
    except Exception:
        # best-effort only
        pass

def _ensure_selection(options: list[str], preferred: list[str] | None) -> list[str]:
    opts = list(options or [])
    if not opts:
        return []
    if not preferred:
        return [opts[0]]
    filtered = [x for x in preferred if x in opts]
    return filtered if filtered else [opts[0]]

def _default_date_range(months: int) -> tuple[date, date]:
    end = date.today()
    # Approximate months by 30-day windows
    start = end - timedelta(days=max(1, months) * 30)
    return start, end

def _format_days(value: float) -> str:
    if value == float("inf") or value >= 1e9:
        return "∞"
    try:
        return f"{float(value):.0f}"
    except Exception:
        return "-"

def _generate_supplier_performance_pdf(
    filtered_data: pd.DataFrame,
    all_groups: list,
    group_field: str,
    group_name: str,
    ytd_start_date: date,
    ytd_end_date: date,
    unique_years: list,
    inventory_costs: pd.DataFrame
) -> BytesIO:
    """Generate a professional PDF report for supplier performance."""
    if not _HAS_REPORTLAB:
        raise ImportError("reportlab is not installed")
    
    from reportlab.lib.units import inch
    from reportlab.platypus import PageBreak, KeepTogether
    import plotly.graph_objects as go
    from plotly.subplots import make_subplots
    import base64
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), topMargin=30, bottomMargin=30, leftMargin=36, rightMargin=36)
    story = []
    styles = getSampleStyleSheet()
    
    # Custom color palette for professional look with pops of color
    header_color = colors.HexColor('#2C3E50')  # Dark blue-grey
    accent_color = colors.HexColor('#3498DB')   # Bright blue
    positive_color = colors.HexColor('#27AE60') # Green
    negative_color = colors.HexColor('#E74C3C') # Red
    
    # Title with accent color
    title_style = styles['Title'].clone('CustomTitle')
    title_style.textColor = header_color
    title_text = f"Supplier Performance Report - {group_name}"
    story.append(Paragraph(title_text, title_style))
    story.append(Spacer(1, 8))
    
    # Summary info
    date_range = f"Date Range: {ytd_start_date.strftime('%m/%d/%Y')} to {ytd_end_date.strftime('%m/%d/%Y')}"
    years_text = f"Years: {', '.join(map(str, unique_years))}"
    story.append(Paragraph(date_range, styles['Normal']))
    story.append(Paragraph(years_text, styles['Normal']))
    story.append(Paragraph(f"Total {group_name} Groups: {len(all_groups)}", styles['Normal']))
    
    # Calculate overall ROI across all groups
    overall_roi_text = ""
    if not inventory_costs.empty and "gross_profit_usd" in filtered_data.columns and "inventory_flag" in filtered_data.columns:
        all_stocking = filtered_data[filtered_data["inventory_flag"] == "Y"].copy()
        if not all_stocking.empty:
            all_inv_cost = inventory_costs[
                inventory_costs["sku"].isin(all_stocking["sku"].unique())
            ]["total_cost"].sum()
            
            if all_inv_cost > 0:
                roi_by_year = []
                for year in unique_years:
                    year_stocking = all_stocking[all_stocking["invoice_date"].dt.year == year]
                    year_gp = year_stocking["gross_profit_usd"].sum()
                    roi = year_gp / all_inv_cost
                    roi_by_year.append(f"{year}: {roi:.2f}x")
                overall_roi_text = "Overall ROI: " + ", ".join(roi_by_year)
    
    if overall_roi_text:
        roi_style = styles['Normal'].clone('ROIStyle')
        roi_style.fontName = 'Helvetica-Bold'
        roi_style.fontSize = 11
        roi_style.textColor = header_color
        story.append(Paragraph(overall_roi_text, roi_style))
    
    story.append(Spacer(1, 12))
    
    # Add ROI summary table for all groups
    if not inventory_costs.empty and "gross_profit_usd" in filtered_data.columns and "inventory_flag" in filtered_data.columns:
        summary_data = []
        
        for group_value in all_groups:
            group_data = filtered_data[filtered_data[group_field] == group_value].copy()
            
            row = [str(group_value)]
            
            # For Salesperson: show average monthly sales instead of ROI
            if group_field == "salesperson_desc":
                monthly_avgs = {}
                for year in unique_years:
                    year_data = group_data[group_data["invoice_date"].dt.year == year]
                    if not year_data.empty:
                        year_data_copy = year_data.copy()
                        year_data_copy["month"] = year_data_copy["invoice_date"].dt.to_period("M")
                        monthly_sales = year_data_copy.groupby("month")["extended_price_usd"].sum()
                        avg_monthly = monthly_sales.mean()
                        monthly_avgs[year] = avg_monthly
                        row.append(f"${avg_monthly:,.0f}")
                    else:
                        row.append("$0")
                
                # Add growth percentage
                if len(unique_years) >= 2:
                    current_avg = monthly_avgs.get(unique_years[-1], 0)
                    prior_avg = monthly_avgs.get(unique_years[-2], 0)
                    if prior_avg > 0:
                        growth_pct = (current_avg / prior_avg) * 100
                        row.append(f"{growth_pct:.1f}%")
                    else:
                        row.append("0.0%")
                
                summary_data.append(row)
            else:
                # For other groupings: show ROI
                stocking_items = group_data[group_data["inventory_flag"] == "Y"].copy()
                
                if not stocking_items.empty:
                    group_inv_cost = inventory_costs[
                        inventory_costs["sku"].isin(stocking_items["sku"].unique())
                    ]["total_cost"].sum()
                    
                    if group_inv_cost > 0:
                        for year in unique_years:
                            year_stocking = stocking_items[stocking_items["invoice_date"].dt.year == year]
                            year_gp = year_stocking["gross_profit_usd"].sum()
                            roi = year_gp / group_inv_cost
                            row.append(f"{roi:.2f}x")
                        
                        summary_data.append(row)
        
        if summary_data:
            # Sort by appropriate column
            if group_field == "salesperson_desc" and len(unique_years) >= 2:
                # Sort by growth % (last column)
                summary_data.sort(key=lambda x: float(x[-1].replace('%', '')), reverse=True)
                
                # Add weighted average row for Salesperson
                overall_monthly_avgs = {}
                for year in unique_years:
                    year_data = filtered_data[filtered_data["invoice_date"].dt.year == year]
                    if not year_data.empty:
                        year_data_copy = year_data.copy()
                        year_data_copy["month"] = year_data_copy["invoice_date"].dt.to_period("M")
                        monthly_sales = year_data_copy.groupby("month")["extended_price_usd"].sum()
                        avg_monthly = monthly_sales.mean()
                        overall_monthly_avgs[year] = avg_monthly
                
                # Create summary row
                summary_row = ["WEIGHTED AVERAGE"]
                for year in unique_years:
                    avg_val = overall_monthly_avgs.get(year, 0)
                    summary_row.append(f"${avg_val:,.0f}")
                
                # Calculate growth
                current_avg = overall_monthly_avgs.get(unique_years[-1], 0)
                prior_avg = overall_monthly_avgs.get(unique_years[-2], 0)
                if prior_avg > 0:
                    growth_pct = (current_avg / prior_avg) * 100
                    summary_row.append(f"{growth_pct:.1f}%")
                else:
                    summary_row.append("0.0%")
                
                summary_data.append(summary_row)
            elif len(unique_years) > 0:
                # Sort by current year (last year) ROI
                summary_data.sort(key=lambda x: float(x[-1].replace('x', '')), reverse=True)
            
            # Create table header based on grouping type
            if group_field == "salesperson_desc":
                summary_header = [group_name] + [f"{year} Avg Monthly Sales" for year in unique_years]
                if len(unique_years) >= 2:
                    summary_header.append("Growth %")
            else:
                summary_header = [group_name] + [f"{year} ROI" for year in unique_years]
            
            summary_table_data = [summary_header] + summary_data
            
            summary_title_style = styles['Heading3'].clone('SummaryTitle')
            summary_title_style.textColor = header_color
            story.append(Paragraph("ROI Summary by Group", summary_title_style))
            story.append(Spacer(1, 6))
            
            from reportlab.platypus import Table as RLTable
            summary_table = RLTable(summary_table_data, repeatRows=1)
            
            # Determine if last row is summary (for Salesperson)
            num_rows = len(summary_table_data)
            has_summary_row = group_field == "salesperson_desc" and len(unique_years) >= 2
            
            table_style = [
                ('BACKGROUND', (0, 0), (-1, 0), header_color),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
            ]
            
            # Add special styling for weighted average row
            if has_summary_row:
                table_style.extend([
                    ('BACKGROUND', (0, num_rows-1), (-1, num_rows-1), colors.lightblue),
                    ('FONTNAME', (0, num_rows-1), (-1, num_rows-1), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, num_rows-1), (-1, num_rows-1), 10),
                    ('LINEABOVE', (0, num_rows-1), (-1, num_rows-1), 1.5, colors.black)
                ])
            
            summary_table.setStyle(table_style)
            story.append(summary_table)
            story.append(Spacer(1, 12))
    
    # Add aggregated price class comparison table for non-Salesperson groupings
    if group_field != "salesperson_desc" and len(unique_years) >= 2:
        current_year = unique_years[-1]
        prior_year = unique_years[-2]
        
        # Filter to current and prior year
        ytd_current_year = filtered_data[filtered_data["invoice_date"].dt.year == current_year].copy()
        ytd_prior_year = filtered_data[filtered_data["invoice_date"].dt.year == prior_year].copy()
        
        # Calculate ROI for each price class
        current_metric_lookup = {}
        prior_metric_lookup = {}
        
        if "inventory_costs" in locals() and not inventory_costs.empty and "gross_profit_usd" in filtered_data.columns:
            # Current year ROI by price class
            current_stocking = ytd_current_year[ytd_current_year["inventory_flag"] == "Y"].copy()
            if not current_stocking.empty:
                current_gp_by_class = current_stocking.groupby("price_class_desc")["gross_profit_usd"].sum()
                
                for pc_desc in current_gp_by_class.index:
                    pc_inv_cost = inventory_costs[
                        (inventory_costs["price_class_desc"].astype(str).str.strip() == str(pc_desc).strip()) &
                        (inventory_costs["price_class_desc"].notna())
                    ]["total_cost"].sum()
                    if pc_inv_cost > 0:
                        current_metric_lookup[pc_desc] = current_gp_by_class[pc_desc] / pc_inv_cost
            
            # Prior year ROI by price class
            prior_stocking = ytd_prior_year[ytd_prior_year["inventory_flag"] == "Y"].copy()
            if not prior_stocking.empty:
                prior_gp_by_class = prior_stocking.groupby("price_class_desc")["gross_profit_usd"].sum()
                
                for pc_desc in prior_gp_by_class.index:
                    pc_inv_cost = inventory_costs[
                        (inventory_costs["price_class_desc"].astype(str).str.strip() == str(pc_desc).strip()) &
                        (inventory_costs["price_class_desc"].notna())
                    ]["total_cost"].sum()
                    if pc_inv_cost > 0:
                        prior_metric_lookup[pc_desc] = prior_gp_by_class[pc_desc] / pc_inv_cost
        
        # Get top 10 for each year
        if not ytd_current_year.empty:
            ytd_current_top10 = ytd_current_year.groupby("price_class_desc")["extended_price_usd"].sum().reset_index()
            ytd_current_top10 = ytd_current_top10.sort_values("extended_price_usd", ascending=False).head(10)
        else:
            ytd_current_top10 = pd.DataFrame(columns=["price_class_desc", "extended_price_usd"])
        
        if not ytd_prior_year.empty:
            ytd_prior_top10 = ytd_prior_year.groupby("price_class_desc")["extended_price_usd"].sum().reset_index()
            ytd_prior_top10 = ytd_prior_top10.sort_values("extended_price_usd", ascending=False).head(10)
        else:
            ytd_prior_top10 = pd.DataFrame(columns=["price_class_desc", "extended_price_usd"])
        
        # Build combined table
        price_class_title_style = styles['Heading3'].clone('PriceClassTitle')
        price_class_title_style.textColor = header_color
        story.append(Paragraph("Aggregated Price Class Comparison", price_class_title_style))
        story.append(Spacer(1, 6))
        
        combined_data = [["Price Class", f"{current_year}", "ROI", "Price Class", f"{prior_year}", "ROI"]]
        
        max_rows = max(len(ytd_current_top10), len(ytd_prior_top10))
        for i in range(max_rows):
            current_class = ytd_current_top10.iloc[i]["price_class_desc"] if i < len(ytd_current_top10) else ""
            current_sales = f"${ytd_current_top10.iloc[i]['extended_price_usd']:,.0f}" if i < len(ytd_current_top10) else ""
            current_metric_val = current_metric_lookup.get(current_class, 0)
            current_metric = f"{current_metric_val:.2f}x" if current_metric_val > 0 else ""
            
            prior_class = ytd_prior_top10.iloc[i]["price_class_desc"] if i < len(ytd_prior_top10) else ""
            prior_sales = f"${ytd_prior_top10.iloc[i]['extended_price_usd']:,.0f}" if i < len(ytd_prior_top10) else ""
            prior_metric_val = prior_metric_lookup.get(prior_class, 0)
            prior_metric = f"{prior_metric_val:.2f}x" if prior_metric_val > 0 else ""
            
            combined_data.append([current_class, current_sales, current_metric, prior_class, prior_sales, prior_metric])
        
        current_top10_total = ytd_current_top10["extended_price_usd"].sum() if not ytd_current_top10.empty else 0
        current_grand_total = ytd_current_year["extended_price_usd"].sum() if not ytd_current_year.empty else 0
        prior_top10_total = ytd_prior_top10["extended_price_usd"].sum() if not ytd_prior_top10.empty else 0
        prior_grand_total = ytd_prior_year["extended_price_usd"].sum() if not ytd_prior_year.empty else 0
        
        combined_data.append(["Top 10 Total", f"${current_top10_total:,.0f}", "", "Top 10 Total", f"${prior_top10_total:,.0f}", ""])
        combined_data.append(["Grand Total", f"${current_grand_total:,.0f}", "", "Grand Total", f"${prior_grand_total:,.0f}", ""])
        
        price_table = RLTable(combined_data, colWidths=[155, 80, 50, 155, 80, 50])
        price_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), header_color),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('ALIGN', (2, 0), (2, -1), 'RIGHT'),
            ('ALIGN', (3, 0), (3, -1), 'LEFT'),
            ('ALIGN', (4, 0), (4, -1), 'RIGHT'),
            ('ALIGN', (5, 0), (5, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('TOPPADDING', (0, 0), (-1, -1), 2),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
            ('BACKGROUND', (0, 1), (2, -3), colors.HexColor('#E3F2FD')),
            ('BACKGROUND', (3, 1), (5, -3), colors.HexColor('#FFF9E1')),
            ('BACKGROUND', (0, -2), (2, -1), colors.HexColor('#B3E5FC')),
            ('BACKGROUND', (3, -2), (5, -1), colors.HexColor('#FFF176')),
            ('FONTNAME', (0, -2), (-1, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('LINEAFTER', (2, 0), (2, -1), 1.5, header_color)
        ]))
        story.append(price_table)
        story.append(Spacer(1, 12))
    
    story.append(PageBreak())
    
    # Process each group (limit to avoid huge PDFs)
    max_groups = min(50, len(all_groups))
    
    for idx, group_value in enumerate(all_groups[:max_groups]):
        # Build content for this group that must stay together
        group_content = []
        
        # Group header with accent color (will add ROI below)
        header_style = styles['Heading2'].clone('CustomHeader')
        header_style.textColor = accent_color
        group_title_base = f"{group_name}: {group_value}"
        
        # Filter data for this group
        group_data = filtered_data[filtered_data[group_field] == group_value].copy()
        
        # Get all unique years in the group data
        all_years = sorted(group_data["invoice_date"].dt.year.unique())
        
        # Create monthly aggregation for chart and analysis
        group_data_copy = group_data.copy()
        group_data_copy["month"] = group_data_copy["invoice_date"].dt.to_period("M").dt.to_timestamp()
        group_data_copy["year"] = group_data_copy["invoice_date"].dt.year
        monthly_data = group_data_copy.groupby(["month", "year"])["extended_price_usd"].sum().reset_index()
        monthly_data["period"] = monthly_data["year"].astype(str)
        
        # Calculate average sales for all groups (for secondary y-axis)
        all_data_monthly = filtered_data.copy()
        all_data_monthly["month"] = all_data_monthly["invoice_date"].dt.to_period("M").dt.to_timestamp()
        # Use weighted average: sum of all sales / count of months (not divided by number of groups)
        # This gives proper weight to high and low performers
        avg_monthly_data = all_data_monthly.groupby("month")["extended_price_usd"].sum().reset_index()
        avg_monthly_data.columns = ["month", "avg_sales"]  # avg_sales is actually total sales per month
        
        # Two-column layout: Chart on left, tables on right
        from reportlab.platypus import Table as RLTable
        
        # LEFT SIDE: Generate chart
        chart_content = []
        if not monthly_data.empty:
            try:
                from datetime import datetime
                
                fig = make_subplots(specs=[[{"secondary_y": True}]])
                
                years_in_data = sorted(monthly_data["period"].unique())
                year_colors = ["#1f77b4", "#ff7f0e", "#2ca02c", "#d62728", "#9467bd"]
                
                # Add group lines (primary y-axis)
                for idx_year, year in enumerate(years_in_data):
                    year_data = monthly_data[monthly_data["period"] == year]
                    fig.add_trace(
                        go.Scatter(
                            x=year_data["month"],
                            y=year_data["extended_price_usd"],
                            name=year,
                            mode="lines+markers",
                            line=dict(color=year_colors[idx_year % len(year_colors)], width=2),
                            marker=dict(size=5)
                        ),
                        secondary_y=False
                    )
                
                # Add average line (secondary y-axis)
                if not avg_monthly_data.empty:
                    fig.add_trace(
                        go.Scatter(
                            x=avg_monthly_data["month"],
                            y=avg_monthly_data["avg_sales"],
                            name="Average",
                            mode="lines+markers",
                            line=dict(color="#8c564b", dash="dash", width=2),
                            marker=dict(size=4)
                        ),
                        secondary_y=True
                    )
                
                # Update axes
                fig.update_xaxes(title_text="Month", title_font=dict(size=10))
                fig.update_yaxes(title_text="Sales (USD)", tickformat="$,.0f", secondary_y=False, title_font=dict(size=10))
                fig.update_yaxes(title_text="Avg Sales (USD)", tickformat="$,.0f", secondary_y=True, title_font=dict(size=10))
                
                fig.update_layout(
                    title="Sales Comparison",
                    title_font=dict(size=11),
                    height=170,
                    width=380,
                    showlegend=True,
                    legend=dict(orientation="h", yanchor="top", y=-0.22, xanchor="center", x=0.5, font=dict(size=8)),
                    margin=dict(l=40, r=25, t=30, b=45)
                )
                
                # Convert to image with higher resolution for clarity
                img_bytes = fig.to_image(format="png", width=380, height=170, engine="kaleido", scale=2)
                from reportlab.platypus import Image as RLImage
                img = RLImage(BytesIO(img_bytes), width=3.8*inch, height=1.7*inch)
                chart_content.append(img)
            except ImportError as ie:
                chart_content.append(Paragraph("<i>Chart unavailable: Restart Streamlit app.</i>", styles['Normal']))
            except Exception as e:
                chart_content.append(Paragraph("<i>Chart unavailable: Restart Streamlit app.</i>", styles['Normal']))
        
        # RIGHT SIDE: Build tables
        tables_content = []
        
        # Performance metrics table (compact)
        if not monthly_data.empty:
            from datetime import datetime
            current_month = pd.Timestamp(datetime.now().replace(day=1))
            
            # Filter out current incomplete month from group data
            complete_months = monthly_data[monthly_data["month"] < current_month].copy()
            
            # Also filter out current month from overall average data
            complete_avg_monthly = avg_monthly_data[avg_monthly_data["month"] < current_month].copy()
            
            if not complete_months.empty and not complete_avg_monthly.empty:
                # Merge group sales with overall total sales for same months
                comparison_data = complete_months.merge(
                    complete_avg_monthly[["month", "avg_sales"]], 
                    on="month", 
                    how="inner"
                )
                
                if not comparison_data.empty:
                    # Calculate this group's percentage of total sales for each month
                    comparison_data["pct_of_total"] = (comparison_data["extended_price_usd"] / comparison_data["avg_sales"] * 100).round(2)
                    comparison_data["deviation_from_avg"] = comparison_data["extended_price_usd"] - comparison_data["avg_sales"]
                    comparison_data["deviation_pct"] = ((comparison_data["extended_price_usd"] - comparison_data["avg_sales"]) / comparison_data["avg_sales"] * 100).round(2)
                    
                    # Best month = highest percentage of total sales
                    best_month = comparison_data.loc[comparison_data["pct_of_total"].idxmax()]
                    # Worst month = lowest percentage of total sales (excluding current month already filtered)
                    worst_month = comparison_data.loc[comparison_data["pct_of_total"].idxmin()]
                    
                    # Calculate overall average percentage of total sales
                    overall_avg_pct = comparison_data["pct_of_total"].mean()
                    
                    # Calculate average for display
                    avg_monthly_sales = complete_months["extended_price_usd"].mean()
                else:
                    avg_monthly_sales = 0
                    overall_avg_pct = 0
                    best_month = None
                    worst_month = None
            else:
                # If no complete months (shouldn't happen), skip metrics
                avg_monthly_sales = 0
                overall_avg_pct = 0
                best_month = None
                worst_month = None
            
            if best_month is not None and worst_month is not None:
                perf_data = [
                    ["Metric", "Value"],
                    ["Avg Monthly", f"${avg_monthly_sales:,.0f}"],
                    ["Avg % of Total", f"{overall_avg_pct:.1f}% of all sales"],
                    ["Best", f"{best_month['month'].strftime('%b %Y')}: {best_month['pct_of_total']:.1f}% of total"],
                    ["Worst", f"{worst_month['month'].strftime('%b %Y')}: {worst_month['pct_of_total']:.1f}% of total"]
                ]
                
                perf_table = Table(perf_data, colWidths=[75, 190])
                perf_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), header_color),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 8),
                    ('FONTSIZE', (0, 1), (-1, -1), 7),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 4),
                    ('TOPPADDING', (0, 1), (-1, -1), 2),
                    ('BOTTOMPADDING', (0, 1), (-1, -1), 2),
                    ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#E8F4F8')),
                    ('BACKGROUND', (0, 2), (-1, 2), colors.HexColor('#E8F4F8')),
                    ('BACKGROUND', (0, 3), (-1, 3), colors.HexColor('#D5F4E6')),  # Green tint for best
                    ('BACKGROUND', (0, 4), (-1, 4), colors.HexColor('#FADBD8')),  # Red tint for worst
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)
                ]))
                tables_content.append(perf_table)
                tables_content.append(Spacer(1, 6))
        
        # Calculate ROI or Growth % metrics
        roi_text = ""
        if group_field == "salesperson_desc":
            # For Salesperson: show Growth % based on average monthly sales
            if len(all_years) >= 2:
                current_year = all_years[-1]
                prior_year = all_years[-2]
                
                # Calculate current year avg monthly sales
                current_data = group_data[group_data["invoice_date"].dt.year == current_year].copy()
                if not current_data.empty:
                    current_data_copy = current_data.copy()
                    current_data_copy["month"] = current_data_copy["invoice_date"].dt.to_period("M")
                    current_monthly = current_data_copy.groupby("month")["extended_price_usd"].sum()
                    current_avg = current_monthly.mean() if len(current_monthly) > 0 else 0
                else:
                    current_avg = 0
                
                # Calculate prior year avg monthly sales
                prior_data = group_data[group_data["invoice_date"].dt.year == prior_year].copy()
                if not prior_data.empty:
                    prior_data_copy = prior_data.copy()
                    prior_data_copy["month"] = prior_data_copy["invoice_date"].dt.to_period("M")
                    prior_monthly = prior_data_copy.groupby("month")["extended_price_usd"].sum()
                    prior_avg = prior_monthly.mean() if len(prior_monthly) > 0 else 0
                else:
                    prior_avg = 0
                
                if prior_avg > 0:
                    growth_pct = (current_avg / prior_avg) * 100
                    roi_text = f" | Growth %: {growth_pct:.1f}%"
        elif not inventory_costs.empty and "gross_profit_usd" in group_data.columns:
            # Get stocking items in this group
            if "inventory_flag" in group_data.columns:
                stocking_items = group_data[group_data["inventory_flag"] == "Y"].copy()
            else:
                stocking_items = pd.DataFrame()
            
            if not stocking_items.empty and "price_class" in stocking_items.columns:
                # Calculate total inventory cost for this group's stocking items
                group_inv_cost = inventory_costs[
                    inventory_costs["sku"].isin(stocking_items["sku"].unique())
                ]["total_cost"].sum()
                
                if group_inv_cost > 0:
                    roi_values = []
                    for year in all_years:
                        year_stocking = stocking_items[stocking_items["invoice_date"].dt.year == year]
                        year_gp = year_stocking["gross_profit_usd"].sum()
                        roi = year_gp / group_inv_cost
                        roi_values.append(f"{year}: {roi:.2f}x")
                    roi_text = " | ROI: " + ", ".join(roi_values)
        
        # Add title with ROI
        group_title = group_title_base + roi_text
        group_content.append(Paragraph(f"<b>{group_title}</b>", header_style))
        group_content.append(Spacer(1, 4))
        
        # Year summary table (compact)
        summary_data = [["Year", "Total"]]
        for year in all_years:
            year_data = group_data[group_data["invoice_date"].dt.year == year]
            grand_total = year_data["extended_price_usd"].sum()
            summary_data.append([str(year), f"${grand_total:,.0f}"])
        overall_total = group_data["extended_price_usd"].sum()
        summary_data.append(["All Years", f"${overall_total:,.0f}"])
        
        summary_table = Table(summary_data, colWidths=[60, 90])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), header_color),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 7),
            ('FONTSIZE', (0, 1), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 1),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
            ('BACKGROUND', (0, 1), (-1, -2), colors.HexColor('#F0F0F0')),
            ('BACKGROUND', (0, -1), (-1, -1), accent_color),
            ('TEXTCOLOR', (0, -1), (-1, -1), colors.whitesmoke),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)
        ]))
        tables_content.append(summary_table)
        
        # Place chart and tables side by side (2 columns)
        side_by_side = [[chart_content, tables_content]]
        layout_table = Table(side_by_side, colWidths=[3.8*inch, 2.8*inch])
        layout_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
            ('RIGHTPADDING', (0, 0), (-1, -1), 10)
        ]))
        group_content.append(layout_table)
        group_content.append(Spacer(1, 4))
        
        # Price Classes Comparison (full width, compact)
        if len(all_years) >= 2:
            current_year = all_years[-1]
            prior_year = all_years[-2]
            
            ytd_current_year = group_data[group_data["invoice_date"].dt.year == current_year].copy()
            ytd_prior_year = group_data[group_data["invoice_date"].dt.year == prior_year].copy()
            
            if "price_class_desc" in group_data.columns:
                price_title_style = styles['Heading3'].clone('PriceTitle')
                price_title_style.fontSize = 9
                group_content.append(Paragraph(f"<b>Top 5 Price Classes Comparison</b>", price_title_style))
                group_content.append(Spacer(1, 2))
                
                # Change header based on grouping type
                if group_field == "salesperson_desc":
                    combined_data = [["Price Class", f"{current_year}", "Growth %", "Price Class", f"{prior_year}", ""]]
                else:
                    combined_data = [["Price Class", f"{current_year}", "ROI", "Price Class", f"{prior_year}", "ROI"]]
                
                # Calculate ROI or Growth % lookup for current and prior years
                current_metric_lookup = {}
                prior_metric_lookup = {}
                
                if group_field == "salesperson_desc":
                    # For Salesperson: Calculate Growth % based on average monthly sales
                    # Current year
                    if not ytd_current_year.empty and "price_class_desc" in ytd_current_year.columns:
                        for pc_desc in ytd_current_year["price_class_desc"].dropna().unique():
                            if str(pc_desc).strip():
                                pc_current_data = ytd_current_year[ytd_current_year["price_class_desc"] == pc_desc].copy()
                                pc_current_data["month"] = pc_current_data["invoice_date"].dt.to_period("M")
                                monthly_sales = pc_current_data.groupby("month")["extended_price_usd"].sum()
                                if len(monthly_sales) > 0:
                                    current_metric_lookup[pc_desc] = monthly_sales.mean()
                    
                    # Prior year  
                    if not ytd_prior_year.empty and "price_class_desc" in ytd_prior_year.columns:
                        for pc_desc in ytd_prior_year["price_class_desc"].dropna().unique():
                            if str(pc_desc).strip():
                                pc_prior_data = ytd_prior_year[ytd_prior_year["price_class_desc"] == pc_desc].copy()
                                pc_prior_data["month"] = pc_prior_data["invoice_date"].dt.to_period("M")
                                monthly_sales = pc_prior_data.groupby("month")["extended_price_usd"].sum()
                                if len(monthly_sales) > 0:
                                    prior_metric_lookup[pc_desc] = monthly_sales.mean()
                
                elif not inventory_costs.empty and "gross_profit_usd" in group_data.columns:
                    # Current year ROI by price class
                    if not ytd_current_year.empty and "inventory_flag" in ytd_current_year.columns:
                        current_stocking = ytd_current_year[ytd_current_year["inventory_flag"] == "Y"].copy()
                        if not current_stocking.empty and "price_class_desc" in current_stocking.columns:
                            # Filter out empty/null price class descriptions
                            current_stocking = current_stocking[
                                current_stocking["price_class_desc"].notna() & 
                                (current_stocking["price_class_desc"].astype(str).str.strip() != "")
                            ].copy()
                            
                            if not current_stocking.empty:
                                current_gp_by_class = current_stocking.groupby("price_class_desc")["gross_profit_usd"].sum()
                                
                                for pc_desc in current_gp_by_class.index:
                                    # Get inventory cost for this price class - match on both desc and code
                                    pc_inv_cost = inventory_costs[
                                        (inventory_costs["price_class_desc"].astype(str).str.strip() == str(pc_desc).strip()) &
                                        (inventory_costs["price_class_desc"].notna())
                                    ]["total_cost"].sum()
                                    if pc_inv_cost > 0:
                                        current_metric_lookup[pc_desc] = current_gp_by_class[pc_desc] / pc_inv_cost
                    
                    # Prior year ROI by price class
                    if not ytd_prior_year.empty and "inventory_flag" in ytd_prior_year.columns:
                        prior_stocking = ytd_prior_year[ytd_prior_year["inventory_flag"] == "Y"].copy()
                        if not prior_stocking.empty and "price_class_desc" in prior_stocking.columns:
                            # Filter out empty/null price class descriptions
                            prior_stocking = prior_stocking[
                                prior_stocking["price_class_desc"].notna() & 
                                (prior_stocking["price_class_desc"].astype(str).str.strip() != "")
                            ].copy()
                            
                            if not prior_stocking.empty:
                                prior_gp_by_class = prior_stocking.groupby("price_class_desc")["gross_profit_usd"].sum()
                                
                                for pc_desc in prior_gp_by_class.index:
                                    pc_inv_cost = inventory_costs[
                                        (inventory_costs["price_class_desc"].astype(str).str.strip() == str(pc_desc).strip()) &
                                        (inventory_costs["price_class_desc"].notna())
                                    ]["total_cost"].sum()
                                    if pc_inv_cost > 0:
                                        prior_metric_lookup[pc_desc] = prior_gp_by_class[pc_desc] / pc_inv_cost
                
                if not ytd_current_year.empty:
                    ytd_current_top10 = ytd_current_year.groupby("price_class_desc")["extended_price_usd"].sum().reset_index()
                    ytd_current_top10 = ytd_current_top10.sort_values("extended_price_usd", ascending=False).head(5)
                else:
                    ytd_current_top10 = pd.DataFrame(columns=["price_class_desc", "extended_price_usd"])
                
                if not ytd_prior_year.empty:
                    ytd_prior_top10 = ytd_prior_year.groupby("price_class_desc")["extended_price_usd"].sum().reset_index()
                    ytd_prior_top10 = ytd_prior_top10.sort_values("extended_price_usd", ascending=False).head(5)
                else:
                    ytd_prior_top10 = pd.DataFrame(columns=["price_class_desc", "extended_price_usd"])
                
                max_rows = max(len(ytd_current_top10), len(ytd_prior_top10))
                for i in range(max_rows):
                    current_class = ytd_current_top10.iloc[i]["price_class_desc"] if i < len(ytd_current_top10) else ""
                    current_sales = f"${ytd_current_top10.iloc[i]['extended_price_usd']:,.0f}" if i < len(ytd_current_top10) else ""
                    
                    # Calculate metric based on grouping type
                    if group_field == "salesperson_desc":
                        # Growth %
                        current_avg = current_metric_lookup.get(current_class, 0)
                        prior_avg = prior_metric_lookup.get(current_class, 0)
                        if prior_avg > 0:
                            current_metric = f"{(current_avg / prior_avg * 100):.1f}%"
                        else:
                            current_metric = ""
                    else:
                        # ROI
                        current_metric_val = current_metric_lookup.get(current_class, 0)
                        current_metric = f"{current_metric_val:.2f}x" if current_metric_val > 0 else ""
                    
                    prior_class = ytd_prior_top10.iloc[i]["price_class_desc"] if i < len(ytd_prior_top10) else ""
                    prior_sales = f"${ytd_prior_top10.iloc[i]['extended_price_usd']:,.0f}" if i < len(ytd_prior_top10) else ""
                    
                    if group_field == "salesperson_desc":
                        # For salesperson, don't show growth % for prior year (always same as current)
                        prior_metric = ""
                    else:
                        # ROI
                        prior_metric_val = prior_metric_lookup.get(prior_class, 0)
                        prior_metric = f"{prior_metric_val:.2f}x" if prior_metric_val > 0 else ""
                    
                    combined_data.append([current_class, current_sales, current_metric, prior_class, prior_sales, prior_metric])
                
                current_top10_total = ytd_current_top10["extended_price_usd"].sum() if not ytd_current_top10.empty else 0
                current_grand_total = ytd_current_year["extended_price_usd"].sum() if not ytd_current_year.empty else 0
                prior_top10_total = ytd_prior_top10["extended_price_usd"].sum() if not ytd_prior_top10.empty else 0
                prior_grand_total = ytd_prior_year["extended_price_usd"].sum() if not ytd_prior_year.empty else 0
                
                combined_data.append(["Top 5 Total", f"${current_top10_total:,.0f}", "", "Top 5 Total", f"${prior_top10_total:,.0f}", ""])
                combined_data.append(["Grand Total", f"${current_grand_total:,.0f}", "", "Grand Total", f"${prior_grand_total:,.0f}", ""])
                
                price_table = Table(combined_data, colWidths=[120, 60, 40, 120, 60, 40])
                price_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), header_color),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                    ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
                    ('ALIGN', (2, 0), (2, -1), 'RIGHT'),
                    ('ALIGN', (3, 0), (3, -1), 'LEFT'),
                    ('ALIGN', (4, 0), (4, -1), 'RIGHT'),
                    ('ALIGN', (5, 0), (5, -1), 'RIGHT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 8),
                    ('FONTSIZE', (0, 1), (-1, -1), 6),
                    ('TOPPADDING', (0, 0), (-1, -1), 1),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
                    ('BACKGROUND', (0, 1), (2, -3), colors.HexColor('#E3F2FD')),
                    ('BACKGROUND', (3, 1), (5, -3), colors.HexColor('#FFF9E1')),
                    ('BACKGROUND', (0, -2), (2, -1), colors.HexColor('#B3E5FC')),
                    ('BACKGROUND', (3, -2), (5, -1), colors.HexColor('#FFF176')),
                    ('FONTNAME', (0, -2), (-1, -1), 'Helvetica-Bold'),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('LINEAFTER', (2, 0), (2, -1), 1.5, header_color)
                ]))
                
                # Place price class table and pie chart side by side for Salesperson
                if group_field == "salesperson_desc" and 'pie_content' in locals() and pie_content:
                    price_with_pie = [[price_table, pie_content[0]]]
                    price_layout = Table(price_with_pie, colWidths=[4.4*inch, 2.0*inch])
                    price_layout.setStyle(TableStyle([
                        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                        ('LEFTPADDING', (0, 0), (-1, -1), 0),
                        ('RIGHTPADDING', (0, 0), (-1, -1), 5)
                    ]))
                    group_content.append(price_layout)
                else:
                    group_content.append(price_table)
        
        # Add account growth analysis for Salesperson grouping
        if group_field == "salesperson_desc" and len(all_years) >= 2:
            group_content.append(Spacer(1, 3))
            
            # Add Item Class 1 pie chart
            if "item_class_1_desc" in group_data.columns:
                try:
                    # Calculate sales by Item Class 1
                    class1_sales = group_data.groupby("item_class_1_desc")["extended_price_usd"].sum().reset_index()
                    class1_sales = class1_sales.sort_values("extended_price_usd", ascending=False)
                    
                    # Take top 6 and group rest as "Other"
                    if len(class1_sales) > 6:
                        top6 = class1_sales.head(6)
                        other_sum = class1_sales.iloc[6:]["extended_price_usd"].sum()
                        if other_sum > 0:
                            other_row = pd.DataFrame([{"item_class_1_desc": "Other", "extended_price_usd": other_sum}])
                            class1_sales = pd.concat([top6, other_row], ignore_index=True)
                        else:
                            class1_sales = top6
                    
                    if not class1_sales.empty:
                        # Professional color palette
                        colors_palette = ['#3b82f6', '#10b981', '#f59e0b', '#ef4444', '#8b5cf6', '#ec4899', '#6366f1']
                        
                        # Create pie chart
                        fig_pie = go.Figure(data=[go.Pie(
                            labels=class1_sales["item_class_1_desc"],
                            values=class1_sales["extended_price_usd"],
                            hole=0.4,
                            textinfo='percent',
                            textposition='inside',
                            textfont=dict(size=9, color='white'),
                            marker=dict(
                                colors=colors_palette[:len(class1_sales)],
                                line=dict(color='white', width=2)
                            ),
                            hovertemplate='<b>%{label}</b><br>$%{value:,.0f}<br>%{percent}<extra></extra>'
                        )])
                        
                        fig_pie.update_layout(
                            title=dict(text="Sales by Category", font=dict(size=10)),
                            height=180,
                            width=260,
                            showlegend=True,
                            legend=dict(
                                orientation="v",
                                yanchor="middle",
                                y=0.5,
                                xanchor="left",
                                x=1.0,
                                font=dict(size=7)
                            ),
                            margin=dict(l=5, r=80, t=25, b=5)
                        )
                        
                        # Convert to image with higher resolution
                        img_bytes_pie = fig_pie.to_image(format="png", width=260, height=180, engine="kaleido", scale=2)
                        img_pie = RLImage(BytesIO(img_bytes_pie), width=2.0*inch, height=1.4*inch)
                        pie_content = [img_pie]
                    else:
                        pie_content = []
                except:
                    pie_content = []
            else:
                pie_content = []
            
            account_title_style = styles['Heading3'].clone('AccountTitle')
            account_title_style.textColor = accent_color
            account_title_style.fontSize = 9
            group_content.append(Paragraph("Account Performance", account_title_style))
            group_content.append(Spacer(1, 2))
            
            current_year = all_years[-1]
            prior_year = all_years[-2]
            
            # Create account identifier (bank_name - account_number)
            if "bank_name" in group_data.columns and "account_number" in group_data.columns:
                group_data_copy = group_data.copy()
                group_data_copy["account_id"] = group_data_copy["bank_name"].astype(str) + " - " + group_data_copy["account_number"].astype(str)
                
                # Calculate average monthly sales for each account by year
                account_growth = []
                for account_id in group_data_copy["account_id"].unique():
                    account_data = group_data_copy[group_data_copy["account_id"] == account_id]
                    
                    # Current year
                    current_data = account_data[account_data["invoice_date"].dt.year == current_year].copy()
                    if not current_data.empty:
                        current_data["month"] = current_data["invoice_date"].dt.to_period("M")
                        current_monthly = current_data.groupby("month")["extended_price_usd"].sum()
                        current_avg = current_monthly.mean() if len(current_monthly) > 0 else 0
                    else:
                        current_avg = 0
                    
                    # Prior year
                    prior_data = account_data[account_data["invoice_date"].dt.year == prior_year].copy()
                    if not prior_data.empty:
                        prior_data["month"] = prior_data["invoice_date"].dt.to_period("M")
                        prior_monthly = prior_data.groupby("month")["extended_price_usd"].sum()
                        prior_avg = prior_monthly.mean() if len(prior_monthly) > 0 else 0
                    else:
                        prior_avg = 0
                    
                    # Calculate growth %
                    if prior_avg > 0:
                        growth_pct = (current_avg / prior_avg) * 100
                        account_growth.append({
                            "Account": account_id,
                            "prior_avg": prior_avg,
                            "current_avg": current_avg,
                            "growth_pct": growth_pct
                        })
                    elif current_avg > 0:
                        # New account with no prior year data
                        account_growth.append({
                            "Account": account_id,
                            "prior_avg": 0,
                            "current_avg": current_avg,
                            "growth_pct": 999.9  # Mark as new
                        })
                
                if account_growth:
                    account_df = pd.DataFrame(account_growth)
                    account_df = account_df.sort_values("growth_pct", ascending=False)
                    
                    # Get top 5 and bottom 5
                    top5 = account_df.head(5)
                    bottom5 = account_df.tail(5)
                    
                    # Combine for chart
                    combined = pd.concat([top5, bottom5])
                    combined = combined.sort_values("growth_pct", ascending=True)
                    
                    # Create horizontal bar chart
                    try:
                        # Shorten account names for better display
                        combined['short_account'] = combined['Account'].apply(lambda x: x[:25] + '...' if len(str(x)) > 25 else str(x))
                        combined['growth_display'] = combined['growth_pct'].apply(lambda x: min(max(x, -100), 300))  # Cap for display
                        
                        fig_acct = go.Figure()
                        
                        # Color code: green for positive, red for negative
                        colors_bar = ['#10b981' if x >= 0 else '#ef4444' for x in combined['growth_display']]
                        
                        fig_acct.add_trace(go.Bar(
                            y=combined['short_account'],
                            x=combined['growth_display'],
                            orientation='h',
                            marker=dict(color=colors_bar),
                            text=combined['growth_pct'].apply(lambda x: 'NEW' if x >= 999 else f'{x:.0f}%'),
                            textposition='outside',
                            textfont=dict(size=7),
                            hovertemplate='<b>%{y}</b><br>Growth: %{text}<extra></extra>'
                        ))
                        
                        fig_acct.update_layout(
                            title=dict(text="Top 5 & Bottom 5 Accounts by Growth", font=dict(size=11)),
                            xaxis_title="Growth %",
                            xaxis=dict(tickfont=dict(size=8)),
                            yaxis=dict(tickfont=dict(size=8)),
                            height=150,
                            width=530,
                            showlegend=False,
                            margin=dict(l=115, r=35, t=30, b=30)
                        )
                        
                        # Convert to image with higher resolution
                        img_bytes_acct = fig_acct.to_image(format="png", width=530, height=150, engine="kaleido", scale=2)
                        img_acct = RLImage(BytesIO(img_bytes_acct), width=5.3*inch, height=1.5*inch)
                        group_content.append(img_acct)
                    except:
                        # Fallback to compact table if chart fails
                        account_table_data = [["Account", f"{prior_year}", f"{current_year}", "Growth %"]]
                        
                        for _, row in combined.iterrows():
                            account_table_data.append([
                                row["Account"][:30],
                                f"${row['prior_avg']:,.0f}",
                                f"${row['current_avg']:,.0f}",
                                "NEW" if row["growth_pct"] >= 999 else f"{row['growth_pct']:.0f}%"
                            ])
                        
                        account_table = RLTable(account_table_data, colWidths=[180, 60, 60, 50])
                        account_table.setStyle(TableStyle([
                            ('BACKGROUND', (0, 0), (-1, 0), header_color),
                            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                            ('FONTSIZE', (0, 0), (-1, 0), 7),
                            ('FONTSIZE', (0, 1), (-1, -1), 6),
                            ('TOPPADDING', (0, 0), (-1, -1), 1),
                            ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
                            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)
                        ]))
                        group_content.append(account_table)
        
        # Wrap group content to keep together on one page
        story.append(KeepTogether(group_content))
        
        # Page break after each group
        if idx + 1 < max_groups:
            story.append(PageBreak())
    
    # Build PDF
    doc.build(story)
    buffer.seek(0)
    return buffer

# Minimal run() to satisfy external callers; build UI when invoked
def run() -> None:
    # Page config is set in the main script (streamlit_app.py) at import time
    # to ensure correct initialization ordering. Just set the title and build.
    st.title("Inventory Dashboard")
    _build()

# Main UI builder (wrapped to fix prior stray indentation issues)
def _build() -> None:
    tabs = st.tabs(["Overview", "Stock Turn", "Load Times", "Launch Dates", "Price List Generator", "Open Orders", "Drops", "Supplier Performance", "Cost Center ROI", "Sales Rep Performance", "Executive Intelligence", "Returns", "CCA"])

    # Hide developer-facing debug controls on Overview by default
    SHOW_OVERVIEW_DEBUG = False


    user_prefs = _load_prefs()

    # Helper: queue details request to run after data is loaded and dialog is defined
    def _queue_details(entity_type: str, entity_key: str) -> None:
        st.session_state["_details_request"] = (entity_type, entity_key)

    with st.sidebar:
        st.header("Filters")
        # Filters form groups changes and applies them on submit only
        with st.form("filters_form", clear_on_submit=False):
            cost_center_options = loaders.load_cost_centers(config.connection_string)
            pref_cc = user_prefs.get("cost_centers", config.default_cost_centers)
            form_cost_centers = st.multiselect(
                "Cost centers",
                options=cost_center_options,
                default=_ensure_selection(cost_center_options, pref_cc),
                help="All metrics will be scoped to the selected cost centers.",
            )

            # Supplier filter (distinct from items filtered by cost centers)
            supplier_options = loaders.load_suppliers(config.connection_string, form_cost_centers)
            pref_suppliers = user_prefs.get("suppliers", [])
            form_suppliers = st.multiselect(
                "Suppliers",
                options=supplier_options,
                default=[],
                help="Filter all metrics and tables to items from these suppliers.",
            )

            # Price classes filter (dynamic by CC + Supplier); use codes as option values
            pc_df = loaders.load_price_classes(config.connection_string)
            items_for_pc = loaders.load_items(config.connection_string, form_cost_centers)
            pc_map: dict[str, str] = {}
            if isinstance(pc_df, pd.DataFrame) and not pc_df.empty and {"price_class", "price_class_desc"}.issubset(set(pc_df.columns)):
                pc_map = {str(c): str(d).strip() for c, d in zip(pc_df["price_class"], pc_df["price_class_desc"])}
            codes_options: list[str] = []
            if isinstance(items_for_pc, pd.DataFrame) and not items_for_pc.empty:
                items_sc = items_for_pc
                if form_suppliers:
                    items_sc = items_sc[items_sc.get("supplier_number").astype(str).isin([str(s) for s in form_suppliers])]
                codes = items_sc.get("price_class").dropna().astype(str)
                codes = codes[codes.str.len() > 0]
                codes_options = sorted(codes.unique().tolist())
            pref_pclasses = user_prefs.get("price_class_codes", [])
            form_price_classes = st.multiselect(
                "Price classes",
                options=codes_options,
                default=[],
                format_func=lambda code: pc_map.get(str(code), str(code)),
                help="Filter all metrics and tables to items in these price classes.",
            )

            # Fixed order date range per requirements: 2025-08-04 through today
            range_start_fixed = date(2025, 8, 4)
            range_end_fixed = date.today()
            st.caption(f"Order date range: {range_start_fixed} to {range_end_fixed} (fixed)")

            pref_ratings = user_prefs.get("sku_ratings", list("ABCD"))
            form_ratings = st.multiselect(
                "SKU rating",
                options=list("ABCD"),
                default=_ensure_selection(list("ABCD"), pref_ratings),
                help="Filter the grids and alerts by the rating tiers (A=top performers).",
            )

            form_sku_search = st.text_input(
                "Search SKU",
                value=user_prefs.get("sku_search", ""),
                help="Filter the SKU table by SKU substring.",
            )

            submitted_filters = st.form_submit_button("Apply filters")

        # After form: handle details buttons which should be outside the form
        # Use last applied suppliers for buttons
        applied = st.session_state.get("_applied_filters")
        if applied is None and submitted_filters is False:
            # initialize applied with current form defaults
            # fixed range per requirements
            start_date, end_date = date(2025, 8, 4), date.today()
            applied = {
                "cost_centers": list(form_cost_centers),
                "suppliers": list(form_suppliers),
                "price_class_codes": list(form_price_classes),
                "date_start": start_date,
                "date_end": end_date,
                "sku_ratings": list(form_ratings),
                "sku_search": form_sku_search,
            }
            st.session_state["_applied_filters"] = applied

        if submitted_filters:
            # Update applied filters and persist prefs
            start_date, end_date = date(2025, 8, 4), date.today()
            applied = {
                "cost_centers": list(form_cost_centers),
                "suppliers": list(form_suppliers),
                "price_class_codes": list(form_price_classes),
                "date_start": start_date,
                "date_end": end_date,
                "sku_ratings": list(form_ratings),
                "sku_search": form_sku_search,
            }
            st.session_state["_applied_filters"] = applied

            # Persist current preferences (exclude date range intentionally?) Keeping consistent with earlier behavior
            current_prefs = {
                "cost_centers": applied["cost_centers"],
                "suppliers": applied["suppliers"],
                "price_class_codes": applied["price_class_codes"],
                "sku_ratings": applied["sku_ratings"],
                "sku_search": applied["sku_search"],
                "defer_heavy": st.session_state.get("defer_heavy", True),
                "run_warmup": st.session_state.get("run_warmup", False),
            }
            _save_prefs(current_prefs)

        # Extract applied values for use below
        applied = st.session_state.get("_applied_filters", {})
        selected_cost_centers = applied.get("cost_centers", [])
        selected_suppliers = applied.get("suppliers", [])
        selected_price_classes = applied.get("price_class_codes", [])
        start_date = applied.get("date_start", date(2025, 8, 4))
        end_date = applied.get("date_end", date.today())
        rating_filter = applied.get("sku_ratings", list("ABCD"))
        sku_search = applied.get("sku_search", "")

        # Details buttons based on applied suppliers/price classes/SKU search
        if selected_suppliers:
            for sup in selected_suppliers[:3]:  # limit to avoid too many controls
                st.button(
                    f"Open Supplier details: {sup}",
                    key=f"supplier_btn_{sup}",
                    on_click=_queue_details,
                    args=("supplier", sup),
                )
        if selected_price_classes:
            # Need price class descriptions for labels
            pc_df = loaders.load_price_classes(config.connection_string)
            pc_map: dict[str, str] = {}
            if isinstance(pc_df, pd.DataFrame) and not pc_df.empty and {"price_class", "price_class_desc"}.issubset(set(pc_df.columns)):
                pc_map = {str(c): str(d).strip() for c, d in zip(pc_df["price_class"], pc_df["price_class_desc"])}
            for pcc in selected_price_classes[:3]:
                label = pc_map.get(str(pcc), str(pcc))
                st.button(
                    f"Open Price Class details: {label}",
                    key=f"pc_btn_{pcc}",
                    on_click=_queue_details,
                    args=("price_class", str(pcc)),
                )
        if sku_search.strip():
            if st.button("Open SKU details"):
                _queue_details("sku", sku_search.strip())

        st.divider()
        st.caption("Performance")
        defer_heavy = st.checkbox(
            "Defer heavy UI builds (faster load)", value=bool(user_prefs.get("defer_heavy", True)),
            help="Skip building large tables/reports until requested. Uncheck to always build."
        )
        run_warmup = st.checkbox(
            "Run warm-up across cost centers (cache 1h)", value=bool(user_prefs.get("run_warmup", False)),
            help="Pre-loads all cost centers into the local caches. Can be slow the first time."
        )
        # Keep in session for access outside sidebar scope
        st.session_state["defer_heavy"] = defer_heavy
        st.session_state["run_warmup"] = run_warmup

        # Seasonality Editor (monthly percentages by cost center)
        with st.expander("Seasonality (monthly % by cost center)"):
            st.caption("Enter monthly percentages for each selected cost center. Values are normalized to sum to 100%.")
            existing = load_seasonality()
            if not selected_cost_centers:
                st.info("Select at least one cost center to edit seasonality.")
            else:
                for cc in selected_cost_centers:
                    st.markdown(f"**Cost Center {cc}**")
                    months = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
                    defaults = existing.get(str(cc), [1.0/12.0]*12)
                    # Render inputs in a 4x3 grid
                    inputs: list[float] = []
                    cols = st.columns(4)
                    for i, m in enumerate(months):
                        col = cols[i % 4]
                        val = col.number_input(
                            f"{m} % ({cc})",
                            min_value=0.0,
                            max_value=100.0,
                            value=float(defaults[i] * 100.0),
                            step=1.0,
                            key=f"season_{cc}_{i}",
                        )
                        inputs.append(val)
                    if st.button(f"Save seasonality for {cc}"):
                        total = sum(inputs)
                        if total <= 0:
                            st.error("Please enter positive percentages.")
                        else:
                            normalized = [v/total for v in inputs]
                            ok = save_for_cost_center(str(cc), normalized)
                            if ok:
                                st.success("Saved.")
                            else:
                                st.error("Failed to save. Check permissions.")

        # Stock Turn Targets per Cost Center (separate expander; not nested)
        with st.expander("Stock Turn Targets (per cost center)"):
            st.caption("Set a target stock turn for each cost center. If not set, the app will use the current average.")
            targets = load_targets()
            cc_for_targets = selected_cost_centers if selected_cost_centers else loaders.load_cost_centers(config.connection_string)
            for cc in cc_for_targets:
                current = targets.get(str(cc))
                val = st.number_input(
                    f"Target for {cc}",
                    min_value=0.0,
                    max_value=1000.0,
                    value=float(current) if current is not None else float(config.stockturn_target),
                    step=0.1,
                    key=f"turn_target_{cc}",
                )
                if st.button(f"Save target for {cc}"):
                    if save_target(str(cc), float(val)):
                        st.success("Saved.")
                    else:
                        st.error("Failed to save target.")

        # Persist Performance prefs (outside form) when changed
        perf_prefs = {
            "defer_heavy": defer_heavy,
            "run_warmup": run_warmup,
        }
        last_perf = st.session_state.get("_last_perf_prefs")
        if last_perf != perf_prefs:
            base = _load_prefs()
            base.update(perf_prefs)
            _save_prefs(base)
            st.session_state["_last_perf_prefs"] = perf_prefs

    @st.cache_data(ttl=config.cache_ttl_seconds, show_spinner=False)
    def _load_dashboard(cost_centers_tuple: tuple[str, ...], suppliers_tuple: tuple[str, ...], price_classes_tuple: tuple[str, ...], start: date, end: date):
        filters = MetricsFilters(cost_centers=list(cost_centers_tuple), start_date=start, end_date=end)
        # Piggy-back suppliers into the filters object via attribute (backward compatible)
        setattr(filters, "suppliers", list(suppliers_tuple))
        # Piggy-back price class codes (IPRCCD)
        setattr(filters, "price_class_codes", list(price_classes_tuple))
        return compute_dashboard_data(filters)

    # Measure compute + additional UI steps for accurate wall-clock totals
    ui_timings: dict[str, float] = {}
    t_compute = perf_counter()
    with st.spinner("Crunching metrics..."):
        dashboard = _load_dashboard(tuple(selected_cost_centers), tuple(selected_suppliers), tuple(selected_price_classes), start_date, end_date)
    ui_timings["compute_dashboard_call"] = perf_counter() - t_compute

    summary = dashboard["summary"]
    sku_metrics = dashboard["sku_metrics"].copy()
    # Keep an unfiltered copy for details dialog and reports that need the full set
    sku_metrics_full = dashboard["sku_metrics"].copy()
    items_df = dashboard.get("items", pd.DataFrame()).copy()
    # Build a set of base SKUs considered "stocking" (IINVEN='Y') using IIXREF when present
    stocking_base_sku_set: set[str] = set()
    try:
        if isinstance(items_df, pd.DataFrame) and not items_df.empty and "sku" in items_df.columns:
            base_series = items_df.get("sku").astype(str)
            if "iixref" in items_df.columns:
                iix = items_df.get("iixref").fillna("").astype(str).str.strip()
                base_series = base_series.where(iix.eq(""), iix)
            stocking_base_sku_set = set(base_series.astype(str).str.strip().str.upper().tolist())
    except Exception:
        stocking_base_sku_set = set()
    sales_orders = dashboard.get("sales_orders", pd.DataFrame()).copy()
    timings = dashboard.get("timings", {})

    # One-time warm-up: snapshot all cost centers (excluding those starting with '1') into history and backorders
    # Run best-effort and cache per session to avoid repeated heavy work
    @st.cache_data(ttl=60*60, show_spinner=False)
    def _warm_all_centers(start: date, end: date):
        try:
            cc_all = [cc for cc in loaders.load_cost_centers(get_config().connection_string) if not str(cc).startswith("1")]
            for cc in cc_all:
                try:
                    _ = _load_dashboard((cc,), tuple(), tuple(), start, end)
                except Exception:
                    continue
        except Exception:
            return False
        return True

    # Warm-up (may run once per hour via cache) – only if enabled in sidebar
    if st.session_state.get("run_warmup", False):
        t_warm = perf_counter()
        with st.spinner("Warming up cached snapshots across cost centers…"):
            _ = _warm_all_centers(start_date, end_date)
        ui_timings["ui_warm_all_centers"] = perf_counter() - t_warm

    # Trends charts removed per request; skip loading local time-series history and series computation

    # Overall caption (compute + other UI work). backend_total is informative only.
    backend_total = float(timings.get("total", 0.0)) if isinstance(timings, dict) else 0.0
    compute_call = float(ui_timings.get("compute_dashboard_call", 0.0))
    other_ui = float(sum(v for k, v in ui_timings.items() if k != "compute_dashboard_call"))
    overall = compute_call + other_ui
    if overall > 0:
        st.caption(f"Loaded in {overall:.2f}s (compute: {compute_call:.2f}s, UI: {other_ui:.2f}s, backend-only: {backend_total:.2f}s)")

    # (Load Times tab moved to end so it can include UI build timings below)

    # -------------------------
    # Details dialog for SKU / Supplier
    # -------------------------
    def _entity_subset(entity_type: str, entity_key: str):
        # Prefer pre-sliced datasets for editor context, but fall back to full datasets if the requested SKU isn’t present
        perf_so = st.session_state.get("_perf_sales_orders")
        perf_po = st.session_state.get("_perf_po_orders")
        perf_items = st.session_state.get("_perf_items_df")
        use_so = perf_so if isinstance(perf_so, pd.DataFrame) else sales_orders
        use_po = perf_po if isinstance(perf_po, pd.DataFrame) else dashboard.get("purchase_orders", pd.DataFrame())
        use_items = perf_items if isinstance(perf_items, pd.DataFrame) else items_df
        # Normalization helpers
        def norm_series(s: pd.Series) -> pd.Series:
            return s.astype(str).str.strip().str.upper()

        if entity_type == "sku":
            sku_list = [entity_key]
        elif entity_type == "supplier":
            sku_list = use_items.loc[norm_series(use_items.get("supplier_number", pd.Series(dtype=str))) == str(entity_key).strip().upper(), "sku"].dropna().unique().tolist()
        elif entity_type == "price_class":
            # Match by price_class code
            code_norm = str(entity_key).strip()
            pc_series = use_items.get("price_class")
            if pc_series is not None:
                mask = norm_series(pc_series) == code_norm.strip().upper()
                sku_list = use_items.loc[mask, "sku"].dropna().unique().tolist()
            else:
                sku_list = []
        else:
            sku_list = []
        # Normalize SKU set
        sku_set = {str(s) for s in sku_list}
        sku_set_norm = {str(s).strip().upper() for s in sku_set}
        # If pre-sliced datasets don’t contain these SKUs, fall back to the full datasets
        def _contains_sku(df: pd.DataFrame) -> bool:
            try:
                if isinstance(df, pd.DataFrame) and not df.empty and "sku" in df.columns and sku_set_norm:
                    present = df.get("sku").astype(str).str.strip().str.upper().isin(sku_set_norm)
                    return bool(present.any())
            except Exception:
                pass
            return False
        if isinstance(perf_so, pd.DataFrame) and not _contains_sku(perf_so):
            use_so = sales_orders
        if isinstance(perf_po, pd.DataFrame) and not _contains_sku(perf_po):
            use_po = dashboard.get("purchase_orders", pd.DataFrame())
        if isinstance(perf_items, pd.DataFrame):
            try:
                if not _contains_sku(perf_items):
                    use_items = items_df
            except Exception:
                use_items = items_df
        # Match metrics by normalized SKU
        if sku_set:
            met_mask = sku_metrics_full.get("sku", pd.Series(dtype=str)).astype(str).str.strip().str.upper().isin(sku_set_norm)
            met = sku_metrics_full.loc[met_mask].copy()
        else:
            met = pd.DataFrame(columns=sku_metrics_full.columns)
        if sku_set and isinstance(use_so, pd.DataFrame) and not use_so.empty:
            so_mask = norm_series(use_so.get("sku", pd.Series(dtype=str))).isin(sku_set_norm)
            so = use_so.loc[so_mask].copy()
        else:
            so = pd.DataFrame(columns=use_so.columns if isinstance(use_so, pd.DataFrame) and not use_so.empty else [])
        po = use_po
        if sku_set and isinstance(po, pd.DataFrame) and not po.empty:
            po_mask = norm_series(po.get("sku", pd.Series(dtype=str))).isin(sku_set_norm)
            po = po.loc[po_mask].copy()
        else:
            po = pd.DataFrame(columns=po.columns if isinstance(po, pd.DataFrame) and not po.empty else [])
        return met, so, po

    def _show_details_body(entity_type: str, entity_key: str):
        met, so, po = _entity_subset(entity_type, entity_key)
        # Special SKU details view matching requested layout
        if entity_type == "sku":
            sku_key = str(entity_key).strip()
            # Top metric: Total inventory (SY) for this SKU
            total_inventory = float(pd.to_numeric(met.get("inventory_sy"), errors="coerce").fillna(0).sum()) if not met.empty else 0.0
            st.metric("Total Inventory (SY)", f"{total_inventory:.2f}")

            # Combined chart: SKU qty over time vs its Price Class qty over time (weekly)
            # Determine price class for this SKU (prefer pre-sliced items)
            pc_code = None
            try:
                # Prefer editor-sliced items; fallback to items_df; finally sku_metrics_full
                source_items = st.session_state.get("_perf_items_df")
                if not isinstance(source_items, pd.DataFrame) or source_items.empty:
                    source_items = items_df
                if isinstance(source_items, pd.DataFrame) and not source_items.empty and {"sku", "price_class"}.issubset(set(source_items.columns)):
                    row_pc = source_items.loc[source_items["sku"].astype(str).str.strip().str.upper() == sku_key.strip().upper(), "price_class"]
                    if not row_pc.empty:
                        pc_code = str(row_pc.iloc[0])
                if pc_code is None and isinstance(sku_metrics_full, pd.DataFrame) and not sku_metrics_full.empty and {"sku","price_class"}.issubset(set(sku_metrics_full.columns)):
                    row_pc2 = sku_metrics_full.loc[sku_metrics_full["sku"].astype(str).str.strip().str.upper() == sku_key.strip().upper(), "price_class"]
                    if not row_pc2.empty:
                        pc_code = str(row_pc2.iloc[0])
            except Exception:
                pc_code = None

            # Weekly series for the clicked SKU
            so_sku = so.copy()
            if not so_sku.empty and "order_entry_date" in so_sku.columns:
                so_sku["order_entry_date"] = pd.to_datetime(so_sku.get("order_entry_date"), errors="coerce")
                so_sku = so_sku.dropna(subset=["order_entry_date"]).set_index("order_entry_date").sort_index()
                weekly_sku = so_sku.resample("W")["quantity_sy"].sum().reset_index()
            else:
                weekly_sku = pd.DataFrame(columns=["order_entry_date", "quantity_sy"])

            # Weekly aggregate for the price class (use ALL sales for the price class,
            # not just the pre-sliced visible SKUs subset)
            weekly_pc = pd.DataFrame(columns=["order_entry_date", "quantity_sy"])
            if pc_code is not None:
                try:
                    # Primary: use items in current context; Fallback: derive from sku_metrics_full
                    pc_skus: list[str] = []
                    source_items = items_df
                    if isinstance(source_items, pd.DataFrame) and not source_items.empty and "price_class" in source_items.columns:
                        pc_skus = source_items.loc[
                            source_items.get("price_class").astype(str).str.strip().str.upper() == pc_code.strip().upper(),
                            "sku"
                        ].dropna().astype(str).unique().tolist()
                    if not pc_skus and isinstance(sku_metrics_full, pd.DataFrame) and not sku_metrics_full.empty and {"price_class","sku"}.issubset(set(sku_metrics_full.columns)):
                        pc_skus = sku_metrics_full.loc[
                            sku_metrics_full.get("price_class").astype(str).str.strip().str.upper() == pc_code.strip().upper(),
                            "sku"
                        ].dropna().astype(str).unique().tolist()
                    # Always use the full sales_orders (not pre-sliced editor cache)
                    source_so = sales_orders
                    if pc_skus and isinstance(source_so, pd.DataFrame) and not source_so.empty:
                        norm_pc = {str(s).strip().upper() for s in pc_skus}
                        sku_col = source_so.get("sku")
                        so_pc = source_so[sku_col.astype(str).str.strip().str.upper().isin(norm_pc)].copy() if sku_col is not None else pd.DataFrame(columns=source_so.columns)
                        if not so_pc.empty:
                            so_pc["order_entry_date"] = pd.to_datetime(so_pc.get("order_entry_date"), errors="coerce")
                            so_pc = so_pc.dropna(subset=["order_entry_date"]).set_index("order_entry_date").sort_index()
                            weekly_pc = so_pc.resample("W")["quantity_sy"].sum().reset_index()
                except Exception:
                    weekly_pc = pd.DataFrame(columns=["order_entry_date", "quantity_sy"])

            if (not weekly_sku.empty) or (not weekly_pc.empty):
                from plotly.subplots import make_subplots
                import plotly.graph_objects as go
                fig = make_subplots(specs=[[{"secondary_y": True}]])
                if not weekly_sku.empty:
                    fig.add_trace(
                        go.Scatter(
                            x=weekly_sku["order_entry_date"],
                            y=weekly_sku["quantity_sy"],
                            mode="lines+markers",
                            name=f"SKU {sku_key}"
                        ),
                        secondary_y=False,
                    )
                if not weekly_pc.empty:
                    fig.add_trace(
                        go.Scatter(
                            x=weekly_pc["order_entry_date"],
                            y=weekly_pc["quantity_sy"],
                            mode="lines+markers",
                            name=f"Price Class {pc_code}"
                        ),
                        secondary_y=True,
                    )
                # Titles and axis labels
                fig.update_layout(title=f"Weekly Sales (SY): {sku_key} vs Price Class {pc_code}")
                fig.update_xaxes(title_text="Week")
                # Prefer descriptive axis labels if both present
                if (not weekly_sku.empty) and (not weekly_pc.empty):
                    fig.update_yaxes(title_text="SKU Quantity (SY)", secondary_y=False)
                    fig.update_yaxes(title_text="Price Class Quantity (SY)", secondary_y=True)
                else:
                    # Single-axis fallback uses the primary axis title
                    fig.update_yaxes(title_text="Quantity (SY)", secondary_y=False)
                st.plotly_chart(fig, use_container_width=True)
            else:
                st.info("No sales found for this SKU or its price class.")
                # Optional debug to help diagnose empty results
                with st.expander("Debug (click to expand)"):
                    st.write({
                        "sku": sku_key,
                        "pc_code": pc_code,
                        "met_rows": 0 if met is None else len(met),
                        "so_rows": 0 if so is None else len(so),
                        "po_rows": 0 if po is None else len(po),
                    })
                    try:
                        st.write("Sample sales_orders rows for this SKU (normalized match):")
                        source_so = st.session_state.get("_perf_sales_orders")
                        if not isinstance(source_so, pd.DataFrame) or source_so.empty:
                            source_so = sales_orders
                        if isinstance(source_so, pd.DataFrame) and not source_so.empty:
                            mask = source_so.get("sku").astype(str).str.strip().str.upper() == sku_key.strip().upper()
                            st.dataframe(source_so.loc[mask].head(10), use_container_width=True, hide_index=True)
                    except Exception:
                        pass

            # Backorders table for this SKU: status 'B' only and selected columns
            st.markdown("### Backorders")
            bo = so.copy()
            if not bo.empty and "detail_line_status" in bo.columns:
                detail = bo.get("detail_line_status").astype(str).str.strip().str.upper()
                bo = bo.loc[detail == "B"].copy()
            else:
                bo = pd.DataFrame(columns=["order_number", "quantity_sy", "actual_ship_date"])
            if not bo.empty:
                cols = [c for c in ["order_number", "quantity_sy", "actual_ship_date"] if c in bo.columns]
                bo_small = bo[cols].copy()
                bo_small.rename(columns={"order_number": "Order #", "quantity_sy": "Qty", "actual_ship_date": "Ship Date"}, inplace=True)
                if "Ship Date" in bo_small.columns:
                    bo_small["Ship Date"] = pd.to_datetime(bo_small["Ship Date"], errors="coerce").dt.date
                st.dataframe(bo_small, use_container_width=True, hide_index=True)
            else:
                st.write("No backorders for this SKU.")

            # Purchase orders table for this SKU with requested columns
            st.markdown("### Purchase Orders")
            po_sku = po.copy()
            if not po_sku.empty:
                cols = [c for c in ["order_number", "quantity_sy", "eta_date"] if c in po_sku.columns]
                po_small = po_sku[cols].copy()
                po_small.rename(columns={"order_number": "Order #", "quantity_sy": "Qty", "eta_date": "ETA"}, inplace=True)
                if "ETA" in po_small.columns:
                    po_small["ETA"] = pd.to_datetime(po_small["ETA"], errors="coerce").dt.date
                st.dataframe(po_small, use_container_width=True, hide_index=True)
            else:
                st.write("No purchase orders for this SKU.")
            return

        if met.empty and so.empty and po.empty:
            st.info("No data found for this selection.")
            return

        # KPIs (same cards, scoped)
        total_avg_daily = pd.to_numeric(met.get("avg_daily_sales_sy"), errors="coerce").fillna(0).sum()
        total_inventory = pd.to_numeric(met.get("inventory_sy"), errors="coerce").fillna(0).sum()
        stock_turn_local = (total_avg_daily * 365.0) / total_inventory if total_inventory > 0 else 0.0
        total_orders_local = pd.to_numeric(met.get("orders_count"), errors="coerce").fillna(0).sum()
        total_backorders_local = pd.to_numeric(met.get("backorder_count"), errors="coerce").fillna(0).sum()
        fill_rate_local = (1 - (total_backorders_local / total_orders_local)) if total_orders_local > 0 else 0.0
        days_inv_local = met["days_of_inventory"].replace([np.inf, -np.inf], np.nan).median()
        inv_qty = pd.to_numeric(met.get("inventory_sy"), errors="coerce").fillna(0)
        inv_age = pd.to_numeric(met.get("inventory_age_days"), errors="coerce")
        mask = (inv_qty > 0) & inv_age.notna()
        avg_age_local = float((inv_qty[mask] * inv_age[mask]).sum() / inv_qty[mask].sum()) if mask.any() else 0.0
        label_map = {"sku": "SKU", "supplier": "Supplier", "price_class": "Price Class"}
        label = label_map.get(entity_type, str(entity_type).title())
        st.markdown(f"**Details for {label}: {entity_key}**")
        # Reflow KPIs into two rows for readability on smaller widths
        r1 = st.columns(3)
        r1[0].metric("Stock Turn", f"{stock_turn_local:.2f}")
        r1[1].metric("Fill Rate", f"{fill_rate_local:.1%}")
        r1[2].metric("Days of Inventory", _format_days(float(days_inv_local) if pd.notna(days_inv_local) else float("inf")))
        r2 = st.columns(2)
        r2[0].metric("Runout Risk", str(int(met.get("runout_risk", pd.Series(dtype=bool)).sum())))
        r2[1].metric("Average Age of Inventory (days)", f"{avg_age_local:.0f}")

        # Visuals
        # 1) Weekly sales quantity (SY)
        if not so.empty:
            so_w = so.copy()
            so_w["order_entry_date"] = pd.to_datetime(so_w["order_entry_date"], errors="coerce")
            so_w = so_w.dropna(subset=["order_entry_date"]).set_index("order_entry_date").sort_index()
            weekly_qty = so_w.resample("W")["quantity_sy"].sum().reset_index()
            fig_sales = px.line(weekly_qty, x="order_entry_date", y="quantity_sy", markers=True, title="Weekly Sales (SY)")
            st.plotly_chart(fig_sales, use_container_width=True)

            # 2) Weekly orders & backorder rate
            weekly_orders = so_w.resample("W")["order_line_id"].nunique().reset_index(name="orders")
            weekly_bo = so_w.resample("W")["backorder_flag"].sum().reset_index(name="backorders")
            wo = weekly_orders.merge(weekly_bo, on="order_entry_date", how="left").fillna(0)
            wo["backorder_rate"] = np.where(wo["orders"] > 0, wo["backorders"] / wo["orders"], 0.0)
            fig_bo = px.bar(wo, x="order_entry_date", y="orders", title="Weekly Orders & Backorder Rate")
            fig_bo.add_scatter(x=wo["order_entry_date"], y=wo["backorder_rate"], mode="lines+markers", name="Backorder Rate", yaxis="y2")
            fig_bo.update_layout(yaxis2=dict(overlaying="y", side="right", tickformat=",.0%", title="Backorder Rate"))
            st.plotly_chart(fig_bo, use_container_width=True)

            # 3) 30-day rolling avg daily sales
            daily_qty = so_w["quantity_sy"].resample("D").sum()
            ads30 = (daily_qty.rolling(30).sum() / 30.0).reset_index().rename(columns={"quantity_sy": "avg_daily_30"})
            fig_ads = px.line(ads30, x="order_entry_date", y="avg_daily_30", markers=False, title="Avg Daily Sales (30-day rolling)")
            st.plotly_chart(fig_ads, use_container_width=True)
        else:
            st.info("No sales orders for this selection.")

        # 4) Incoming PO quantity by ETA week
        po_df = po.copy()
        if not po_df.empty and "eta_date" in po_df.columns:
            po_df["eta_date"] = pd.to_datetime(po_df["eta_date"], errors="coerce")
            # Keep ETAs within ±12 months from today to avoid placeholder far-future dates
            today_ts = pd.Timestamp(date.today())
            min_eta = today_ts - pd.DateOffset(months=12)
            max_eta = today_ts + pd.DateOffset(months=12)
            mask_eta = po_df["eta_date"].between(min_eta, max_eta, inclusive="both")
            po_df = po_df.loc[mask_eta]
            if not po_df.empty:
                po_df = po_df.dropna(subset=["eta_date"]).set_index("eta_date").sort_index()
                weekly_po = po_df.resample("W")["quantity_sy"].sum().reset_index()
                if not weekly_po.empty:
                    fig_po = px.bar(weekly_po, x="eta_date", y="quantity_sy", title="Incoming PO (SY) by ETA Week (±12 months)")
                    st.plotly_chart(fig_po, use_container_width=True)
                else:
                    st.info("No purchase orders within ±12 months for this selection.")
            else:
                st.info("No purchase orders within ±12 months for this selection.")
        else:
            st.info("No purchase orders for this selection.")

        # Tables
        st.markdown("### Orders")
        if not so.empty:
            orders_small = so[[
                "order_number", "sku", "quantity_sy", "unit_of_measure", "order_entry_date", "actual_ship_date", "detail_line_status", "backorder_flag"
            ]].copy()
            # Attach description where available
            if "sku_description" in items_df.columns:
                desc_map = items_df.set_index("sku")["sku_description"]
                orders_small.insert(2, "description", orders_small["sku"].map(desc_map))
            orders_small["order_entry_date"] = pd.to_datetime(orders_small["order_entry_date"]).dt.date
            if "actual_ship_date" in orders_small.columns:
                orders_small["actual_ship_date"] = pd.to_datetime(orders_small["actual_ship_date"]).dt.date
            st.dataframe(orders_small, use_container_width=True, hide_index=True)
        else:
            st.write("No orders")

        st.markdown("### Purchase Orders")
        if not po.empty:
            po_small = po[[
                "order_number", "sku", "quantity_sy", "unit_of_measure", "eta_date", "supplier_number"
            ]].copy()
            if "sku_description" in items_df.columns:
                desc_map = items_df.set_index("sku")["sku_description"]
                po_small.insert(2, "description", po_small["sku"].map(desc_map))
            po_small["eta_date"] = pd.to_datetime(po_small["eta_date"]).dt.date
            st.dataframe(po_small, use_container_width=True, hide_index=True)
        else:
            st.write("No purchase orders")

    # If a details request was queued earlier (from sidebar), open a dialog and render it now
    details_req = st.session_state.pop("_details_request", None)
    if details_req:
        etype, ekey = details_req
        dlg = getattr(st, "dialog", None)
        if callable(dlg):
            @dlg("Details")
            def _details_dialog():
                _show_details_body(etype, ekey)
        else:
            # Fallback: inline render
            st.subheader("Details")
            _show_details_body(etype, ekey)

    # Build triggers for details

    # -------------------------
    # Stock Turn Report (Tab 2)
    # -------------------------
    with tabs[1]:
        st.subheader("Stock Turn Report")
        # Option: Use last full month for all MTD calculations (exclude the current partial month)
        try:
            _pref_full_months = bool(user_prefs.get("stockturn_full_months", False))
        except Exception:
            _pref_full_months = False
        use_full_months = st.checkbox(
            "Use last full month for MTD",
            value=_pref_full_months,
            help=(
                "When enabled, the MTD columns, MTD turns, and MTD fill rate are computed over the last "
                "completed calendar month instead of the current partial month."
            ),
            key="st_full_month_mode",
        )
        # Persist preference and set a session flag for consistency outside this tab
        if use_full_months != _pref_full_months:
            try:
                user_prefs["stockturn_full_months"] = bool(use_full_months)
                _save_prefs(user_prefs)
            except Exception:
                pass
        st.session_state["stockturn_use_full_months"] = bool(use_full_months)
        if use_full_months:
            st.caption("MTD values use the last completed calendar month.")
        st.markdown("**Stock Turn Date Range**")
        sdc1, sdc2 = st.columns(2)
        with sdc1:
            stock_start = st.date_input(
                "Start date",
                value=date(2025, 8, 4),
                key="stock_turn_start_date",
            )
        with sdc2:
            stock_end = st.date_input(
                "End date",
                value=date.today(),
                key="stock_turn_end_date",
            )
        if stock_end < stock_start:
            st.warning("Stock Turn end date is before start date. Swapping the range.")
            stock_start, stock_end = stock_end, stock_start
        if st.session_state.get("defer_heavy", False):
            st.info("Report build deferred. Uncheck 'Defer heavy UI builds' in the sidebar to generate.")
        else:
            t_report = perf_counter()

            # Prepare range + MTD windows relative to Stock Turn date range
            ref_end = stock_end or date.today()
            ytd_start = stock_start or date(ref_end.year, 1, 1)
            # MTD: either current month-to-date, or last full month if toggled
            if use_full_months:
                prev_year = ref_end.year if ref_end.month > 1 else (ref_end.year - 1)
                prev_month = (ref_end.month - 1) if ref_end.month > 1 else 12
                mtd_start = date(prev_year, prev_month, 1)
                mtd_end = date(prev_year, prev_month, calendar.monthrange(prev_year, prev_month)[1])
            else:
                mtd_start = date(ref_end.year, ref_end.month, 1)
                mtd_end = ref_end
            if stock_end and mtd_end > stock_end:
                mtd_end = stock_end

            def _sum_units(df: pd.DataFrame, start_d: date, end_d: date) -> pd.Series:
                if df.empty:
                    return pd.Series(dtype="float64")
                orders = df.copy()
                orders["order_entry_date"] = pd.to_datetime(orders.get("order_entry_date"), errors="coerce")
                mask = orders["order_entry_date"].between(pd.to_datetime(start_d), pd.to_datetime(end_d), inclusive="both")
                return orders.loc[mask].groupby("sku")["quantity_sy"].sum()

            units_ytd = _sum_units(sales_orders, ytd_start, ref_end).rename("units_ytd_sy")
            units_mtd = _sum_units(sales_orders, mtd_start, mtd_end).rename("units_mtd_sy")
            # Convert Series to DataFrames with explicit 'sku' column for safe merges
            def _series_to_df(s: pd.Series, name: str) -> pd.DataFrame:
                if not isinstance(s, pd.Series):
                    return pd.DataFrame(columns=["sku", name])
                s = s.copy()
                s.name = name
                s.index.name = "sku"
                return s.reset_index()
            units_ytd_df = _series_to_df(units_ytd, "units_ytd_sy")
            units_mtd_df = _series_to_df(units_mtd, "units_mtd_sy")

            # Compute per-SKU range order/backorder counts for Fill Rate (distinct order lines)
            orders_range_counts = pd.Series(dtype="float64")
            backorders_range_counts = pd.Series(dtype="float64")
            try:
                if isinstance(sales_orders, pd.DataFrame) and not sales_orders.empty:
                    so_range = sales_orders.copy()
                    so_range["order_entry_date"] = pd.to_datetime(so_range.get("order_entry_date"), errors="coerce")
                    range_mask = so_range["order_entry_date"].between(pd.to_datetime(ytd_start), pd.to_datetime(ref_end), inclusive="both")
                    so_range = so_range.loc[range_mask].copy()
                    if "backorder_flag" not in so_range.columns and "detail_line_status" in so_range.columns:
                        detail_col = so_range.get("detail_line_status").fillna("").astype(str).str.strip().str.upper()
                        so_range["backorder_flag"] = detail_col.isin(["R", "B"])
                    if "order_line_id" in so_range.columns:
                        so_range = so_range.sort_values("order_entry_date").drop_duplicates(subset=["order_line_id"], keep="first")
                    grp_orders = so_range.groupby("sku")
                    orders_range_counts = grp_orders.size().rename("orders_count").astype(float)
                    backorders_range_counts = so_range.loc[so_range.get("backorder_flag") == True].groupby("sku").size().rename("backorder_count").astype(float)
            except Exception:
                pass

            # Compute per-SKU MTD order/backorder counts for MTD Fill Rate (distinct order lines)
            orders_mtd_counts = pd.Series(dtype="float64")
            backorders_mtd_counts = pd.Series(dtype="float64")
            try:
                if isinstance(sales_orders, pd.DataFrame) and not sales_orders.empty:
                    so_mtd_cnt = sales_orders.copy()
                    so_mtd_cnt["order_entry_date"] = pd.to_datetime(so_mtd_cnt.get("order_entry_date"), errors="coerce")
                    mtd_mask_cnt = so_mtd_cnt["order_entry_date"].between(pd.to_datetime(mtd_start), pd.to_datetime(mtd_end), inclusive="both")
                    so_mtd_cnt = so_mtd_cnt.loc[mtd_mask_cnt].copy()
                    # Ensure backorder_flag exists (rebuild if needed)
                    if "backorder_flag" not in so_mtd_cnt.columns and "detail_line_status" in so_mtd_cnt.columns:
                        detail_col = so_mtd_cnt.get("detail_line_status").fillna("").astype(str).str.strip().str.upper()
                        so_mtd_cnt["backorder_flag"] = detail_col.isin(["R", "B"])
                    # Distinct lines
                    if "order_line_id" in so_mtd_cnt.columns:
                        so_mtd_cnt = so_mtd_cnt.sort_values("order_entry_date").drop_duplicates(subset=["order_line_id"], keep="first")
                    grp_orders = so_mtd_cnt.groupby("sku")
                    orders_mtd_counts = grp_orders.size().rename("orders_count_mtd").astype(float)
                    backorders_mtd_counts = so_mtd_cnt.loc[so_mtd_cnt.get("backorder_flag") == True].groupby("sku").size().rename("backorder_count_mtd").astype(float)
            except Exception:
                pass

            # Join price class description from items (do NOT filter by IIXREF for description availability)
            sku_info = items_df[[c for c in ["sku", "price_class_desc", "sku_description"] if c in items_df.columns]].drop_duplicates()
            base = sku_metrics.merge(sku_info, on="sku", how="left")

            base = base.merge(units_ytd_df, on="sku", how="left").merge(units_mtd_df, on="sku", how="left")
            base["units_ytd_sy"] = pd.to_numeric(base.get("units_ytd_sy"), errors="coerce").fillna(0.0)
            base["units_mtd_sy"] = pd.to_numeric(base.get("units_mtd_sy"), errors="coerce").fillna(0.0)

            # Attach MTD order/backorder counts
            if not orders_mtd_counts.empty:
                try:
                    odf = orders_mtd_counts.copy()
                    if isinstance(odf, pd.Series):
                        odf.index.name = "sku"
                        odf = odf.reset_index()
                    base = base.merge(odf, on="sku", how="left")
                except Exception:
                    pass
            if not backorders_mtd_counts.empty:
                try:
                    bdf = backorders_mtd_counts.copy()
                    if isinstance(bdf, pd.Series):
                        bdf.index.name = "sku"
                        bdf = bdf.reset_index()
                    base = base.merge(bdf, on="sku", how="left")
                except Exception:
                    pass
            # Ensure columns exist before coercion to avoid scalar to_numeric returning numpy.float64
            if "orders_count_mtd" not in base.columns:
                base["orders_count_mtd"] = 0.0
            if "backorder_count_mtd" not in base.columns:
                base["backorder_count_mtd"] = 0.0
            base["orders_count_mtd"] = pd.to_numeric(base["orders_count_mtd"], errors="coerce").fillna(0.0)
            base["backorder_count_mtd"] = pd.to_numeric(base["backorder_count_mtd"], errors="coerce").fillna(0.0)

            # Attach range orders/backorders and recompute avg_daily_sales_sy for the range
            for col in ["orders_count", "backorder_count"]:
                if col in base.columns:
                    base.drop(columns=[col], inplace=True)
            if not orders_range_counts.empty:
                try:
                    odf_range = orders_range_counts.copy()
                    if isinstance(odf_range, pd.Series):
                        odf_range.index.name = "sku"
                        odf_range = odf_range.reset_index()
                    base = base.merge(odf_range, on="sku", how="left")
                except Exception:
                    pass
            if not backorders_range_counts.empty:
                try:
                    bdf_range = backorders_range_counts.copy()
                    if isinstance(bdf_range, pd.Series):
                        bdf_range.index.name = "sku"
                        bdf_range = bdf_range.reset_index()
                    base = base.merge(bdf_range, on="sku", how="left")
                except Exception:
                    pass
            if "orders_count" not in base.columns:
                base["orders_count"] = 0.0
            if "backorder_count" not in base.columns:
                base["backorder_count"] = 0.0
            base["orders_count"] = pd.to_numeric(base["orders_count"], errors="coerce").fillna(0.0)
            base["backorder_count"] = pd.to_numeric(base["backorder_count"], errors="coerce").fillna(0.0)

            days_in_range = max((pd.Timestamp(ref_end) - pd.Timestamp(ytd_start)).days + 1, 1)
            avg_daily_range = units_ytd / float(days_in_range)
            avg_daily_df = _series_to_df(avg_daily_range, "avg_daily_sales_sy")
            base = base.merge(avg_daily_df, on="sku", how="left", suffixes=("", "_range"))
            if "avg_daily_sales_sy_range" in base.columns:
                base["avg_daily_sales_sy"] = base["avg_daily_sales_sy_range"].fillna(0.0)
                base.drop(columns=["avg_daily_sales_sy_range"], inplace=True)

            inv_range = pd.to_numeric(base.get("inventory_sy"), errors="coerce").fillna(0.0)
            avg_daily_range_vals = pd.to_numeric(base.get("avg_daily_sales_sy"), errors="coerce").fillna(0.0)
            base["days_of_inventory"] = np.where(avg_daily_range_vals > 0, inv_range / avg_daily_range_vals, float("inf"))

            # Compute per-SKU fill rate when possible
            if "orders_count" in base.columns and "backorder_count" in base.columns:
                oc = pd.to_numeric(base.get("orders_count"), errors="coerce").fillna(0.0)
                bc = pd.to_numeric(base.get("backorder_count"), errors="coerce").fillna(0.0)
                with np.errstate(divide="ignore", invalid="ignore"):
                    fr = 1.0 - np.where(oc > 0, bc / np.maximum(oc, 1e-9), 0.0)
                # clamp 0..1
                base["fill_rate"] = np.clip(fr, 0.0, 1.0)
            else:
                base["fill_rate"] = np.nan

            # Recompute SKU ratings based on filtered orders_count
            try:
                from app.services.sku_rating import assign_sku_ratings
                base = assign_sku_ratings(
                    base.reset_index(drop=True),
                    value_column="orders_count",
                    rating_column="sku_rating",
                    count_column="orders_count",
                )
            except Exception:
                pass

            # MTD fill rate (distinct order lines MTD)
            try:
                oc_mtd = pd.to_numeric(base.get("orders_count_mtd"), errors="coerce").fillna(0.0)
                bc_mtd = pd.to_numeric(base.get("backorder_count_mtd"), errors="coerce").fillna(0.0)
                with np.errstate(divide="ignore", invalid="ignore"):
                    fr_mtd = 1.0 - np.where(oc_mtd > 0, bc_mtd / np.maximum(oc_mtd, 1e-9), 0.0)
                base["mtd_fill_rate"] = np.clip(fr_mtd, 0.0, 1.0)
            except Exception:
                base["mtd_fill_rate"] = np.nan

            # Compute YTD/MTD stock turn at SKU level
            # YTD Turn aligned with top metric: (avg_daily_sales_sy * 365) / inventory
            # MTD Turn: project month from MTD sales so far, annualize: (units_mtd * (days_in_month/elapsed_days) * 12) / inventory
            if use_full_months:
                days_in_month = calendar.monthrange(mtd_start.year, mtd_start.month)[1]
                elapsed_days = days_in_month
            else:
                days_in_month = calendar.monthrange(ref_end.year, ref_end.month)[1]
                elapsed_days = max(ref_end.day, 1)

            inv = pd.to_numeric(base.get("inventory_sy"), errors="coerce").fillna(0.0)
            avg_daily = pd.to_numeric(base.get("avg_daily_sales_sy"), errors="coerce").fillna(0.0)
            base["ytd_turn"] = np.where(inv > 0, (avg_daily * 365.0) / np.maximum(inv, 1e-6), 0.0)
            mtd_est_month = base["units_mtd_sy"] * (days_in_month / elapsed_days)
            base["mtd_turn"] = np.where(inv > 0, (mtd_est_month * 12.0) / np.maximum(inv, 1e-6), 0.0)

            # Select and rename display columns
            disp_cols = [
                "price_class_desc", "sku", "sku_description", "sku_rating", "units_ytd_sy", "units_mtd_sy", "inventory_sy",
                "on_order_sy", "ytd_turn", "mtd_turn", "fill_rate", "mtd_fill_rate", "days_of_inventory", "inventory_age_days",
            ]
            existing_disp = [c for c in disp_cols if c in base.columns]
            report = base[existing_disp].copy()
            # Default sort: by price class then ascending by units_mtd_sy, then sku
            if "price_class_desc" in report.columns:
                sort_cols = ["price_class_desc"]
                if "units_mtd_sy" in report.columns:
                    sort_cols.append("units_mtd_sy")
                sort_cols.append("sku")
                report.sort_values(sort_cols, ascending=[True] * len(sort_cols), inplace=True, kind="mergesort")

            # Merge avg_daily for group-level days_of_inventory in totals
            # Enrich for PDF: ensure we carry avg daily and counts for fill rate (rating already in 'report')
            range_cols = [c for c in ["sku", "avg_daily_sales_sy", "orders_count", "backorder_count", "sku_rating"] if c in base.columns or c == "sku"]
            rpt = report.merge(base[range_cols], on="sku", how="left")
            # If a previous merge created suffixed rating columns, coalesce to 'sku_rating'
            if "sku_rating" not in rpt.columns:
                for cand in ("sku_rating_x", "sku_rating_y"):
                    if cand in rpt.columns:
                        rpt["sku_rating"] = rpt[cand]
                        break

            # Compute price-class ratings by comparing classes against each other
            pc_rating_map: dict[str, str] = {}
            try:
                if "orders_count" in base.columns and "price_class_desc" in base.columns and "sku" in base.columns:
                    # Use average orders per SKU (sum/orders_count divided by # of SKUs in that price class)
                    tmp = base[["price_class_desc", "orders_count", "sku"]].copy()
                    tmp["orders_count"] = pd.to_numeric(tmp.get("orders_count"), errors="coerce").fillna(0.0)
                    grp = tmp.groupby("price_class_desc", dropna=False).agg(
                        total_orders=("orders_count", "sum"),
                        sku_count=("sku", "nunique"),
                    ).reset_index()
                    grp["sku_count"] = grp["sku_count"].replace(0, np.nan)
                    grp["avg_orders_per_sku"] = grp["total_orders"] / grp["sku_count"]
                    # Normalize key for NaNs
                    grp["pc_key"] = grp["price_class_desc"].apply(lambda x: str(x) if pd.notna(x) else "(No Price Class)")
                    # Assign ratings based on average orders per SKU to neutralize class size effects
                    from app.services.sku_rating import assign_sku_ratings
                    pc_rated = assign_sku_ratings(
                        grp.rename(columns={"avg_orders_per_sku": "orders_val"}),
                        value_column="orders_val",
                        rating_column="pc_rating",
                    )
                    pc_rating_map = {row["pc_key"]: row["pc_rating"] for _, row in pc_rated.iterrows()}
            except Exception:
                pc_rating_map = {}

            # PDF export helper
            def _build_stock_turn_pdf(df_pdf: pd.DataFrame) -> bytes:
                if not _HAS_REPORTLAB:
                    raise RuntimeError("PDF export requires 'reportlab' package. Please install it.")
                from reportlab.lib import colors  # Import colors within function scope
                buffer = BytesIO()
                doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), leftMargin=24, rightMargin=24, topMargin=24, bottomMargin=24)
                styles = getSampleStyleSheet()
                story: list = []

                # Footer callback with date/time and page number
                def _footer(canv, doc_obj):
                    from datetime import datetime
                    canv.saveState()
                    txt = f"{datetime.now().strftime('%Y-%m-%d %H:%M')}  ·  Page {canv.getPageNumber()}"
                    canv.setFont("Helvetica", 8)
                    canv.drawString(doc_obj.leftMargin, 12, txt)
                    canv.restoreState()

                def fmt(x):
                    try:
                        if x is None or pd.isna(x):
                            return ""
                        if np.isinf(x):
                            return "∞"
                        return f"{float(x):,.2f}"
                    except Exception:
                        return str(x)

                # Build a proper-case description column from INAME (items.sku_description)
                try:
                    if "sku_description" in df_pdf.columns and "desc" not in df_pdf.columns:
                        def _proper(s) -> str:
                            try:
                                t = str(s).strip()
                                if not t:
                                    return ""
                                # Basic proper case; keep all-caps acronyms if 2 or fewer chars
                                parts = t.lower().split()
                                out = [p.upper() if len(p) <= 2 and p.isalpha() else (p.capitalize()) for p in parts]
                                return " ".join(out)
                            except Exception:
                                return str(s)
                        df_pdf = df_pdf.copy()
                        df_pdf["desc"] = df_pdf["sku_description"].apply(_proper)
                    # Fallback: fill desc from items_df mapping if still missing/blank
                    try:
                        if "desc" not in df_pdf.columns:
                            df_pdf = df_pdf.copy()
                            df_pdf["desc"] = ""
                        missing = df_pdf["desc"].isna() | (df_pdf["desc"].astype(str).str.strip() == "")
                        if missing.any():
                            if isinstance(items_df, pd.DataFrame) and not items_df.empty and {"sku", "sku_description"}.issubset(items_df.columns):
                                _map = items_df.set_index("sku")["sku_description"]
                                fill_vals = df_pdf.loc[missing, "sku"].map(_map)
                                # apply proper-case
                                fill_vals = fill_vals.apply(lambda v: "" if pd.isna(v) else " ".join([p.upper() if len(p) <= 2 and p.isalpha() else p.capitalize() for p in str(v).strip().lower().split()]))
                                df_pdf.loc[missing, "desc"] = fill_vals
                    except Exception:
                        pass
                except Exception:
                    pass

                cols = [
                    ("sku", "sku"),
                    ("desc", "desc"),
                    ("sku_rating", "rating"),
                    ("units_ytd_sy", "ytd"),
                    ("units_mtd_sy", "mtd"),
                    ("inventory_sy", "inv"),
                    ("on_order_sy", "PO"),
                    ("ytd_turn", "turn_ytd"),
                    ("mtd_turn", "turn_mtd"),
                    ("fill_rate", "fill%"),
                    ("mtd_fill_rate", "fill%_mtd"),
                    ("days_of_inventory", "DOI"),
                    ("inventory_age_days", "Age(d)"),
                ]

                # Rating thresholds header based on orders_count across the dataset (mirrors quartile logic)
                try:
                    import math as _math
                    orders_series = pd.to_numeric(df_pdf.get("orders_count"), errors="coerce").fillna(0.0) if "orders_count" in df_pdf.columns else pd.Series(dtype=float)
                    pos = orders_series[orders_series > 0].sort_values(ascending=False)
                    if len(pos) > 0:
                        n = len(pos)
                        a_cut = max(1, _math.ceil(0.25 * n))
                        b_cut = max(a_cut, _math.ceil(0.50 * n))
                        c_cut = max(b_cut, _math.ceil(0.75 * n))
                        a_min = float(pos.iloc[a_cut - 1]) if a_cut - 1 < len(pos) else float(pos.iloc[-1])
                        b_min = float(pos.iloc[b_cut - 1]) if b_cut - 1 < len(pos) else float(pos.iloc[-1])
                        c_min = float(pos.iloc[c_cut - 1]) if c_cut - 1 < len(pos) else float(pos.iloc[-1])
                        txt = f"Rating thresholds (orders_count): A ≥ {a_min:.0f}, B ≥ {b_min:.0f}, C ≥ {c_min:.0f}, D = others (incl. 0)."
                    else:
                        txt = "Rating thresholds: All SKUs are D (no positive orders in period)."
                    story.append(Paragraph(txt, styles["Normal"]))
                    story.append(Spacer(1, 8))
                except Exception:
                    pass

                # Track sum of inventory_sy from section Total rows (only for sections with >1 SKU)
                section_totals_inventory_sum = 0.0

                # Price-class sections
                if "price_class_desc" in df_pdf.columns and not df_pdf.empty:
                    for desc, grp in df_pdf.groupby("price_class_desc", sort=True):
                        desc_display = str(desc) if pd.notna(desc) else "(No Price Class)"
                        # Compute group-level fill rate early so it can appear in the heading
                        try:
                            total_orders_grp_h = pd.to_numeric(grp.get("orders_count"), errors="coerce").fillna(0).sum() if "orders_count" in grp.columns else 0.0
                            total_backorders_grp_h = pd.to_numeric(grp.get("backorder_count"), errors="coerce").fillna(0).sum() if "backorder_count" in grp.columns else 0.0
                            fill_rate_grp_h = (1 - (total_backorders_grp_h / total_orders_grp_h)) if total_orders_grp_h > 0 else 0.0
                            fill_rate_txt = f"{fill_rate_grp_h:.1%}"
                        except Exception:
                            fill_rate_txt = ""
                        # Append overall rating and fill rate next to the heading
                        try:
                            desc_key = desc_display
                            pc_rate_heading = pc_rating_map.get(desc_key, "") if isinstance(pc_rating_map, dict) else ""
                        except Exception:
                            pc_rate_heading = ""
                        if pc_rate_heading and fill_rate_txt:
                            title_text = f"{desc_display} ({pc_rate_heading} - {fill_rate_txt})"
                        elif pc_rate_heading:
                            title_text = f"{desc_display} ({pc_rate_heading})"
                        elif fill_rate_txt:
                            title_text = f"{desc_display} ({fill_rate_txt})"
                        else:
                            title_text = desc_display
                        title = Paragraph(title_text, styles["Heading3"])
                        story.append(title)

                        # Totals per group (same as on-screen)
                        total_units_ytd = pd.to_numeric(grp.get("units_ytd_sy"), errors="coerce").fillna(0).sum()
                        total_units_mtd = pd.to_numeric(grp.get("units_mtd_sy"), errors="coerce").fillna(0).sum()
                        total_inventory = pd.to_numeric(grp.get("inventory_sy"), errors="coerce").fillna(0).sum()
                        total_avg_daily = pd.to_numeric(grp.get("avg_daily_sales_sy"), errors="coerce").fillna(0).sum()
                        total_on_order = pd.to_numeric(grp.get("on_order_sy"), errors="coerce").fillna(0).sum() if "on_order_sy" in grp.columns else 0.0
                        opening_m_grp = total_inventory + total_units_mtd
                        opening_y_grp = total_inventory + total_units_ytd
                        denom_m_grp = (opening_m_grp + total_inventory) / 2.0
                        denom_y_grp = (opening_y_grp + total_inventory) / 2.0
                        mtd_turn_grp = (total_units_mtd / denom_m_grp) if denom_m_grp > 0 else 0.0
                        ytd_turn_grp = (total_units_ytd / denom_y_grp) if denom_y_grp > 0 else 0.0
                        days_inv_grp = (total_inventory / total_avg_daily) if total_avg_daily > 0 else float("inf")

                        # Sort rows by MTD units ascending for this section
                        grp_sorted = grp
                        if "units_mtd_sy" in grp.columns:
                            grp_sorted = grp.sort_values(by=["units_mtd_sy"], ascending=True, kind="mergesort")

                        # Ensure sku_rating is populated for rows in PDF; recompute from orders_count if missing
                        try:
                            need_rating = ("sku_rating" not in grp_sorted.columns) or grp_sorted["sku_rating"].isna().all()
                        except Exception:
                            need_rating = True
                        if need_rating and "orders_count" in grp_sorted.columns:
                            try:
                                from app.services.sku_rating import assign_sku_ratings as _assign_pdf_ratings
                                grp_sorted = _assign_pdf_ratings(
                                    grp_sorted.reset_index(drop=True),
                                    value_column="orders_count",
                                    rating_column="sku_rating",
                                    count_column="orders_count",
                                )
                            except Exception:
                                pass

                        # Group-level fill rate if counts available
                        total_orders_grp = pd.to_numeric(grp.get("orders_count"), errors="coerce").fillna(0).sum() if "orders_count" in grp.columns else 0.0
                        total_backorders_grp = pd.to_numeric(grp.get("backorder_count"), errors="coerce").fillna(0).sum() if "backorder_count" in grp.columns else 0.0
                        fill_rate_grp = (1 - (total_backorders_grp / total_orders_grp)) if total_orders_grp > 0 else 0.0

                        # Weighted average inventory age for group
                        inv_qty_grp = pd.to_numeric(grp.get("inventory_sy"), errors="coerce").fillna(0)
                        inv_age_grp = pd.to_numeric(grp.get("inventory_age_days"), errors="coerce") if "inventory_age_days" in grp.columns else pd.Series(dtype=float)
                        mask_age = (inv_qty_grp > 0) & inv_age_grp.notna()
                        avg_age_grp = float((inv_qty_grp[mask_age] * inv_age_grp[mask_age]).sum() / inv_qty_grp[mask_age].sum()) if mask_age.any() else 0.0

                        header = [c for _, c in cols]
                        data = [header]
                        # Collect conditional style commands for low fill rate highlighting
                        red_cmds = []
                        thresholds = {"A": 0.90, "B": 0.80, "C": 0.70}
                        for i, (_, row) in enumerate(grp_sorted.iterrows()):
                            rating_val = str(row.get("sku_rating", "")).strip().upper()
                            try:
                                fr_val = float(pd.to_numeric(row.get("fill_rate"), errors="coerce"))
                            except Exception:
                                fr_val = float("nan")
                            try:
                                fr_mtd_val = float(pd.to_numeric(row.get("mtd_fill_rate"), errors="coerce"))
                            except Exception:
                                fr_mtd_val = float("nan")
                            try:
                                age_val = float(pd.to_numeric(row.get("inventory_age_days"), errors="coerce"))
                            except Exception:
                                age_val = float("nan")
                            # Build data row
                            data.append([
                                str(row.get("sku", "")),
                                str(row.get("desc", row.get("sku_description", ""))),
                                str(row.get("sku_rating", "")),
                                fmt(row.get("units_ytd_sy")),
                                fmt(row.get("units_mtd_sy")),
                                fmt(row.get("inventory_sy")),
                                fmt(row.get("on_order_sy")),
                                fmt(row.get("ytd_turn")),
                                fmt(row.get("mtd_turn")),
                                f"{fr_val:.1%}" if pd.notna(fr_val) else "",
                                f"{fr_mtd_val:.1%}" if pd.notna(fr_mtd_val) else "",
                                fmt(row.get("days_of_inventory")),
                                fmt(row.get("inventory_age_days")),
                            ])
                            # If threshold triggered, color the Fill% cells (columns 9 and 10) red for this row (offset +1 for header)
                            thr = thresholds.get(rating_val)
                            if thr is not None and pd.notna(fr_val) and fr_val < thr:
                                red_cmds.append(("TEXTCOLOR", (9, 1 + i), (9, 1 + i), colors.red))
                            if thr is not None and pd.notna(fr_mtd_val) and fr_mtd_val < thr:
                                red_cmds.append(("TEXTCOLOR", (10, 1 + i), (10, 1 + i), colors.red))
                            # Shade Age(d) cell light red if > 300 days (column index 12)
                            if pd.notna(age_val) and age_val > 300:
                                red_cmds.append(("BACKGROUND", (12, 1 + i), (12, 1 + i), colors.HexColor("#ffe5e5")))
                        # totals row (suppress when there is only one SKU in this price class)
                        if len(grp_sorted.index) > 1:
                            desc_key = str(desc) if pd.notna(desc) else "(No Price Class)"
                            pc_rate = pc_rating_map.get(desc_key, "")
                            # accumulate inventory_sy for this section total
                            try:
                                section_totals_inventory_sum += float(total_inventory)
                            except Exception:
                                pass
                            data.append([
                                "Total",
                                "",  # desc (blank for totals)
                                pc_rate,
                                fmt(total_units_ytd),
                                fmt(total_units_mtd),
                                fmt(total_inventory),
                                fmt(total_on_order),
                                fmt(ytd_turn_grp),
                                fmt(mtd_turn_grp),
                                f"{fill_rate_grp:.1%}",
                                "",  # leave group total MTD fill blank (optional: could compute separately)
                                fmt(days_inv_grp),
                                fmt(avg_age_grp),
                            ])
                            # Shade Age(d) in totals if > 300
                            try:
                                if float(avg_age_grp) > 300:
                                    # totals row index is the last row added
                                    totals_row_idx = len(data) - 1
                                    red_cmds.append(("BACKGROUND", (12, totals_row_idx), (12, totals_row_idx), colors.HexColor("#ffe5e5")))
                            except Exception:
                                pass

                        tbl = Table(data, repeatRows=1)
                        base_style = [
                            ("GRID", (0,0), (-1,-1), 0.25, colors.black),
                            ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#f0f0f0")),
                            ("ALIGN", (2,1), (-1,-1), "RIGHT"),
                            ("ALIGN", (0,0), (1,-1), "LEFT"),
                            ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
                            ("FONTNAME", (0,-1), (-1,-1), "Helvetica-Bold"),
                            ("FONTSIZE", (0,0), (-1,-1), 7),
                            ("TOPPADDING", (0,0), (-1,-1), 1.5),
                            ("BOTTOMPADDING", (0,0), (-1,-1), 1.5),
                        ]
                        tbl.setStyle(TableStyle(base_style + red_cmds))
                        story.append(tbl)
                        # tighter section break
                        story.append(Spacer(1, 10))

                # Grand totals section
                if not df_pdf.empty:
                    # Deduplicate by SKU to avoid any accidental double-counting introduced by merges
                    df_all = df_pdf.copy()
                    if "sku" in df_all.columns:
                        df_all = df_all.drop_duplicates(subset=["sku"])  # keep first occurrence per SKU

                    # Aggregate totals across unique SKUs
                    total_units_ytd_all = pd.to_numeric(df_all.get("units_ytd_sy"), errors="coerce").fillna(0).sum()
                    total_units_mtd_all = pd.to_numeric(df_all.get("units_mtd_sy"), errors="coerce").fillna(0).sum()
                    total_inventory_all = pd.to_numeric(df_all.get("inventory_sy"), errors="coerce").fillna(0).sum()
                    total_on_order_all = pd.to_numeric(df_all.get("on_order_sy"), errors="coerce").fillna(0).sum() if "on_order_sy" in df_all.columns else 0.0
                    total_avg_daily_all = pd.to_numeric(df_all.get("avg_daily_sales_sy"), errors="coerce").fillna(0).sum()

                    # Prefer using the sum of section totals for inventory grand total if available
                    grand_inventory_all = section_totals_inventory_sum if section_totals_inventory_sum > 0 else total_inventory_all

                    # Align grand-total turn formulas with on-screen metrics
                    # YTD Turn: (sum(avg_daily) * 365) / inventory_grand
                    ytd_turn_all = ((total_avg_daily_all * 365.0) / grand_inventory_all) if grand_inventory_all > 0 else 0.0
                    # MTD Turn: project month from MTD so far, annualize, divide by inventory_grand
                    try:
                        mtd_est_month_all = total_units_mtd_all * (days_in_month / max(elapsed_days, 1))
                    except Exception:
                        # Fallback if context variables are unavailable for any reason
                        mtd_est_month_all = total_units_mtd_all
                    mtd_turn_all = (mtd_est_month_all * 12.0 / grand_inventory_all) if grand_inventory_all > 0 else 0.0
                    days_inv_all = (grand_inventory_all / total_avg_daily_all) if total_avg_daily_all > 0 else float("inf")

                    # Grand total fill rate
                    total_orders_all = pd.to_numeric(df_all.get("orders_count"), errors="coerce").fillna(0).sum() if "orders_count" in df_all.columns else 0.0
                    total_backorders_all = pd.to_numeric(df_all.get("backorder_count"), errors="coerce").fillna(0).sum() if "backorder_count" in df_all.columns else 0.0
                    fill_rate_all = (1 - (total_backorders_all / total_orders_all)) if total_orders_all > 0 else 0.0

                    # Weighted inventory age across all (by inventory quantity)
                    inv_qty_all = pd.to_numeric(df_all.get("inventory_sy"), errors="coerce").fillna(0)
                    inv_age_all = pd.to_numeric(df_all.get("inventory_age_days"), errors="coerce") if "inventory_age_days" in df_all.columns else pd.Series(dtype=float)
                    mask_all = (inv_qty_all > 0) & inv_age_all.notna()
                    avg_age_all = float((inv_qty_all[mask_all] * inv_age_all[mask_all]).sum() / inv_qty_all[mask_all].sum()) if mask_all.any() else 0.0

                    story.append(Paragraph("Grand Totals", styles["Heading3"]))
                    data_g = [[
                        "units_ytd_sy", "units_mtd_sy", "inventory_sy", "PO", "ytd_turn", "mtd_turn", "fill_rate", "days_of_inventory", "inventory_age_days"
                    ], [
                        fmt(total_units_ytd_all), fmt(total_units_mtd_all), fmt(grand_inventory_all), fmt(total_on_order_all), fmt(ytd_turn_all), fmt(mtd_turn_all), f"{fill_rate_all:.1%}", fmt(days_inv_all), fmt(avg_age_all)
                    ]]
                    tbl_g = Table(data_g, repeatRows=1)
                    tbl_g.setStyle(TableStyle([
                        ("GRID", (0,0), (-1,-1), 0.25, colors.black),
                        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#f0f0f0")),
                        ("ALIGN", (0,1), (-1,-1), "RIGHT"),
                        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
                        ("FONTNAME", (0,1), (-1,1), "Helvetica-Bold"),
                        ("FONTSIZE", (0,0), (-1,-1), 8),
                        ("TOPPADDING", (0,0), (-1,-1), 1.5),
                        ("BOTTOMPADDING", (0,0), (-1,-1), 1.5),
                    ]))
                    story.append(tbl_g)

                    # Also display the sum of section Total inventories (for comparison with grand total)
                    try:
                        data_s = [["Sum of section totals (inventory_sy)"], [fmt(section_totals_inventory_sum)]]
                        tbl_s = Table(data_s, repeatRows=1)
                        tbl_s.setStyle(TableStyle([
                            ("GRID", (0,0), (-1,-1), 0.25, colors.black),
                            ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#f0f0f0")),
                            ("ALIGN", (0,1), (-1,-1), "RIGHT"),
                            ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
                            ("FONTNAME", (0,1), (-1,1), "Helvetica-Bold"),
                            ("FONTSIZE", (0,0), (-1,-1), 8),
                            ("TOPPADDING", (0,0), (-1,-1), 1.5),
                            ("BOTTOMPADDING", (0,0), (-1,-1), 1.5),
                        ]))
                        story.append(Spacer(1, 6))
                        story.append(tbl_s)
                    except Exception:
                        pass

                doc.build(story, onFirstPage=_footer, onLaterPages=_footer)
                buffer.seek(0)
                return buffer.read()

            # Offer PDF download
            if _HAS_REPORTLAB:
                try:
                    pdf_bytes = _build_stock_turn_pdf(rpt)
                    st.download_button(
                        label="Download Stock Turn (PDF)",
                        data=pdf_bytes,
                        file_name="stock_turn_report.pdf",
                        mime="application/pdf",
                        type="primary",
                    )
                except Exception as e:
                    import traceback
                    import sys
                    print(f"\n{'='*80}", file=sys.stderr)
                    print(f"ERROR BUILDING PDF: {type(e).__name__}", file=sys.stderr)
                    print(f"Message: {str(e)}", file=sys.stderr)
                    traceback.print_exc(file=sys.stderr)
                    print(f"{'='*80}\n", file=sys.stderr)
                    st.warning(f"Unable to build PDF: {e}")
            else:
                st.caption("Install 'reportlab' to enable PDF export (pip install reportlab)")

            # Render per price class with totals
            def _render_group(group_df: pd.DataFrame, title: str) -> None:
                # Totals and weighted metrics for the group
                total_units_ytd = pd.to_numeric(group_df.get("units_ytd_sy"), errors="coerce").fillna(0).sum()
                total_units_mtd = pd.to_numeric(group_df.get("units_mtd_sy"), errors="coerce").fillna(0).sum()
                total_inventory = pd.to_numeric(group_df.get("inventory_sy"), errors="coerce").fillna(0).sum()
                total_avg_daily = pd.to_numeric(group_df.get("avg_daily_sales_sy"), errors="coerce").fillna(0).sum() if "avg_daily_sales_sy" in group_df.columns else 0.0

                # Aggregate YTD/MTD turns at group level
                # YTD Turn matches top metric using avg daily
                ytd_turn_grp = ((total_avg_daily * 365.0) / total_inventory) if total_inventory > 0 else 0.0
                # MTD Turn projects month from MTD so far
                mtd_est_month_grp = total_units_mtd * (days_in_month / elapsed_days)
                mtd_turn_grp = (mtd_est_month_grp * 12.0 / total_inventory) if total_inventory > 0 else 0.0
                days_inv_grp = (total_inventory / total_avg_daily) if total_avg_daily > 0 else float("inf")

                # Weighted average age by inventory
                inv_qty = pd.to_numeric(group_df.get("inventory_sy"), errors="coerce").fillna(0)
                inv_age = pd.to_numeric(group_df.get("inventory_age_days"), errors="coerce")
                mask = (inv_qty > 0) & inv_age.notna()
                avg_age_grp = float((inv_qty[mask] * inv_age[mask]).sum() / inv_qty[mask].sum()) if mask.any() else 0.0

                # Display group
                st.markdown(f"### {title}")
                display_cols = [
                    c for c in ["sku", "sku_description", "sku_rating", "units_ytd_sy", "units_mtd_sy", "inventory_sy", "on_order_sy", "ytd_turn", "mtd_turn", "fill_rate", "days_of_inventory", "inventory_age_days"]
                    if c in group_df.columns
                ]
                df_disp = group_df[display_cols].copy()
                # Ensure ascending sort by units_mtd_sy
                if "units_mtd_sy" in df_disp.columns:
                    df_disp = df_disp.sort_values(by=["units_mtd_sy"], ascending=True, kind="mergesort")
                num_cols = df_disp.select_dtypes(include=["number"]).columns
                df_disp[num_cols] = df_disp[num_cols].round(2)
                # Rename on_order_sy for display consistency with SKU Performance
                if "on_order_sy" in df_disp.columns:
                    df_disp.rename(columns={"on_order_sy": "PO"}, inplace=True)
                # Editable table with Details checkbox like the SKU Performance table
                if "sku" in df_disp.columns:
                    df_edit = df_disp.copy()
                    st_select_col = "__open_details_st__"
                    if st_select_col not in df_edit.columns:
                        df_edit.insert(0, st_select_col, False)
                    st_version = int(st.session_state.get("_stock_turn_editor_version", 0))
                    st_key = f"stock_turn_editor_{str(title)}_{st_version}"
                    edited = st.data_editor(
                        df_edit,
                        use_container_width=True,
                        hide_index=True,
                        num_rows="fixed",
                        column_config={
                            st_select_col: st.column_config.CheckboxColumn(
                                "Details",
                                help="Open details for this SKU",
                                default=False,
                            )
                        },
                        key=st_key,
                    )
                    try:
                        chosen = edited[edited[st_select_col] == True]
                        if not chosen.empty:
                            chosen_sku = str(chosen.iloc[0].get("sku", "")).strip()
                            if chosen_sku:
                                dlg = getattr(st, "dialog", None)
                                if callable(dlg):
                                    @dlg("Details")
                                    def _st_details_dialog():
                                        _show_details_body("sku", chosen_sku)
                                    _st_details_dialog()
                                else:
                                    st.subheader("Details")
                                    _show_details_body("sku", chosen_sku)
                                st.session_state["_stock_turn_editor_version"] = st_version + 1
                    except Exception:
                        pass
                else:
                    st.dataframe(df_disp, use_container_width=True, hide_index=True)

                # Totals row
                # Compute group-level fill rate if available as the mean of per-SKU fill rates weighted equally
                # (Exact order-weighted fill rate isn't available here without orders/backorders columns.)
                fill_rate_grp_col = None
                if "fill_rate" in group_df.columns:
                    try:
                        fr_vals = pd.to_numeric(group_df.get("fill_rate"), errors="coerce")
                        if fr_vals.notna().any():
                            fill_rate_grp_col = f"{float(fr_vals.mean()):.1%}"
                    except Exception:
                        fill_rate_grp_col = None

                # Price class rating for this group (compare this price class to others by total orders_count)
                pc_desc_key = str(title) if pd.notna(title) else "(No Price Class)"
                pc_totals_rating = None
                try:
                    # Reuse previously computed map if present
                    if 'pc_rating_map' in locals() and isinstance(pc_rating_map, dict):
                        pc_totals_rating = pc_rating_map.get(pc_desc_key)
                except Exception:
                    pc_totals_rating = None

                total_on_order_grp = pd.to_numeric(group_df.get("on_order_sy"), errors="coerce").fillna(0).sum() if "on_order_sy" in group_df.columns else 0.0
                totals_row = {
                    "sku": "Total",
                    "sku_rating": pc_totals_rating if pc_totals_rating is not None else np.nan,
                    "units_ytd_sy": total_units_ytd,
                    "units_mtd_sy": total_units_mtd,
                    "inventory_sy": total_inventory,
                    "on_order_sy": total_on_order_grp,
                    "ytd_turn": ytd_turn_grp,
                    "mtd_turn": mtd_turn_grp,
                    "fill_rate": fill_rate_grp_col if fill_rate_grp_col is not None else np.nan,
                    "days_of_inventory": days_inv_grp,
                    "inventory_age_days": avg_age_grp,
                }
                tot_display = pd.DataFrame([totals_row])
                # Ensure all display columns exist in totals frame to avoid KeyError
                missing_cols = [c for c in display_cols if c not in tot_display.columns]
                for mc in missing_cols:
                    tot_display[mc] = np.nan
                tot_display = tot_display.reindex(columns=display_cols)
                num_cols = tot_display.select_dtypes(include=["number"]).columns
                tot_display[num_cols] = tot_display[num_cols].round(2)
                # Rename on_order_sy for display
                if "on_order_sy" in tot_display.columns:
                    tot_display.rename(columns={"on_order_sy": "PO"}, inplace=True)
                st.dataframe(tot_display, use_container_width=True, hide_index=True)

            if "price_class_desc" in report.columns and not report.empty:
                # Merge avg_daily_sales_sy for group days_of_inventory calc
                report = report.merge(sku_metrics[["sku", "avg_daily_sales_sy"]], on="sku", how="left")
                section_totals_inventory_sum_screen = 0.0
                for desc, grp in report.groupby("price_class_desc", sort=True):
                    _render_group(grp, desc if pd.notna(desc) else "(No Price Class)")
                    # accumulate only when the section would display a Total row (more than one SKU)
                    try:
                        if len(grp.index) > 1:
                            section_totals_inventory_sum_screen += float(pd.to_numeric(grp.get("inventory_sy"), errors="coerce").fillna(0).sum())
                    except Exception:
                        pass
            else:
                st.info("No price class data available to build the report.")

            # Grand totals across all
            if not report.empty:
                total_units_ytd_all = pd.to_numeric(report.get("units_ytd_sy"), errors="coerce").fillna(0).sum()
                total_units_mtd_all = pd.to_numeric(report.get("units_mtd_sy"), errors="coerce").fillna(0).sum()
                total_inventory_all = pd.to_numeric(report.get("inventory_sy"), errors="coerce").fillna(0).sum()
                total_on_order_all = pd.to_numeric(report.get("on_order_sy"), errors="coerce").fillna(0).sum() if "on_order_sy" in report.columns else 0.0
                total_avg_daily_all = pd.to_numeric(base.get("avg_daily_sales_sy"), errors="coerce").fillna(0).sum()
                # Prefer using the sum of section totals for inventory grand total if available
                try:
                    grand_inventory_all_screen = section_totals_inventory_sum_screen if section_totals_inventory_sum_screen > 0 else total_inventory_all
                except Exception:
                    grand_inventory_all_screen = total_inventory_all
                # YTD Turn grand total aligned with top metric
                ytd_turn_all = ((total_avg_daily_all * 365.0) / grand_inventory_all_screen) if grand_inventory_all_screen > 0 else 0.0
                # MTD Turn grand total projected from MTD so far
                mtd_est_month_all = total_units_mtd_all * (days_in_month / elapsed_days)
                mtd_turn_all = (mtd_est_month_all * 12.0 / grand_inventory_all_screen) if grand_inventory_all_screen > 0 else 0.0
                days_inv_all = (grand_inventory_all_screen / total_avg_daily_all) if total_avg_daily_all > 0 else float("inf")

                # Weighted age across all
                inv_qty_all = pd.to_numeric(base.get("inventory_sy"), errors="coerce").fillna(0)
                inv_age_all = pd.to_numeric(base.get("inventory_age_days"), errors="coerce")
                mask_all = (inv_qty_all > 0) & inv_age_all.notna()
                avg_age_all = float((inv_qty_all[mask_all] * inv_age_all[mask_all]).sum() / inv_qty_all[mask_all].sum()) if mask_all.any() else 0.0

                st.markdown("### Grand Totals")
                grand = pd.DataFrame([
                    {
                        "units_ytd_sy": total_units_ytd_all,
                        "units_mtd_sy": total_units_mtd_all,
                        "inventory_sy": grand_inventory_all_screen,
                        "on_order_sy": total_on_order_all,
                        "ytd_turn": ytd_turn_all,
                        "mtd_turn": mtd_turn_all,
                        "fill_rate": (1 - (pd.to_numeric(base.get("backorder_count"), errors="coerce").fillna(0).sum() / max(pd.to_numeric(base.get("orders_count"), errors="coerce").fillna(0).sum(), 1e-9))) if ("orders_count" in base.columns and "backorder_count" in base.columns) else np.nan,
                        "days_of_inventory": days_inv_all,
                        "inventory_age_days": avg_age_all,
                    }
                ])
                num_cols = grand.select_dtypes(include=["number"]).columns
                grand[num_cols] = grand[num_cols].round(2)
                # Rename on_order_sy for display
                if "on_order_sy" in grand.columns:
                    grand.rename(columns={"on_order_sy": "PO"}, inplace=True)
                st.dataframe(grand, use_container_width=True, hide_index=True)

                # Show sum of section totals inventory for comparison
                try:
                    st.caption(f"Sum of section totals (inventory_sy): {section_totals_inventory_sum_screen:,.2f}")
                except Exception:
                    pass

                # Persist grand totals components for reuse in Key Metrics
                try:
                    st.session_state["_st_grand_inventory"] = float(grand_inventory_all_screen)
                    st.session_state["_st_total_avg_daily"] = float(total_avg_daily_all)
                    st.session_state["_st_total_units_mtd"] = float(total_units_mtd_all)
                    st.session_state["_st_total_units_ytd"] = float(total_units_ytd_all)
                    st.session_state["_st_days_in_month"] = int(days_in_month)
                    st.session_state["_st_elapsed_days"] = int(elapsed_days)
                except Exception:
                    pass

            ui_timings["ui_build_stock_turn_report"] = perf_counter() - t_report

    if rating_filter:
        sku_metrics = sku_metrics[sku_metrics.get("sku_rating").isin(rating_filter)]

    # Compute YTD and MTD Stock Turn to match Stock Turn Grand Totals
    def _compute_kpi_turns() -> tuple[float, float]:
        try:
            gi = float(st.session_state.get("_st_grand_inventory", float("nan")))
            tad = float(st.session_state.get("_st_total_avg_daily", float("nan")))
            um = float(st.session_state.get("_st_total_units_mtd", float("nan")))
            uy = float(st.session_state.get("_st_total_units_ytd", float("nan")))
            dim = int(st.session_state.get("_st_days_in_month", 0))
            ed = int(st.session_state.get("_st_elapsed_days", 0))
            if not np.isfinite(gi) or gi <= 0 or not np.isfinite(tad):
                raise ValueError("missing session grand totals")
            ytd_turn = (tad * 365.0) / gi if gi > 0 else 0.0
            mtd_est_month = um * (dim / max(ed, 1)) if dim and ed else um
            mtd_turn = (mtd_est_month * 12.0 / gi) if gi > 0 else 0.0
            return float(ytd_turn), float(mtd_turn)
        except Exception:
            # Fallback: compute from current data (may not exactly match Stock Turn tab if not built)
            # Build grand inventory as sum of section totals (groups with >1 SKU) if possible
            try:
                gi = float("nan")
                if {"sku", "inventory_sy"}.issubset(sku_metrics.columns) and {"sku", "price_class_desc"}.issubset(items_df.columns):
                    tmp = sku_metrics[["sku", "inventory_sy"]].merge(
                        items_df[["sku", "price_class_desc"]], on="sku", how="left"
                    )
                    grp = tmp.groupby("price_class_desc").agg(
                        inv_sum=("inventory_sy", "sum"), sku_count=("sku", "nunique")
                    ).reset_index()
                    gi = float(grp.loc[grp["sku_count"] > 1, "inv_sum"].sum())
                if not np.isfinite(gi) or gi <= 0:
                    gi = float(pd.to_numeric(sku_metrics.get("inventory_sy"), errors="coerce").fillna(0).sum())

                tad = float(pd.to_numeric(sku_metrics.get("avg_daily_sales_sy"), errors="coerce").fillna(0).sum())
                # Approximate units_mtd and units_ytd from sales_orders
                from datetime import date as _date
                ref_end = end_date or _date.today()
                ytd_start = _date(ref_end.year, 1, 1)
                mtd_start = _date(ref_end.year, ref_end.month, 1)
                so = sales_orders.copy()
                if not so.empty and "order_entry_date" in so.columns:
                    so["order_entry_date"] = pd.to_datetime(so.get("order_entry_date"), errors="coerce")
                    um = float(so.loc[so["order_entry_date"].between(pd.to_datetime(mtd_start), pd.to_datetime(ref_end), inclusive="both"), "quantity_sy"].sum())
                    uy = float(so.loc[so["order_entry_date"].between(pd.to_datetime(ytd_start), pd.to_datetime(ref_end), inclusive="both"), "quantity_sy"].sum())
                else:
                    um = uy = 0.0
                import calendar as _cal
                dim = _cal.monthrange(ref_end.year, ref_end.month)[1]
                ed = max(ref_end.day, 1)
                ytd_turn = (tad * 365.0) / gi if gi > 0 else 0.0
                mtd_est_month = um * (dim / max(ed, 1))
                mtd_turn = (mtd_est_month * 12.0 / gi) if gi > 0 else 0.0
                return float(ytd_turn), float(mtd_turn)
            except Exception:
                return 0.0, 0.0

    ytd_turn_kpi, mtd_turn_kpi = _compute_kpi_turns()

    st.subheader("Key Metrics")
    # Add an extra slot for MTD Fill Rate
    metric_cols = st.columns(6)
    # Show saved per-CC target if available; otherwise indicate no target in red
    saved_target_text = None
    delta_color = "normal"
    try:
        if selected_cost_centers and len(selected_cost_centers) == 1:
            cc = str(selected_cost_centers[0])
            cc_target = get_target_for_cc(cc)
            if cc_target is not None:
                saved_target_text = f"Target {float(cc_target):.2f}"
            else:
                saved_target_text = "No target"
                delta_color = "inverse"  # attempt to render in red
        else:
            # Multiple or no CC context: treat as no specific target
            saved_target_text = "No target"
            delta_color = "inverse"
    except Exception:
        saved_target_text = "No target"
        delta_color = "inverse"

    metric_cols[0].metric(
        "YTD Turn",
        f"{ytd_turn_kpi:.2f}",
        delta=saved_target_text,
        delta_color=delta_color,
    )
    metric_cols[1].metric("MTD Turn", f"{mtd_turn_kpi:.2f}")
    # Dynamic backorder rate (Fill Rate counterpart) based on currently filtered SKUs
    total_orders = pd.to_numeric(sku_metrics.get("orders_count"), errors="coerce").fillna(0).sum()
    total_backorders = pd.to_numeric(sku_metrics.get("backorder_count"), errors="coerce").fillna(0).sum()
    dynamic_backorder_rate = (total_backorders / total_orders) if total_orders > 0 else 0.0
    # Show Fill Rate as requested (complement of backorder rate) but do not change underlying numbers
    dynamic_fill_rate = (1 - dynamic_backorder_rate) if total_orders > 0 else 0.0
    metric_cols[2].metric("Fill Rate", f"{dynamic_fill_rate:.1%}")

    # Month-To-Date Fill Rate: compute from sales_orders within current month or last full month (if toggled)
    try:
        from datetime import date as _date
        if isinstance(sales_orders, pd.DataFrame) and not sales_orders.empty and "order_entry_date" in sales_orders.columns:
            ref_end = end_date or _date.today()
            use_full_months_flag = bool(st.session_state.get("stockturn_use_full_months", False))
            if use_full_months_flag:
                _prev_year = ref_end.year if ref_end.month > 1 else (ref_end.year - 1)
                _prev_month = (ref_end.month - 1) if ref_end.month > 1 else 12
                mtd_start_calc = _date(_prev_year, _prev_month, 1)
                mtd_end_calc = _date(_prev_year, _prev_month, calendar.monthrange(_prev_year, _prev_month)[1])
            else:
                mtd_start_calc = _date(ref_end.year, ref_end.month, 1)
                mtd_end_calc = ref_end
            so_mtd = sales_orders.copy()
            so_mtd["order_entry_date"] = pd.to_datetime(so_mtd.get("order_entry_date"), errors="coerce")
            mask_mtd = so_mtd["order_entry_date"].between(pd.to_datetime(mtd_start_calc), pd.to_datetime(mtd_end_calc), inclusive="both")
            so_mtd = so_mtd.loc[mask_mtd].copy()
            # If backorder_flag missing (should be present from metrics_service), rebuild a quick flag
            if "backorder_flag" not in so_mtd.columns and "detail_line_status" in so_mtd.columns:
                detail_col = so_mtd.get("detail_line_status").fillna("").astype(str).str.strip().str.upper()
                so_mtd["backorder_flag"] = detail_col.isin(["R", "B"])
            # Distinct order lines by order_line_id when available to mirror metrics_service aggregation
            if "order_line_id" in so_mtd.columns:
                so_mtd = so_mtd.sort_values("order_entry_date").drop_duplicates(subset=["order_line_id"], keep="first")
            total_orders_mtd = len(so_mtd) if "order_line_id" not in so_mtd.columns else so_mtd["order_line_id"].nunique()
            if total_orders_mtd <= 0:
                mtd_fill_rate = float("nan")
            else:
                backorders_mtd = so_mtd.loc[so_mtd.get("backorder_flag") == True]
                # Use number of backordered order lines as numerator as per main fill rate logic
                backorders_mtd_count = len(backorders_mtd) if "order_line_id" not in so_mtd.columns else backorders_mtd["order_line_id"].nunique()
                mtd_fill_rate = 1 - (backorders_mtd_count / max(total_orders_mtd, 1))
        else:
            mtd_fill_rate = float("nan")
    except Exception:
        mtd_fill_rate = float("nan")
    metric_cols[3].metric(
        "MTD Fill Rate" if not bool(st.session_state.get("stockturn_use_full_months", False)) else "Last Full Month Fill Rate",
        "--" if not np.isfinite(mtd_fill_rate) else f"{mtd_fill_rate:.1%}"
    )

    metric_cols[4].metric("Days of Inventory", _format_days(summary.get("days_of_inventory", float("inf"))))
    metric_cols[5].metric("Runout Risk", str(summary.get("runout_sku_count", 0)))
    # Replace SKUs Tracked with Average Age of Inventory (weighted)
    # Weighted average at SKU level has already been computed; for the key metric,
    # compute an overall weighted average across filtered SKUs using inventory_sy as weights.
    inv_qty = pd.to_numeric(sku_metrics.get("inventory_sy"), errors="coerce").fillna(0)
    inv_age = pd.to_numeric(sku_metrics.get("inventory_age_days"), errors="coerce")
    mask = (inv_qty > 0) & inv_age.notna()
    overall_avg_age = float((inv_qty[mask] * inv_age[mask]).sum() / inv_qty[mask].sum()) if mask.any() else 0.0
    # Move Average Age below if you want to keep five metrics; here we keep it implicit to avoid layout overflow

    # Responsive charts (same as details dialog), scoped to the current filtered SKU set
    try:
        current_skus = set(sku_metrics.get("sku", pd.Series(dtype=str)).astype(str))
    except Exception:
        current_skus = set()
    st.subheader("Sales and PO activity")
    so_curr = sales_orders.copy()
    if current_skus:
        so_curr = so_curr[so_curr.get("sku").astype(str).isin(current_skus)]
    if not so_curr.empty:
        so_w = so_curr.copy()
        so_w["order_entry_date"] = pd.to_datetime(so_w.get("order_entry_date"), errors="coerce")
        so_w = so_w.dropna(subset=["order_entry_date"]).set_index("order_entry_date").sort_index()

        # Layout 2x2
        row1 = st.columns(2)
        # 1) Weekly sales quantity (SY)
        weekly_qty = so_w.resample("W")["quantity_sy"].sum().reset_index()
        fig_sales = px.line(weekly_qty, x="order_entry_date", y="quantity_sy", markers=True, title="Weekly Sales (SY)")
        row1[0].plotly_chart(fig_sales, use_container_width=True)

        # 2) Weekly orders & backorder rate
        weekly_orders = so_w.resample("W")["order_line_id"].nunique().reset_index(name="orders")
        weekly_bo = so_w.resample("W")["backorder_flag"].sum().reset_index(name="backorders")
        wo = weekly_orders.merge(weekly_bo, on="order_entry_date", how="left").fillna(0)
        wo["backorder_rate"] = np.where(wo["orders"] > 0, wo["backorders"] / wo["orders"], 0.0)
        fig_bo = px.bar(wo, x="order_entry_date", y="orders", title="Weekly Orders & Backorder Rate")
        fig_bo.add_scatter(x=wo["order_entry_date"], y=wo["backorder_rate"], mode="lines+markers", name="Backorder Rate", yaxis="y2")
        fig_bo.update_layout(yaxis2=dict(overlaying="y", side="right", tickformat=",.0%", title="Backorder Rate"))
        row1[1].plotly_chart(fig_bo, use_container_width=True)

        # Row 2
        row2 = st.columns(2)
        # 3) 30-day rolling avg daily sales
        daily_qty = so_w["quantity_sy"].resample("D").sum()
        ads30 = (daily_qty.rolling(30).sum() / 30.0).reset_index().rename(columns={"quantity_sy": "avg_daily_30"})
        fig_ads = px.line(ads30, x="order_entry_date", y="avg_daily_30", markers=False, title="Avg Daily Sales (30-day rolling)")
        row2[0].plotly_chart(fig_ads, use_container_width=True)
    else:
        st.info("No sales orders for the current filter selection.")

    # 4) Incoming PO quantity by ETA week
    po_curr = dashboard.get("purchase_orders", pd.DataFrame()).copy()
    if current_skus:
        po_curr = po_curr[po_curr.get("sku").astype(str).isin(current_skus)]
    if not po_curr.empty and "eta_date" in po_curr.columns:
        po_curr["eta_date"] = pd.to_datetime(po_curr.get("eta_date"), errors="coerce")
        # Filter to ETAs within ±12 months of today
        today_ts = pd.Timestamp(date.today())
        min_eta = today_ts - pd.DateOffset(months=12)
        max_eta = today_ts + pd.DateOffset(months=12)
        mask_eta = po_curr["eta_date"].between(min_eta, max_eta, inclusive="both")
        po_curr = po_curr.loc[mask_eta]
        if not po_curr.empty:
            po_curr = po_curr.dropna(subset=["eta_date"]).set_index("eta_date").sort_index()
            weekly_po = po_curr.resample("W")["quantity_sy"].sum().reset_index()
            if not weekly_po.empty:
                fig_po = px.bar(weekly_po, x="eta_date", y="quantity_sy", title="Incoming PO (SY) by ETA Week (±12 months)")
                # place to the right of ADS chart or alone if no sales
                if 'row2' in locals():
                    row2[1].plotly_chart(fig_po, use_container_width=True)
                else:
                    st.plotly_chart(fig_po, use_container_width=True)
            else:
                st.info("No purchase orders within ±12 months for the current filter selection.")
        else:
            st.info("No purchase orders within ±12 months for the current filter selection.")
    else:
        st.info("No purchase orders for the current filter selection.")

    # Trends charts removed per request

    # Preserve the current filtered set before narrowing to runout-only
    filtered_before_runout = sku_metrics.copy()

    t_sku = perf_counter()
    st.subheader("SKU Performance")
    # Scenario control: adjust reorder demand up/down by a whole-number percent.
    try:
        _pref_growth_pct = int(user_prefs.get("sku_growth_factor_pct", 0))
    except Exception:
        _pref_growth_pct = 0
    try:
        _pref_min_usable_roll_sy = float(user_prefs.get("sku_min_usable_roll_sy", 0.0))
    except Exception:
        _pref_min_usable_roll_sy = 0.0
    gcol1, gcol2, gcol3 = st.columns([3, 1, 2])
    with gcol1:
        growth_pct = int(
            st.number_input(
                "Growth factor (%)",
                min_value=-200,
                max_value=500,
                value=_pref_growth_pct,
                step=1,
                key="sku_growth_factor_pct_input",
                help="Whole-number percent applied to demand for planning reorder quantities. Example: 1 = +1%, -1 = -1%.",
            )
        )
    with gcol2:
        st.write("")
        st.write("")
        reset_growth = st.button("Reset to Default", key="sku_growth_factor_reset")
    with gcol3:
        min_usable_roll_sy = float(
            st.number_input(
                "Exclude rolls below (SY)",
                min_value=0.0,
                max_value=10000.0,
                value=float(_pref_min_usable_roll_sy),
                step=1.0,
                format="%.2f",
                key="sku_min_usable_roll_sy_input",
                help="Inventory from roll rows below this SY size is excluded from SKU Performance DR/Reorder inventory math.",
            )
        )
    if reset_growth:
        growth_pct = 0
        st.session_state["sku_growth_factor_pct_input"] = 0
    if growth_pct != _pref_growth_pct:
        try:
            user_prefs["sku_growth_factor_pct"] = int(growth_pct)
            _save_prefs(user_prefs)
        except Exception:
            pass
    if abs(min_usable_roll_sy - _pref_min_usable_roll_sy) > 1e-9:
        try:
            user_prefs["sku_min_usable_roll_sy"] = float(min_usable_roll_sy)
            _save_prefs(user_prefs)
        except Exception:
            pass
    growth_scale = max(0.0, 1.0 + (float(growth_pct) / 100.0))
    if growth_pct == 0:
        st.caption("Growth factor: default (no adjustment).")
    elif growth_pct > 0:
        st.caption(f"Growth factor applied: +{growth_pct}% (higher reorder demand).")
    else:
        st.caption(f"Growth factor applied: {growth_pct}% (lower reorder demand).")
    if min_usable_roll_sy > 0:
        st.caption(f"Inventory filter: excluding roll rows below {min_usable_roll_sy:.2f} SY in SKU Performance reorder calculations.")
    else:
        st.caption("Inventory filter: none (all roll sizes included).")

    # Build scenario demand columns used only for planning math (DR/Reorder/list inclusion).
    if not sku_metrics.empty:
        base_avg = pd.to_numeric(sku_metrics.get("avg_daily_sales_sy"), errors="coerce").fillna(0.0)
        base_target = pd.to_numeric(sku_metrics.get("target_on_hand_sy"), errors="coerce").fillna(0.0)
        sku_metrics["adj_avg_daily_sales_sy"] = base_avg * growth_scale
        sku_metrics["adj_target_on_hand_sy"] = base_target * growth_scale

        # Recompute scenario reorder so list membership can expand/shrink based on the growth factor.
        inv_s = pd.to_numeric(sku_metrics.get("inventory_sy"), errors="coerce").fillna(0.0)
        assumed_s = pd.to_numeric(sku_metrics.get("assumed_on_order_sy"), errors="coerce").fillna(0.0)
        bo_s = pd.to_numeric(sku_metrics.get("backorder_qty_sy"), errors="coerce").fillna(0.0)
        net_s = inv_s + assumed_s
        safety_s = sku_metrics["adj_avg_daily_sales_sy"] * 7.0
        sku_metrics["reorder_qty_sy"] = np.maximum(
            (sku_metrics["adj_target_on_hand_sy"] + safety_s) - net_s, 0.0
        ) + bo_s

    # Prefer to show at-risk SKUs (runout_risk). We'll keep rows with Reorder > 0,
    # but also include explicitly searched SKUs and any stocking items with zero inventory.
    # Add days-until-runout using net inventory (inventory + on order) and avg daily sales
    if not sku_metrics.empty:
        _avg = pd.to_numeric(sku_metrics.get("adj_avg_daily_sales_sy"), errors="coerce").fillna(0.0)
        _net = pd.to_numeric(sku_metrics.get("net_inventory_sy"), errors="coerce").fillna(0.0)
        _bo = pd.to_numeric(sku_metrics.get("backorder_qty_sy"), errors="coerce").fillna(0.0)
        _net_after_bo = np.maximum(_net - _bo, 0.0)
        sku_metrics["days_until_runout"] = np.where(_avg > 0, _net_after_bo / _avg, float("inf"))
        # Only show SKUs where days_until_runout < min(lead_time*2, 120)
        lt = pd.to_numeric(sku_metrics.get("lead_time_days"), errors="coerce")
        dur = pd.to_numeric(sku_metrics.get("days_until_runout"), errors="coerce")
        threshold = np.minimum(lt * 2, 120)
        mask_lt = dur < threshold
    else:
        mask_lt = pd.Series(False, index=sku_metrics.index)
    # Keep rows with positive reorder; also keep searched SKUs and stocking items with zero inventory AND zero PO
    rq = pd.to_numeric(sku_metrics.get("reorder_qty_sy"), errors="coerce").fillna(0.0)
    mask_any_reorder = rq > 0
    try:
        search_mask_metrics = sku_metrics.get("sku", pd.Series(dtype=str)).astype(str).str.contains(sku_search, case=False, na=False) if sku_search else pd.Series(False, index=sku_metrics.index)
    except Exception:
        search_mask_metrics = pd.Series(False, index=sku_metrics.index)
    try:
        stocking_mask_metrics = sku_metrics.get("sku", pd.Series(dtype=str)).astype(str).str.strip().str.upper().isin(stocking_base_sku_set)
    except Exception:
        stocking_mask_metrics = pd.Series(False, index=sku_metrics.index)
    zero_inv_metrics = pd.to_numeric(sku_metrics.get("inventory_sy"), errors="coerce").fillna(0.0) <= 0.0
    # Determine zero PO using pending PO when available, else on_order_sy
    po_pending_metrics = pd.to_numeric(sku_metrics.get("po_pending_qty"), errors="coerce") if "po_pending_qty" in sku_metrics.columns else pd.Series(np.nan, index=sku_metrics.index)
    on_order_metrics = pd.to_numeric(sku_metrics.get("on_order_sy"), errors="coerce").fillna(0.0)
    po_eff_metrics = po_pending_metrics.fillna(on_order_metrics).fillna(0.0)
    zero_po_metrics = po_eff_metrics <= 0.0
    mask_show_always_metrics = search_mask_metrics | (stocking_mask_metrics & zero_inv_metrics & zero_po_metrics)
    # When searching, show ALL matching SKUs; otherwise apply the reorder/stocking logic
    if sku_search:
        sku_metrics = sku_metrics.loc[search_mask_metrics].copy()
    else:
        sku_metrics = sku_metrics.loc[mask_any_reorder | mask_show_always_metrics].copy()
    if not sku_metrics.empty:
        sku_metrics.sort_values(
            by=["runout_risk", "cover_required_sy", "reorder_qty_sy", "avg_daily_sales_sy"],
            ascending=[False, False, False, False],
            inplace=True,
        )

    # Compute roll statistics per SKU
    # - avg_roll_size_sy = inventory_sy / distinct roll count
    # - max_roll_size_sy = largest single roll (inventory_sy) for that SKU
    try:
        inv_all = dashboard.get("inventory", pd.DataFrame())
        roll_counts = None
        max_roll_sizes = None
        # Table-only inventory override map for CC 010/012 after excluding low-available rolls (<240)
        # and optionally excluding tiny roll quantities by user threshold.
        inv_override_sum_by_sku = None
        if isinstance(inv_all, pd.DataFrame) and not inv_all.empty and "sku" in inv_all.columns:
            inv_ok = inv_all.copy()
            # Prefer positive inventory rows; tolerate missing roll_number
            inv_ok["inventory_sy"] = pd.to_numeric(inv_ok.get("inventory_sy"), errors="coerce").fillna(0.0)
            inv_ok = inv_ok[inv_ok["inventory_sy"] > 0]
            if "roll_number" in inv_ok.columns:
                # Count distinct roll numbers per SKU
                roll_counts = (
                    inv_ok.dropna(subset=["sku"]).assign(roll_number=inv_ok.get("roll_number").astype(str))
                    .groupby(inv_ok.get("sku").astype(str))[["roll_number"]]
                    .nunique()["roll_number"]
                )
            else:
                # Fallback: count rows per SKU
                roll_counts = inv_ok.groupby(inv_ok.get("sku").astype(str)).size()

            # Compute max roll size per SKU from per-row inventory_sy, excluding RET locations and status_code containing 'I'
            try:
                # Identify location column if present and exclude rows starting with 'RET'
                rloc_col = (
                    "RLOC1" if "RLOC1" in inv_ok.columns else (
                        "rloc1" if "rloc1" in inv_ok.columns else (
                            "RCLOC1" if "RCLOC1" in inv_ok.columns else (
                                "rcloc1" if "rcloc1" in inv_ok.columns else (
                                    "location" if "location" in inv_ok.columns else (
                                        "Location" if "Location" in inv_ok.columns else None
                                    )
                                )
                            )
                        )
                    )
                )
                # Identify status_code/RCODE@ column if present and exclude rows where it contains 'I'
                rcode_col = None
                for _cand in ["status_code", "Status_Code", "status", "RCODE@", "rcode@", "RCODE", "rcode"]:
                    if _cand in inv_ok.columns:
                        rcode_col = _cand
                        break
                mask_ret = inv_ok.get(rloc_col).astype(str).str.upper().str.startswith("RET") if rloc_col is not None else False
                mask_rcode_i = inv_ok.get(rcode_col).astype(str).str.upper().str.strip().str.contains("I", na=False) if rcode_col is not None else False
                mask_small_roll = inv_ok.get("inventory_sy", pd.Series(index=inv_ok.index, dtype=float)).astype(float) < float(min_usable_roll_sy) if float(min_usable_roll_sy) > 0 else False
                mask_exclude = mask_ret | mask_rcode_i | mask_small_roll
                inv_for_max = inv_ok.loc[~mask_exclude].copy()
                max_roll_sizes = (
                    inv_for_max.dropna(subset=["sku"]).groupby(inv_for_max.get("sku").astype(str))["inventory_sy"].max()
                )
            except Exception:
                max_roll_sizes = None

            # Build a per-SKU sum that excludes rolls with Available < 240 for cost centers 010 and 012 only.
            # This affects ONLY the displayed Inv column in the table below; it does not change metrics elsewhere.
            try:
                # Map SKU -> cost center using items_df; normalize to 3-digit strings (e.g., '010').
                cc_series = pd.Series(dtype=str)
                if isinstance(items_df, pd.DataFrame) and not items_df.empty and {"sku", "cost_center"}.issubset(items_df.columns):
                    cc_series = (
                        items_df[["sku", "cost_center"]]
                        .dropna(subset=["sku"])
                        .assign(cost_center=lambda d: d["cost_center"].astype(str).str.strip().str.zfill(3))
                        .set_index("sku")["cost_center"]
                    )
                inv_tmp = inv_all.copy()
                inv_tmp["sku"] = inv_tmp.get("sku").astype(str)
                # Attach cost center for each roll row
                if not cc_series.empty:
                    inv_tmp["__cc__"] = inv_tmp["sku"].map(cc_series)
                else:
                    inv_tmp["__cc__"] = None
                # Always exclude RET locations and status_code containing 'I'
                try:
                    rloc_col = (
                        "RLOC1" if "RLOC1" in inv_tmp.columns else (
                            "rloc1" if "rloc1" in inv_tmp.columns else (
                                "RCLOC1" if "RCLOC1" in inv_tmp.columns else (
                                    "rcloc1" if "rcloc1" in inv_tmp.columns else (
                                        "location" if "location" in inv_tmp.columns else (
                                            "Location" if "Location" in inv_tmp.columns else None
                                        )
                                    )
                                )
                            )
                        )
                    )
                    rcode_col = None
                    for _cand in ["status_code", "Status_Code", "status", "RCODE@", "rcode@", "RCODE", "rcode"]:
                        if _cand in inv_tmp.columns:
                            rcode_col = _cand
                            break
                    mask_ret = inv_tmp.get(rloc_col).astype(str).str.upper().str.startswith("RET") if rloc_col is not None else False
                    mask_rcode_i = inv_tmp.get(rcode_col).astype(str).str.upper().str.strip().str.contains("I", na=False) if rcode_col is not None else False
                    mask_exclude = mask_ret | mask_rcode_i
                    inv_tmp = inv_tmp.loc[~mask_exclude].copy()
                except Exception:
                    pass
                # Optionally exclude tiny rolls globally before CC-specific low-available logic.
                if float(min_usable_roll_sy) > 0:
                    inv_tmp = inv_tmp.loc[
                        pd.to_numeric(inv_tmp.get("inventory_sy"), errors="coerce").fillna(0.0) >= float(min_usable_roll_sy)
                    ].copy()
                # Exclusion criteria: CC in {010,012} AND Available < 240
                avail = pd.to_numeric(inv_tmp.get("available_quantity"), errors="coerce")
                cc_mask = inv_tmp["__cc__"].astype(str).str.zfill(3).isin(["010", "012"])
                low_avail = avail < 240
                # Keep all rows except those matching both conditions
                inv_tmp = inv_tmp.loc[~(cc_mask & low_avail)].copy()
                # Sum per SKU using already-computed inventory_sy per roll
                inv_override_sum_by_sku = (
                    inv_tmp.dropna(subset=["sku"]).groupby("sku")["inventory_sy"].sum()
                )
            except Exception:
                inv_override_sum_by_sku = None
        # Precompute per-SKU sum excluding RET and status_code containing 'I' (for all SKUs)
        try:
            inv_nr = inv_all.copy()
            rloc_col = (
                "RLOC1" if "RLOC1" in inv_nr.columns else (
                    "rloc1" if "rloc1" in inv_nr.columns else (
                        "RCLOC1" if "RCLOC1" in inv_nr.columns else (
                            "rcloc1" if "rcloc1" in inv_nr.columns else (
                                "location" if "location" in inv_nr.columns else (
                                    "Location" if "Location" in inv_nr.columns else None
                                )
                            )
                        )
                    )
                )
            )
            rcode_col = None
            for _cand in ["status_code", "Status_Code", "status", "RCODE@", "rcode@", "RCODE", "rcode"]:
                if _cand in inv_nr.columns:
                    rcode_col = _cand
                    break
            mask_ret_all = inv_nr.get(rloc_col).astype(str).str.upper().str.startswith("RET") if rloc_col is not None else False
            mask_rcode_all = inv_nr.get(rcode_col).astype(str).str.upper().str.strip().str.contains("I", na=False) if rcode_col is not None else False
            mask_exclude_all = mask_ret_all | mask_rcode_all
            inv_nr = inv_nr.loc[~mask_exclude_all].copy()
            if float(min_usable_roll_sy) > 0:
                inv_nr = inv_nr.loc[
                    pd.to_numeric(inv_nr.get("inventory_sy"), errors="coerce").fillna(0.0) >= float(min_usable_roll_sy)
                ].copy()
            inv_exclude_ret_sum_by_sku = (
                inv_nr.dropna(subset=["sku"]).groupby(inv_nr.get("sku").astype(str))["inventory_sy"].sum()
            )
        except Exception:
            inv_exclude_ret_sum_by_sku = None
        if roll_counts is not None:
            # Align to sku_metrics and compute avg roll size
            sku_key_series = sku_metrics.get("sku", pd.Series(dtype=str)).astype(str)
            counts_aligned = sku_key_series.map(roll_counts).fillna(0).astype(float)
            inv_sy = pd.to_numeric(sku_metrics.get("inventory_sy"), errors="coerce").fillna(0.0)
            sku_metrics["avg_roll_size_sy"] = np.where(counts_aligned > 0, inv_sy / counts_aligned, np.nan)
        else:
            sku_metrics["avg_roll_size_sy"] = np.nan

        # Align and attach max roll size
        if max_roll_sizes is not None:
            sku_key_series = sku_metrics.get("sku", pd.Series(dtype=str)).astype(str)
            sku_metrics["max_roll_size_sy"] = sku_key_series.map(max_roll_sizes).astype(float)
        else:
            sku_metrics["max_roll_size_sy"] = np.nan
    except Exception:
        sku_metrics["avg_roll_size_sy"] = np.nan
        sku_metrics["max_roll_size_sy"] = np.nan

    display_columns = [
        "sku",
        "sku_description",
        "sku_rating",
        "quantity_sy",
        "inventory_sy",
        "max_roll_size_sy",  # largest single roll
        # removed average roll size from table per request
        "on_order_sy",
        "partial_received_po",
        "backorder_qty_sy",  # show total backorders (SY) as BO
        "avg_daily_sales_sy",
        "jstock",
        "lead_time_days",
        # "days_of_inventory",  # hidden per request
        "days_until_runout",
        "reorder_qty_sy",
    ]
    available_columns = [col for col in display_columns if col in sku_metrics.columns]
    # Start from the metrics frame but override inventory to exclude RET locations globally,
    # then apply CC 010/012-specific low-available filter; recompute dependent fields with filtered inventory
    sku_table = sku_metrics[available_columns].copy()
    try:
        # Determine which rows in the table belong to CC 010/012
        cc_map_for_table = sku_metrics.set_index("sku")["cost_center"].astype(str).str.strip().str.zfill(3) if "cost_center" in sku_metrics.columns else pd.Series(dtype=str)
        # Apply RET exclusion override for all rows when available
        if inv_exclude_ret_sum_by_sku is not None and not isinstance(inv_exclude_ret_sum_by_sku, pd.DataFrame):
            override_all = sku_table["sku"].map(inv_exclude_ret_sum_by_sku).astype(float)
            if "inventory_sy" in sku_table.columns:
                sku_table.loc[:, "inventory_sy"] = override_all.fillna(0.0)
        if inv_override_sum_by_sku is not None and not isinstance(inv_override_sum_by_sku, pd.DataFrame):
            # Map per-SKU overridden inventory sums to the table rows
            override_vals = sku_table["sku"].map(inv_override_sum_by_sku).astype(float)
            if not cc_map_for_table.empty and "inventory_sy" in sku_table.columns:
                cc_mask_tbl = sku_table["sku"].map(cc_map_for_table).isin(["010", "012"])
                # For CC 010/012 rows, replace inventory_sy with the filtered sum (NaN -> original value)
                sku_table.loc[cc_mask_tbl, "inventory_sy"] = override_vals.loc[cc_mask_tbl].fillna(sku_table.loc[cc_mask_tbl, "inventory_sy"]) 

        # Recompute table-dependent fields (DR and Reorder) for all rows using filtered inventory
        metrics_by_sku = sku_metrics.set_index("sku")
        avg_map = metrics_by_sku.get("adj_avg_daily_sales_sy") if "adj_avg_daily_sales_sy" in metrics_by_sku.columns else (
            metrics_by_sku.get("avg_daily_sales_sy") if "avg_daily_sales_sy" in metrics_by_sku.columns else pd.Series(dtype=float)
        )
        assumed_map = metrics_by_sku.get("assumed_on_order_sy") if "assumed_on_order_sy" in metrics_by_sku.columns else pd.Series(dtype=float)
        target_map = metrics_by_sku.get("adj_target_on_hand_sy") if "adj_target_on_hand_sy" in metrics_by_sku.columns else (
            metrics_by_sku.get("target_on_hand_sy") if "target_on_hand_sy" in metrics_by_sku.columns else pd.Series(dtype=float)
        )
        bo_map = metrics_by_sku.get("backorder_qty_sy") if "backorder_qty_sy" in metrics_by_sku.columns else pd.Series(dtype=float)
        avg_vals_all = sku_table["sku"].map(avg_map).astype(float)
        assumed_vals_all = sku_table["sku"].map(assumed_map).astype(float).fillna(0.0)
        target_vals_all = sku_table["sku"].map(target_map).astype(float)
        bo_vals_all = sku_table["sku"].map(bo_map).astype(float).fillna(0.0)
        inv_filtered_all = sku_table["inventory_sy"].astype(float)
        net_filtered_all = inv_filtered_all + assumed_vals_all
        net_after_bo_all = np.maximum(net_filtered_all - bo_vals_all, 0.0)
        days_all = np.where(avg_vals_all > 0, net_after_bo_all / np.maximum(avg_vals_all, 1e-9), float("inf"))
        if "days_until_runout" in sku_table.columns:
            sku_table.loc[:, "days_until_runout"] = pd.to_numeric(days_all, errors="coerce")
        safety_all = avg_vals_all * 7.0
        rq_all = np.maximum((target_vals_all + safety_all) - net_filtered_all, 0.0) + bo_vals_all
        if "reorder_qty_sy" in sku_table.columns:
            sku_table.loc[:, "reorder_qty_sy"] = pd.to_numeric(rq_all, errors="coerce")
    except Exception:
        # Best-effort; fall back to original values on any issue
        pass
    # Final enforcement: keep rows with positive reorder OR searched SKUs OR stocking items with zero inventory AND zero PO
    try:
        rq_final = pd.to_numeric(sku_table.get("reorder_qty_sy"), errors="coerce").fillna(0.0)
        try:
            search_mask_tbl = sku_table.get("sku", pd.Series(dtype=str)).astype(str).str.contains(sku_search, case=False, na=False) if sku_search else pd.Series(False, index=sku_table.index)
        except Exception:
            search_mask_tbl = pd.Series(False, index=sku_table.index)
        try:
            stocking_mask_tbl = sku_table.get("sku", pd.Series(dtype=str)).astype(str).str.strip().str.upper().isin(stocking_base_sku_set)
        except Exception:
            stocking_mask_tbl = pd.Series(False, index=sku_table.index)
        zero_inv_tbl = pd.to_numeric(sku_table.get("inventory_sy"), errors="coerce").fillna(0.0) <= 0.0
        # on_order_sy in table has already been overwritten by po_pending_qty when available
        zero_po_tbl = pd.to_numeric(sku_table.get("on_order_sy"), errors="coerce").fillna(0.0) <= 0.0
        mask_show_always_tbl = search_mask_tbl | (stocking_mask_tbl & zero_inv_tbl & zero_po_tbl)
        sku_table = sku_table.loc[(rq_final > 0) | mask_show_always_tbl].copy()
    except Exception:
        pass
    numeric_columns = sku_table.select_dtypes(include=["number"]).columns
    sku_table[numeric_columns] = sku_table[numeric_columns].round(2)
    # Prefer OPENPO_D pending quantity for PO display if present
    try:
        if "po_pending_qty" in sku_metrics.columns and "on_order_sy" in sku_metrics.columns:
            # Build mapping and overwrite in the table copy prior to renaming
            pend_map = sku_metrics.set_index("sku")["po_pending_qty"].astype(float)
            sku_table["on_order_sy"] = sku_table["sku"].map(pend_map).fillna(sku_table["on_order_sy"]).fillna(0.0)
    except Exception:
        pass
    # Apply table-stage filter: keep (BO > PO OR DR < min(LT*2, 120)) AND Reorder > 0,
    # but always include searched SKUs and stocking items with zero inventory AND zero PO
    try:
        bo_vals = pd.to_numeric(sku_table.get("backorder_qty_sy"), errors="coerce").fillna(0.0)
        po_vals = pd.to_numeric(sku_table.get("on_order_sy"), errors="coerce").fillna(0.0)
        lt_vals = pd.to_numeric(sku_table.get("lead_time_days"), errors="coerce")
        dr_vals = pd.to_numeric(sku_table.get("days_until_runout"), errors="coerce")
        thresh = np.minimum(lt_vals * 2.0, 120.0)
        mask_bo_gt_po = bo_vals > po_vals
        mask_dr_lt = dr_vals < thresh
        reorder_vals = pd.to_numeric(sku_table.get("reorder_qty_sy"), errors="coerce").fillna(0.0)
        mask_reorder = reorder_vals > 0
        try:
            search_mask_tbl2 = sku_table.get("sku", pd.Series(dtype=str)).astype(str).str.contains(sku_search, case=False, na=False) if sku_search else pd.Series(False, index=sku_table.index)
        except Exception:
            search_mask_tbl2 = pd.Series(False, index=sku_table.index)
        try:
            stocking_mask_tbl2 = sku_table.get("sku", pd.Series(dtype=str)).astype(str).str.strip().str.upper().isin(stocking_base_sku_set)
        except Exception:
            stocking_mask_tbl2 = pd.Series(False, index=sku_table.index)
        zero_inv_tbl2 = pd.to_numeric(sku_table.get("inventory_sy"), errors="coerce").fillna(0.0) <= 0.0
        zero_po_tbl2 = pd.to_numeric(sku_table.get("on_order_sy"), errors="coerce").fillna(0.0) <= 0.0
        mask_show_always_tbl2 = search_mask_tbl2 | (stocking_mask_tbl2 & zero_inv_tbl2 & zero_po_tbl2)
        sku_table = sku_table.loc[((mask_bo_gt_po | mask_dr_lt) & mask_reorder) | mask_show_always_tbl2].copy()
    except Exception:
        # If any issue, fall back to existing Reorder > 0 enforcement (already applied earlier)
        pass
    # Rename columns for display: drop _sy suffixes, rename lead_time_days and sku_rating
    rename_map = {
        "sku_description": "desc",
        "quantity_sy": "qty",
        "inventory_sy": "Inv",
        "max_roll_size_sy": "Max",
        # avg_roll_size_sy removed from view
        "backorder_qty_sy": "BO",
        "on_order_sy": "PO",
        "partial_received_po": "partial",
        "avg_daily_sales_sy": "ADS",
        "reorder_qty_sy": "Reorder",
        "lead_time_days": "LT",
        "sku_rating": "rating",
        "jstock": "Jstk",
        "days_until_runout": "DR",
    }
    sku_table.rename(columns=rename_map, inplace=True)

    # Ensure column order keeps Max and avg next to Inv
    try:
        desired_order = [
            c for c in [
                "sku",
                "desc",
                "rating",
                "qty",
                "Inv",
                "Max",
                "PO",
                "partial",
                "BO",
                "ADS",
                "Jstk",
                "LT",
                "DR",
                "Reorder",
            ] if c in sku_table.columns
        ]
        sku_table = sku_table[desired_order]
    except Exception:
        pass

    # Replace the static table with an editable view that includes a per-row Details selector
    if not sku_table.empty:
        # Persist a copy for combined export (Excel multi-sheet)
        try:
            st.session_state["_export_sku_performance"] = sku_table.copy()
        except Exception:
            st.session_state["_export_sku_performance"] = sku_table
        sku_table_edit = sku_table.copy()
        # Ensure a boolean column for selecting details
        select_col = "__open_details__"
        if select_col not in sku_table_edit.columns:
            sku_table_edit.insert(0, select_col, False)
        editor_version = int(st.session_state.get("_editor_version", 0))
        editor_key = f"sku_perf_editor_{editor_version}"
        edited = st.data_editor(
            sku_table_edit,
            use_container_width=True,
            hide_index=True,
            num_rows="fixed",
            column_config={
                select_col: st.column_config.CheckboxColumn(
                    "Details",
                    help="Open details for this SKU",
                    default=False,
                )
            },
            key=editor_key,
        )
        # Pre-slice and cache datasets for the SKUs visible in this table for faster details rendering
        try:
            visible_skus = set(sku_table_edit.get("sku", pd.Series(dtype=str)).astype(str))
            # Items slice
            if isinstance(items_df, pd.DataFrame) and not items_df.empty:
                st.session_state["_perf_items_df"] = items_df[items_df.get("sku").astype(str).isin(visible_skus)].copy()
            # Sales orders slice (already filtered by applied date range earlier)
            if isinstance(sales_orders, pd.DataFrame) and not sales_orders.empty:
                so_slice = sales_orders[sales_orders.get("sku").astype(str).isin(visible_skus)].copy()
                st.session_state["_perf_sales_orders"] = so_slice
            # Purchase orders slice
            po_all = dashboard.get("purchase_orders", pd.DataFrame())
            if isinstance(po_all, pd.DataFrame) and not po_all.empty:
                po_slice = po_all[po_all.get("sku").astype(str).isin(visible_skus)].copy()
                st.session_state["_perf_po_orders"] = po_slice
        except Exception:
            pass

        if SHOW_OVERVIEW_DEBUG:
            with st.expander("Debug: Reorder inputs (per SKU)"):
                try:
                    # Show core inputs used in reorder calculation for quick verification
                    debug_cols = [
                        c for c in [
                            "sku","inventory_sy","on_order_sy","net_inventory_sy",
                            "assumed_on_order_sy","partial_received_po","backorder_qty_sy",
                            "avg_daily_sales_sy","lead_time_days","seasonal_window_multiplier",
                            "cover_required_sy","jstock","target_on_hand_sy","reorder_qty_sy"
                        ] if c in sku_metrics.columns
                    ]
                    st.dataframe(sku_metrics[debug_cols].round(2), use_container_width=True, hide_index=True)
                except Exception:
                    st.write("Unable to render debug inputs.")
        # If any row is checked, open the first one's details and immediately schedule uncheck
        try:
            chosen = edited[edited[select_col] == True]
            if not chosen.empty:
                chosen_sku = str(chosen.iloc[0].get("sku", "")).strip()
                if chosen_sku:
                    dlg = getattr(st, "dialog", None)
                    if callable(dlg):
                        @dlg("Details")
                        def _details_dialog_inline():
                            _show_details_body("sku", chosen_sku)
                        _details_dialog_inline()
                    else:
                        st.subheader("Details")
                        _show_details_body("sku", chosen_sku)
                    # Bump editor version so the checkbox appears unchecked after the dialog is closed.
                    st.session_state["_editor_version"] = editor_version + 1
        except Exception:
            pass
    else:
        st.info("No SKUs currently flagged as low stock based on lead time coverage.")

    # Companion table: A/B-rated SKUs not shown above, with small max roll size.
    st.markdown("#### A/B Rated SKUs Missing from SKU Performance (Max < 75)")
    st.caption(
        "Surfaces A/B items where inventory may be split across many small rolls that are less usable in practice."
    )
    try:
        shown_sku_norm = set(
            sku_table.get("sku", pd.Series(dtype=str)).astype(str).str.strip().str.upper().tolist()
        )
    except Exception:
        shown_sku_norm = set()

    try:
        ab_missing = sku_metrics.copy()
        if not ab_missing.empty:
            rating_vals = (
                ab_missing.get("sku_rating", pd.Series(index=ab_missing.index, dtype=str))
                .astype(str)
                .str.strip()
                .str.upper()
            )
            max_vals = pd.to_numeric(ab_missing.get("max_roll_size_sy"), errors="coerce")
            sku_norm_vals = (
                ab_missing.get("sku", pd.Series(index=ab_missing.index, dtype=str))
                .astype(str)
                .str.strip()
                .str.upper()
            )
            mask_ab = rating_vals.isin(["A", "B"])
            mask_small_max = max_vals.lt(75)
            mask_not_in_main = ~sku_norm_vals.isin(shown_sku_norm)
            ab_missing = ab_missing.loc[mask_ab & mask_small_max & mask_not_in_main].copy()

            if "po_pending_qty" in sku_metrics.columns and "on_order_sy" in ab_missing.columns:
                try:
                    pend_map_ab = sku_metrics.set_index("sku")["po_pending_qty"].astype(float)
                    ab_missing["on_order_sy"] = (
                        ab_missing["sku"].map(pend_map_ab).fillna(ab_missing["on_order_sy"]).fillna(0.0)
                    )
                except Exception:
                    pass

            numeric_cols_ab = ab_missing.select_dtypes(include=["number"]).columns
            ab_missing[numeric_cols_ab] = ab_missing[numeric_cols_ab].round(2)
            ab_missing.rename(columns=rename_map, inplace=True)

            desired_cols_ab = [
                c
                for c in [
                    "sku",
                    "desc",
                    "rating",
                    "qty",
                    "Inv",
                    "Max",
                    "PO",
                    "partial",
                    "BO",
                    "ADS",
                    "Jstk",
                    "LT",
                    "DR",
                    "Reorder",
                ]
                if c in ab_missing.columns
            ]
            ab_missing = ab_missing[desired_cols_ab]

            sort_cols_ab = [c for c in ["rating", "Max", "Reorder", "sku"] if c in ab_missing.columns]
            if sort_cols_ab:
                sort_asc = [True, True, False, True][: len(sort_cols_ab)]
                ab_missing = ab_missing.sort_values(sort_cols_ab, ascending=sort_asc, kind="mergesort")
        else:
            ab_missing = pd.DataFrame()
    except Exception:
        ab_missing = pd.DataFrame()

    if ab_missing.empty:
        st.info("No A/B-rated SKUs were found outside the current SKU Performance table with Max < 100.")
    else:
        try:
            st.session_state["_export_sku_performance_missing_ab"] = ab_missing.copy()
        except Exception:
            st.session_state["_export_sku_performance_missing_ab"] = ab_missing

        ab_edit = ab_missing.copy()
        missing_select_col = "__open_details_missing__"
        if missing_select_col not in ab_edit.columns:
            ab_edit.insert(0, missing_select_col, False)

        missing_editor_version = int(st.session_state.get("_editor_version_missing", 0))
        missing_editor_key = f"sku_perf_missing_editor_{missing_editor_version}"
        edited_missing = st.data_editor(
            ab_edit,
            use_container_width=True,
            hide_index=True,
            num_rows="fixed",
            column_config={
                missing_select_col: st.column_config.CheckboxColumn(
                    "Details",
                    help="Open details for this SKU",
                    default=False,
                )
            },
            key=missing_editor_key,
        )

        try:
            chosen_missing = edited_missing[edited_missing[missing_select_col] == True]
            if not chosen_missing.empty:
                chosen_sku_missing = str(chosen_missing.iloc[0].get("sku", "")).strip()
                if chosen_sku_missing:
                    dlg = getattr(st, "dialog", None)
                    if callable(dlg):
                        @dlg("Details")
                        def _details_dialog_missing_inline():
                            _show_details_body("sku", chosen_sku_missing)

                        _details_dialog_missing_inline()
                    else:
                        st.subheader("Details")
                        _show_details_body("sku", chosen_sku_missing)
                    st.session_state["_editor_version_missing"] = missing_editor_version + 1
        except Exception:
            pass
    ui_timings["ui_build_sku_performance"] = perf_counter() - t_sku

    if not sku_metrics.empty:
        # Download as CSV with a filename prompt and a brief column description
        with st.expander("Download SKU Performance as CSV"):
            default_name = "sku_performance.csv"
            fname = st.text_input("File name", value=default_name, help="Enter a name for the CSV file.")
            if not fname.strip().lower().endswith(".csv"):
                fname = (fname.strip() or default_name).rstrip(".") + ".csv"
            # Column descriptions for the abbreviated headers
            col_desc = [
                "# Inventory Dashboard - SKU Performance",
                "# Columns:",
                "# sku: Item number",
                "# desc: SKU description",
                "# rating: SKU rating tier (A-D)",
                "# qty: Total shipped quantity over the analysis window (SY)",
                "# Inv: On-hand inventory (SY). Excludes RET locations and status_code containing 'I' globally. For CC 010/012 rows, also excludes rolls with Available < 240.",
                "# Max: Largest single roll size (SY)",
                "# PO: Effective purchase orders pending (SY). Prefers OPENPO_D pending; falls back to legacy on-order.",
                "# partial: Partially received PO quantity (SY)",
                "# ADS: Average daily sales (SY/day)",
                "# Jstk: JSTOCK target (SY)",
                "# LT: Lead time in days",
                    "# DR: Days until runout = (net inventory - BO) / ADS",
                "# Reorder: Recommended reorder qty (SY) from seasonal/JSTOCK target + safety",
            ]
            try:
                csv_body = sku_table.to_csv(index=False)
                csv_with_desc = ("\n".join(col_desc) + "\n" + csv_body).encode("utf-8")
            except Exception:
                csv_with_desc = sku_table.to_csv(index=False).encode("utf-8")
            st.download_button(
                label="Download as CSV",
                data=csv_with_desc,
                file_name=fname,
                mime="text/csv",
            )

        if SHOW_OVERVIEW_DEBUG:
            with st.expander("Debug: Lookup a SKU (inventory and aliases)"):
                q = st.text_input("Enter SKU to inspect", key="dbg_lookup_sku")
                if q:
                    qn = str(q).strip()
                    try:
                        inv_df = dashboard.get("inventory", pd.DataFrame())
                        items_df = dashboard.get("items", pd.DataFrame())
                        st.write("Raw inventory rows for SKU after alias mapping:")
                        st.dataframe(inv_df[inv_df.get("sku").astype(str) == qn].head(50), use_container_width=True, hide_index=True)
                        if "iixref" in items_df.columns:
                            st.write("Alias map rows (IIXREF not blank):")
                            st.dataframe(items_df[items_df.get("iixref").astype(str).str.strip() != ""][
                                ["sku","iixref","sku_description"]
                            ].head(50), use_container_width=True, hide_index=True)
                    except Exception:
                        st.write("Unable to show lookup data.")
                # Optional: search by roll numbers
                rn = st.text_input("Enter roll numbers (comma-separated)", key="dbg_lookup_rolls")
                if rn:
                    try:
                        inv_df = dashboard.get("inventory", pd.DataFrame())
                        rolls = [s.strip() for s in rn.split(",") if s.strip()]
                        if not inv_df.empty and "roll_number" in inv_df.columns and rolls:
                            mask = inv_df["roll_number"].astype(str).isin(rolls)
                            st.write("Matching roll rows:")
                            cols = [c for c in [
                                "roll_number","sku","available_quantity","unit_of_measure","inventory_sy","location","status_code","receive_date"
                            ] if c in inv_df.columns]
                            st.dataframe(inv_df.loc[mask, cols].head(200), use_container_width=True, hide_index=True)
                    except Exception:
                        st.write("Unable to search by roll numbers.")
    else:
        st.info("No SKUs currently flagged as low stock based on lead time coverage.")

    # -------------------------------
    # Potential Over-Ordered Items
    # -------------------------------
    if st.session_state.get("defer_heavy", False):
        st.info("Over-Ordered table deferred. Uncheck 'Defer heavy UI builds' in the sidebar to generate.")
    else:
        t_over = perf_counter()
        over_df = filtered_before_runout.copy()
        if not over_df.empty:
            # Align Over-Ordered inventory with SKU Performance logic:
            # For CC 010/012, use filtered roll-sum inventory (excluding Available < 240) when available.
            try:
                # inv_override_sum_by_sku is prepared earlier for the SKU Performance table
                if (
                    'sku' in over_df.columns
                    and 'cost_center' in over_df.columns
                    and inv_override_sum_by_sku is not None
                    and not isinstance(inv_override_sum_by_sku, pd.DataFrame)
                ):
                    cc_norm_oo = over_df['cost_center'].astype(str).str.strip().str.zfill(3)
                    cc_mask_oo = cc_norm_oo.isin(['010', '012'])
                    override_vals_oo = over_df['sku'].astype(str).map(inv_override_sum_by_sku).astype(float)
                    # Replace inventory only for targeted CC rows, fill NaN with 0 to match SKU table behavior
                    over_df.loc[cc_mask_oo, 'inventory_sy'] = override_vals_oo.loc[cc_mask_oo].fillna(0.0)
            except Exception:
                pass
            # Ensure numeric types
            inv = pd.to_numeric(over_df.get("inventory_sy"), errors="coerce").fillna(0.0)
            # Prefer OPENPO_D pending quantity when > 0; otherwise fall back to on_order_sy
            on_order_sy = pd.to_numeric(over_df.get("on_order_sy"), errors="coerce").fillna(0.0)
            po_pending = pd.to_numeric(over_df.get("po_pending_qty"), errors="coerce") if "po_pending_qty" in over_df.columns else None
            if isinstance(po_pending, pd.Series):
                po_pending = po_pending.fillna(0.0)
                on_order = np.where(po_pending > 0, po_pending, on_order_sy)
            else:
                on_order = on_order_sy
            cover = pd.to_numeric(over_df.get("cover_required_sy"), errors="coerce").fillna(0.0)
            backorders = pd.to_numeric(over_df.get("backorder_count"), errors="coerce").fillna(0.0)
            avg_daily = pd.to_numeric(over_df.get("avg_daily_sales_sy"), errors="coerce").fillna(0.0)
            # Use effective PO for net inventory (do NOT add partials here to avoid double-count)
            net_inv = inv + on_order
            # Update the displayed net_inventory_sy to reflect the effective PO
            over_df["net_inventory_sy"] = net_inv
            excess_total = (net_inv - cover).clip(lower=0)
            over_amount = np.minimum(on_order, excess_total)
            over_df["over_order_sy"] = over_amount
            # Exclude new items: require positive avg daily sales
            mask_over = (on_order > 0) & (backorders == 0) & (avg_daily > 0) & (over_df["over_order_sy"] > 0.01)
            over_df = over_df.loc[mask_over].copy()
            if not over_df.empty:
                # Prepare display, showing the effective PO value (aligned to filtered index)
                po_pending_f = pd.to_numeric(over_df.get("po_pending_qty"), errors="coerce") if "po_pending_qty" in over_df.columns else None
                on_order_sy_f = pd.to_numeric(over_df.get("on_order_sy"), errors="coerce").fillna(0.0)
                if isinstance(po_pending_f, pd.Series):
                    po_pending_f = po_pending_f.fillna(0.0)
                    po_disp_arr = np.where(po_pending_f > 0, po_pending_f, on_order_sy_f)
                else:
                    po_disp_arr = on_order_sy_f.values
                po_display_series = pd.Series(po_disp_arr, index=over_df.index, dtype=float)
                over_df["__po_display__"] = po_display_series

                # Days of supply including on-order (recompute aligned series after setting display PO)
                avg_daily_f = pd.to_numeric(over_df.get("avg_daily_sales_sy"), errors="coerce").fillna(0.0)
                inv_f = pd.to_numeric(over_df.get("inventory_sy"), errors="coerce").fillna(0.0)
                net_inv_f = inv_f + po_display_series
                over_df["days_of_supply"] = np.where(avg_daily_f > 0, net_inv_f / avg_daily_f, float("inf"))

                # Additional CC-specific filters for 010/012/013
                try:
                    cc_series = over_df.get("cost_center")
                    if cc_series is not None:
                        cc_norm = cc_series.astype(str).str.strip().str.zfill(3)
                        cc_mask = cc_norm.isin(["010", "012", "013"])
                        if cc_mask.any():
                            lt = pd.to_numeric(over_df.get("lead_time_days"), errors="coerce").fillna(0.0)
                            po_disp = pd.to_numeric(over_df.get("__po_display__"), errors="coerce").fillna(0.0)
                            dos = pd.to_numeric(over_df.get("days_of_supply"), errors="coerce").fillna(0.0)
                            inv0 = inv_f  # already numeric and aligned
                            partial_f = pd.to_numeric(over_df.get("partial_received_po"), errors="coerce").fillna(0.0)
                            eq_partial = np.isclose(po_disp, partial_f, atol=0.01)
                            keep_cc = (
                                (lt < 90.0)
                                & (~((inv0 <= 100.0) & (po_disp < 300.0)))
                                & (dos >= 100.0)
                                & (po_disp >= 100.0)
                                & (~eq_partial)
                            )
                            # For rows outside targeted CCs, keep as-is; apply filter only within CC mask
                            keep_mask = (~cc_mask) | (cc_mask & keep_cc)
                            over_df = over_df.loc[keep_mask].copy()
                            # Also shrink helper Series to match new index
                            po_display_series = po_display_series.reindex(over_df.index)
                            avg_daily_f = avg_daily_f.reindex(over_df.index)
                            inv_f = inv_f.reindex(over_df.index)
                except Exception:
                    pass
                over_df.sort_values("over_order_sy", ascending=False, inplace=True)
                display_cols_over = [
                    "sku", "sku_description", "sku_rating", "over_order_sy", "__po_display__", "partial_received_po",
                    "quantity_sy", "inventory_sy", "avg_daily_sales_sy",
                    "lead_time_days", "cover_required_sy", "net_inventory_sy", "days_of_supply"
                ]
                display_cols_over = [c for c in display_cols_over if c in over_df.columns]
                over_table = over_df[display_cols_over].copy()
                num_cols_over = over_table.select_dtypes(include=["number"]).columns
                over_table[num_cols_over] = over_table[num_cols_over].round(2)
                # Rename columns to match SKU Performance view
                rename_over = {"__po_display__": "PO", "partial_received_po": "partial", "quantity_sy": "qty"}
                over_table.rename(columns=rename_over, inplace=True)
                st.subheader("Potential Over-Ordered Items")
                st.dataframe(over_table, use_container_width=True, hide_index=True)
                # Persist a copy for combined export (Excel multi-sheet)
                try:
                    st.session_state["_export_over_ordered"] = over_table.copy()
                except Exception:
                    st.session_state["_export_over_ordered"] = over_table
                # Download as CSV with filename prompt and brief column descriptions
                with st.expander("Download Potential Over-Ordered as CSV"):
                    default_name = "potential_over_ordered.csv"
                    fname_po = st.text_input("File name", value=default_name, help="Enter a name for the CSV file.", key="po_download_name")
                    if not fname_po.strip().lower().endswith(".csv"):
                        fname_po = (fname_po.strip() or default_name).rstrip(".") + ".csv"
                    col_desc_po = [
                        "# Inventory Dashboard - Potential Over-Ordered Items",
                        "# Columns:",
                        "# sku: Item number",
                        "# sku_description: SKU description",
                        "# sku_rating: SKU rating tier (A-D)",
                        "# over_order_sy: Estimated over-order amount (SY) = min(PO, max(net - cover, 0))",
                        "# qty: Same as SKU Performance qty (SY)",
                        "# PO: Effective purchase orders pending (SY). Prefers OPENPO_D pending; falls back to legacy on-order.",
                        "# partial: Partially received PO quantity (SY)",
                        "# inventory_sy: On-hand inventory (SY)",
                        "# avg_daily_sales_sy: Average daily sales (SY/day)",
                        "# lead_time_days: Lead time in days",
                        "# cover_required_sy: Seasonal/JSTOCK cover requirement (SY)",
                        "# net_inventory_sy: inventory + effective PO (SY)",
                        "# days_of_supply: (inventory + effective PO) / ADS",
                    ]
                    try:
                        po_csv_body = over_table.to_csv(index=False)
                        po_csv_with_desc = ("\n".join(col_desc_po) + "\n" + po_csv_body).encode("utf-8")
                    except Exception:
                        po_csv_with_desc = over_table.to_csv(index=False).encode("utf-8")
                    st.download_button(
                        label="Download as CSV",
                        data=po_csv_with_desc,
                        file_name=fname_po,
                        mime="text/csv",
                        key="po_download_btn",
                    )
                # Combined Excel download for both tables on separate sheets
                with st.expander("Download SKU Performance + Over-Ordered (Excel)"):
                    default_xlsx = "inventory_dashboard_export.xlsx"
                    fname_xlsx = st.text_input(
                        "File name",
                        value=default_xlsx,
                        help="Excel file with two sheets: 'SKU Performance' and 'Potential Over-Ordered'.",
                        key="both_download_name",
                    )
                    if not fname_xlsx.strip().lower().endswith((".xlsx", ".xlsm")):
                        fname_xlsx = (fname_xlsx.strip() or default_xlsx).rstrip(".") + ".xlsx"
                    # Fetch latest tables from session
                    sku_export = st.session_state.get("_export_sku_performance")
                    over_export = st.session_state.get("_export_over_ordered")
                    if sku_export is None and over_export is None:
                        st.info("No tables available to export.")
                    else:
                        try:
                            import pandas as _pd
                            from io import BytesIO as _BytesIO
                            buf = _BytesIO()
                            with _pd.ExcelWriter(buf, engine="openpyxl") as writer:
                                # Sheet 1: SKU Performance with brief descriptions
                                if sku_export is not None and isinstance(sku_export, _pd.DataFrame) and not sku_export.empty:
                                    perf_desc = [
                                        "Inventory Dashboard - SKU Performance",
                                        "Columns:",
                                        "sku: Item number",
                                        "desc: SKU description",
                                        "rating: SKU rating tier (A-D)",
                                        "qty: Total shipped quantity (SY) over the selected window",
                                        "Inv: On-hand inventory (SY)",
                                        "Max: Largest single roll size (SY)",
                                        "avg: Average roll size (SY) = inventory / roll count",
                                        "PO: Purchase orders pending (SY). Prefers OPENPO_D pending; falls back to legacy on-order.",
                                        "partial: Partially received PO quantity (SY)",
                                        "ADS: Average daily sales (SY/day)",
                                        "Jstk: JSTOCK target (SY)",
                                        "LT: Lead time in days",
                                        "DR: Days until runout = net inventory / ADS",
                                        "Reorder: Recommended reorder qty (SY)",
                                    ]
                                    _pd.DataFrame({"Notes": perf_desc}).to_excel(
                                        writer, sheet_name="SKU Performance", index=False, header=False
                                    )
                                    start_row_perf = len(perf_desc) + 1
                                    sku_export.to_excel(writer, sheet_name="SKU Performance", index=False, startrow=start_row_perf)
                                # Sheet 2: Potential Over-Ordered with brief descriptions
                                if over_export is not None and isinstance(over_export, _pd.DataFrame) and not over_export.empty:
                                    over_desc = [
                                        "Inventory Dashboard - Potential Over-Ordered Items",
                                        "Columns:",
                                        "sku: Item number",
                                        "sku_description: SKU description",
                                        "sku_rating: SKU rating tier (A-D)",
                                        "over_order_sy: Over-order amount (SY) = min(PO, max(net - cover, 0))",
                                        "qty: Same as SKU Performance qty (SY)",
                                        "PO: Purchase orders pending (SY). Prefers OPENPO_D pending; falls back to legacy on-order.",
                                        "partial: Partially received PO quantity (SY)",
                                        "inventory_sy: On-hand inventory (SY)",
                                        "avg_daily_sales_sy: Average daily sales (SY/day)",
                                        "lead_time_days: Lead time in days",
                                        "cover_required_sy: Seasonal/JSTOCK cover requirement (SY)",
                                        "net_inventory_sy: inventory + effective PO (SY)",
                                        "days_of_supply: (inventory + effective PO) / ADS",
                                    ]
                                    _pd.DataFrame({"Notes": over_desc}).to_excel(
                                        writer, sheet_name="Potential Over-Ordered", index=False, header=False
                                    )
                                    start_row_over = len(over_desc) + 1
                                    over_export.to_excel(writer, sheet_name="Potential Over-Ordered", index=False, startrow=start_row_over)
                            data_bytes = buf.getvalue()
                        except Exception as _ex:
                            data_bytes = b""
                            st.error(f"Failed to build Excel export: {_ex}")
                        st.download_button(
                            label="Download both as Excel",
                            data=data_bytes,
                            file_name=fname_xlsx,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="both_download_btn",
                            disabled=(len(data_bytes) == 0),
                        )
            else:
                # Diagnostic when nothing qualifies
                st.info("No items currently meet the Over-Ordered criteria (positive PO and excess vs. cover).")

            # Global view: ignore sidebar filters and evaluate potential over-ordered items across all cost centers.
            try:
                _all_cc = tuple(loaders.load_cost_centers(config.connection_string))
            except Exception:
                _all_cc = tuple()

            over_all_df = pd.DataFrame()
            try:
                if _all_cc:
                    _dashboard_all = _load_dashboard(_all_cc, tuple(), tuple(), start_date, end_date)
                    over_all_df = _dashboard_all.get("sku_metrics", pd.DataFrame()).copy()
            except Exception:
                over_all_df = pd.DataFrame()

            if not over_all_df.empty:
                inv_all_tbl = pd.to_numeric(over_all_df.get("inventory_sy"), errors="coerce").fillna(0.0)
                on_order_sy_all_tbl = pd.to_numeric(over_all_df.get("on_order_sy"), errors="coerce").fillna(0.0)
                po_pending_all_tbl = pd.to_numeric(over_all_df.get("po_pending_qty"), errors="coerce") if "po_pending_qty" in over_all_df.columns else None
                if isinstance(po_pending_all_tbl, pd.Series):
                    po_pending_all_tbl = po_pending_all_tbl.fillna(0.0)
                    on_order_all_tbl = np.where(po_pending_all_tbl > 0, po_pending_all_tbl, on_order_sy_all_tbl)
                else:
                    on_order_all_tbl = on_order_sy_all_tbl

                cover_all_tbl = pd.to_numeric(over_all_df.get("cover_required_sy"), errors="coerce").fillna(0.0)
                backorders_all_tbl = pd.to_numeric(over_all_df.get("backorder_count"), errors="coerce").fillna(0.0)
                avg_daily_all_tbl = pd.to_numeric(over_all_df.get("avg_daily_sales_sy"), errors="coerce").fillna(0.0)

                net_all_tbl = inv_all_tbl + on_order_all_tbl
                over_all_df["net_inventory_sy"] = net_all_tbl
                excess_all_tbl = (net_all_tbl - cover_all_tbl).clip(lower=0)
                over_all_df["over_order_sy"] = np.minimum(on_order_all_tbl, excess_all_tbl)

                mask_all_tbl = (
                    (on_order_all_tbl > 0)
                    & (backorders_all_tbl == 0)
                    & (avg_daily_all_tbl > 0)
                    & (over_all_df["over_order_sy"] > 0.01)
                )
                over_all_df = over_all_df.loc[mask_all_tbl].copy()

                if not over_all_df.empty:
                    po_pending_all_disp = pd.to_numeric(over_all_df.get("po_pending_qty"), errors="coerce") if "po_pending_qty" in over_all_df.columns else None
                    on_order_sy_all_disp = pd.to_numeric(over_all_df.get("on_order_sy"), errors="coerce").fillna(0.0)
                    if isinstance(po_pending_all_disp, pd.Series):
                        po_pending_all_disp = po_pending_all_disp.fillna(0.0)
                        po_all_disp_arr = np.where(po_pending_all_disp > 0, po_pending_all_disp, on_order_sy_all_disp)
                    else:
                        po_all_disp_arr = on_order_sy_all_disp.values
                    po_all_display_series = pd.Series(po_all_disp_arr, index=over_all_df.index, dtype=float)
                    over_all_df["__po_display__"] = po_all_display_series

                    avg_daily_all_f = pd.to_numeric(over_all_df.get("avg_daily_sales_sy"), errors="coerce").fillna(0.0)
                    inv_all_f = pd.to_numeric(over_all_df.get("inventory_sy"), errors="coerce").fillna(0.0)
                    over_all_df["days_of_supply"] = np.where(
                        avg_daily_all_f > 0,
                        (inv_all_f + po_all_display_series) / avg_daily_all_f,
                        float("inf"),
                    )

                    try:
                        cc_series_all = over_all_df.get("cost_center")
                        if cc_series_all is not None:
                            cc_norm_all = cc_series_all.astype(str).str.strip().str.zfill(3)
                            cc_mask_all = cc_norm_all.isin(["010", "012", "013"])
                            if cc_mask_all.any():
                                lt_all = pd.to_numeric(over_all_df.get("lead_time_days"), errors="coerce").fillna(0.0)
                                po_all = pd.to_numeric(over_all_df.get("__po_display__"), errors="coerce").fillna(0.0)
                                dos_all = pd.to_numeric(over_all_df.get("days_of_supply"), errors="coerce").fillna(0.0)
                                partial_all = pd.to_numeric(over_all_df.get("partial_received_po"), errors="coerce").fillna(0.0)
                                eq_partial_all = np.isclose(po_all, partial_all, atol=0.01)
                                keep_cc_all = (
                                    (lt_all < 90.0)
                                    & (~((inv_all_f <= 100.0) & (po_all < 300.0)))
                                    & (dos_all >= 100.0)
                                    & (po_all >= 100.0)
                                    & (~eq_partial_all)
                                )
                                keep_mask_all = (~cc_mask_all) | (cc_mask_all & keep_cc_all)
                                over_all_df = over_all_df.loc[keep_mask_all].copy()
                    except Exception:
                        pass

                    # Merge shadow SKUs that reference another SKU in description into the
                    # referenced/base SKU so history and planning metrics are treated as one item.
                    try:
                        if {"sku", "sku_description"}.issubset(over_all_df.columns):
                            _m = over_all_df.copy()
                            _m["sku"] = _m["sku"].astype(str).str.strip().str.upper()
                            _m["sku_description"] = _m["sku_description"].fillna("").astype(str).str.strip()

                            _desc_items_map = {}
                            _valid_skus = set(_m["sku"].astype(str).str.strip().str.upper().tolist())
                            try:
                                _items_all_df = _dashboard_all.get("items", pd.DataFrame()) if isinstance(_dashboard_all, dict) else pd.DataFrame()
                                if not _items_all_df.empty and {"sku", "sku_description"}.issubset(_items_all_df.columns):
                                    _tmp_map = _items_all_df[["sku", "sku_description"]].copy()
                                    _tmp_map["sku"] = _tmp_map["sku"].astype(str).str.strip().str.upper()
                                    _tmp_map["sku_description"] = _tmp_map["sku_description"].fillna("").astype(str).str.strip()
                                    _tmp_map = _tmp_map[_tmp_map["sku"].ne("")]
                                    _tmp_map = _tmp_map[_tmp_map["sku_description"].ne("")]
                                    _valid_skus |= set(_tmp_map["sku"].astype(str).str.strip().str.upper().tolist())
                                    _tmp_map = _tmp_map[["sku", "sku_description"]]
                                    if not _tmp_map.empty:
                                        _desc_items_map = dict(_tmp_map.drop_duplicates(subset=["sku"]).values.tolist())
                            except Exception:
                                _desc_items_map = {}

                            # Detect any SKU token referenced in description, not only "USE CODE ...".
                            _ref_list: list[str] = []
                            for _cur_sku, _desc in zip(_m["sku"].tolist(), _m["sku_description"].tolist()):
                                _cur = str(_cur_sku).strip().upper()
                                _txt = str(_desc).upper()
                                _tokens = [t.strip().upper() for t in re.findall(r"[A-Z0-9_-]{5,}", _txt)]
                                _ref = ""
                                for _t in _tokens:
                                    if _t in _valid_skus and _t != _cur:
                                        _ref = _t
                                        break
                                _ref_list.append(_ref)

                            _ref_sku = pd.Series(_ref_list, index=_m.index, dtype=str)
                            _has_ref = _ref_sku.ne("")
                            _m["canonical_sku"] = _m["sku"]
                            _m.loc[_has_ref, "canonical_sku"] = _ref_sku.loc[_has_ref]

                            _m["sku"] = _m["canonical_sku"].astype(str).str.strip().str.upper()
                            _m.drop(columns=["canonical_sku"], inplace=True)

                            # Track how many source rows were collapsed into each canonical SKU.
                            _src_counts = _m.groupby("sku").size().rename("_source_count")

                            def _pick_desc(series: pd.Series) -> str:
                                for _v in series.tolist():
                                    _s = str(_v).strip()
                                    _tokens = [t.strip().upper() for t in re.findall(r"[A-Z0-9_-]{5,}", _s.upper())]
                                    _has_ref_in_desc = any((_t in _valid_skus) for _t in _tokens)
                                    if _s and (not _has_ref_in_desc):
                                        return _s
                                for _v in series.tolist():
                                    _s = str(_v).strip()
                                    if _s:
                                        return _s
                                return ""

                            _sum_cols = [
                                "over_order_sy", "__po_display__", "partial_received_po", "quantity_sy",
                                "inventory_sy", "avg_daily_sales_sy", "lead_time_days", "cover_required_sy",
                                "net_inventory_sy",
                            ]
                            _sum_cols = [c for c in _sum_cols if c in _m.columns]

                            _agg = {c: "sum" for c in _sum_cols}
                            _agg["sku_description"] = _pick_desc
                            if "sku_rating" in _m.columns:
                                _agg["sku_rating"] = lambda s: " / ".join(sorted({str(v).strip() for v in s if str(v).strip() and str(v).strip().lower() != "none"})) or "None"

                            _m = _m.groupby("sku", as_index=False).agg(_agg)
                            _m = _m.merge(_src_counts.reset_index(), on="sku", how="left")
                            _m["Merged"] = np.where(pd.to_numeric(_m.get("_source_count"), errors="coerce").fillna(1).astype(int) > 1, "Y", "")

                            if _desc_items_map:
                                _desc_fill = _m["sku"].map(_desc_items_map)
                                _desc_has_ref = _m["sku_description"].fillna("").astype(str).str.upper().map(
                                    lambda _s: any((_t in _valid_skus) for _t in re.findall(r"[A-Z0-9_-]{5,}", str(_s)))
                                )
                                _m["sku_description"] = _m["sku_description"].where(
                                    ~(_desc_has_ref | _m["sku_description"].fillna("").eq("")),
                                    _desc_fill,
                                )
                                _m["sku_description"] = _m["sku_description"].fillna("")

                            if {"net_inventory_sy", "avg_daily_sales_sy"}.issubset(_m.columns):
                                _avg = pd.to_numeric(_m.get("avg_daily_sales_sy"), errors="coerce").fillna(0.0)
                                _net = pd.to_numeric(_m.get("net_inventory_sy"), errors="coerce").fillna(0.0)
                                _m["days_of_supply"] = np.where(_avg > 0, _net / _avg, float("inf"))

                            _m.drop(columns=["_source_count"], inplace=True, errors="ignore")

                            over_all_df = _m.copy()
                    except Exception:
                        pass

                    # Sort by highest PO order number per SKU (for sort only), then by over_order_sy.
                    # Also attach newest order_entry_date per SKU for display.
                    try:
                        _po_all_orders = _dashboard_all.get("purchase_orders", pd.DataFrame()) if isinstance(_dashboard_all, dict) else pd.DataFrame()
                        if not _po_all_orders.empty and {"sku", "order_number"}.issubset(_po_all_orders.columns):
                            _po_agg_cols = ["sku", "order_number"]
                            _has_date = "order_entry_date" in _po_all_orders.columns
                            if _has_date:
                                _po_agg_cols.append("order_entry_date")
                            _po_tmp = _po_all_orders[_po_agg_cols].copy()
                            _po_tmp["order_number"] = pd.to_numeric(_po_tmp["order_number"], errors="coerce")
                            # Build per-SKU max order number
                            _po_max_ord = (
                                _po_tmp.groupby("sku", as_index=False)["order_number"]
                                .max()
                                .rename(columns={"order_number": "__max_order_num__"})
                            )
                            over_all_df = over_all_df.merge(_po_max_ord, on="sku", how="left")
                            # Build per-SKU max (newest) order_entry_date
                            if _has_date:
                                _po_tmp["order_entry_date"] = pd.to_datetime(_po_tmp["order_entry_date"], errors="coerce")
                                _po_max_date = (
                                    _po_tmp.groupby("sku", as_index=False)["order_entry_date"]
                                    .max()
                                    .rename(columns={"order_entry_date": "order_date"})
                                )
                                over_all_df = over_all_df.merge(_po_max_date, on="sku", how="left")
                                # Format as date only for display
                                over_all_df["order_date"] = pd.to_datetime(over_all_df["order_date"], errors="coerce").dt.date
                            over_all_df.sort_values(
                                ["__max_order_num__", "over_order_sy"],
                                ascending=[False, False],
                                inplace=True,
                            )
                            over_all_df.drop(columns=["__max_order_num__"], inplace=True, errors="ignore")
                        else:
                            over_all_df.sort_values("over_order_sy", ascending=False, inplace=True)
                    except Exception:
                        over_all_df.sort_values("over_order_sy", ascending=False, inplace=True)
                    display_cols_over_all = [
                        "sku", "sku_description", "sku_rating", "over_order_sy", "__po_display__",
                        "quantity_sy", "inventory_sy", "avg_daily_sales_sy",
                        "lead_time_days", "cover_required_sy", "net_inventory_sy", "days_of_supply", "order_date"
                    ]
                    display_cols_over_all = [c for c in display_cols_over_all if c in over_all_df.columns]
                    over_all_table = over_all_df[display_cols_over_all].copy()
                    num_cols_over_all = over_all_table.select_dtypes(include=["number"]).columns
                    over_all_table[num_cols_over_all] = over_all_table[num_cols_over_all].round(2)
                    over_all_table.rename(
                        columns={"__po_display__": "PO", "quantity_sy": "qty"},
                        inplace=True,
                    )
                    st.subheader("Potential Over-Ordered Items (All Cost Centers)")
                    st.caption(
                        "This table ignores sidebar filters and always shows all qualifying over-ordered items. "
                        "Sorted by highest PO order number (highest to lowest), then by over-order amount."
                    )
                    st.dataframe(over_all_table, use_container_width=True, hide_index=True)
                else:
                    st.info("No global over-ordered items currently meet the criteria across all cost centers.")
        ui_timings["ui_build_over_ordered"] = perf_counter() - t_over

    # -------------------------
    # Build a Truck
    # -------------------------
    st.subheader("Build a Truck")
    st.caption(
        "Plan a bulk order for one or more price classes. "
        "Select the product(s) and enter a total quantity target — the system distributes "
        "that quantity across SKUs proportionally based on current reorder demand, "
        "using the same logic as SKU Performance."
    )
    try:
        # Build price class option list from items_df
        _bat_options: list[str] = []
        _bat_labels: dict[str, str] = {}
        if not items_df.empty and "price_class" in items_df.columns:
            _pc_cols = ["price_class"] + (["price_class_desc"] if "price_class_desc" in items_df.columns else [])
            _bat_pc_rows = (
                items_df[_pc_cols]
                .dropna(subset=["price_class"])
                .drop_duplicates(subset=["price_class"])
                .copy()
            )
            _bat_pc_rows["price_class"] = _bat_pc_rows["price_class"].astype(str).str.strip()
            _bat_pc_rows = _bat_pc_rows[_bat_pc_rows["price_class"] != ""].sort_values("price_class")
            for _, _r in _bat_pc_rows.iterrows():
                _code = str(_r["price_class"])
                _desc = str(_r.get("price_class_desc", "")).strip() if "price_class_desc" in _bat_pc_rows.columns else ""
                _bat_options.append(_code)
                _bat_labels[_code] = f"{_code} — {_desc}" if _desc else _code

        _bat_pref_pcs: list = [str(p).strip() for p in user_prefs.get("bat_price_classes", []) if str(p).strip()]
        _bat_state_pcs: list = [
            str(p).strip() for p in st.session_state.get("bat_price_classes", []) if str(p).strip()
        ]
        _bat_persisted = sorted(set(_bat_pref_pcs).union(set(_bat_state_pcs)))
        # Preserve selected classes when current cost-center filters do not include them.
        for _pc in _bat_persisted:
            if _pc not in _bat_options:
                _bat_options.append(_pc)
                _bat_labels[_pc] = f"{_pc} (not in current filters)"
        _bat_pref_qty: float = float(user_prefs.get("bat_target_qty", 0.0))

        _bat_col1, _bat_col2 = st.columns([3, 1])
        with _bat_col1:
            bat_selected_pcs = st.multiselect(
                "Product (Price Class)",
                options=_bat_options,
                default=[p for p in _bat_pref_pcs if p in _bat_options],
                format_func=lambda x: _bat_labels.get(x, x),
                key="bat_price_classes",
                help="Choose one or more price classes to plan a bulk order for.",
            )
        with _bat_col2:
            bat_target_qty = st.number_input(
                "Target Total (SY)",
                min_value=0.0,
                value=_bat_pref_qty,
                step=50.0,
                format="%.1f",
                key="bat_target_qty_input",
                help="Total SY you want to order across all SKUs in the selected price class(es).",
            )

        # Persist selections to user_prefs
        if bat_selected_pcs != _bat_pref_pcs or bat_target_qty != _bat_pref_qty:
            try:
                user_prefs["bat_price_classes"] = bat_selected_pcs
                user_prefs["bat_target_qty"] = float(bat_target_qty)
                _save_prefs(user_prefs)
            except Exception:
                pass

        if bat_selected_pcs and bat_target_qty > 0:
            try:
                # --- Resolve SKUs for the selected price classes ---
                _bat_pc_norm = {str(p).strip().upper() for p in bat_selected_pcs}
                _bat_sku_set: set[str] = set()
                if not items_df.empty and "price_class" in items_df.columns:
                    _bat_pc_mask = items_df["price_class"].astype(str).str.strip().str.upper().isin(_bat_pc_norm)
                    _bat_sku_set = set(
                        items_df.loc[_bat_pc_mask, "sku"].astype(str).str.strip().str.upper().dropna().tolist()
                    )

                if not _bat_sku_set:
                    st.info("No SKUs found for the selected price class(es).")
                else:
                    # Use the full (unfiltered by rating) metrics so all SKUs in the price class are shown
                    _bat_m = sku_metrics_full.copy()
                    _bat_m["sku"] = _bat_m.get("sku", pd.Series(dtype=str)).astype(str).str.strip()
                    _bat_m_mask = _bat_m.get("sku", pd.Series(dtype=str)).astype(str).str.strip().str.upper().isin(_bat_sku_set)
                    _bat_m = _bat_m.loc[_bat_m_mask].copy()

                    # Include every SKU from the selected price classes, even if it has sparse or no metrics yet.
                    _bat_item_cols = [
                        c for c in ["sku", "sku_description", "price_class", "price_class_desc", "cost_center"]
                        if c in items_df.columns
                    ]
                    _bat_items_sel = items_df.loc[_bat_pc_mask, _bat_item_cols].copy() if _bat_item_cols else pd.DataFrame()
                    if not _bat_items_sel.empty:
                        _bat_items_sel["sku"] = _bat_items_sel["sku"].astype(str).str.strip()
                        _bat_items_sel = _bat_items_sel[_bat_items_sel["sku"] != ""].drop_duplicates(subset=["sku"])
                        _bat_m = _bat_items_sel.set_index("sku").join(
                            _bat_m.set_index("sku"), how="left", rsuffix="_metric"
                        ).reset_index()
                        for _merge_col in ["sku_description", "price_class", "price_class_desc", "cost_center"]:
                            _metric_col = f"{_merge_col}_metric"
                            if _metric_col in _bat_m.columns:
                                if _merge_col in _bat_m.columns:
                                    _bat_m[_merge_col] = _bat_m[_merge_col].fillna(_bat_m[_metric_col])
                                else:
                                    _bat_m[_merge_col] = _bat_m[_metric_col]
                                _bat_m.drop(columns=[_metric_col], inplace=True)

                    if _bat_m.empty:
                        st.info("No inventory metrics found for the selected price class(es). These SKUs may not have sales history yet.")
                    else:
                        # --- Apply same inventory overrides as SKU Performance ---
                        # (RET exclusion + RCODE@ filter, and CC 010/012 low-available filter)
                        _bat_inv = pd.to_numeric(_bat_m.get("inventory_sy"), errors="coerce").fillna(0.0).copy()
                        try:
                            _bat_exc = inv_exclude_ret_sum_by_sku
                            if _bat_exc is not None and not isinstance(_bat_exc, pd.DataFrame):
                                _bat_inv = _bat_m["sku"].map(_bat_exc).astype(float).fillna(_bat_inv)
                        except NameError:
                            pass
                        try:
                            _bat_ovr = inv_override_sum_by_sku
                            if _bat_ovr is not None and not isinstance(_bat_ovr, pd.DataFrame) and "cost_center" in sku_metrics_full.columns:
                                _bat_cc_full = sku_metrics_full.set_index("sku")["cost_center"].astype(str).str.strip().str.zfill(3)
                                _bat_cc_vals = _bat_m["sku"].map(_bat_cc_full)
                                _bat_cc_mask_tbl = _bat_cc_vals.isin(["010", "012"])
                                _bat_ovr_mapped = _bat_m["sku"].map(_bat_ovr).astype(float)
                                _bat_inv.loc[_bat_cc_mask_tbl] = _bat_ovr_mapped.loc[_bat_cc_mask_tbl].fillna(_bat_inv.loc[_bat_cc_mask_tbl])
                        except NameError:
                            pass
                        _bat_m = _bat_m.copy()
                        _bat_m["inventory_sy"] = _bat_inv.values

                        # --- Apply growth factor (same scaling as SKU Performance) ---
                        try:
                            _bat_gs = float(growth_scale)
                        except NameError:
                            _bat_gs = 1.0
                        _bat_avg = pd.to_numeric(_bat_m.get("avg_daily_sales_sy"), errors="coerce").fillna(0.0) * _bat_gs
                        _bat_tgt = pd.to_numeric(_bat_m.get("target_on_hand_sy"), errors="coerce").fillna(0.0) * _bat_gs
                        _bat_assumed = pd.to_numeric(_bat_m.get("assumed_on_order_sy"), errors="coerce").fillna(0.0)
                        _bat_bo = pd.to_numeric(_bat_m.get("backorder_qty_sy"), errors="coerce").fillna(0.0)
                        _bat_net = _bat_inv + _bat_assumed
                        _bat_safety = _bat_avg * 7.0
                        _bat_rq = np.maximum((_bat_tgt + _bat_safety) - _bat_net, 0.0) + _bat_bo
                        _bat_dur = np.where(
                            _bat_avg > 0,
                            np.maximum(_bat_net - _bat_bo, 0.0) / np.maximum(_bat_avg, 1e-9),
                            float("inf"),
                        )
                        _bat_m["adj_avg_daily_sales_sy"] = _bat_avg.values
                        _bat_m["reorder_qty_sy"] = _bat_rq.values
                        _bat_m["days_until_runout"] = _bat_dur

                        # Prefer OPENPO_D pending qty for PO display
                        try:
                            if "po_pending_qty" in _bat_m.columns:
                                _bat_m["on_order_sy"] = (
                                    _bat_m["po_pending_qty"].astype(float)
                                    .fillna(_bat_m.get("on_order_sy", pd.Series(0.0, index=_bat_m.index)))
                                    .fillna(0.0)
                                )
                        except Exception:
                            pass

                        # Attach max_roll_size_sy from pre-computed data when available
                        try:
                            if "max_roll_size_sy" not in _bat_m.columns:
                                _bat_mrs = max_roll_sizes
                                if _bat_mrs is not None:
                                    _bat_m["max_roll_size_sy"] = _bat_m["sku"].map(_bat_mrs).astype(float)
                                else:
                                    _bat_inv_raw = inv_all.copy()
                                    _bat_inv_raw["inventory_sy"] = pd.to_numeric(
                                        _bat_inv_raw.get("inventory_sy"), errors="coerce"
                                    ).fillna(0.0)
                                    _bat_inv_raw = _bat_inv_raw[_bat_inv_raw["inventory_sy"] > 0]
                                    if not _bat_inv_raw.empty and "sku" in _bat_inv_raw.columns:
                                        _bat_m["max_roll_size_sy"] = _bat_m["sku"].map(
                                            _bat_inv_raw.groupby("sku")["inventory_sy"].max()
                                        ).astype(float)
                        except (NameError, Exception):
                            pass

                        # --- Lowest-DR-first allocation ---
                        # Raise the weakest coverage SKU(s) first; as their DR catches up to the next SKU,
                        # bring that SKU into the active pool and continue until the requested total is allocated.
                        _bat_alloc = np.zeros(len(_bat_m), dtype=float)
                        _bat_fallback_used = False
                        _bat_remaining = float(bat_target_qty)
                        _bat_avg_np = np.asarray(_bat_avg, dtype=float)
                        _bat_dur_np = np.asarray(_bat_dur, dtype=float)
                        _bat_reorder_np = np.asarray(_bat_rq, dtype=float)

                        _bat_active = np.where(np.isfinite(_bat_dur_np) & (_bat_avg_np > 0))[0]
                        if len(_bat_active) > 0 and _bat_remaining > 0:
                            _bat_rank = sorted(_bat_active.tolist(), key=lambda i: (_bat_dur_np[i], -_bat_avg_np[i], str(_bat_m.iloc[i].get("sku", ""))))
                            _bat_pool: list[int] = []
                            _bat_level = None
                            for _idx in _bat_rank:
                                if _bat_level is None:
                                    _bat_pool.append(_idx)
                                    _bat_level = float(_bat_dur_np[_idx])
                                    continue
                                _bat_next = float(_bat_dur_np[_idx])
                                _bat_pool_avg = float(_bat_avg_np[_bat_pool].sum())
                                _bat_needed = max(_bat_next - float(_bat_level), 0.0) * _bat_pool_avg
                                if _bat_pool_avg <= 0:
                                    break
                                if _bat_needed > 0 and _bat_remaining < _bat_needed:
                                    _bat_raise = _bat_remaining / _bat_pool_avg
                                    for _pool_idx in _bat_pool:
                                        _bat_alloc[_pool_idx] += _bat_avg_np[_pool_idx] * _bat_raise
                                    _bat_remaining = 0.0
                                    _bat_level += _bat_raise
                                    break
                                if _bat_needed > 0:
                                    for _pool_idx in _bat_pool:
                                        _bat_alloc[_pool_idx] += _bat_avg_np[_pool_idx] * (_bat_next - float(_bat_level))
                                    _bat_remaining -= _bat_needed
                                _bat_pool.append(_idx)
                                _bat_level = _bat_next

                            if _bat_remaining > 0 and _bat_pool:
                                _bat_pool_avg = float(_bat_avg_np[_bat_pool].sum())
                                if _bat_pool_avg > 0:
                                    _bat_raise = _bat_remaining / _bat_pool_avg
                                    for _pool_idx in _bat_pool:
                                        _bat_alloc[_pool_idx] += _bat_avg_np[_pool_idx] * _bat_raise
                                    _bat_remaining = 0.0

                        if _bat_remaining > 0:
                            _bat_fallback_used = True
                            _bat_weights = np.maximum(_bat_reorder_np, 0.0)
                            if float(_bat_weights.sum()) <= 0:
                                _bat_weights = np.maximum(_bat_avg_np, 0.0)
                            if float(_bat_weights.sum()) <= 0:
                                _bat_weights = np.ones(len(_bat_m), dtype=float)
                            _bat_alloc += (_bat_weights / float(_bat_weights.sum())) * _bat_remaining
                            _bat_remaining = 0.0

                        _bat_m["allocated_sy"] = np.round(_bat_alloc, 2)
                        _bat_post_net = _bat_net + _bat_m["allocated_sy"].astype(float)
                        _bat_post_dur = np.where(
                            _bat_avg > 0,
                            np.maximum(_bat_post_net - _bat_bo, 0.0) / np.maximum(_bat_avg, 1e-9),
                            float("inf"),
                        )
                        _bat_m["post_build_days_until_runout"] = _bat_post_dur

                        # --- Build display table (same columns as SKU Performance + Allocated) ---
                        _bat_display_cols = [
                            "sku", "sku_description", "sku_rating", "quantity_sy",
                            "inventory_sy", "max_roll_size_sy", "on_order_sy",
                            "partial_received_po", "backorder_qty_sy",
                            "adj_avg_daily_sales_sy", "jstock", "lead_time_days",
                            "days_until_runout", "reorder_qty_sy", "allocated_sy", "post_build_days_until_runout",
                        ]
                        _bat_avail = [c for c in _bat_display_cols if c in _bat_m.columns]
                        _bat_tbl = _bat_m[_bat_avail].copy()
                        _bat_num_cols = _bat_tbl.select_dtypes(include=["number"]).columns
                        _bat_tbl[_bat_num_cols] = _bat_tbl[_bat_num_cols].round(2)
                        _bat_rename_map = {
                            "sku_description": "desc",
                            "quantity_sy": "qty",
                            "inventory_sy": "Inv",
                            "max_roll_size_sy": "Max",
                            "backorder_qty_sy": "BO",
                            "on_order_sy": "PO",
                            "partial_received_po": "partial",
                            "adj_avg_daily_sales_sy": "ADS",
                            "reorder_qty_sy": "Reorder",
                            "lead_time_days": "LT",
                            "sku_rating": "rating",
                            "jstock": "Jstk",
                            "days_until_runout": "DR",
                            "allocated_sy": "Allocated",
                            "post_build_days_until_runout": "Post DR",
                        }
                        _bat_tbl.rename(columns=_bat_rename_map, inplace=True)
                        _bat_col_order = [
                            c for c in [
                                "sku", "desc", "rating", "qty", "Inv", "Max",
                                "PO", "partial", "BO", "ADS", "Jstk", "LT", "DR",
                                "Reorder", "Allocated", "Post DR",
                            ] if c in _bat_tbl.columns
                        ]
                        _bat_tbl = _bat_tbl[_bat_col_order]
                        if "Allocated" in _bat_tbl.columns:
                            _bat_tbl = _bat_tbl.sort_values("Allocated", ascending=False)

                        _bat_alloc_total = float(_bat_tbl["Allocated"].sum()) if "Allocated" in _bat_tbl.columns else 0.0
                        _bat_note = " Coverage-balanced by lowest DR first."
                        if _bat_fallback_used:
                            _bat_note += " Residual quantity was spread using demand weights for rows without usable DR."
                        st.caption(
                            f"Allocating **{bat_target_qty:,.1f} SY** across **{len(_bat_tbl)} SKUs**.{_bat_note} "
                            f"Total allocated: {_bat_alloc_total:,.2f} SY."
                        )
                        st.dataframe(_bat_tbl, use_container_width=True, hide_index=True)

                        try:
                            _bat_csv = _bat_tbl.to_csv(index=False).encode("utf-8")
                            st.download_button(
                                label="Download as CSV",
                                data=_bat_csv,
                                file_name="build_a_truck.csv",
                                mime="text/csv",
                                key="bat_download_csv",
                            )
                        except Exception:
                            pass

            except Exception as _bat_ex:
                st.warning(f"Unable to compute truck planning: {_bat_ex}")
        elif bat_selected_pcs:
            st.info("Enter a target quantity greater than 0 to see the allocation plan.")
    except Exception as _bat_outer_ex:
        st.warning(f"Build a Truck section error: {_bat_outer_ex}")

    # -------------------------
    # Watch List (always show selected SKUs)
    # -------------------------
    st.markdown("#### Watch List")
    st.caption("Add SKUs you always want visible here. This list is saved and can be updated anytime.")

    try:
        _all_skus = []
        if isinstance(items_df, pd.DataFrame) and not items_df.empty and "sku" in items_df.columns:
            _all_skus = sorted(
                {
                    str(s).strip().upper()
                    for s in items_df.get("sku").dropna().astype(str).tolist()
                    if str(s).strip() != ""
                }
            )

        _pref_watch_raw = user_prefs.get("sku_watch_list", [])
        _pref_watch = [str(s).strip().upper() for s in _pref_watch_raw if str(s).strip() != ""]
        _state_watch = [str(s).strip().upper() for s in st.session_state.get("watch_list_selected", []) if str(s).strip()]
        _watch_persisted = sorted(set(_pref_watch).union(set(_state_watch)))
        # Keep selected SKUs in the picker even when current filters exclude them.
        if _watch_persisted:
            _all_skus = sorted(set(_all_skus).union(set(_watch_persisted)))

        if "watch_list_selected" not in st.session_state:
            st.session_state["watch_list_selected"] = _pref_watch

        wcol1, wcol2 = st.columns([4, 2])
        with wcol1:
            _watch_selected_ui = st.multiselect(
                "Watch list SKUs",
                options=_all_skus,
                default=[s for s in st.session_state.get("watch_list_selected", []) if s in _all_skus],
                key="watch_list_selected",
                help="Selected SKUs always appear in the Watch List table below.",
            )
        with wcol2:
            _watch_add_text = st.text_input(
                "Add SKU",
                value="",
                key="watch_add_sku_input",
                help="Enter one SKU or comma-separated SKUs to add.",
            )
            _wb1, _wb2 = st.columns(2)
            _watch_add_btn = _wb1.button("Add", key="watch_add_btn")
            _watch_clear_btn = _wb2.button("Clear", key="watch_clear_btn")

        if _watch_add_btn:
            _tokens = [t.strip().upper() for t in str(_watch_add_text).split(",") if t.strip()]
            _existing = [str(s).strip().upper() for s in st.session_state.get("watch_list_selected", []) if str(s).strip()]
            _merged = sorted(set(_existing).union(set(_tokens)))
            st.session_state["watch_list_selected"] = _merged
            st.session_state["watch_add_sku_input"] = ""
            st.rerun()

        if _watch_clear_btn:
            st.session_state["watch_list_selected"] = []
            st.rerun()

        _watch_selected = [str(s).strip().upper() for s in st.session_state.get("watch_list_selected", []) if str(s).strip()]

        if _watch_selected != _pref_watch:
            try:
                user_prefs["sku_watch_list"] = _watch_selected
                _save_prefs(user_prefs)
            except Exception:
                pass

        if _watch_selected:
            _watch_m = sku_metrics_full.copy() if isinstance(sku_metrics_full, pd.DataFrame) else pd.DataFrame()
            if not _watch_m.empty:
                _watch_m["sku"] = _watch_m.get("sku", pd.Series(dtype=str)).astype(str).str.strip().str.upper()

            _watch_items_cols = [c for c in ["sku", "sku_description", "cost_center"] if c in items_df.columns]
            _watch_items = pd.DataFrame({"sku": _watch_selected})
            if _watch_items_cols:
                _watch_items = _watch_items.merge(
                    items_df[_watch_items_cols].copy().assign(sku=lambda d: d["sku"].astype(str).str.strip().str.upper()),
                    on="sku",
                    how="left",
                )

            if not _watch_m.empty:
                _watch_m = _watch_items.set_index("sku").join(_watch_m.set_index("sku"), how="left", rsuffix="_metric").reset_index()
                if "sku_description_metric" in _watch_m.columns:
                    if "sku_description" in _watch_m.columns:
                        _watch_m["sku_description"] = _watch_m["sku_description"].fillna(_watch_m["sku_description_metric"])
                    else:
                        _watch_m["sku_description"] = _watch_m["sku_description_metric"]
                    _watch_m.drop(columns=["sku_description_metric"], inplace=True)
            else:
                _watch_m = _watch_items.copy()

            # Align inventory to SKU Performance behavior when these maps are available.
            _watch_inv = pd.to_numeric(_watch_m.get("inventory_sy"), errors="coerce").fillna(0.0).copy()
            try:
                if inv_exclude_ret_sum_by_sku is not None and not isinstance(inv_exclude_ret_sum_by_sku, pd.DataFrame):
                    _watch_inv = _watch_m["sku"].map(inv_exclude_ret_sum_by_sku).astype(float).fillna(_watch_inv)
            except Exception:
                pass
            try:
                if inv_override_sum_by_sku is not None and not isinstance(inv_override_sum_by_sku, pd.DataFrame):
                    if "cost_center" in _watch_m.columns:
                        _watch_cc = _watch_m["cost_center"].astype(str).str.strip().str.zfill(3)
                        _watch_cc_mask = _watch_cc.isin(["010", "012"])
                        _watch_ovr = _watch_m["sku"].map(inv_override_sum_by_sku).astype(float)
                        _watch_inv.loc[_watch_cc_mask] = _watch_ovr.loc[_watch_cc_mask].fillna(_watch_inv.loc[_watch_cc_mask])
            except Exception:
                pass
            _watch_m["inventory_sy"] = _watch_inv

            try:
                _watch_gs = float(growth_scale)
            except Exception:
                _watch_gs = 1.0
            _watch_avg = pd.to_numeric(_watch_m.get("avg_daily_sales_sy"), errors="coerce").fillna(0.0) * _watch_gs
            _watch_tgt = pd.to_numeric(_watch_m.get("target_on_hand_sy"), errors="coerce").fillna(0.0) * _watch_gs
            _watch_assumed = pd.to_numeric(_watch_m.get("assumed_on_order_sy"), errors="coerce").fillna(0.0)
            _watch_bo = pd.to_numeric(_watch_m.get("backorder_qty_sy"), errors="coerce").fillna(0.0)
            _watch_net = _watch_inv + _watch_assumed
            _watch_dur = np.where(
                _watch_avg > 0,
                np.maximum(_watch_net - _watch_bo, 0.0) / np.maximum(_watch_avg, 1e-9),
                float("inf"),
            )
            _watch_rq = np.maximum((_watch_tgt + (_watch_avg * 7.0)) - _watch_net, 0.0) + _watch_bo
            _watch_m["adj_avg_daily_sales_sy"] = _watch_avg.values
            _watch_m["days_until_runout"] = _watch_dur
            _watch_m["reorder_qty_sy"] = _watch_rq

            try:
                if "po_pending_qty" in _watch_m.columns:
                    _watch_m["on_order_sy"] = (
                        _watch_m["po_pending_qty"].astype(float)
                        .fillna(_watch_m.get("on_order_sy", pd.Series(0.0, index=_watch_m.index)))
                        .fillna(0.0)
                    )
            except Exception:
                pass

            if "max_roll_size_sy" not in _watch_m.columns:
                try:
                    if max_roll_sizes is not None:
                        _watch_m["max_roll_size_sy"] = _watch_m["sku"].map(max_roll_sizes).astype(float)
                except Exception:
                    _watch_m["max_roll_size_sy"] = np.nan

            _watch_cols = [
                "sku",
                "sku_description",
                "sku_rating",
                "quantity_sy",
                "inventory_sy",
                "max_roll_size_sy",
                "on_order_sy",
                "partial_received_po",
                "backorder_qty_sy",
                "adj_avg_daily_sales_sy",
                "jstock",
                "lead_time_days",
                "days_until_runout",
                "reorder_qty_sy",
            ]
            _watch_tbl = _watch_m[[c for c in _watch_cols if c in _watch_m.columns]].copy()
            _watch_rename = {
                "sku_description": "desc",
                "sku_rating": "rating",
                "quantity_sy": "qty",
                "inventory_sy": "Inv",
                "max_roll_size_sy": "Max",
                "on_order_sy": "PO",
                "partial_received_po": "partial",
                "backorder_qty_sy": "BO",
                "adj_avg_daily_sales_sy": "ADS",
                "jstock": "Jstk",
                "lead_time_days": "LT",
                "days_until_runout": "DR",
                "reorder_qty_sy": "Reorder",
            }
            _watch_tbl.rename(columns=_watch_rename, inplace=True)
            _watch_order = [
                c for c in [
                    "sku", "desc", "rating", "qty", "Inv", "Max", "PO", "partial", "BO",
                    "ADS", "Jstk", "LT", "DR", "Reorder"
                ] if c in _watch_tbl.columns
            ]
            _watch_tbl = _watch_tbl[_watch_order]

            _watch_num = _watch_tbl.select_dtypes(include=["number"]).columns
            _watch_tbl[_watch_num] = _watch_tbl[_watch_num].round(2)
            st.dataframe(_watch_tbl, use_container_width=True, hide_index=True)
        else:
            st.info("Watch list is empty. Add SKUs above to keep them always visible here.")
    except Exception as _watch_ex:
        st.warning(f"Watch list error: {_watch_ex}")

    # -------------------------
    # Load Times (Tab 3)
    # -------------------------
    with tabs[2]:
        st.subheader("Load Times")
        have_backend = isinstance(timings, dict) and len(timings) > 0
        have_ui = len(ui_timings) > 0
        if not have_backend and not have_ui:
            st.info("Timing details will appear after the next load.")
        else:
            backend_items = [(f"backend:{k}", float(v)) for k, v in timings.items() if k != "total"] if have_backend else []
            ui_items = [(f"ui:{k}", float(v)) for k, v in ui_timings.items()]
            combined = backend_items + ui_items
            combined.sort(key=lambda x: x[1], reverse=True)
            df_t = pd.DataFrame(combined, columns=["step", "seconds"]) if combined else pd.DataFrame(columns=["step", "seconds"])
            df_t["seconds"] = pd.to_numeric(df_t["seconds"], errors="coerce").fillna(0.0)

            backend_total = float(timings.get("total", sum(v for _, v in backend_items))) if have_backend else 0.0
            compute_call = float(ui_timings.get("compute_dashboard_call", 0.0))
            other_ui_total = float(sum(v for k, v in ui_timings.items() if k != "compute_dashboard_call"))
            overall_total = compute_call + other_ui_total
            st.metric("Overall Load Time", f"{overall_total:.2f}s")
            st.caption(f"compute: {compute_call:.2f}s, other UI: {other_ui_total:.2f}s, backend-only: {backend_total:.2f}s")
            st.dataframe(df_t, use_container_width=True, hide_index=True)

    # Reorder Recommendations hidden per request

    # Backorder Tracker hidden per request

    backorder_history = dashboard.get("backorder_history")
    if backorder_history is not None and not backorder_history.empty:
        with st.expander("Historical backorders"):
            history_display = backorder_history.copy()
            # Attach description where available
            if "sku_description" in items_df.columns:
                desc_map = items_df.set_index("sku")["sku_description"]
                # Insert after sku column if present
                if "sku" in history_display.columns:
                    history_display.insert(history_display.columns.get_loc("sku") + 1, "description", history_display["sku"].map(desc_map))
            datetime_cols = ["first_detected_at", "last_seen_at", "resolved_at"]
            for col in datetime_cols:
                if col in history_display.columns:
                    history_display[col] = pd.to_datetime(history_display[col], errors='coerce', utc=True).dt.tz_convert(None)
            if "order_entry_date" in history_display.columns:
                history_display["order_entry_date"] = pd.to_datetime(history_display["order_entry_date"], errors='coerce').dt.date
            st.dataframe(history_display, use_container_width=True, hide_index=True)
            csv_history = history_display.to_csv(index=False).encode('utf-8')
            st.download_button(
                label="Download backorder history",
                data=csv_history,
                file_name="backorder_history.csv",
                mime="text/csv",
            )

    with st.expander("Purchase orders in flight"):
        purchase_orders = dashboard["purchase_orders"]
        if purchase_orders.empty:
            st.write("No purchase orders found for the selected filters.")
        else:
            po_display = purchase_orders[[
                "sku",
                "order_number",
                "quantity_sy",
                "unit_of_measure",
                "eta_date",
                "supplier_number",
            ]].copy()
            if "sku_description" in items_df.columns:
                desc_map = items_df.set_index("sku")["sku_description"]
                po_display.insert(1, "description", po_display["sku"].map(desc_map))
            po_display["eta_date"] = pd.to_datetime(po_display["eta_date"]).dt.date
            # Rename for clarity: show standardized SY quantity
            po_display.rename(columns={"quantity_sy": "PO (SY)"}, inplace=True)
            st.dataframe(po_display, use_container_width=True, hide_index=True)

    with st.expander("Raw sales orders"):
        sales_orders = dashboard["sales_orders"]
        if sales_orders.empty:
            st.write("No sales orders found for the selected filters.")
        else:
            sales_display = sales_orders[[
                "order_number",
                "sku",
                "quantity_ordered",
                "unit_of_measure",
                "order_entry_date",
                "actual_ship_date",
                "backorder_flag",
            ]].copy()
            if "sku_description" in items_df.columns:
                desc_map = items_df.set_index("sku")["sku_description"]
                sales_display.insert(2, "description", sales_display["sku"].map(desc_map))
            sales_display["order_entry_date"] = pd.to_datetime(sales_display["order_entry_date"]).dt.date
            sales_display["actual_ship_date"] = pd.to_datetime(sales_display["actual_ship_date"]).dt.date
            st.dataframe(sales_display, use_container_width=True, hide_index=True)

    # -------------------------
    # Launch Dates (Tab 4)
    # -------------------------
    with tabs[3]:
        st.subheader("Price Class Launch Dates")
        st.caption("If a launch date is blank, calculations default to 2025-08-04.")
        # Load saved launch dates (used to populate launch_date values)
        saved = load_launch_dates()
        # Determine relevant price classes based on current filters
        pcs_in_items: list[str] = []
        if not items_df.empty and "price_class" in items_df.columns:
            pcs_in_items = items_df.get("price_class").dropna().astype(str).unique().tolist()
        # Respect explicit sidebar selection when present; otherwise use those in filtered items
        filtered_pcs = selected_price_classes if selected_price_classes else pcs_in_items
        filtered_pcs = sorted({str(pc) for pc in filtered_pcs})
        if not filtered_pcs:
            st.info("No price classes match the current filters.")
        else:
            # Map descriptions
            pc_df = loaders.load_price_classes(config.connection_string)
            pc_map = {}
            if isinstance(pc_df, pd.DataFrame) and not pc_df.empty and {"price_class", "price_class_desc"}.issubset(set(pc_df.columns)):
                pc_map = {str(c): str(d) for c, d in zip(pc_df["price_class"], pc_df["price_class_desc"])}
            rows = []
            for code in filtered_pcs:
                desc = pc_map.get(str(code), "")
                dt = saved.get(str(code)) or "2025-08-04"
                rows.append({
                    "price_class": code,
                    "price_class_desc": desc,
                    "launch_date": dt,
                })
            df_launch = pd.DataFrame(rows)
            if not df_launch.empty:
                # Display as dates
                if "launch_date" in df_launch.columns:
                    df_launch["launch_date"] = pd.to_datetime(df_launch["launch_date"], errors="coerce").dt.date
                # Order by description then code
                sort_cols = [c for c in ["price_class_desc", "price_class"] if c in df_launch.columns]
                if sort_cols:
                    df_launch.sort_values(sort_cols, inplace=True, na_position="last")
                st.dataframe(df_launch, use_container_width=True, hide_index=True)
                csv_ld = df_launch.to_csv(index=False).encode("utf-8")
                st.download_button("Download launch dates", data=csv_ld, file_name="launch_dates.csv", mime="text/csv")

    # -------------------------
    # Price List Generator (Tab 5)
    # -------------------------
    with tabs[4]:
        st.subheader("Price List Generator")
        st.caption("Build an Excel template with static and dynamic fields. Use placeholders like {SKU}, {PRICE_CLASS}, or your own variables like {MARGIN} in the header and detail sections.")

        # If a template was just loaded, options were staged under 'plg_restore_options'.
        # Apply them BEFORE creating any widgets to avoid Streamlit state mutation errors.
        _pending = st.session_state.get("plg_restore_options")
        if isinstance(_pending, dict):
            opts = st.session_state.pop("plg_restore_options")
            try:
                if opts.get("mode") in ["Structured builder", "Advanced (legacy)"]:
                    st.session_state["plg_mode"] = opts.get("mode")
                if opts.get("detail_mode") in ["Item", "Price Class"]:
                    st.session_state["plg_detail_mode_struct"] = opts.get("detail_mode")
                flt = opts.get("filters", {}) if isinstance(opts.get("filters"), dict) else {}
                if "manufacturer" in flt:
                    st.session_state["plg_filter_mf"] = list(flt.get("manufacturer") or [])
                if "product_line" in flt:
                    st.session_state["plg_filter_pl"] = list(flt.get("product_line") or [])
                if "supplier" in flt:
                    st.session_state["plg_filter_sp"] = list(flt.get("supplier") or [])
                if "iinven" in flt:
                    st.session_state["plg_filter_inv"] = list(flt.get("iinven") or [])
                # Restore exclude filters if present
                flt_ex = opts.get("filters_exclude", {}) if isinstance(opts.get("filters_exclude"), dict) else {}
                if "manufacturer" in flt_ex:
                    st.session_state["plg_exclude_mf"] = list(flt_ex.get("manufacturer") or [])
                if "product_line" in flt_ex:
                    st.session_state["plg_exclude_pl"] = list(flt_ex.get("product_line") or [])
                if "supplier" in flt_ex:
                    st.session_state["plg_exclude_sp"] = list(flt_ex.get("supplier") or [])
                if "iinven" in flt_ex:
                    st.session_state["plg_exclude_inv"] = list(flt_ex.get("iinven") or [])
                # Restore additional options
                if isinstance(opts.get("col_labels"), list):
                    st.session_state["plg_col_labels"] = opts.get("col_labels")
                if "include_labels_excel" in opts:
                    st.session_state["plg_include_labels_excel"] = bool(opts.get("include_labels_excel"))
                if isinstance(opts.get("sort_keys"), list):
                    st.session_state["plg_sort_keys"] = opts.get("sort_keys")
                if isinstance(opts.get("logos"), list):
                    st.session_state["plg_logos"] = opts.get("logos")
                if "exclude_single_pc" in opts:
                    st.session_state["plg_exclude_single_pc"] = bool(opts.get("exclude_single_pc"))
                if "exclude_pc_threshold" in opts:
                    try:
                        st.session_state["plg_exclude_pc_threshold"] = int(opts.get("exclude_pc_threshold") or 3)
                    except Exception:
                        st.session_state["plg_exclude_pc_threshold"] = 3
            except Exception:
                pass

        # Source data needed to fill dynamic fields from ITEM table
        # Use all items (including non-inventory) for Price List Generator so Stocking and previews include everything
        try:
            items_src = loaders.load_items_all(config.connection_string, selected_cost_centers).copy()
        except Exception:
            items_src = dashboard.get("items", pd.DataFrame()).copy()
        if items_src.empty:
            st.info("No items available. Load Overview first or adjust filters.")
        else:
            # Normalize column names we commonly use
            items_src.rename(columns={
                "sku": "SKU",
                "sku_description": "DESCRIPTION",
                "manufacturer": "MANUFACTURER",
                "product_line": "PRODUCT_LINE",
                "supplier_number": "SUPPLIER",
                "price_class": "PRICE_CLASS",
                "cost_center": "COST_CENTER",
                "item_pattern": "ITEM_PATTERN",
                # Ensure ITEM table helpers are present
                "inventory_flag": "IINVEN",
                "item_width_inches": "ITEM_WIDTH_INCHES",
            }, inplace=True)

            # Attach price class description for ItemVar use
            try:
                pc_df = loaders.load_price_classes(config.connection_string)
                if isinstance(pc_df, pd.DataFrame) and not pc_df.empty and {"price_class","price_class_desc"}.issubset(pc_df.columns):
                    pc_map = {str(c): str(d) for c, d in zip(pc_df["price_class"], pc_df["price_class_desc"])}
                    items_src["PRICE_CLASS_DESC"] = items_src.get("PRICE_CLASS").astype(str).map(lambda x: pc_map.get(x, None))
            except Exception:
                pass

            # Filter scope selection
            with st.expander("Filters", expanded=True):
                f_cols = st.columns(4)
                mf_opts = sorted(items_src.get("MANUFACTURER", pd.Series(dtype=str)).dropna().astype(str).unique().tolist())
                pl_opts = sorted(items_src.get("PRODUCT_LINE", pd.Series(dtype=str)).dropna().astype(str).unique().tolist())
                sp_opts = sorted(items_src.get("SUPPLIER", pd.Series(dtype=str)).dropna().astype(str).unique().tolist())
                inv_opts = sorted(
                    [v for v in items_src.get("IINVEN", pd.Series(dtype=str)).dropna().astype(str).str.upper().unique().tolist() if str(v).strip()]
                )
                mf_sel = f_cols[0].multiselect(
                    "Manufacturer",
                    mf_opts,
                    default=[v for v in st.session_state.get("plg_filter_mf", []) if v in mf_opts],
                    key="plg_filter_mf",
                )
                pl_sel = f_cols[1].multiselect(
                    "Product Line",
                    pl_opts,
                    default=[v for v in st.session_state.get("plg_filter_pl", []) if v in pl_opts],
                    key="plg_filter_pl",
                )
                sp_sel = f_cols[2].multiselect(
                    "Supplier",
                    sp_opts,
                    default=[v for v in st.session_state.get("plg_filter_sp", []) if v in sp_opts],
                    key="plg_filter_sp",
                )
                inv_sel = f_cols[3].multiselect(
                    "Inventory (IINVEN)",
                    inv_opts or ["Y","N"],
                    default=[v for v in st.session_state.get("plg_filter_inv", []) if (not inv_opts or v in inv_opts)],
                    key="plg_filter_inv",
                )
                # Exclude filters row
                st.caption("Exclude (optional)")
                e_cols = st.columns(4)
                mf_ex = e_cols[0].multiselect(
                    "Manufacturer",
                    mf_opts,
                    default=[v for v in st.session_state.get("plg_exclude_mf", []) if v in mf_opts],
                    key="plg_exclude_mf",
                )
                pl_ex = e_cols[1].multiselect(
                    "Product Line",
                    pl_opts,
                    default=[v for v in st.session_state.get("plg_exclude_pl", []) if v in pl_opts],
                    key="plg_exclude_pl",
                )
                sp_ex = e_cols[2].multiselect(
                    "Supplier",
                    sp_opts,
                    default=[v for v in st.session_state.get("plg_exclude_sp", []) if v in sp_opts],
                    key="plg_exclude_sp",
                )
                inv_ex = e_cols[3].multiselect(
                    "Inventory (IINVEN)",
                    inv_opts or ["Y","N"],
                    default=[v for v in st.session_state.get("plg_exclude_inv", []) if (not inv_opts or v in inv_opts)],
                    key="plg_exclude_inv",
                )

                base = items_src.copy()
                if mf_sel:
                    base = base[base.get("MANUFACTURER").astype(str).isin(mf_sel)]
                if pl_sel:
                    base = base[base.get("PRODUCT_LINE").astype(str).isin(pl_sel)]
                if sp_sel:
                    base = base[base.get("SUPPLIER").astype(str).isin(sp_sel)]
                if inv_sel:
                    base = base[base.get("IINVEN").astype(str).str.upper().isin([str(v).upper() for v in inv_sel])]
                # Apply exclusions
                if mf_ex:
                    base = base[~base.get("MANUFACTURER").astype(str).isin(mf_ex)]
                if pl_ex:
                    base = base[~base.get("PRODUCT_LINE").astype(str).isin(pl_ex)]
                if sp_ex:
                    base = base[~base.get("SUPPLIER").astype(str).isin(sp_ex)]
                if inv_ex:
                    base = base[~base.get("IINVEN").astype(str).str.upper().isin([str(v).upper() for v in inv_ex])]
                # Option: exclude price classes that only have a single item
                # Exclude price classes with N or fewer items
                exclude_pc_enabled = st.checkbox(
                    "Exclude price classes with N or fewer items",
                    value=bool(st.session_state.get("plg_exclude_single_pc", True)),
                    key="plg_exclude_single_pc",
                )
                exclude_threshold = st.number_input(
                    "N (items per price class)", min_value=1, max_value=50,
                    value=int(st.session_state.get("plg_exclude_pc_threshold", 3)), step=1,
                    key="plg_exclude_pc_threshold",
                )
                if exclude_pc_enabled and not base.empty:
                    try:
                        thr = int(st.session_state.get("plg_exclude_pc_threshold", 3))

                        # 0) Exclude ITEM_PATTERN (IPATT) clusters with < thr items across the current scope
                        try:
                            ip_series = base.get("ITEM_PATTERN").fillna("").astype(str)
                            # Count only non-empty IPATTs
                            mask_ip = ip_series.str.strip() != ""
                            if mask_ip.any():
                                ip_counts = (
                                    base.loc[mask_ip]
                                        .assign(_sku=base.loc[mask_ip].get("SKU").astype(str), _ipatt=ip_series.loc[mask_ip].str.strip())
                                        .groupby("_ipatt", dropna=False)["_sku"].nunique()
                                )
                                keep_ip = set([str(ix) for ix, cnt in ip_counts.items() if cnt and cnt >= thr])
                                base = base.loc[~mask_ip | ip_series.str.strip().isin(keep_ip)]
                        except Exception:
                            pass

                        # 1) Exclude price classes with < thr items in the remaining scope
                        pc_counts = (
                            base.assign(_sku=base.get("SKU").astype(str))
                                .groupby("PRICE_CLASS", dropna=False)["_sku"].nunique()
                        )
                        # First criterion: PC must have at least 'thr' items in current scope
                        pc_pass_1 = set([str(ix) for ix, cnt in pc_counts.items() if cnt and cnt >= thr])
                        scoped = base[base.get("PRICE_CLASS").astype(str).isin(pc_pass_1)]

                        # 2) Within each passing PC, require at least 'thr' items sharing the same ITEM_PATTERN
                        valid_pc = set()
                        if not scoped.empty:
                            for pc, g in scoped.groupby("PRICE_CLASS", dropna=False):
                                patt_series = g.get("ITEM_PATTERN")
                                if patt_series is None:
                                    continue
                                gp = (
                                    g.assign(_sku=g.get("SKU").astype(str), _ipatt=patt_series.fillna("").astype(str))
                                     .query("_ipatt != ''")
                                     .groupby("_ipatt", dropna=False)["_sku"].nunique()
                                )
                                if not gp.empty and int(gp.max()) >= thr:
                                    valid_pc.add(str(pc))
                        base = base[base.get("PRICE_CLASS").astype(str).isin(valid_pc)]
                    except Exception:
                        pass

            # Helper: derive per-SKU style names from DESCRIPTION within each PRICE_CLASS
            def _norm_txt(s: str) -> str:
                try:
                    import re as _re
                    s2 = str(s or "")
                    s2 = _re.sub(r"[^A-Za-z0-9 ]+", " ", s2)
                    s2 = " ".join(s2.split())
                    return s2.strip().lower()
                except Exception:
                    return str(s or "").strip().lower()

            def _common_prefix_str(strings: list[str]) -> str:
                if not strings:
                    return ""
                # Character-level common prefix (case-sensitive to preserve original case in slice)
                shortest = min(strings, key=len)
                for i in range(len(shortest)):
                    ch = shortest[i]
                    for s in strings:
                        if s[i] != ch:
                            return shortest[:i]
                return shortest

            def _derive_style_map(df: pd.DataFrame) -> dict:
                """Derive a per-SKU style with concise naming:
                - Group by ITEM_PATTERN (IPATT), and for each group, compute a common starting phrase across INAMEs
                  by taking the longest common token prefix (alphabetic tokens), with a sensible max length (up to 4 tokens)
                - If the first token mixes letters+digits (e.g., SP2900), use that single token regardless of group commonality
                - Append an early gauge number (>=20) only if the base doesn't already contain digits
                - Special-case: if base is "Outer Banks" and "Plus" appears early, append "Plus"
                Sizes at the tail like 12/15 are ignored.
                """
                style_by_sku: dict[str, str] = {}
                if df.empty:
                    return style_by_sku
                try:
                    import re as _re
                    from collections import defaultdict
                    # Include COST_CENTER when present to enable center-specific rules
                    cols = [c for c in ["SKU","DESCRIPTION","PRICE_CLASS","ITEM_PATTERN","COST_CENTER"] if c in df.columns]
                    tmp = df[cols].dropna(subset=["SKU"]).copy()
                    tmp["SKU"] = tmp["SKU"].astype(str)
                    tmp["DESCRIPTION"] = tmp["DESCRIPTION"].astype(str)
                    tmp["ITEM_PATTERN"] = tmp.get("ITEM_PATTERN", "").fillna("").astype(str)

                    def _alpha_num_tokens(s: str) -> list[str]:
                        # Keep alpha or numeric tokens (no punctuation)
                        return _re.findall(r"[A-Za-z0-9]+", s or "")
                    def _alpha_tokens(s: str) -> list[str]:
                        return _re.findall(r"[A-Za-z]+", s or "")
                    def _has_letters_and_digits(tok: str) -> bool:
                        return any(c.isalpha() for c in tok) and any(c.isdigit() for c in tok)

                    def _alpha_tokens_seq(s: str) -> list[str]:
                        return _re.findall(r"[A-Za-z]+", s or "")
                    # Build common token prefix per ITEM_PATTERN
                    common_prefix_by_ipatt: dict[str, list[str]] = {}
                    for patt, grp in tmp.groupby("ITEM_PATTERN", dropna=False):
                        tokens_list = []
                        for _, rr in grp.iterrows():
                            tokens_list.append(_alpha_tokens_seq(str(rr.get("DESCRIPTION", ""))))
                        # Longest common prefix across lists (case-insensitive compare, preserve original case of first row)
                        if not tokens_list:
                            continue
                        # Normalize to lowercase for comparison
                        lower_lists = [[t.lower() for t in lst] for lst in tokens_list]
                        lcp: list[str] = []
                        for idx in range(min(len(lst) for lst in lower_lists)):
                            cand = lower_lists[0][idx]
                            if all(idx < len(lst) and lst[idx] == cand for lst in lower_lists):
                                lcp.append(cand)
                            else:
                                break
                        # Cap prefix length to avoid over-long styles (keep up to 4 words)
                        if lcp:
                            common_prefix_by_ipatt[str(patt)] = lcp[:4]

                    # Build concise style per SKU
                    for _, r in tmp.iterrows():
                        desc_orig = str(r.get("DESCRIPTION", ""))
                        toks_all = _alpha_num_tokens(desc_orig)
                        # Base selection:
                        # - If the first token mixes letters+digits (e.g., SP2900), use only that token
                        # - Else, use common prefix from ITEM_PATTERN if available; otherwise fall back to first two alpha tokens
                        alpha = _alpha_tokens(desc_orig)
                        ipatt = str(r.get("ITEM_PATTERN", ""))
                        if toks_all and _has_letters_and_digits(toks_all[0]):
                            base_tokens = [toks_all[0]]
                        else:
                            cp = common_prefix_by_ipatt.get(ipatt)
                            if cp and len(cp) >= 1:
                                # Use the prefix, but map back to the original-case tokens of this row where indices align
                                base_tokens = []
                                lower_alpha = [t.lower() for t in alpha]
                                for i, tok_lc in enumerate(cp):
                                    if i < len(lower_alpha) and lower_alpha[i] == tok_lc:
                                        base_tokens.append(alpha[i])
                                    else:
                                        # if mismatch in this row, stop prefix to avoid dragging unrelated words
                                        break
                                if not base_tokens:
                                    base_tokens = alpha[:2] if alpha else (toks_all[:1] if toks_all else [])
                            else:
                                base_tokens = alpha[:2] if alpha else (toks_all[:1] if toks_all else [])
                        # If the base came out as two words but ITEM_PATTERN matches the start of the first token,
                        # collapse to just the family token (e.g., MALA -> MALAGA CURATOR => MALAGA)
                        try:
                            if len(base_tokens) >= 2 and ipatt:
                                first_lc = base_tokens[0].lower()
                                patt_lc = ipatt.lower()
                                # Allowlist two-word families we never collapse
                                allowed_two_word = {"outer banks", "posh bio"}
                                base_phrase_lc_chk = " ".join([t.lower() for t in base_tokens[:2]])
                                if base_phrase_lc_chk not in allowed_two_word:
                                    if first_lc.startswith(patt_lc) or patt_lc.startswith(first_lc[:len(patt_lc)]):
                                        base_tokens = [base_tokens[0]]
                        except Exception:
                            pass
                        # Small version digit (e.g., "2" in "BARRINGTON 2") only when tied to ITEM_PATTERN suffix
                        version_digit = None
                        try:
                            m_ver = _re.search(r"(\d+)$", ipatt)
                            if m_ver:
                                vd = m_ver.group(1)
                                if vd.isdigit() and 1 <= int(vd) <= 9:
                                    version_digit = vd
                        except Exception:
                            version_digit = None
                        # Detect end-of-description dimension pattern like 19.69x19.69
                        has_dim_at_end = False
                        try:
                            if _re.search(r"(\d+(?:\.\d+)?)\s*[xX]\s*(\d+(?:\.\d+)?)\s*$", desc_orig.strip()):
                                has_dim_at_end = True
                        except Exception:
                            has_dim_at_end = False
                        # Early gauge among first 6 tokens, limited to two-digit values 20..99
                        # (avoids appending 3+ digit color/codes like 551/679)
                        gauge = None
                        for t in toks_all[:6]:
                            if t.isdigit() and len(t) == 2:
                                try:
                                    val = int(t)
                                except Exception:
                                    val = -1
                                if 20 <= val <= 99 and not has_dim_at_end:
                                    gauge = t
                                    break
                        style_tokens = list(base_tokens)
                        # Append "Plus" if base is exactly "Outer Banks" and PLUS appears early
                        base_phrase_lc = " ".join([t.lower() for t in base_tokens])
                        if base_phrase_lc == "outer banks":
                            if any(t.lower() == "plus" for t in toks_all[:6]):
                                style_tokens.append("Plus")
                        # Append version indicator when appropriate:
                        # - If the base already has a Roman numeral (e.g., II), do NOT append the Arabic digit
                        # - Else, and only if the base has no digits, append the Arabic digit (e.g., 2)
                        # Do not add a version digit when description includes feet/inches measurements like 13'2" or 12'
                        has_measure = False
                        try:
                            if _re.search(r"\b\d+\s*(?:ft|')\s*\d*(?:in|\")?\b", desc_orig.lower()):
                                has_measure = True
                        except Exception:
                            has_measure = False
                        if version_digit and not any(ch.isdigit() for ch in "".join(style_tokens)) and not has_measure:
                            try:
                                _roman_map = {1: "I", 2: "II", 3: "III", 4: "IV", 5: "V", 6: "VI", 7: "VII", 8: "VIII", 9: "IX"}
                                vv = int(version_digit)
                                roman_tok = _roman_map.get(vv)
                            except Exception:
                                roman_tok = None
                            base_has_roman = False
                            if roman_tok:
                                # Check existing tokens for roman already present
                                base_has_roman = any(t.lower() == roman_tok.lower() for t in style_tokens)
                            if not base_has_roman:
                                if version_digit not in style_tokens:
                                    style_tokens.append(version_digit)
                        # Only append gauge if base doesn't already contain digits (to avoid SP2900 + 4000 => SP2900 4000)
                        if gauge and (gauge not in style_tokens) and not (style_tokens and any(ch.isdigit() for ch in style_tokens[0])):
                            style_tokens.append(gauge)
                        style_by_sku[str(r["SKU"])] = " ".join(style_tokens)
                except Exception:
                    return style_by_sku
                return style_by_sku

            style_map_by_sku = _derive_style_map(base)

            def _effective_pc_desc_for_record(rec: dict) -> str:
                # If a precomputed style hint is present (e.g., when iterating by style within a price class), prefer it
                try:
                    style_hint = str(rec.get("STYLE_DERIVED", ""))
                    if style_hint.strip():
                        return style_hint.strip()
                except Exception:
                    pass
                pc_desc = str(rec.get("PRICE_CLASS_DESC", "") or "").strip()
                desc = str(rec.get("DESCRIPTION", "") or "").strip()
                sku = str(rec.get("SKU", ""))
                # If we have a derived style and it's longer (e.g., adds gauge) than PC desc, prefer it
                derived = style_map_by_sku.get(sku, "").strip()
                if derived:
                    # New: if derived is a single concise token, and PC desc simply adds extra words
                    # that are not part of a known two-word family, prefer the concise derived token.
                    try:
                        import re as _re
                        allowed_two_word = {"posh bio", "outer banks"}
                        d_toks = _re.findall(r"[A-Za-z0-9]+", derived)
                        pc_toks = _re.findall(r"[A-Za-z0-9]+", pc_desc)
                        first_two_pc = " ".join([t.lower() for t in pc_toks[:2]])
                        if len(d_toks) == 1:
                            # If PC desc starts with the derived token but is longer, and not whitelisted two-word
                            if _norm_txt(pc_desc).startswith(_norm_txt(derived)) and len(pc_toks) >= 2 and first_two_pc not in allowed_two_word:
                                return derived
                    except Exception:
                        pass
                    try:
                        if pc_desc and _norm_txt(derived).startswith(_norm_txt(pc_desc)) and len(derived) > len(pc_desc):
                            return derived
                    except Exception:
                        if len(derived) > len(pc_desc):
                            return derived
                    # Prefer derived if it contains an early gauge and PC desc does not
                    import re as _re
                    has_num = bool(_re.search(r"\b([2-9][0-9])\b", derived))
                    has_num_pc = bool(_re.search(r"\b([2-9][0-9])\b", pc_desc))
                    if has_num and not has_num_pc:
                        return derived
                # If PC desc aligns with the beginning of DESCRIPTION but misses an early gauge, append it
                if pc_desc and _norm_txt(desc).startswith(_norm_txt(pc_desc)):
                    import re as _re
                    tokens = _re.findall(r"[A-Za-z0-9]+", desc)
                    base_tokens = _re.findall(r"[A-Za-z]+", pc_desc)
                    # Look for the first numeric token just after the base tokens
                    if tokens and base_tokens and len(tokens) > len(base_tokens):
                        nxt = tokens[len(base_tokens)]
                        if nxt.isdigit() and len(nxt) == 2:
                            try:
                                val = int(nxt)
                            except Exception:
                                val = -1
                            # Only treat two-digit values between 20 and 99 as gauge; ignore large codes like 15004
                            if 20 <= val <= 99:
                                return f"{pc_desc} {nxt}"
                    return pc_desc
                # Otherwise use derived style from the item's own cluster
                sku = str(rec.get("SKU", ""))
                style = style_map_by_sku.get(sku, "").strip()
                if style:
                    return style
                # Last resort: first two tokens of description or the PC desc
                toks = [t for t in desc.split() if t]
                return (" ".join(toks[:2]) if toks else pc_desc) or pc_desc

            # Mode selection: Structured builder or Advanced (legacy)
            mode_ui = st.radio("Mode", ["Structured builder", "Advanced (legacy)"], horizontal=True, key="plg_mode")

            if mode_ui == "Structured builder":
                st.markdown("### Static variables")
                st.caption("Add key/value pairs you can reference in dynamic variables, e.g. {MARGIN}.")
                # Manage as rows for easy editing
                static_rows = st.session_state.get("plg_static_rows")
                if static_rows is None:
                    # initialize from existing dict if present
                    sv = st.session_state.get("plg_static_vars", {"MARGIN": "0.40", "TERMS": "NET30"})
                    static_rows = [{"name": k, "value": str(v)} for k, v in sv.items()]
                static_df = pd.DataFrame(static_rows or [], columns=["name", "value"]).astype({"name": "string", "value": "string"})
                static_ed = st.data_editor(
                    static_df,
                    num_rows="dynamic",
                    use_container_width=True,
                    column_config={
                        "name": st.column_config.TextColumn("Name", required=True),
                        "value": st.column_config.TextColumn("Value"),
                    },
                    key="plg_static_editor",
                )
                # Persist
                ed_rows_static = static_ed.to_dict(orient="records")
                st.session_state["plg_static_rows"] = ed_rows_static
                static_vars = {r["name"]: r.get("value", "") for r in ed_rows_static if r.get("name")}
                st.session_state["plg_static_vars"] = static_vars

                st.markdown("### Dynamic variables")
                st.caption("Add variables your system understands. Use ITEM fields and static vars with {FIELD} placeholders.")
                # Persisted dynamic vars helpers
                from pathlib import Path as _Path
                def _dyn_file():
                    return _Path("history") / "price_dynamic_vars.json"
                def _dyn_load():
                    p = _dyn_file()
                    if p.exists():
                        try:
                            import json as _json
                            return _json.loads(p.read_text(encoding="utf-8")) or []
                        except Exception:
                            return []
                    return []
                def _dyn_save(rows: list[dict]):
                    try:
                        import json as _json
                        # Sanitize before saving in case external edits introduced invalid entries
                        def _sanitize_dyn_rows(rows_in):
                            clean = []
                            if not isinstance(rows_in, list):
                                return clean
                            for x in rows_in:
                                if isinstance(x, dict):
                                    name = str(x.get("name", "")).strip()
                                    tmpl = x.get("template", "")
                                    width = x.get("width_chars")
                                elif isinstance(x, str):
                                    name = x.strip(); tmpl = x; width = None
                                else:
                                    continue
                                if not name:
                                    continue
                                try:
                                    width = int(width) if (width is not None and int(width) > 0) else None
                                except Exception:
                                    width = None
                                clean.append({"name": name, "template": str(tmpl), "width_chars": width})
                            # Dedupe by name, keep first
                            seen = set(); out = []
                            for r in clean:
                                if r["name"] in seen:
                                    continue
                                seen.add(r["name"]); out.append(r)
                            return out
                        rows = _sanitize_dyn_rows(rows)
                        p = _dyn_file()
                        p.parent.mkdir(parents=True, exist_ok=True)
                        p.write_text(_json.dumps(rows, indent=2), encoding="utf-8")
                    except Exception:
                        pass

                if "plg_dyn_rows" not in st.session_state:
                    # Start with sensible defaults
                    base_rows = [
                        {"name": "SKU_ONLY", "template": "{SKU}", "width_chars": None},
                        {"name": "DESC_ONLY", "template": "{DESCRIPTION}", "width_chars": None},
                    ]
                    # Merge in persisted ones (override duplicates so widths/templates from disk win)
                    try:
                        persisted = _dyn_load()
                    except Exception:
                        persisted = []
                    try:
                        # Build by-name map from base
                        by_name = {str(r["name"]): dict(r) for r in base_rows if r.get("name")}
                        for r in (persisted or []):
                            if isinstance(r, dict) and r.get("name"):
                                nm = str(r.get("name"))
                                by_name[nm] = {"name": nm, "template": r.get("template", ""), "width_chars": r.get("width_chars")}
                            elif isinstance(r, str):
                                nm = r; by_name[nm] = {"name": nm, "template": r, "width_chars": None}
                        # Preserve original base order, then append any extras from persisted
                        ordered = []
                        seen = set()
                        for r in base_rows:
                            nm = str(r.get("name"))
                            if nm in by_name and nm not in seen:
                                ordered.append(by_name[nm]); seen.add(nm)
                        for nm, row in by_name.items():
                            if nm not in seen:
                                ordered.append(row)
                        st.session_state["plg_dyn_rows"] = ordered
                    except Exception:
                        st.session_state["plg_dyn_rows"] = base_rows
                add_var_clicked = st.button("Add dynamic variable", key="plg_add_dyn_btn")
                if add_var_clicked:
                    dlg = getattr(st, "dialog", None)
                    if callable(dlg):
                        @dlg("Add Dynamic Variable")
                        def _dlg_add_var():
                            with st.form("plg_add_var_form"):
                                vname = st.text_input("Variable name", key="plg_new_var_name")
                                vtmpl = st.text_area("Template (use {FIELD} placeholders)", key="plg_new_var_tmpl")
                                vwidth = st.number_input("Width (characters, optional)", min_value=0, max_value=200, value=0, step=1, key="plg_new_var_width")
                                ok = st.form_submit_button("Save")
                                if ok and vname.strip():
                                    rows = st.session_state.get("plg_dyn_rows", [])
                                    # replace if exists
                                    rows = [r for r in rows if r.get("name") != vname.strip()]
                                    rows.append({"name": vname.strip(), "template": vtmpl, "width_chars": (int(vwidth) if int(vwidth) > 0 else None)})
                                    st.session_state["plg_dyn_rows"] = rows
                                    _dyn_save(rows)
                                    # Let Streamlit rerun naturally; avoids tab resets
                        _dlg_add_var()
                    else:
                        with st.expander("Add Dynamic Variable", expanded=True):
                            vname = st.text_input("Variable name", key="plg_new_var_name_f")
                            vtmpl = st.text_area("Template (use {FIELD} placeholders)", key="plg_new_var_tmpl_f")
                            vwidth = st.number_input("Width (characters, optional)", min_value=0, max_value=200, value=0, step=1, key="plg_new_var_width_f")
                            if st.button("Save variable", key="plg_save_var_btn") and vname.strip():
                                rows = st.session_state.get("plg_dyn_rows", [])
                                rows = [r for r in rows if r.get("name") != vname.strip()]
                                rows.append({"name": vname.strip(), "template": vtmpl, "width_chars": (int(vwidth) if int(vwidth) > 0 else None)})
                                st.session_state["plg_dyn_rows"] = rows
                                _dyn_save(rows)
                # Manage current variables with edit/delete
                # Sanitize dynamic rows before rendering to avoid non-dict entries
                def _sanitize_dyn_rows(rows_in):
                    clean = []
                    if not isinstance(rows_in, list):
                        return clean
                    for x in rows_in:
                        if isinstance(x, dict):
                            name = str(x.get("name", "")).strip()
                            tmpl = x.get("template", "")
                            width = x.get("width_chars")
                        elif isinstance(x, str):
                            name = x.strip(); tmpl = x; width = None
                        else:
                            continue
                        if not name:
                            continue
                        try:
                            width = int(width) if (width is not None and int(width) > 0) else None
                        except Exception:
                            width = None
                        clean.append({"name": name, "template": str(tmpl), "width_chars": width})
                    # Dedupe by name, keep first
                    seen = set(); out = []
                    for r in clean:
                        if r["name"] in seen:
                            continue
                        seen.add(r["name"]); out.append(r)
                    return out
                dyn_rows = _sanitize_dyn_rows(st.session_state.get("plg_dyn_rows", []))
                st.session_state["plg_dyn_rows"] = dyn_rows
                if dyn_rows:
                    st.markdown("Current dynamic variables")
                    for i, r in enumerate(list(dyn_rows)):
                        c1, c2, c3, c4 = st.columns([1.2, 4, 1, 1])
                        c1.write(r.get("name", ""))
                        c2.code(r.get("template", ""))
                        if c3.button("Edit", key=f"plg_dyn_edit_{i}"):
                            dlg = getattr(st, "dialog", None)
                            def _save_edit(new_name: str, new_tmpl: str, new_width: int, idx: int):
                                rows = st.session_state.get("plg_dyn_rows", [])
                                # if name changed, ensure uniqueness by replacing the row at idx
                                if 0 <= idx < len(rows):
                                    rows[idx] = {"name": new_name.strip(), "template": new_tmpl, "width_chars": (int(new_width) if int(new_width) > 0 else None)}
                                    st.session_state["plg_dyn_rows"] = rows
                                    _dyn_save(rows)
                            if callable(dlg):
                                @dlg("Edit Dynamic Variable")
                                def _dlg_edit_var():
                                    with st.form(f"plg_edit_var_form_{i}"):
                                        rows_local = st.session_state.get("plg_dyn_rows", [])
                                        row = rows_local[i] if 0 <= i < len(rows_local) else {}
                                        if not isinstance(row, dict):
                                            if isinstance(row, str):
                                                row = {"name": row, "template": row, "width_chars": None}
                                            else:
                                                row = {"name": "", "template": "", "width_chars": None}
                                        vname_e = st.text_input("Variable name", value=str(row.get("name", "")))
                                        vtmpl_e = st.text_area("Template", value=str(row.get("template", "")))
                                        vwidth_e = st.number_input("Width (characters, optional)", min_value=0, max_value=200, value=int(row.get("width_chars") or 0), step=1)
                                        ok_e = st.form_submit_button("Save")
                                        if ok_e and vname_e.strip():
                                            _save_edit(vname_e, vtmpl_e, vwidth_e, i)
                                            # Close dialog and refresh
                                            st.session_state["_plg_dyn_edit_done"] = True
                                            st.rerun()
                                _dlg_edit_var()
                            else:
                                with st.expander(f"Edit {r.get('name','')}", expanded=True):
                                    # Use fresh row from session in case prior list had malformed entries
                                    rows_local = st.session_state.get("plg_dyn_rows", [])
                                    row = rows_local[i] if 0 <= i < len(rows_local) else {}
                                    if not isinstance(row, dict):
                                        if isinstance(row, str):
                                            row = {"name": row, "template": row, "width_chars": None}
                                        else:
                                            row = {"name": "", "template": "", "width_chars": None}
                                    vname_e = st.text_input("Variable name", value=str(row.get("name", "")), key=f"plg_edit_name_{i}")
                                    vtmpl_e = st.text_area("Template", value=str(row.get("template", "")), key=f"plg_edit_tmpl_{i}")
                                    vwidth_e = st.number_input("Width (characters, optional)", min_value=0, max_value=200, value=int(row.get("width_chars") or 0), step=1, key=f"plg_edit_width_{i}")
                                    if st.button("Save", key=f"plg_edit_save_{i}") and vname_e.strip():
                                        _save_edit(vname_e, vtmpl_e, vwidth_e, i)
                                        st.session_state["_plg_dyn_edit_done"] = True
                                        st.rerun()
                        if c4.button("Delete", key=f"plg_dyn_del_{i}"):
                            rows = st.session_state.get("plg_dyn_rows", [])
                            if 0 <= i < len(rows):
                                rows.pop(i)
                                st.session_state["plg_dyn_rows"] = rows
                                _dyn_save(rows)
                dyn_map = {r["name"]: r.get("template", "") for r in st.session_state.get("plg_dyn_rows", []) if r.get("name")}

                # ---------------- Simple field builder ----------------
                st.markdown("### Field Builder (simple)")
                st.caption("Add fields one by one: choose Header or Detail, pick type (Static / Dynamic / ItemVar / Price Class Description / Stocking / Widths / Text), enter value/name when applicable, and the target column number.")

                # Template persistence helpers
                from pathlib import Path as _Path
                def _templates_file() -> _Path:
                    return _Path("history") / "price_templates.json"
                def _load_templates() -> list[dict]:
                    p = _templates_file()
                    if p.exists():
                        try:
                            import json as _json
                            return _json.loads(p.read_text(encoding="utf-8")) or []
                        except Exception:
                            return []
                    return []
                def _save_templates(templates: list[dict]):
                    try:
                        import json as _json
                        p = _templates_file()
                        p.parent.mkdir(parents=True, exist_ok=True)
                        p.write_text(_json.dumps(templates, indent=2), encoding="utf-8")
                    except Exception:
                        pass

                # Manage fields state
                if "plg_fields" not in st.session_state:
                    st.session_state["plg_fields"] = []  # each: {section, column, type, value}

                # Add Field dialog
                add_field = st.button("Add field", key="plg_add_field_btn")
                if add_field:
                    dlg = getattr(st, "dialog", None)
                    if callable(dlg):
                        @dlg("Add Field")
                        def _dlg_add_field():
                            section = st.selectbox("Place in", ["Header", "Detail"], key="plg_field_section")
                            if section == "Header":
                                position = st.number_input("Row #", min_value=1, step=1, value=1, key="plg_field_row")
                            else:
                                position = st.number_input("Column #", min_value=1, step=1, value=1, key="plg_field_col")
                            ftype = st.selectbox("Type", ["Static", "Dynamic", "ItemVar", "Price Class Description", "Stocking", "Widths", "Text"], key="plg_field_type")
                            field_value = ""
                            preview = ""
                            # Build item field options mapping (friendly -> record key)
                            rename_map = {
                                "sku": "SKU",
                                "sku_description": "DESCRIPTION",
                                "manufacturer": "MANUFACTURER",
                                "product_line": "PRODUCT_LINE",
                                "supplier_number": "SUPPLIER",
                                "price_class": "PRICE_CLASS",
                                "cost_center": "COST_CENTER",
                            }
                            base_sample = base.head(1).to_dict(orient="records")[0] if not base.empty else {}
                            if ftype == "Static":
                                stat_vars = st.session_state.get("plg_static_vars") or {}
                                stat_names = sorted(list(stat_vars.keys()))
                                field_value = st.selectbox(
                                    "Static variable",
                                    options=stat_names if stat_names else [""],
                                    index=0,
                                    key="plg_field_static",
                                )
                                preview = str(stat_vars.get(field_value, ""))
                            elif ftype == "Dynamic":
                                choices = [""] + sorted(list(dyn_map.keys()))
                                field_value = st.selectbox("Dynamic variable", choices, index=0, key="plg_field_dynamic")
                                if field_value:
                                    preview = dyn_map.get(field_value, "")
                            elif ftype == "ItemVar":  # ItemVar
                                # Present both friendly and raw keys
                                raw_cols = dashboard.get("items", pd.DataFrame()).columns.tolist()
                                options = [("ItemNumber (SKU)", "SKU")] + [
                                    (c, rename_map.get(c, c.upper())) for c in raw_cols
                                ]
                                labels = [lbl for lbl, _ in options]
                                idx = st.selectbox("Item field", options=list(range(len(labels))), format_func=lambda i: labels[i], key="plg_field_itemvar")
                                field_value = options[idx][1]
                                preview = str(base_sample.get(field_value, ""))
                            elif ftype == "Price Class Description":
                                # Prefer aligned PC description; fall back to style derived from DESCRIPTION cluster
                                field_value = ""
                                preview = _effective_pc_desc_for_record(base_sample)
                            elif ftype == "Stocking":
                                # Per-item: IINVEN == 'Y' -> 'S', else ''
                                field_value = ""
                                try:
                                    inv_flag = str(base_sample.get("IINVEN", "")).strip().upper()
                                    if inv_flag == "Y":
                                        preview = "S"
                                    else:
                                        preview = ""
                                except Exception:
                                    preview = ""
                            elif ftype == "Text":
                                field_value = st.text_input("Text", value="", key="plg_field_text")
                                preview = field_value
                            else:  # Widths
                                # Two most common non-zero widths in inches converted to feet.inches and joined with ' & '
                                field_value = ""
                                try:
                                    pc = str(base_sample.get("PRICE_CLASS", ""))
                                    grp = items_src[items_src.get("PRICE_CLASS").astype(str) == pc]
                                    w = pd.to_numeric(grp.get("ITEM_WIDTH_INCHES"), errors="coerce").dropna()
                                    w = w[w > 0]
                                    if not w.empty:
                                        vc = w.value_counts().sort_values(ascending=False)
                                        # tie-break by smaller width: sort_index on filtered top n later
                                        top_widths = (
                                            vc.to_frame("cnt")
                                            .assign(width=vc.index)
                                            .sort_values(["cnt","width"], ascending=[False, True])
                                        )["width"].head(2).tolist()
                                        def _fmt(inches: float) -> str:
                                            feet = int(inches // 12)
                                            ins = int(round(inches - feet * 12))
                                            if ins == 12:
                                                feet += 1
                                                ins = 0
                                            return f"{feet}.{ins:02d}"
                                        formatted = [
                                            _fmt(float(x)) for x in top_widths
                                        ]
                                        preview = formatted[0] if len(formatted) == 1 else f"{formatted[0]} & {formatted[1]}"
                                    else:
                                        preview = ""
                                except Exception:
                                    preview = ""
                            st.caption(f"Preview: {preview}")
                            if st.button("Save field"):
                                rec = {
                                    "section": section,
                                    "type": ftype,
                                    "value": field_value or "",
                                }
                                if section == "Header":
                                    rec["row"] = int(position)
                                else:
                                    rec["column"] = int(position)
                                st.session_state["plg_fields"].append(rec)
                                # Rerun to refresh Current fields immediately; tabs typically persist active state
                                st.rerun()
                        _dlg_add_field()
                    else:
                        st.info("Dialog not available; please update Streamlit to enable modal dialogs.")

                # Fields table with remove buttons
                if st.session_state["plg_fields"]:
                    st.markdown("Current fields")
                    for i, fld in enumerate(list(st.session_state["plg_fields"])):
                        c1, c2, c3, c4, c5 = st.columns([1,1.5,1.2,3,1])
                        c1.write(f"{fld.get('section')}")
                        if fld.get("section") == "Header":
                            r = fld.get("row") or fld.get("column")
                            c2.write(f"Row {r}")
                        else:
                            coln = int(fld.get('column', 0) or 0)
                            # Show custom label if set
                            labels_list = st.session_state.get("plg_col_labels", []) or []
                            label = labels_list[coln-1] if 0 < coln <= len(labels_list) else ""
                            c2.write(f"Col {coln}{(' ('+label+')') if label else ''}")
                        c3.write(fld.get("type"))
                        # Preview/editor for each row
                        ftype = str(fld.get("type"))
                        if ftype == "Text":
                            # Editable free-text field bound to this row's value
                            key = f"plg_field_text_value_{i}"
                            current_val = str(fld.get("value", ""))
                            new_val = c4.text_input("", value=current_val, key=key)
                            if new_val != current_val:
                                st.session_state["plg_fields"][i]["value"] = new_val
                        elif ftype == "Static":
                            # Allow typing over the static value or variable name; if it matches a static var, it resolves, else literal
                            key = f"plg_field_static_value_{i}"
                            current_val = str(fld.get("value", ""))
                            new_val = c4.text_input("", value=current_val, key=key)
                            if new_val != current_val:
                                st.session_state["plg_fields"][i]["value"] = new_val
                        else:
                            prev_txt = ""
                            if ftype == "Dynamic":
                                prev_txt = dyn_map.get(fld.get("value", ""), "")
                            elif ftype == "ItemVar":
                                prev_txt = str(base.head(1).to_dict(orient="records")[0].get(fld.get("value"), "")) if not base.empty else ""
                            elif ftype == "Price Class Description":
                                prev_txt = _effective_pc_desc_for_record(base.head(1).to_dict(orient="records")[0]) if not base.empty else ""
                            elif ftype == "Stocking":
                                try:
                                    sample = base.head(1).to_dict(orient="records")[0] if not base.empty else {}
                                    inv_flag = str(sample.get("IINVEN", "")).strip().upper()
                                    prev_txt = "S" if inv_flag == "Y" else ""
                                except Exception:
                                    prev_txt = ""
                            elif ftype == "Widths":
                                try:
                                    sample = base.head(1).to_dict(orient="records")[0] if not base.empty else {}
                                    pc = str(sample.get("PRICE_CLASS", ""))
                                    prev_txt = str(pc_widths_map.get(pc, ""))
                                except Exception:
                                    prev_txt = ""
                            c4.write(prev_txt)
                        if c5.button("Remove", key=f"plg_rm_field_{i}"):
                            st.session_state["plg_fields"].pop(i)
                            st.rerun()

                # Template save/load
                st.markdown("### Template")
                templates = _load_templates()
                existing_names = [t.get("name") for t in templates if t.get("name")]
                sel = st.selectbox("Load existing", options=[""] + existing_names, key="plg_tpl_sel")
                col_tpl = st.columns(3)
                tpl_name = col_tpl[0].text_input("Template name", value=st.session_state.get("plg_tpl_name", ""), key="plg_tpl_name_in")
                if col_tpl[1].button("Load") and sel:
                    tpl = next((t for t in templates if t.get("name") == sel), None)
                    if tpl:
                        st.session_state["plg_tpl_name"] = tpl.get("name", "")
                        st.session_state["plg_fields"] = tpl.get("fields", [])
                        # Merge dynamic vars from template with current ones to preserve width_chars if missing in template
                        incoming = tpl.get("dynamic_vars", []) or []
                        current = st.session_state.get("plg_dyn_rows", []) or []
                        try:
                            # Build map from current by name
                            cur_map = {}
                            for x in current:
                                if isinstance(x, dict) and x.get("name"):
                                    cur_map[str(x.get("name"))] = x
                                elif isinstance(x, str):
                                    cur_map[x] = {"name": x, "template": x, "width_chars": None}
                            merged = []
                            seen = set()
                            for y in incoming:
                                if isinstance(y, dict) and y.get("name"):
                                    nm = str(y.get("name"))
                                    base = dict(y)
                                elif isinstance(y, str):
                                    nm = y
                                    base = {"name": y, "template": y, "width_chars": None}
                                else:
                                    continue
                                # If template row lacks width, inherit from current
                                if not base.get("width_chars") and nm in cur_map:
                                    w = cur_map[nm].get("width_chars")
                                    if w:
                                        base["width_chars"] = w
                                if nm not in seen:
                                    merged.append(base); seen.add(nm)
                            # Include any current rows that aren't in the template list
                            for nm, row in cur_map.items():
                                if nm not in seen:
                                    merged.append(row)
                            st.session_state["plg_dyn_rows"] = merged
                        except Exception:
                            st.session_state["plg_dyn_rows"] = incoming or current
                        st.session_state["plg_static_vars"] = tpl.get("static_vars", st.session_state.get("plg_static_vars", {}))
                        # Stage PLG UI options to apply before widget creation on next run
                        opts = tpl.get("options", {}) if isinstance(tpl.get("options"), dict) else {}
                        st.session_state["plg_restore_options"] = {
                            "mode": opts.get("mode"),
                            "detail_mode": opts.get("detail_mode"),
                            "filters": opts.get("filters", {}),
                            "filters_exclude": opts.get("filters_exclude", {}),
                            "col_labels": opts.get("col_labels", []),
                            "include_labels_excel": opts.get("include_labels_excel", False),
                            "sort_keys": opts.get("sort_keys", []),
                            "logos": opts.get("logos", []),
                            "exclude_single_pc": opts.get("exclude_single_pc", False),
                            "exclude_pc_threshold": opts.get("exclude_pc_threshold", 3),
                        }
                        st.rerun()
                if col_tpl[2].button("Save"):
                    name = tpl_name.strip() or sel
                    if name:
                        payload = {
                            "name": name,
                            "fields": st.session_state.get("plg_fields", []),
                            "dynamic_vars": st.session_state.get("plg_dyn_rows", []),
                            "static_vars": st.session_state.get("plg_static_vars", {}),
                            "version": 1,
                            "options": {
                                "mode": st.session_state.get("plg_mode", "Structured builder"),
                                "detail_mode": st.session_state.get("plg_detail_mode_struct", "Item"),
                                "filters": {
                                    "manufacturer": st.session_state.get("plg_filter_mf", []),
                                    "product_line": st.session_state.get("plg_filter_pl", []),
                                    "supplier": st.session_state.get("plg_filter_sp", []),
                                    "iinven": st.session_state.get("plg_filter_inv", []),
                                },
                                "filters_exclude": {
                                    "manufacturer": st.session_state.get("plg_exclude_mf", []),
                                    "product_line": st.session_state.get("plg_exclude_pl", []),
                                    "supplier": st.session_state.get("plg_exclude_sp", []),
                                    "iinven": st.session_state.get("plg_exclude_inv", []),
                                },
                                "col_labels": st.session_state.get("plg_col_labels", []),
                                "include_labels_excel": bool(st.session_state.get("plg_include_labels_excel", False)),
                                "sort_keys": st.session_state.get("plg_sort_keys", []),
                                "logos": st.session_state.get("plg_logos", []),
                                "exclude_single_pc": bool(st.session_state.get("plg_exclude_single_pc", False)),
                                "exclude_pc_threshold": int(st.session_state.get("plg_exclude_pc_threshold", 3)),
                            },
                        }
                        updated = [t for t in templates if t.get("name") != name]
                        updated.append(payload)
                        _save_templates(updated)
                        st.session_state["plg_tpl_name"] = name
                        st.success(f"Template '{name}' saved")

                # Column labels editor (for Detail columns)
                # Determine max detail columns currently used
                max_col_d = max([f.get("column", 0) for f in st.session_state.get("plg_fields", []) if f.get("section") == "Detail"] + [0])
                if max_col_d > 0:
                    with st.expander("Column labels (Detail)", expanded=False):
                        labels = st.session_state.get("plg_col_labels")
                        if not isinstance(labels, list) or len(labels) < max_col_d:
                            # grow list to needed size with empty defaults
                            labels = (labels or []) + [""] * (max_col_d - len(labels or []))
                        cols_lbl = st.columns(min(max_col_d, 4))
                        for i in range(max_col_d):
                            idx = i % len(cols_lbl)
                            with cols_lbl[idx]:
                                labels[i] = st.text_input(f"Col {i+1} name", value=labels[i], key=f"plg_col_label_{i+1}")
                        st.session_state["plg_col_labels"] = labels
                        st.session_state["plg_include_labels_excel"] = st.checkbox("Include labels row in Excel", value=bool(st.session_state.get("plg_include_labels_excel", False)))

                # Sort editor (for Detail rows)
                if max_col_d > 0:
                    with st.expander("Sort (Detail)", expanded=False):
                        labels = st.session_state.get("plg_col_labels") if isinstance(st.session_state.get("plg_col_labels"), list) else []
                        col_choices = ["(none)"] + [
                            (labels[i] if i < len(labels) and str(labels[i]).strip() else f"Col{i+1}")
                            for i in range(max_col_d)
                        ]
                        sort_keys = st.session_state.get("plg_sort_keys")
                        if not isinstance(sort_keys, list):
                            sort_keys = []
                        new_sort = []
                        for lvl in range(3):
                            c1, c2 = st.columns([3,1])
                            pre_col = sort_keys[lvl]["col"] if lvl < len(sort_keys) and isinstance(sort_keys[lvl], dict) else 0
                            pre_lab = "(none)" if pre_col == 0 else (labels[pre_col-1] if pre_col-1 < len(labels) and str(labels[pre_col-1]).strip() else f"Col{pre_col}")
                            sel = c1.selectbox(f"Level {lvl+1} column", options=col_choices, index=(col_choices.index(pre_lab) if pre_lab in col_choices else 0), key=f"plg_sort_col_{lvl+1}")
                            desc = c2.checkbox("Desc", value=(not sort_keys[lvl]["asc"]) if lvl < len(sort_keys) and isinstance(sort_keys[lvl], dict) else False, key=f"plg_sort_desc_{lvl+1}")
                            if sel != "(none)":
                                try:
                                    col_idx = col_choices.index(sel)  # 1..max_col_d
                                    col_num = col_idx
                                except Exception:
                                    col_num = 0
                                if 1 <= col_num <= max_col_d:
                                    new_sort.append({"col": col_num, "asc": not desc})
                        st.session_state["plg_sort_keys"] = new_sort

                # Logos editor (up to 3), saved with template
                with st.expander("Logos (up to 3)", expanded=False):
                    logos = st.session_state.get("plg_logos")
                    if not isinstance(logos, list):
                        logos = []
                    # Existing logos list with position controls and remove buttons
                    for i in range(len(logos)):
                        lg = logos[i]
                        c1, c2, c3, c4 = st.columns([1.2, 2, 3, 1])
                        c1.write(f"Logo {i+1}")
                        # Position selector
                        current_pos = str(lg.get("position", "Left"))
                        try:
                            idx = ["Left","Center","Right"].index(current_pos if current_pos in ["Left","Center","Right"] else "Left")
                        except Exception:
                            idx = 0
                        pos = c2.selectbox("Position", options=["Left","Center","Right"], index=idx, key=f"plg_logo_pos_{i}")
                        logos[i]["position"] = pos
                        # Preview image
                        try:
                            import base64 as _b64
                            data_b64 = lg.get("data_b64", "")
                            if data_b64:
                                c3.image(_b64.b64decode(data_b64), use_column_width=True)
                        except Exception:
                            pass
                        if c4.button("Remove", key=f"plg_logo_rm_{i}"):
                            logos.pop(i)
                            st.session_state["plg_logos"] = logos
                            st.rerun()
                    # Add new logo controls
                    c_add = st.columns(3)
                    with c_add[0]:
                        pos_new = st.selectbox("New logo position", ["Left","Center","Right"], key="plg_logo_new_pos")
                    with c_add[1]:
                        up = st.file_uploader("Upload logo (PNG/JPG)", type=["png","jpg","jpeg"], key="plg_logo_new_file")
                    with c_add[2]:
                        if st.button("Add logo") and up is not None:
                            try:
                                import base64 as _b64
                                data = up.read()
                                logos.append({"position": pos_new, "data_b64": _b64.b64encode(data).decode("ascii"), "name": getattr(up, 'name', 'logo')})
                                logos = logos[:3]  # keep at most 3
                                st.session_state["plg_logos"] = logos
                                st.success("Logo added")
                            except Exception:
                                st.warning("Could not read the uploaded image.")
                    st.session_state["plg_logos"] = logos

                # Detail scope
                detail_scope = st.radio("Detail rows by", ["Item", "Price Class"], horizontal=True, key="plg_detail_mode_struct")

                # Template renderer re-used
                def render_template(t: str, stat: dict, row: dict) -> str:
                    env = {**stat, **row}
                    try:
                        return str(t).format(**env)
                    except Exception:
                        return str(t)

                # Helper maps for PC-level fields
                def _stocking_for_pc_map(df: pd.DataFrame) -> dict:
                    out = {}
                    if df.empty:
                        return out
                    try:
                        tmp = df.copy()
                        tmp["PRICE_CLASS"] = tmp.get("PRICE_CLASS").astype(str)
                        inv = tmp.get("IINVEN").astype(str).str.upper().str.strip()
                        tmp["IINVEN_Y"] = (inv == "Y").astype(int)
                        grp = tmp.groupby("PRICE_CLASS")["IINVEN_Y"].agg(["sum","count"]).reset_index()
                        for _, r in grp.iterrows():
                            y = int(r["sum"])
                            c = int(r["count"])
                            # Mark Stocking at the price-class level if ANY item in the class is stocking
                            out[str(r["PRICE_CLASS"])]= "S" if c > 0 and y > 0 else ""
                    except Exception:
                        pass
                    return out

                def _widths_for_pc_map(df: pd.DataFrame) -> dict:
                    out = {}
                    if df.empty:
                        return out
                    try:
                        tmp = df.copy()
                        tmp["PRICE_CLASS"] = tmp.get("PRICE_CLASS").astype(str)
                        w = pd.to_numeric(tmp.get("ITEM_WIDTH_INCHES"), errors="coerce")
                        tmp["ITEM_WIDTH_INCHES"] = w
                        def _fmt(inches: float) -> str:
                            feet = int(inches // 12)
                            ins = int(round(inches - feet * 12))
                            if ins == 12:
                                feet += 1
                                ins = 0
                            return f"{feet}.{ins:02d}"
                        for pc, g in tmp.groupby("PRICE_CLASS"):
                            ww = pd.to_numeric(g.get("ITEM_WIDTH_INCHES"), errors="coerce").dropna()
                            ww = ww[ww > 0]
                            if ww.empty:
                                out[str(pc)] = ""
                                continue
                            vc = ww.value_counts().sort_values(ascending=False)
                            top_widths = (
                                vc.to_frame("cnt").assign(width=vc.index)
                                .sort_values(["cnt","width"], ascending=[False, True])
                            )["width"].head(2).tolist()
                            formatted = [_fmt(float(x)) for x in top_widths]
                            out[str(pc)] = formatted[0] if len(formatted)==1 else f"{formatted[0]} & {formatted[1]}"
                    except Exception:
                        pass
                    return out

                pc_stocking_map = _stocking_for_pc_map(base)
                pc_widths_map = _widths_for_pc_map(base)

                # Build header (single wide column) based on fields
                max_row_hdr = max([(f.get("row") or f.get("column") or 0) for f in st.session_state.get("plg_fields", []) if f.get("section") == "Header"] + [0])
                header_rows = [""] * max(1, max_row_hdr)
                for fld in st.session_state.get("plg_fields", []):
                    if fld.get("section") != "Header":
                        continue
                    row_idx = int((fld.get("row") or fld.get("column") or 1)) - 1
                    if fld.get("type") == "Static":
                        _name = fld.get("value", "")
                        _stat = static_vars or {}
                        sval = _stat[_name] if _name in _stat else _name
                        header_rows[row_idx:row_idx+1] = [str(sval)]
                    elif fld.get("type") == "Dynamic":
                        tmpl = dyn_map.get(fld.get("value", ""), "")
                        sample = base.head(1).to_dict(orient="records")[0] if not base.empty else {}
                        t_eff = tmpl
                        if "ItemNumber" in t_eff:
                            sku_rep = str(sample.get("SKU", ""))
                            t_eff = t_eff.replace("ItemNumber", sku_rep)
                        header_rows[row_idx:row_idx+1] = [render_template(t_eff, static_vars, sample)]
                    elif fld.get("type") == "ItemVar":
                        sample = base.head(1).to_dict(orient="records")[0] if not base.empty else {}
                        header_rows[row_idx:row_idx+1] = [str(sample.get(fld.get("value", ""), ""))]
                    elif fld.get("type") == "Price Class Description":
                        sample = base.head(1).to_dict(orient="records")[0] if not base.empty else {}
                        header_rows[row_idx:row_idx+1] = [_effective_pc_desc_for_record(sample)]
                    elif fld.get("type") == "Stocking":
                        sample = base.head(1).to_dict(orient="records")[0] if not base.empty else {}
                        if detail_scope == "Price Class":
                            pc = str(sample.get("PRICE_CLASS", ""))
                            header_rows[row_idx:row_idx+1] = [str(pc_stocking_map.get(pc, ""))]
                        else:
                            inv_flag = str(sample.get("IINVEN", "")).strip().upper()
                            header_rows[row_idx:row_idx+1] = ["S" if inv_flag == "Y" else ""]
                    elif fld.get("type") == "Text":
                        header_rows[row_idx:row_idx+1] = [str(fld.get("value", ""))]
                    else:  # Widths
                        sample = base.head(1).to_dict(orient="records")[0] if not base.empty else {}
                        pc = str(sample.get("PRICE_CLASS", ""))
                        header_rows[row_idx:row_idx+1] = [str(pc_widths_map.get(pc, ""))]

                # Build detail rows from fields
                detail_rows = []
                if not base.empty:
                    # Respect the user's selected scope; do not auto-force Price Class
                    effective_scope = detail_scope
                    if effective_scope == "Item":
                        iter_rows = base.to_dict(orient="records")
                    else:
                        # Build one detail row per DISTINCT style within each Price Class
                        grp = base.groupby("PRICE_CLASS", dropna=False)
                        iter_rows = []
                        for pc, g in grp:
                            if g.empty:
                                continue
                            try:
                                gg = g.copy()
                                gg["_SKU_STR_"] = gg.get("SKU").astype(str)
                                gg["_IPATT_"] = gg.get("ITEM_PATTERN").fillna("").astype(str)
                                # Prefer grouping by ITEM_PATTERN (IPATT): exactly one row per IPATT
                                by_ipatt = gg["_IPATT_"].str.strip() != ""
                                if by_ipatt.any():
                                    # Build cluster keys that collapse IPATTs which only differ by a 1–2 digit suffix
                                    sub = gg.loc[by_ipatt].copy()
                                    ip_list = sub["_IPATT_"].astype(str).str.strip().unique().tolist()
                                    # Pre-compute cluster bases: for each IPATT s, find the shortest other IPATT p
                                    # such that s starts with p and the remainder is 1–2 digits (e.g., s=RZ3015, p=RZ30)
                                    def _cluster_base(s: str, candidates: list[str]) -> str:
                                        s = str(s)
                                        best = s
                                        for p in candidates:
                                            if p == s:
                                                continue
                                            if s.startswith(p):
                                                rest = s[len(p):]
                                                if rest.isdigit() and 1 <= len(rest) <= 2:
                                                    # prefer the shortest base prefix
                                                    if len(p) < len(best):
                                                        best = p
                                        return best
                                    cluster_map = {s: _cluster_base(s, ip_list) for s in ip_list}
                                    sub["_IP_CLUS_"] = sub["_IPATT_"].map(lambda s: cluster_map.get(str(s), str(s)))
                                    for ip_key, g_ip in sub.groupby("_IP_CLUS_", dropna=False):
                                        # pick representative row; prefer exact IPATT == cluster base if present
                                        try:
                                            row = g_ip.loc[g_ip["_IPATT_"].astype(str) == str(ip_key)].iloc[0]
                                        except Exception:
                                            row = g_ip.iloc[0]
                                        rec = {k: (row[k] if k in gg.columns else None) for k in gg.columns}
                                        rec["PRICE_CLASS"] = pc
                                        sku_rep = str(row.get("SKU", ""))
                                        rec["STYLE_DERIVED"] = str(style_map_by_sku.get(sku_rep, "")).strip()
                                        rec.setdefault("SKU", sku_rep)
                                        rec.setdefault("DESCRIPTION", str(row.get("DESCRIPTION", f"Price Class {pc}")))
                                        iter_rows.append(rec)
                                    # Also consider any rows missing IPATT: collapse them to a single fallback row
                                    g_no_ip = gg.loc[~by_ipatt]
                                    if not g_no_ip.empty:
                                        row = g_no_ip.iloc[0]
                                        rec = {k: (row[k] if k in gg.columns else None) for k in gg.columns}
                                        rec["PRICE_CLASS"] = pc
                                        sku_rep = str(row.get("SKU", ""))
                                        rec["STYLE_DERIVED"] = str(style_map_by_sku.get(sku_rep, "")).strip()
                                        rec.setdefault("SKU", sku_rep)
                                        rec.setdefault("DESCRIPTION", str(row.get("DESCRIPTION", f"Price Class {pc}")))
                                        iter_rows.append(rec)
                                else:
                                    # If no IPATT available, dedupe by derived style as before
                                    gg["STYLE_DERIVED"] = gg["_SKU_STR_"].map(lambda s: style_map_by_sku.get(s, "")).fillna("")
                                    style_rows = {}
                                    for _, row in gg.iterrows():
                                        sty = str(row.get("STYLE_DERIVED", ""))
                                        if not sty or not str(sty).strip():
                                            continue
                                        key = str(sty).strip().lower()
                                        if key not in style_rows:
                                            style_rows[key] = (sty, row)
                                    if not style_rows:
                                        row = gg.iloc[0]
                                        rec = {k: (row[k] if k in gg.columns else None) for k in gg.columns}
                                        rec["PRICE_CLASS"] = pc
                                        rec.setdefault("SKU", str(row.get("SKU", "")))
                                        rec.setdefault("DESCRIPTION", f"Price Class {pc}")
                                        iter_rows.append(rec)
                                    else:
                                        for key, (sty, row) in style_rows.items():
                                            rec = {k: (row[k] if k in gg.columns else None) for k in gg.columns}
                                            rec["PRICE_CLASS"] = pc
                                            rec["STYLE_DERIVED"] = str(sty).strip()
                                            rec.setdefault("SKU", str(row.get("SKU", "")))
                                            rec.setdefault("DESCRIPTION", str(row.get("DESCRIPTION", f"Price Class {pc}")))
                                            iter_rows.append(rec)
                            except Exception:
                                # Safe fallback to one row per price class on any error
                                row = g.iloc[0]
                                rec = {k: (row[k] if k in g.columns else None) for k in g.columns}
                                rec["PRICE_CLASS"] = pc
                                rec.setdefault("SKU", str(row.get("SKU", "")))
                                rec.setdefault("DESCRIPTION", f"Price Class {pc}")
                                iter_rows.append(rec)
                    # Determine detail row width
                    # Post-processing: Disambiguate duplicate Style names across iter_rows
                    try:
                        # Build base style text for each row using the same resolver as rendering
                        styles_base = []  # list of (index, base_style)
                        for idx, rec in enumerate(iter_rows):
                            try:
                                base_style = _effective_pc_desc_for_record(rec)
                            except Exception:
                                base_style = str(rec.get("PRICE_CLASS_DESC", "")) or str(rec.get("DESCRIPTION", ""))
                            styles_base.append((idx, str(base_style).strip()))
                        # Group rows by normalized base style
                        from collections import defaultdict as _dd
                        buckets = _dd(list)
                        def _norm_key(s: str) -> str:
                            try:
                                return str(s or "").strip().lower()
                            except Exception:
                                return str(s or "").lower()
                        for idx, txt in styles_base:
                            if txt:
                                buckets[_norm_key(txt)].append(idx)
                        # For each group with duplicates, append manufacturer or widths
                        for norm_key, idx_list in buckets.items():
                            if len(idx_list) <= 1:
                                # Singletons: keep base style as display
                                i0 = idx_list[0]
                                rec0 = iter_rows[i0]
                                rec0["STYLE_DISPLAY"] = styles_base[i0][1]
                                continue
                            # Collect manufacturers present in this group
                            mfs = set()
                            for i in idx_list:
                                mf = str(iter_rows[i].get("MANUFACTURER", "")).strip()
                                if mf:
                                    mfs.add(mf.upper())
                            multiple_mf = len(mfs) > 1
                            for i in idx_list:
                                rec = iter_rows[i]
                                base_txt = styles_base[i][1]
                                if not base_txt:
                                    continue
                                disp = base_txt
                                if multiple_mf:
                                    mf = str(rec.get("MANUFACTURER", "")).strip()
                                    if mf:
                                        disp = f"{base_txt} ({mf})"
                                else:
                                    # Same manufacturer: append width string for the price class if available
                                    pc = str(rec.get("PRICE_CLASS", ""))
                                    wtxt = str(pc_widths_map.get(pc, ""))
                                    if wtxt:
                                        # Avoid duplicating width if it's already present (case-insensitive)
                                        if wtxt.lower() not in base_txt.lower():
                                            disp = f"{base_txt} {wtxt}"
                                rec["STYLE_DISPLAY"] = disp
                        # Ensure any rows not covered (e.g., empty style) still have STYLE_DISPLAY set
                        for idx, txt in styles_base:
                            if "STYLE_DISPLAY" not in iter_rows[idx]:
                                iter_rows[idx]["STYLE_DISPLAY"] = txt
                    except Exception:
                        # On any error, skip disambiguation gracefully
                        pass

                    # Determine detail row width
                    max_col_d = max([f.get("column", 0) for f in st.session_state.get("plg_fields", []) if f.get("section") == "Detail"] + [0])
                    for rec in iter_rows:
                        row_vals = [""] * max(1, max_col_d)
                        for fld in st.session_state.get("plg_fields", []):
                            if fld.get("section") != "Detail":
                                continue
                            col_idx = int(fld.get("column", 1)) - 1
                            if fld.get("type") == "Static":
                                _name = fld.get("value", "")
                                _stat = static_vars or {}
                                row_vals[col_idx] = str(_stat[_name] if _name in _stat else _name)
                            elif fld.get("type") == "Dynamic":
                                tmpl = dyn_map.get(fld.get("value", ""), "")
                                t_eff = tmpl
                                if "ItemNumber" in t_eff:
                                    sku_rep = str(rec.get("SKU", ""))
                                    t_eff = t_eff.replace("ItemNumber", sku_rep)
                                row_vals[col_idx] = render_template(t_eff, static_vars, rec)
                            elif fld.get("type") == "ItemVar":
                                fld_name = str(fld.get("value", ""))
                                if fld_name.upper() == "PRICE_CLASS_DESC" and rec.get("STYLE_DISPLAY"):
                                    row_vals[col_idx] = str(rec.get("STYLE_DISPLAY", ""))
                                else:
                                    row_vals[col_idx] = str(rec.get(fld_name, ""))
                            elif fld.get("type") == "Price Class Description":
                                # Use computed display style when available
                                txt_style = str(rec.get("STYLE_DISPLAY", "")).strip() or _effective_pc_desc_for_record(rec)
                                try:
                                    cc = str(rec.get("COST_CENTER", "")).strip()
                                    if cc in ("013","13"):
                                        # If description ends with a dimension (e.g., 19.69x19.69), insert it before (Tile)
                                        import re as _re
                                        desc = str(rec.get("DESCRIPTION", ""))
                                        m_dim = _re.search(r"(\d+(?:\.\d+)?)\s*[xX]\s*(\d+(?:\.\d+)?)\s*$", desc.strip())
                                        dim_txt = None
                                        if m_dim:
                                            a, b = m_dim.group(1), m_dim.group(2)
                                            dim_txt = f"{a}x{b}"
                                        if dim_txt:
                                            txt_style = f"{txt_style} {dim_txt} (Tile)"
                                        else:
                                            txt_style = f"{txt_style} (Tile)"
                                except Exception:
                                    pass
                                row_vals[col_idx] = txt_style
                            elif fld.get("type") == "Stocking":
                                # For Item scope, show per-item stocking. For Price Class scope, use the per-class map
                                if effective_scope == "Item":
                                    inv_flag = str(rec.get("IINVEN", "")).strip().upper()
                                    row_vals[col_idx] = "S" if inv_flag == "Y" else ""
                                else:
                                    pc = str(rec.get("PRICE_CLASS", ""))
                                    row_vals[col_idx] = str(pc_stocking_map.get(pc, ""))
                            elif fld.get("type") == "Text":
                                row_vals[col_idx] = str(fld.get("value", ""))
                            else:  # Widths
                                pc = str(rec.get("PRICE_CLASS", ""))
                                row_vals[col_idx] = str(pc_widths_map.get(pc, ""))
                        detail_rows.append(row_vals)

                # DataFrames for preview
                # Build DataFrames for preview
                # Detail columns use custom labels when provided; header is a single wide column (one column many rows)
                if detail_rows:
                    ncols = len(detail_rows[0])
                    labels = st.session_state.get("plg_col_labels") if isinstance(st.session_state.get("plg_col_labels"), list) else []
                    detail_cols = [
                        (labels[i] if i < len(labels) and str(labels[i]).strip() else f"Col{i+1}")
                        for i in range(ncols)
                    ]
                else:
                    detail_cols = []
                detail_df = pd.DataFrame(detail_rows, columns=detail_cols) if detail_cols else pd.DataFrame()
                # Apply preview sort
                if not detail_df.empty:
                    skeys = st.session_state.get("plg_sort_keys")
                    if isinstance(skeys, list) and skeys:
                        by_cols = []
                        ascending = []
                        for sk in skeys:
                            try:
                                idx = int(sk.get("col", 0)) - 1
                                if 0 <= idx < len(detail_cols):
                                    by_cols.append(detail_cols[idx])
                                    ascending.append(bool(sk.get("asc", True)))
                            except Exception:
                                continue
                        if by_cols:
                            try:
                                detail_df = detail_df.sort_values(by=by_cols, ascending=ascending, na_position="last", kind="mergesort")
                            except Exception:
                                pass
                header_df = pd.DataFrame({"Col1": header_rows}) if header_rows else pd.DataFrame()

                st.markdown("### Preview & Export")
                # Render logos above the header preview
                logos = st.session_state.get("plg_logos")
                if isinstance(logos, list) and logos:
                    cols_logo = st.columns(3)
                    # For each position, ensure consistent ordering
                    pos_buckets = {"left": [], "center": [], "right": []}
                    for lg in logos:
                        try:
                            pos = str(lg.get("position", "Left")).lower()
                            data_b64 = lg.get("data_b64", "")
                            if not data_b64:
                                continue
                            pos_key = "left" if pos.startswith("l") else ("center" if pos.startswith("c") or pos.startswith("m") else "right")
                            pos_buckets[pos_key].append(lg)
                        except Exception:
                            continue
                    import base64 as _b64
                    for pkey, col_idx in [("left",0),("center",1),("right",2)]:
                        for lg in pos_buckets.get(pkey, []):
                            try:
                                cols_logo[col_idx].image(_b64.b64decode(lg.get("data_b64","")), use_column_width=True)
                            except Exception:
                                pass
                st.markdown("#### Header (preview)")
                if not header_df.empty:
                    st.dataframe(header_df, use_container_width=True, hide_index=True)
                else:
                    st.write("(no header)")
                st.markdown("#### Detail (preview)")
                if not detail_df.empty:
                    st.dataframe(detail_df, use_container_width=True, hide_index=True)
                else:
                    st.write("(no detail rows)")

                # Export filename input (persisted)
                default_fname = st.session_state.get("plg_export_filename", "template.xlsx")
                fname = st.text_input("Export file name", value=default_fname, key="plg_export_filename")
                # Normalise extension
                try:
                    if isinstance(fname, str) and not fname.lower().endswith(".xlsx"):
                        fname = f"{fname}.xlsx"
                        st.session_state["plg_export_filename"] = fname
                except Exception:
                    pass

                def export_excel_bytes(hdf: pd.DataFrame, ddf: pd.DataFrame) -> bytes:
                    from io import BytesIO
                    out = BytesIO()
                    with pd.ExcelWriter(out, engine="openpyxl") as writer:
                        # Reserve top rows for logos if any are configured
                        logos = st.session_state.get("plg_logos")
                        image_rows = 6 if (isinstance(logos, list) and len(logos) > 0) else 0
                        start_row = image_rows
                        sheet_name = "PriceList"
                        header_rows_written = 0
                        if not hdf.empty:
                            hdf.to_excel(writer, index=False, header=False, sheet_name=sheet_name, startrow=start_row)
                            header_rows_written = len(hdf)
                            start_row += header_rows_written
                        if not ddf.empty:
                            if start_row > 0:
                                start_row += 1  # blank line between header and detail
                            # Apply sort to export (using current labels -> column mapping)
                            try:
                                skeys = st.session_state.get("plg_sort_keys")
                                if isinstance(skeys, list) and skeys and isinstance(ddf, pd.DataFrame):
                                    labels = list(ddf.columns)
                                    by_cols = []
                                    ascending = []
                                    for sk in skeys:
                                        idx = int(sk.get("col", 0)) - 1
                                        if 0 <= idx < len(labels):
                                            by_cols.append(labels[idx])
                                            ascending.append(bool(sk.get("asc", True)))
                                    if by_cols:
                                        ddf = ddf.sort_values(by=by_cols, ascending=ascending, na_position="last", kind="mergesort")
                            except Exception:
                                pass
                            # Optional labels row
                            include_labels = bool(st.session_state.get("plg_include_labels_excel", False))
                            if include_labels:
                                labels_row = pd.DataFrame([list(ddf.columns)])
                                labels_row.to_excel(writer, index=False, header=False, sheet_name=sheet_name, startrow=start_row)
                                start_row += 1
                            ddf.to_excel(writer, index=False, header=False, sheet_name=sheet_name, startrow=start_row)

                        # Post-process: merge header rows across detail width and center text
                        try:
                            from openpyxl.utils import get_column_letter
                            from openpyxl.styles import Alignment
                            wb = writer.book
                            ws = wb[sheet_name]
                            detail_cols_count = ddf.shape[1] if isinstance(ddf, pd.DataFrame) else 0
                            max_cols = max(1, detail_cols_count)
                            last_col_letter = get_column_letter(max_cols)
                            # Merge each header row A{r} : {last_col}{r}
                            for i in range(header_rows_written):
                                row_num = image_rows + 1 + i  # offset by image rows
                                rng = f"A{row_num}:{last_col_letter}{row_num}"
                                if max_cols > 1:
                                    ws.merge_cells(rng)
                                cell = ws[f"A{row_num}"]
                                cell.alignment = Alignment(horizontal="center")
                            # Insert logos at top row across Left/Center/Right positions
                            if isinstance(logos, list) and logos:
                                try:
                                    from openpyxl.drawing.image import Image as XLImage
                                    from PIL import Image as PILImage
                                    from io import BytesIO as _BIO
                                    import base64 as _b64
                                    # Determine anchor columns
                                    mid_col_letter = get_column_letter(int(max_cols/2) if max_cols > 2 else 2)
                                    anchors = {"left": "A1", "center": f"{mid_col_letter}1", "right": f"{last_col_letter}1"}
                                    for lg in logos:
                                        try:
                                            pos = str(lg.get("position", "Left")).lower()
                                            key = "left" if pos.startswith("l") else ("center" if pos.startswith("c") or pos.startswith("m") else "right")
                                            data_b64 = lg.get("data_b64", "")
                                            if not data_b64:
                                                continue
                                            img_bytes = _b64.b64decode(data_b64)
                                            pil_img = PILImage.open(_BIO(img_bytes))
                                            xl_img = XLImage(pil_img)
                                            ws.add_image(xl_img, anchors.get(key, "A1"))
                                        except Exception:
                                            continue
                                except Exception:
                                    # Pillow may be missing; skip embedding logos in Excel
                                    pass
                            # Column widths for detail section (only)
                            # Strategy:
                            # - Identify which detail columns contain Dynamic fields; set those to fixed width 14 chars
                            # - For non-dynamic columns: auto-fit to the longest text in that detail column, or the column label if wider, plus one space
                            # - Finally, set page setup to fit all columns to one printed page width (Letter 8.5x11)
                            try:
                                fields = st.session_state.get("plg_fields", [])
                                # Map detail column -> flag if any Dynamic field assigned
                                dyn_col: dict[int, bool] = {}
                                for f in fields:
                                    if f.get("section") == "Detail" and f.get("type") == "Dynamic":
                                        col = int(f.get("column", 0))
                                        if col > 0:
                                            dyn_col[col] = True
                                # Column labels (whether or not we include a labels row in Excel)
                                labels_list = st.session_state.get("plg_col_labels", []) or []
                                # Determine header offset (top rows used by logos + header + optional blank line + optional labels row)
                                detail_start_excel_row = image_rows + header_rows_written + (1 if header_rows_written > 0 else 0) + (1 if bool(st.session_state.get("plg_include_labels_excel", False)) else 0) + 1
                                for ci in range(1, max_cols + 1):
                                    col_letter = get_column_letter(ci)
                                    # Dynamic columns: fixed 14 characters wide
                                    if dyn_col.get(ci, False):
                                        ws.column_dimensions[col_letter].width = 14.0
                                    else:
                                        # Auto-fit: compute max text length within detail rows for this column
                                        max_len = 0
                                        try:
                                            for r_idx in range(ddf.shape[0]):
                                                val = ddf.iloc[r_idx, ci-1]
                                                txt = "" if val is None else str(val)
                                                if len(txt) > max_len:
                                                    max_len = len(txt)
                                        except Exception:
                                            max_len = 10
                                        # Include header label width if present
                                        label_len = 0
                                        try:
                                            if 0 <= (ci-1) < len(labels_list):
                                                label_txt = str(labels_list[ci-1] or "")
                                                label_len = len(label_txt)
                                        except Exception:
                                            label_len = 0
                                        target = max(max_len, label_len) + 1
                                        ws.column_dimensions[col_letter].width = float(max(6, min(target, 60)))
                                # Ensure the printed page fits width-wise for Letter paper
                                try:
                                    ws.page_setup.fitToWidth = 1
                                    ws.page_setup.fitToHeight = 0
                                    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
                                    ws.sheet_properties.pageSetUpPr.fitToPage = True
                                except Exception:
                                    pass
                            except Exception:
                                pass
                        except Exception:
                            pass
                    out.seek(0)
                    return out.read()

                if st.button("Generate Excel", type="primary"):
                    xls = export_excel_bytes(header_df, detail_df)
                    dl_name = st.session_state.get("plg_export_filename", "template.xlsx")
                    try:
                        if not isinstance(dl_name, str) or not dl_name.strip():
                            dl_name = "template.xlsx"
                        if not dl_name.lower().endswith(".xlsx"):
                            dl_name = f"{dl_name}.xlsx"
                    except Exception:
                        dl_name = "template.xlsx"
                    st.download_button("Download " + dl_name, data=xls, file_name=dl_name, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

            else:
                # Advanced (legacy) free-form templates retained for power users
                # Static variables editor
                st.markdown("### Static variables (legacy)")
                st.caption("Define constants you can reference in templates with {VARNAME}.")
                import json, ast
                static_default = st.session_state.get("plg_static_vars", {"MARGIN": "0.40", "TERMS": "NET30"})
                static_json = st.text_area("Static variables (JSON)", value=json.dumps(static_default, indent=2), height=160, key="plg_static_json")
                try:
                    static_vars = json.loads(static_json)
                except Exception:
                    try:
                        static_vars = ast.literal_eval(static_json)
                    except Exception:
                        static_vars = st.session_state.get("plg_static_vars", {})
                        st.warning("Unable to parse JSON. Using last valid static variables.")
                if not isinstance(static_vars, dict):
                    static_vars = st.session_state.get("plg_static_vars", {})
                    st.warning("Static variables must be a JSON object (key:value map). Using last valid values.")
                st.session_state["plg_static_vars"] = static_vars

                # Template editors
                st.markdown("### Templates (legacy)")
                col_t = st.columns(2)
                header_tmpl = col_t[0].text_area(
                    "Header template (one or more lines)",
                    value=st.session_state.get("plg_header_tmpl", "PRICE_LIST\nEFFECTIVE_DATE,{EFFECTIVE_DATE}\nSUPPLIER,{SUPPLIER}"),
                    height=160,
                    key="plg_header_tmpl_in",
                )
                detail_tmpl = col_t[1].text_area(
                    "Detail template (single row template)",
                    value=st.session_state.get("plg_detail_tmpl", "{SKU},{DESCRIPTION},{PRICE_CLASS},{MANUFACTURER},{PRODUCT_LINE},{SUPPLIER},{MARGIN}"),
                    height=160,
                    key="plg_detail_tmpl_in",
                )
                st.session_state["plg_header_tmpl"] = header_tmpl
                st.session_state["plg_detail_tmpl"] = detail_tmpl

                # Helper: render a single template with dict + row context
                def render_template(t: str, stat: dict, row: dict) -> str:
                    env = {**{k: v for k, v in stat.items()}, **{k: v for k, v in row.items()}}
                    try:
                        return t.format(**env)
                    except KeyError as e:
                        missing = str(e).strip("{}")
                        return t.replace("{"+missing+"}", "")
                    except Exception:
                        return t

                # Show detected tokens in templates to help authoring
                import string as _string
                def detect_tokens(t: str) -> list[str]:
                    toks = []
                    for lit, fld, fmt, conv in _string.Formatter().parse(t):
                        if fld:
                            toks.append(fld)
                    seen = set(); out = []
                    for x in toks:
                        if x not in seen:
                            seen.add(x); out.append(x)
                    return out
                tok_cols = st.columns(2)
                tok_cols[0].markdown("Detected header tokens: " + ", ".join(detect_tokens(header_tmpl)) if header_tmpl else "")
                tok_cols[1].markdown("Detected detail tokens: " + ", ".join(detect_tokens(detail_tmpl)) if detail_tmpl else "")

                # Build header rows
                st.markdown("### Preview & Export")
                import csv
                from io import StringIO
                header_filled = render_template(header_tmpl, static_vars, base.head(1).to_dict(orient="records")[0] if not base.empty else {})
                header_lines = [ln for ln in header_filled.splitlines() if ln.strip()]
                header_rows = []
                for ln in header_lines:
                    try:
                        reader = csv.reader(StringIO(ln))
                        header_rows.extend(list(reader))
                    except Exception:
                        header_rows.append([ln])

                # Detail rows
                detail_mode_legacy = st.radio("Detail rows by", ["Item", "Price Class"], horizontal=True, key="plg_detail_mode_legacy")
                detail_rows = []
                if not base.empty:
                    if detail_mode_legacy == "Item":
                        iter_rows = base.to_dict(orient="records")
                    else:
                        grp = base.groupby("PRICE_CLASS", dropna=False)
                        iter_rows = []
                        for pc, g in grp:
                            if not g.empty:
                                rec = {k: (g[k].iloc[0] if k in g.columns else None) for k in g.columns}
                                rec["PRICE_CLASS"] = pc
                                rec.setdefault("SKU", "")
                                rec.setdefault("DESCRIPTION", f"Price Class {pc}")
                                iter_rows.append(rec)
                    for rec in iter_rows:
                        row_str = render_template(detail_tmpl, static_vars, rec)
                        try:
                            reader = csv.reader(StringIO(row_str))
                            parsed = list(reader)
                            if parsed:
                                detail_rows.append(parsed[0])
                            else:
                                detail_rows.append([row_str])
                        except Exception:
                            detail_rows.append([row_str])

                def rows_to_df(rows: list[list[str]]) -> pd.DataFrame:
                    if not rows:
                        return pd.DataFrame()
                    width = max(len(r) for r in rows)
                    norm = [r + [""] * (width - len(r)) for r in rows]
                    cols = [f"Col{i+1}" for i in range(width)]
                    return pd.DataFrame(norm, columns=cols)

                header_df = rows_to_df(header_rows)
                detail_df = rows_to_df(detail_rows)

                st.markdown("#### Header (preview)")
                if not header_df.empty:
                    st.dataframe(header_df, use_container_width=True, hide_index=True)
                else:
                    st.write("(no header rows)")
                st.markdown("#### Detail (preview)")
                if not detail_df.empty:
                    st.dataframe(detail_df, use_container_width=True, hide_index=True)
                else:
                    st.write("(no detail rows)")

                def export_excel_bytes(hdf: pd.DataFrame, ddf: pd.DataFrame) -> bytes:
                    from io import BytesIO
                    out = BytesIO()
                    with pd.ExcelWriter(out, engine="openpyxl") as writer:
                        start_row = 0
                        if not hdf.empty:
                            hdf.to_excel(writer, index=False, header=False, sheet_name="PriceList", startrow=start_row)
                            start_row += len(hdf) + 1
                        if not ddf.empty:
                            ddf.to_excel(writer, index=False, header=False, sheet_name="PriceList", startrow=start_row)
                    out.seek(0)
                    return out.read()

                if st.button("Generate Excel", type="primary"):
                    xls = export_excel_bytes(header_df, detail_df)
                    st.download_button("Download template.xlsx", data=xls, file_name="price_list_template.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


    # -------------------------
    # Open Orders (Tab 6)
    # -------------------------
    with tabs[5]:
        st.subheader("Open Orders")
        # Only respond to the Cost Center selection from the sidebar; ignore other filters
        cc_sel = selected_cost_centers
        if not cc_sel:
            st.info("Select at least one cost center in the sidebar to view Open Orders.")
        else:
            # Order Date filter: default to today, exact match
            default_dt = date.today()
            od = st.date_input("Order Date", value=default_dt, help="Filters by the order entry date (exact day). Defaults to today.")

            # Load Open Orders for selected cost centers and chosen date
            try:
                oo_df = loaders.load_open_orders(config.connection_string, cc_sel, od)
            except Exception as e:
                st.error(f"Failed to load Open Orders: {e}")
                oo_df = pd.DataFrame()

            if oo_df.empty:
                st.info("No orders found for the selected date and cost center.")
            else:
                # Prepare display with requested column order and headings
                disp = oo_df.copy()
                # Normalize/format
                if "order_ship_date" in disp.columns:
                    disp["order_ship_date"] = pd.to_datetime(disp["order_ship_date"], errors="coerce").dt.date
                # Determine backorders: DETAIL_LINE_STATUS == 'B'
                bo_mask_src = pd.Series(False, index=disp.index)
                try:
                    if "detail_line_status" in disp.columns:
                        bo_mask_src = disp["detail_line_status"].fillna("").astype(str).str.strip().str.upper().eq("B")
                except Exception:
                    bo_mask_src = pd.Series(False, index=disp.index)
                # Column mapping: data field -> heading
                col_map = {
                    "line_gpp_with_funds": "Gp",
                    "order_reference": "Ref",
                    "item_mfgr_color_pat": "Sku",
                    "item_desc_1": "Desc",
                    "quantity_ordered": "Qty",
                    "order_ship_date": "ShipDate",
                    "price_per_um": "Price",
                    "cost_per_um": "Cost",
                    "unit_of_measure": "UoM",
                    "extended_price_no_funds": "Total",
                    "bank_name2": "Account",
                    "salesperson_desc": "Rep",
                }
                order = [
                    "line_gpp_with_funds",
                    "order_reference",
                    "item_mfgr_color_pat",
                    "item_desc_1",
                    "quantity_ordered",
                    "order_ship_date",
                    "price_per_um",
                    "cost_per_um",
                    "unit_of_measure",
                    "extended_price_no_funds",
                    "bank_name2",
                    "salesperson_desc",
                ]
                present = [c for c in order if c in disp.columns]
                table = disp[present].rename(columns=col_map)
                # Format numeric columns
                for c in ["Qty"]:
                    if c in table.columns:
                        table[c] = pd.to_numeric(table[c], errors="coerce").fillna(0.0)
                for c in ["Price", "Cost", "Total", "Gp"]:
                    if c in table.columns:
                        table[c] = pd.to_numeric(table[c], errors="coerce")
                # Sort by GP ascending (lowest at the top)
                if "Gp" in table.columns:
                    table = table.sort_values(by=["Gp"], ascending=True, kind="mergesort")
                # Build row-highlighting styler for backorders
                try:
                    bo_mask_table = bo_mask_src.reindex(table.index).fillna(False)
                    def _bo_style(row):
                        if bo_mask_table.loc[row.name]:
                            return ["color: #b30000; background-color: #ffe5e5" for _ in row]
                        return ["" for _ in row]
                    styled = table.style.apply(_bo_style, axis=1)
                except Exception:
                    styled = table

                # Highlight negative values in the Total column in yellow (cell-level)
                try:
                    if "Total" in table.columns:
                        # Ensure we have a Styler instance
                        try:
                            from pandas.io.formats.style import Styler  # type: ignore
                            is_styler = isinstance(styled, Styler)
                        except Exception:
                            is_styler = hasattr(styled, "applymap")
                        if not is_styler:
                            styled = table.style

                        def _neg_yellow(v):
                            try:
                                return "background-color: #fff3b0" if float(v) < 0 else ""
                            except Exception:
                                return ""
                        styled = styled.applymap(_neg_yellow, subset=["Total"])  # type: ignore[attr-defined]
                except Exception:
                    # If styling fails for any reason, fall back silently
                    pass
                # Expand the visible height to show at least ~20 rows without scrolling
                est_row_h = 32  # px per row (approximate)
                base_h = 60     # header/margins
                target_rows = max(20, len(table))
                height_px = min(1400, max(600, base_h + est_row_h * target_rows))
                st.dataframe(styled, use_container_width=True, hide_index=True, height=int(height_px))

                # Footer total for the Amount/Total column
                sum_total = 0.0
                try:
                    sum_total = float(pd.to_numeric(oo_df.get("extended_price_no_funds"), errors="coerce").fillna(0.0).sum())
                except Exception:
                    sum_total = 0.0
                st.metric("Total", f"${sum_total:,.2f}")

    # -------------------------
    # Drops (Tab 7)
    # -------------------------
    with tabs[6]:
        st.subheader("Dropped Items")
        st.caption("View items that have been discontinued (marked with DI in IPOL1, IPOL2, or IPOL3 fields).")
        
        # Respond to cost center filters from the sidebar
        cc_sel = selected_cost_centers
        if not cc_sel:
            st.info("Select at least one cost center in the sidebar to view dropped items.")
        else:
            # Date range filter for discontinued date
            col1, col2 = st.columns(2)
            with col1:
                drop_start_date = st.date_input(
                    "Start Date", 
                    value=date.today() - timedelta(days=365),
                    help="Filter dropped items by discontinued date (start of range)"
                )
            with col2:
                drop_end_date = st.date_input(
                    "End Date",
                    value=date.today(),
                    help="Filter dropped items by discontinued date (end of range)"
                )

            # Load dropped items for selected cost centers and date range
            try:
                drops_df = loaders.load_dropped_items(
                    config.connection_string, 
                    cc_sel, 
                    drop_start_date, 
                    drop_end_date
                )
            except Exception as e:
                st.error(f"Failed to load dropped items: {e}")
                drops_df = pd.DataFrame()

            if drops_df.empty:
                st.info("No dropped items found for the selected date range and cost centers.")
            else:
                # Prepare display with requested columns
                disp = drops_df.copy()
                
                # Normalize column names
                disp.rename(columns={
                    "sku": "SKU",
                    "sku_description": "DESCRIPTION",
                    "manufacturer": "MANUFACTURER",
                    "price_class": "PRICE_CLASS",
                    "product_line": "PRODUCT_LINE",
                    "cost_center": "COST_CENTER",
                }, inplace=True)
                
                # Extract style by clustering descriptions within each PRICE_CLASS
                import re as _re
                
                def extract_styles_by_clustering(df: pd.DataFrame) -> pd.Series:
                    """Find style sub-groups within each price class by clustering similar prefixes"""
                    style_map = {}
                    
                    df_work = df.copy()
                    df_work["PRICE_CLASS"] = df_work["PRICE_CLASS"].fillna("NO_PC").astype(str)
                    
                    for pc, group in df_work.groupby("PRICE_CLASS", dropna=False):
                        descriptions = group["DESCRIPTION"].dropna().astype(str).str.strip().tolist()
                        skus = group["SKU"].tolist()
                        
                        if not descriptions or len(descriptions) <= 1:
                            # Single or no item - extract first word
                            if descriptions:
                                words = _re.findall(r'\b([A-Za-z][A-Za-z0-9]*)\b', descriptions[0])
                                style_name = words[0].upper() if words else str(skus[0])[:10].upper()
                            else:
                                style_name = str(skus[0])[:10].upper()
                            for sku in skus:
                                style_map[sku] = style_name
                            continue
                        
                        # Multiple items - cluster by first word
                        clusters = {}  # first_word -> list of (description, sku)
                        for i, desc in enumerate(descriptions):
                            words = _re.findall(r'\b([A-Za-z][A-Za-z0-9]*)\b', desc)
                            if words:
                                first_word = words[0].upper()
                                if first_word not in clusters:
                                    clusters[first_word] = []
                                clusters[first_word].append((desc, skus[i]))
                            else:
                                # No words found, use SKU
                                sku_key = str(skus[i])[:10].upper()
                                if sku_key not in clusters:
                                    clusters[sku_key] = []
                                clusters[sku_key].append((desc, skus[i]))
                        
                        # For each cluster, find common prefix
                        for first_word, items in clusters.items():
                            cluster_descs = [item[0] for item in items]
                            cluster_skus = [item[1] for item in items]
                            
                            if len(cluster_descs) == 1:
                                # Single item in cluster - use first word
                                style_name = first_word
                            else:
                                # Find common character prefix across all in cluster
                                prefix = cluster_descs[0]
                                for desc in cluster_descs[1:]:
                                    match_len = 0
                                    for j in range(min(len(prefix), len(desc))):
                                        if prefix[j].upper() == desc[j].upper():
                                            match_len = j + 1
                                        else:
                                            break
                                    prefix = prefix[:match_len]
                                
                                # Trim to word boundary
                                prefix = prefix.strip()
                                words = _re.findall(r'\b([A-Za-z][A-Za-z0-9]*)\b', prefix)
                                if words:
                                    style_name = " ".join(words).upper()
                                else:
                                    style_name = first_word
                            
                            # Assign style to all SKUs in this cluster
                            for sku in cluster_skus:
                                style_map[sku] = style_name
                    
                    return df["SKU"].map(style_map).fillna("NOSTYLE")
                
                disp["Style"] = extract_styles_by_clustering(disp)
                
                # Group by Style (NOT price class) - multiple styles can exist within one price class
                # Sort by discontinued date first to get most recent per style
                if "discontinued_date" in disp.columns:
                    disp_sorted = disp.sort_values(by=["discontinued_date"], ascending=False, kind="mergesort")
                else:
                    disp_sorted = disp
                
                # Take first record per unique STYLE (most recent discontinued date)
                # This ensures each unique style gets one record, even if multiple styles share a price class
                grouped = disp_sorted.groupby("Style", dropna=False).first().reset_index()
                
                # Format discontinued date
                if "discontinued_date" in grouped.columns:
                    grouped["Discontinued Date"] = pd.to_datetime(grouped["discontinued_date"], errors="coerce").dt.date
                
                # Column mapping for display - no SKU, no Description
                col_map = {
                    "Style": "Style",
                    "MANUFACTURER": "Manufacturer",
                    "Discontinued Date": "Discontinued Date",
                }
                
                # Select and order columns for display
                display_cols = ["Style", "MANUFACTURER", "Discontinued Date"]
                present = [c for c in display_cols if c in grouped.columns]
                table = grouped[present].rename(columns={k: v for k, v in col_map.items() if k in present})
                
                # Sort by Manufacturer, then by Style
                sort_cols = []
                if "Manufacturer" in table.columns:
                    sort_cols.append("Manufacturer")
                if "Style" in table.columns:
                    sort_cols.append("Style")
                if sort_cols:
                    table = table.sort_values(by=sort_cols, ascending=True, kind="mergesort")
                
                # Add "Include in Download" checkbox column
                # Load saved exclusions from user prefs
                drops_exclusions = user_prefs.get("drops_exclusions", [])
                table.insert(0, "Include in Download", True)
                
                # Mark excluded styles as False
                if "Style" in table.columns:
                    table["Include in Download"] = table["Style"].apply(lambda s: s not in drops_exclusions)
                
                # Display editable table with checkboxes
                st.caption("Check/uncheck items to include or exclude from download")
                edited_table = st.data_editor(
                    table,
                    use_container_width=True,
                    hide_index=True,
                    height=600,
                    column_config={
                        "Include in Download": st.column_config.CheckboxColumn(
                            "Include in Download",
                            help="Uncheck to exclude from download",
                            default=True,
                        )
                    },
                    disabled=["Style", "Manufacturer", "Discontinued Date"],  # Only checkbox is editable
                )
                
                # Update exclusions based on user changes
                if edited_table is not None:
                    # Build new exclusions list from unchecked items
                    new_exclusions = []
                    if "Style" in edited_table.columns and "Include in Download" in edited_table.columns:
                        for idx, row in edited_table.iterrows():
                            if not row["Include in Download"]:
                                new_exclusions.append(row["Style"])
                    
                    # Save if changed
                    if set(new_exclusions) != set(drops_exclusions):
                        user_prefs["drops_exclusions"] = new_exclusions
                        _save_prefs(user_prefs)
                
                # Show counts
                total_count = len(edited_table)
                included_count = edited_table["Include in Download"].sum() if "Include in Download" in edited_table.columns else total_count
                excluded_count = total_count - included_count
                
                col1, col2, col3 = st.columns(3)
                col1.metric("Total Dropped Styles", total_count)
                col2.metric("Included in Download", int(included_count))
                col3.metric("Excluded from Download", int(excluded_count))
                
                # Download button - create professional Excel export (only included items)
                if st.button("Download Dropped Items Report", type="primary"):
                    try:
                        # Filter to only included items
                        download_df = edited_table[edited_table["Include in Download"] == True].copy()
                        download_df = download_df.drop(columns=["Include in Download"])  # Remove checkbox column from export
                        
                        if download_df.empty:
                            st.warning("No items selected for download. Please check at least one item.")
                        else:
                            # Import openpyxl styles
                            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
                            
                            # Create Excel file with professional formatting
                            output = BytesIO()
                            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                                # Write data to Excel starting at row 4 (leaving room for title)
                                export_df = download_df.copy()
                                export_df.to_excel(writer, sheet_name="Dropped Items", index=False, startrow=3)
                                
                                # Get workbook and worksheet for formatting
                                workbook = writer.book
                                worksheet = workbook["Dropped Items"]
                                
                                # Add title and metadata at the top
                                worksheet["A1"] = "Dropped Items Report"
                                worksheet["A2"] = f"Generated: {date.today().strftime('%B %d, %Y')}"
                                worksheet["A3"] = f"Date Range: {drop_start_date.strftime('%m/%d/%Y')} to {drop_end_date.strftime('%m/%d/%Y')}"
                                
                                # Format title (bold, larger font)
                                title_font = Font(name="Arial", size=14, bold=True)
                                worksheet["A1"].font = title_font
                                worksheet["A2"].font = Font(name="Arial", size=10, italic=True)
                                worksheet["A3"].font = Font(name="Arial", size=10, italic=True)
                                
                                # Format header row (row 4) with blue background and white text
                                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                                header_font_white = Font(name="Arial", size=11, bold=True, color="FFFFFF")
                                
                                for cell in worksheet[4]:
                                    if cell.value:  # Only format cells with content
                                        cell.fill = header_fill
                                        cell.font = header_font_white
                                        cell.alignment = Alignment(horizontal="center", vertical="center")
                                
                                # Format data rows
                                data_font = Font(name="Arial", size=10)
                                for row in worksheet.iter_rows(min_row=5, max_row=worksheet.max_row):
                                    for cell in row:
                                        if cell.value is not None:
                                            cell.font = data_font
                                            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=False)
                                
                                # Auto-adjust column widths
                                for column_cells in worksheet.columns:
                                    length = max(len(str(cell.value or "")) for cell in column_cells)
                                    worksheet.column_dimensions[column_cells[0].column_letter].width = min(50, max(12, length + 2))
                                
                                # Add borders to all data cells (header + data)
                                thin_border = Border(
                                    left=Side(style='thin', color="000000"),
                                    right=Side(style='thin', color="000000"),
                                    top=Side(style='thin', color="000000"),
                                    bottom=Side(style='thin', color="000000")
                                )
                                
                                for row in worksheet.iter_rows(min_row=4, max_row=worksheet.max_row):
                                    for cell in row:
                                        if cell.value is not None:
                                            cell.border = thin_border
                            
                            output.seek(0)
                            filename = f"Dropped_Items_{drop_start_date.strftime('%Y%m%d')}_{drop_end_date.strftime('%Y%m%d')}.xlsx"
                            
                            st.download_button(
                                label="📥 Download Excel File",
                                data=output.getvalue(),
                                file_name=filename,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                        
                    except Exception as e:
                        st.error(f"Failed to generate Excel export: {e}")

    # ========================================================
    # TAB 8: Supplier Performance
    # ========================================================
    with tabs[7]:
        st.header("Supplier Performance - Year Over Year")
        st.markdown("View supplier sales YTD vs Prior Year with top items breakdown.")

        # Date range selector for supplier performance
        supp_perf_col1, supp_perf_col2 = st.columns(2)
        with supp_perf_col1:
            # Default to start of current year for YTD, or load from preferences
            current_year = date.today().year
            default_start = date(current_year, 1, 1)
            saved_start = user_prefs.get("supp_perf_start_date")
            if saved_start:
                try:
                    # Parse saved date string (format: YYYY-MM-DD)
                    saved_start = date.fromisoformat(saved_start)
                except Exception:
                    saved_start = default_start
            else:
                saved_start = default_start
            
            supp_perf_start = st.date_input(
                "Start Date",
                value=saved_start,
                key="supp_perf_start_date"
            )
        with supp_perf_col2:
            supp_perf_end = st.date_input(
                "End Date",
                value=date.today(),  # Always default to today
                key="supp_perf_end_date"
            )

        # Group By selector
        st.markdown("**Group By:**")
        group_by_field = st.selectbox(
            "Select grouping field",
            options=[
                ("Supplier", "supplier_number"),
                ("Item Class 1", "item_class_1_desc"),
                ("Item Class 2", "item_class_2_desc"),
                ("Item Class 3", "item_class_3_desc"),
                ("Salesperson", "salesperson_desc")
            ],
            format_func=lambda x: x[0],
            key="supp_perf_group_by",
            index=0  # Default to Supplier
        )
        selected_group_field = group_by_field[1]  # Get the field name
        group_display_name = group_by_field[0]  # Get the display name

        if st.button("Load Supplier Performance Data", type="primary"):
            with st.spinner("Loading supplier performance data..."):
                try:
                    # Save start date to preferences
                    user_prefs["supp_perf_start_date"] = supp_perf_start.isoformat()
                    _save_prefs(user_prefs)
                    # Load data with applied filters
                    supp_data = loaders.load_supplier_performance(
                        config.connection_string,
                        cost_centers=selected_cost_centers if selected_cost_centers else None,
                        start_date=None,  # We'll filter both years together
                        end_date=None
                    )
                    
                    # Load inventory costs for ROI calculation
                    inventory_costs = loaders.load_inventory_costs(config.connection_string)

                    if supp_data.empty:
                        st.warning("No supplier performance data found for the selected filters.")
                    else:
                        # Store in session state
                        st.session_state["supp_perf_data"] = supp_data
                        st.session_state["inventory_costs"] = inventory_costs
                        st.session_state["supp_perf_start"] = supp_perf_start
                        st.session_state["supp_perf_end"] = supp_perf_end
                        st.session_state["supp_perf_group_field"] = selected_group_field
                        st.session_state["supp_perf_group_name"] = group_display_name
                        st.success(f"Loaded {len(supp_data):,} order lines grouped by {group_display_name}.")

                except Exception as e:
                    st.error(f"Failed to load supplier performance data: {e}")

        # Display charts if data is loaded
        if "supp_perf_data" in st.session_state and not st.session_state["supp_perf_data"].empty:
            supp_data = st.session_state["supp_perf_data"].copy()
            ytd_start_date = st.session_state.get("supp_perf_start", ytd_start)
            ytd_end_date = st.session_state.get("supp_perf_end", date.today())
            group_field = st.session_state.get("supp_perf_group_field", "supplier_number")
            group_name = st.session_state.get("supp_perf_group_name", "Supplier")

            # Ensure invoice_date is datetime for filtering
            if "invoice_date" in supp_data.columns:
                supp_data["invoice_date"] = pd.to_datetime(supp_data["invoice_date"])
                # Remove any null dates
                supp_data = supp_data[supp_data["invoice_date"].notna()].copy()
                
                if supp_data.empty:
                    st.error("All dates were null after parsing. Check the invoice_date_raw format in the database.")
                    st.stop()
            else:
                st.error("Invoice date not found in data.")
                st.stop()

            # Filter data to only the selected date range
            date_mask = (supp_data["invoice_date"] >= pd.Timestamp(ytd_start_date)) & \
                       (supp_data["invoice_date"] <= pd.Timestamp(ytd_end_date))

            filtered_data = supp_data[date_mask].copy()
            
            # Show data summary for debugging
            st.info(f"📊 Data Summary: Total rows loaded: {len(supp_data):,} | Rows after date filter: {len(filtered_data):,} | Grouped by: {group_name} | Cost centers: {selected_cost_centers if selected_cost_centers else 'All'}")

            if filtered_data.empty:
                st.warning(f"No data found for the selected date range: {ytd_start_date.strftime('%Y-%m-%d')} to {ytd_end_date.strftime('%Y-%m-%d')}")
            else:
                # Get list of groups with data and sort by total sales (descending)
                all_groups = list(filtered_data[group_field].unique())
                
                # Calculate total sales per group for sorting
                group_totals = filtered_data.groupby(group_field)["extended_price_usd"].sum().sort_values(ascending=False)
                all_groups = [g for g in group_totals.index if g in all_groups]

                if not all_groups:
                    st.info(f"No {group_name} groups found in the data.")
                else:
                    # Get unique years in the filtered data
                    unique_years = sorted(filtered_data["invoice_date"].dt.year.unique())
                    years_display = ", ".join(map(str, unique_years))
                    
                    st.markdown(f"### {group_name}: {len(all_groups)} total")
                    st.markdown(f"**Date Range:** {ytd_start_date.strftime('%m/%d/%Y')} to {ytd_end_date.strftime('%m/%d/%Y')} (Years: {years_display})")
                    
                    # Calculate and display overall ROI
                    if "inventory_costs" in st.session_state and not st.session_state["inventory_costs"].empty:
                        inventory_costs = st.session_state["inventory_costs"]
                        if "gross_profit_usd" in filtered_data.columns and "inventory_flag" in filtered_data.columns:
                            all_stocking = filtered_data[filtered_data["inventory_flag"] == "Y"].copy()
                            if not all_stocking.empty:
                                all_inv_cost = inventory_costs[
                                    inventory_costs["sku"].isin(all_stocking["sku"].unique())
                                ]["total_cost"].sum()
                                
                                if all_inv_cost > 0:
                                    roi_metrics = []
                                    for year in unique_years:
                                        year_stocking = all_stocking[all_stocking["invoice_date"].dt.year == year]
                                        year_gp = year_stocking["gross_profit_usd"].sum()
                                        roi = year_gp / all_inv_cost
                                        roi_metrics.append(f"{year}: **{roi:.2f}x**")
                                    
                                    st.markdown(f"**Overall ROI:** {', '.join(roi_metrics)}")
                    
                    # Generate ROI summary table for all groups
                    st.markdown("### 📊 ROI Summary by Group")
                    if "inventory_costs" in st.session_state and not st.session_state["inventory_costs"].empty:
                        inventory_costs = st.session_state["inventory_costs"]
                        if "gross_profit_usd" in filtered_data.columns and "inventory_flag" in filtered_data.columns:
                            summary_data = []
                            
                            for group_value in all_groups:
                                group_data = filtered_data[filtered_data[group_field] == group_value].copy()
                                
                                row = {group_name: group_value}
                                
                                # For Salesperson: show average monthly sales instead of ROI
                                if group_field == "salesperson_desc":
                                    monthly_avgs = {}
                                    for year in unique_years:
                                        year_data = group_data[group_data["invoice_date"].dt.year == year]
                                        if not year_data.empty:
                                            year_data_copy = year_data.copy()
                                            year_data_copy["month"] = year_data_copy["invoice_date"].dt.to_period("M")
                                            monthly_sales = year_data_copy.groupby("month")["extended_price_usd"].sum()
                                            avg_monthly = monthly_sales.mean()
                                            monthly_avgs[year] = avg_monthly
                                            row[f"{year} Avg Monthly Sales"] = avg_monthly
                                        else:
                                            row[f"{year} Avg Monthly Sales"] = 0
                                    
                                    # Calculate growth percentage
                                    if len(unique_years) >= 2:
                                        current_avg = monthly_avgs.get(unique_years[-1], 0)
                                        prior_avg = monthly_avgs.get(unique_years[-2], 0)
                                        if prior_avg > 0:
                                            growth_pct = (current_avg / prior_avg) * 100
                                            row["Growth %"] = growth_pct
                                        else:
                                            row["Growth %"] = 0
                                    
                                    summary_data.append(row)
                                else:
                                    # For other groupings: show ROI
                                    stocking_items = group_data[group_data["inventory_flag"] == "Y"].copy()
                                    
                                    if not stocking_items.empty:
                                        group_inv_cost = inventory_costs[
                                            inventory_costs["sku"].isin(stocking_items["sku"].unique())
                                        ]["total_cost"].sum()
                                        
                                        if group_inv_cost > 0:
                                            for year in unique_years:
                                                year_stocking = stocking_items[stocking_items["invoice_date"].dt.year == year]
                                                year_gp = year_stocking["gross_profit_usd"].sum()
                                                roi = year_gp / group_inv_cost
                                                row[f"{year} ROI"] = roi
                                            
                                            summary_data.append(row)
                            
                            if summary_data:
                                summary_df = pd.DataFrame(summary_data)
                                
                                # Add weighted average row for Salesperson grouping
                                if group_field == "salesperson_desc" and len(unique_years) >= 2:
                                    # Calculate weighted average across all salespeople
                                    overall_monthly_avgs = {}
                                    for year in unique_years:
                                        year_data = filtered_data[filtered_data["invoice_date"].dt.year == year]
                                        if not year_data.empty:
                                            year_data_copy = year_data.copy()
                                            year_data_copy["month"] = year_data_copy["invoice_date"].dt.to_period("M")
                                            monthly_sales = year_data_copy.groupby("month")["extended_price_usd"].sum()
                                            avg_monthly = monthly_sales.mean()
                                            overall_monthly_avgs[year] = avg_monthly
                                    
                                    # Create summary row
                                    summary_row = {group_name: "WEIGHTED AVERAGE"}
                                    for year in unique_years:
                                        summary_row[f"{year} Avg Monthly Sales"] = overall_monthly_avgs.get(year, 0)
                                    
                                    # Calculate growth
                                    current_avg = overall_monthly_avgs.get(unique_years[-1], 0)
                                    prior_avg = overall_monthly_avgs.get(unique_years[-2], 0)
                                    if prior_avg > 0:
                                        growth_pct = (current_avg / prior_avg) * 100
                                        summary_row["Growth %"] = growth_pct
                                    else:
                                        summary_row["Growth %"] = 0
                                    
                                    # Append summary row
                                    summary_df = pd.concat([summary_df, pd.DataFrame([summary_row])], ignore_index=True)
                                
                                # Sort by appropriate column
                                if group_field == "salesperson_desc" and "Growth %" in summary_df.columns:
                                    # Don't sort the last row (weighted average)
                                    if len(summary_df) > 1:
                                        sorted_df = summary_df.iloc[:-1].sort_values("Growth %", ascending=False)
                                        summary_df = pd.concat([sorted_df, summary_df.iloc[-1:]], ignore_index=True)
                                elif len(unique_years) > 0:
                                    current_year_col = f"{unique_years[-1]} ROI"
                                    if current_year_col in summary_df.columns:
                                        summary_df = summary_df.sort_values(current_year_col, ascending=False)
                                
                                # Format for display
                                display_summary = summary_df.copy()
                                for col in display_summary.columns:
                                    if "ROI" in col:
                                        display_summary[col] = display_summary[col].apply(lambda x: f"{x:.2f}x")
                                    elif "Avg Monthly Sales" in col:
                                        display_summary[col] = display_summary[col].apply(lambda x: f"${x:,.0f}")
                                    elif col == "Growth %":
                                        display_summary[col] = display_summary[col].apply(lambda x: f"{x:.1f}%")
                                
                                st.dataframe(display_summary, use_container_width=True, hide_index=True, height=400)
                                
                                # Store in session state for PDF
                                st.session_state["roi_summary_df"] = summary_df
                    
                    # Add aggregated price class comparison table for non-Salesperson groupings
                    if group_field != "salesperson_desc" and len(unique_years) >= 2:
                        st.markdown("### Aggregated Price Class Comparison")
                        st.markdown("Top 10 price classes aggregated across all groups")
                        
                        current_year = unique_years[-1]
                        prior_year = unique_years[-2]
                        
                        # Filter to current and prior year
                        ytd_current_year = filtered_data[filtered_data["invoice_date"].dt.year == current_year].copy()
                        ytd_prior_year = filtered_data[filtered_data["invoice_date"].dt.year == prior_year].copy()
                        
                        # Calculate ROI for each price class
                        current_metric_lookup = {}
                        prior_metric_lookup = {}
                        
                        if "inventory_costs" in st.session_state and not st.session_state["inventory_costs"].empty:
                            inventory_costs = st.session_state["inventory_costs"]
                            if "gross_profit_usd" in filtered_data.columns:
                                # Current year ROI by price class
                                current_stocking = ytd_current_year[ytd_current_year["inventory_flag"] == "Y"].copy()
                                if not current_stocking.empty:
                                    current_gp_by_class = current_stocking.groupby("price_class_desc")["gross_profit_usd"].sum()
                                    
                                    for pc_desc in current_gp_by_class.index:
                                        pc_inv_cost = inventory_costs[
                                            (inventory_costs["price_class_desc"].astype(str).str.strip() == str(pc_desc).strip()) &
                                            (inventory_costs["price_class_desc"].notna())
                                        ]["total_cost"].sum()
                                        if pc_inv_cost > 0:
                                            current_metric_lookup[pc_desc] = current_gp_by_class[pc_desc] / pc_inv_cost
                                
                                # Prior year ROI by price class
                                prior_stocking = ytd_prior_year[ytd_prior_year["inventory_flag"] == "Y"].copy()
                                if not prior_stocking.empty:
                                    prior_gp_by_class = prior_stocking.groupby("price_class_desc")["gross_profit_usd"].sum()
                                    
                                    for pc_desc in prior_gp_by_class.index:
                                        pc_inv_cost = inventory_costs[
                                            (inventory_costs["price_class_desc"].astype(str).str.strip() == str(pc_desc).strip()) &
                                            (inventory_costs["price_class_desc"].notna())
                                        ]["total_cost"].sum()
                                        if pc_inv_cost > 0:
                                            prior_metric_lookup[pc_desc] = prior_gp_by_class[pc_desc] / pc_inv_cost
                        
                        # Get top 10 for each year
                        if not ytd_current_year.empty:
                            ytd_current_top10 = ytd_current_year.groupby("price_class_desc")["extended_price_usd"].sum().reset_index()
                            ytd_current_top10 = ytd_current_top10.sort_values("extended_price_usd", ascending=False).head(10)
                        else:
                            ytd_current_top10 = pd.DataFrame(columns=["price_class_desc", "extended_price_usd"])
                        
                        if not ytd_prior_year.empty:
                            ytd_prior_top10 = ytd_prior_year.groupby("price_class_desc")["extended_price_usd"].sum().reset_index()
                            ytd_prior_top10 = ytd_prior_top10.sort_values("extended_price_usd", ascending=False).head(10)
                        else:
                            ytd_prior_top10 = pd.DataFrame(columns=["price_class_desc", "extended_price_usd"])
                        
                        # Create side-by-side tables
                        col1, col2 = st.columns(2)
                        
                        with col1:
                            st.markdown(f"**{current_year}**")
                            if not ytd_current_top10.empty:
                                display_current = ytd_current_top10.copy()
                                display_current.columns = ["Price Class", "Sales (USD)"]
                                display_current["ROI"] = display_current["Price Class"].map(current_metric_lookup).fillna(0)
                                display_current["Sales (USD)"] = display_current["Sales (USD)"].apply(lambda x: f"${x:,.0f}")
                                display_current["ROI"] = display_current["ROI"].apply(lambda x: f"{x:.2f}x" if x > 0 else "")
                                
                                # Make price classes clickable
                                for idx, row in display_current.iterrows():
                                    price_class = row["Price Class"]
                                    if st.button(f"{price_class}", key=f"agg_current_{price_class}", use_container_width=True):
                                        st.session_state["selected_price_class"] = price_class
                                        st.session_state["show_price_class_detail"] = True
                                        st.rerun()
                                
                                # Totals
                                current_top10_total = ytd_current_top10["extended_price_usd"].sum()
                                current_grand_total = ytd_current_year["extended_price_usd"].sum()
                                st.markdown(f"**Top 10 Total:** ${current_top10_total:,.0f}")
                                st.markdown(f"**Grand Total:** ${current_grand_total:,.0f}")
                            else:
                                st.info("No data available")
                        
                        with col2:
                            st.markdown(f"**{prior_year}**")
                            if not ytd_prior_top10.empty:
                                display_prior = ytd_prior_top10.copy()
                                display_prior.columns = ["Price Class", "Sales (USD)"]
                                display_prior["ROI"] = display_prior["Price Class"].map(prior_metric_lookup).fillna(0)
                                display_prior["Sales (USD)"] = display_prior["Sales (USD)"].apply(lambda x: f"${x:,.0f}")
                                display_prior["ROI"] = display_prior["ROI"].apply(lambda x: f"{x:.2f}x" if x > 0 else "")
                                
                                # Make price classes clickable
                                for idx, row in display_prior.iterrows():
                                    price_class = row["Price Class"]
                                    if st.button(f"{price_class}", key=f"agg_prior_{price_class}", use_container_width=True):
                                        st.session_state["selected_price_class"] = price_class
                                        st.session_state["show_price_class_detail"] = True
                                        st.rerun()
                                
                                # Totals
                                prior_top10_total = ytd_prior_top10["extended_price_usd"].sum()
                                prior_grand_total = ytd_prior_year["extended_price_usd"].sum()
                                st.markdown(f"**Top 10 Total:** ${prior_top10_total:,.0f}")
                                st.markdown(f"**Grand Total:** ${prior_grand_total:,.0f}")
                            else:
                                st.info("No data available")
                    
                    # Display rolling average chart for selected price class
                    if st.session_state.get("show_price_class_detail") and st.session_state.get("selected_price_class"):
                        selected_pc = st.session_state["selected_price_class"]
                        
                        st.markdown("---")
                        st.markdown(f"### 📈 Rolling Average Sales: {selected_pc}")
                        
                        # Close button
                        if st.button("✖ Close Detail View", key="close_price_class_detail"):
                            st.session_state["show_price_class_detail"] = False
                            st.session_state["selected_price_class"] = None
                            st.rerun()
                        
                        # Filter data for selected price class
                        pc_data = filtered_data[filtered_data["price_class_desc"] == selected_pc].copy()
                        
                        if pc_data.empty:
                            st.warning(f"No data available for price class: {selected_pc}")
                        else:
                            # Ensure we have invoice_date as datetime
                            pc_data["invoice_date"] = pd.to_datetime(pc_data["invoice_date"])
                            
                            # Get the end date (most recent date in data)
                            max_date = pc_data["invoice_date"].max()
                            
                            # Define two periods:
                            # Period 1: Last 12 months (from max_date going back 12 months)
                            # Period 2: 12 months prior to period 1 (months 13-24 before max_date)
                            period1_start = max_date - pd.DateOffset(months=12)
                            period1_end = max_date
                            period2_start = max_date - pd.DateOffset(months=24)
                            period2_end = period1_start
                            
                            st.info(f"Period 1 (Last 12 months): {period1_start.strftime('%Y-%m-%d')} to {period1_end.strftime('%Y-%m-%d')}")
                            st.info(f"Period 2 (Prior 12 months): {period2_start.strftime('%Y-%m-%d')} to {period2_end.strftime('%Y-%m-%d')}")
                            
                            # Function to calculate rolling 12-month average
                            def calculate_rolling_12month_avg(data, start_date, end_date):
                                # Filter to period
                                period_data = data[(data["invoice_date"] >= start_date) & (data["invoice_date"] <= end_date)].copy()
                                
                                if period_data.empty:
                                    return pd.DataFrame(columns=["date", "rolling_avg"])
                                
                                # Aggregate by day
                                daily_sales = period_data.groupby(period_data["invoice_date"].dt.date)["extended_price_usd"].sum().reset_index()
                                daily_sales.columns = ["date", "sales"]
                                daily_sales["date"] = pd.to_datetime(daily_sales["date"])
                                daily_sales = daily_sales.sort_values("date")
                                
                                # Create a complete date range for the period
                                date_range = pd.date_range(start=start_date, end=end_date, freq='D')
                                full_df = pd.DataFrame({"date": date_range})
                                full_df = full_df.merge(daily_sales, on="date", how="left")
                                full_df["sales"] = full_df["sales"].fillna(0)
                                
                                # Calculate rolling 3-week (21-day) average
                                full_df["rolling_avg"] = full_df["sales"].rolling(window=21, min_periods=1).mean()
                                
                                return full_df[["date", "rolling_avg"]]
                            
                            # Calculate rolling averages for both periods
                            period1_rolling = calculate_rolling_12month_avg(pc_data, period1_start, period1_end)
                            period2_rolling = calculate_rolling_12month_avg(pc_data, period2_start, period2_end)
                            
                            # Prepare data for plotting
                            import plotly.graph_objects as go
                            
                            fig = go.Figure()
                            
                            if not period1_rolling.empty:
                                fig.add_trace(go.Scatter(
                                    x=period1_rolling["date"],
                                    y=period1_rolling["rolling_avg"],
                                    name="Last 12 Months Rolling Avg",
                                    mode="lines",
                                    line=dict(color="#1f77b4", width=2)
                                ))
                            
                            if not period2_rolling.empty:
                                fig.add_trace(go.Scatter(
                                    x=period2_rolling["date"],
                                    y=period2_rolling["rolling_avg"],
                                    name="Prior 12 Months Rolling Avg",
                                    mode="lines",
                                    line=dict(color="#ff7f0e", width=2)
                                ))
                            
                            fig.update_layout(
                                title=f"Rolling 12-Month Average Sales: {selected_pc}",
                                xaxis_title="Date",
                                yaxis_title="Rolling Average Sales (USD)",
                                yaxis_tickformat="$,.0f",
                                height=500,
                                hovermode="x unified",
                                legend=dict(
                                    orientation="h",
                                    yanchor="bottom",
                                    y=1.02,
                                    xanchor="right",
                                    x=1
                                )
                            )
                            
                            st.plotly_chart(fig, use_container_width=True, key=f"agg_rolling_avg_{selected_pc}")
                            
                            # Show summary statistics
                            col1, col2 = st.columns(2)
                            with col1:
                                if not period1_rolling.empty:
                                    avg1 = period1_rolling["rolling_avg"].mean()
                                    max1 = period1_rolling["rolling_avg"].max()
                                    min1 = period1_rolling["rolling_avg"].min()
                                    st.markdown("**Last 12 Months Statistics:**")
                                    st.markdown(f"- Average: ${avg1:,.2f}")
                                    st.markdown(f"- Maximum: ${max1:,.2f}")
                                    st.markdown(f"- Minimum: ${min1:,.2f}")
                            
                            with col2:
                                if not period2_rolling.empty:
                                    avg2 = period2_rolling["rolling_avg"].mean()
                                    max2 = period2_rolling["rolling_avg"].max()
                                    min2 = period2_rolling["rolling_avg"].min()
                                    st.markdown("**Prior 12 Months Statistics:**")
                                    st.markdown(f"- Average: ${avg2:,.2f}")
                                    st.markdown(f"- Maximum: ${max2:,.2f}")
                                    st.markdown(f"- Minimum: ${min2:,.2f}")
                    
                    # Export to PDF button
                    if st.button("📄 Export to PDF", type="secondary"):
                        with st.spinner("Generating PDF report..."):
                            try:
                                inventory_costs = st.session_state.get("inventory_costs", pd.DataFrame())
                                pdf_buffer = _generate_supplier_performance_pdf(
                                    filtered_data=filtered_data,
                                    all_groups=all_groups,
                                    group_field=group_field,
                                    group_name=group_name,
                                    ytd_start_date=ytd_start_date,
                                    ytd_end_date=ytd_end_date,
                                    unique_years=unique_years,
                                    inventory_costs=inventory_costs
                                )
                                
                                # Offer download
                                st.download_button(
                                    label="Download PDF Report",
                                    data=pdf_buffer.getvalue(),
                                    file_name=f"supplier_performance_{group_name.lower().replace(' ', '_')}_{ytd_start_date.strftime('%Y%m%d')}_to_{ytd_end_date.strftime('%Y%m%d')}.pdf",
                                    mime="application/pdf",
                                    type="primary"
                                )
                                st.success("PDF generated successfully!")
                            except Exception as e:
                                st.error(f"Failed to generate PDF: {str(e)}")
                                import traceback
                                st.code(traceback.format_exc())

                    # Calculate average sales across all groups by month for the selected period
                    def calculate_average_monthly(df):
                        if df.empty:
                            return pd.DataFrame(columns=["month", "sales", "period"])
                        df_copy = df.copy()
                        df_copy["month"] = df_copy["invoice_date"].dt.to_period("M").dt.to_timestamp()
                        # Group by month, calculate total sales
                        monthly = df_copy.groupby(["month"])["extended_price_usd"].sum().reset_index()
                        # Calculate average by dividing by number of groups
                        num_groups = df_copy[group_field].nunique()
                        monthly["sales"] = monthly["extended_price_usd"] / num_groups if num_groups > 0 else 0
                        monthly["period"] = "Average"
                        monthly = monthly[["month", "sales", "period"]]
                        return monthly
                    
                    # Use filtered data for average calculation
                    avg_monthly = calculate_average_monthly(filtered_data)

                    # Process each group
                    for group_value in all_groups:
                        try:
                            st.markdown("---")
                            
                            # Filter data for this group from the filtered_data
                            group_data = filtered_data[filtered_data[group_field] == group_value].copy()
                            
                            # Calculate ROI or Growth % for this group
                            roi_text = ""
                            if "inventory_costs" in st.session_state and not st.session_state["inventory_costs"].empty:
                                inventory_costs = st.session_state["inventory_costs"]
                                if "gross_profit_usd" in group_data.columns:
                                    # Get unique years
                                    all_years = sorted(group_data["invoice_date"].dt.year.unique())
                                    
                                    if group_field == "salesperson_desc":
                                        # For Salesperson: show Growth % based on average monthly sales
                                        if len(all_years) >= 2:
                                            current_year = all_years[-1]
                                            prior_year = all_years[-2]
                                            
                                            # Calculate current year avg monthly sales
                                            current_data = group_data[group_data["invoice_date"].dt.year == current_year].copy()
                                            if not current_data.empty:
                                                current_data["month"] = current_data["invoice_date"].dt.to_period("M")
                                                current_monthly = current_data.groupby("month")["extended_price_usd"].sum()
                                                current_avg = current_monthly.mean() if len(current_monthly) > 0 else 0
                                            else:
                                                current_avg = 0
                                            
                                            # Calculate prior year avg monthly sales
                                            prior_data = group_data[group_data["invoice_date"].dt.year == prior_year].copy()
                                            if not prior_data.empty:
                                                prior_data["month"] = prior_data["invoice_date"].dt.to_period("M")
                                                prior_monthly = prior_data.groupby("month")["extended_price_usd"].sum()
                                                prior_avg = prior_monthly.mean() if len(prior_monthly) > 0 else 0
                                            else:
                                                prior_avg = 0
                                            
                                            if prior_avg > 0:
                                                growth_pct = (current_avg / prior_avg) * 100
                                                roi_text = f" | Growth %: {growth_pct:.1f}%"
                                    else:
                                        # For other groupings: show ROI
                                        # Get stocking items in this group
                                        if "inventory_flag" in group_data.columns:
                                            stocking_items = group_data[group_data["inventory_flag"] == "Y"].copy()
                                        else:
                                            stocking_items = pd.DataFrame()
                                        
                                        if not stocking_items.empty and "price_class" in stocking_items.columns:
                                            # Calculate total inventory cost for this group's stocking items
                                            group_inv_cost = inventory_costs[
                                                inventory_costs["sku"].isin(stocking_items["sku"].unique())
                                            ]["total_cost"].sum()
                                            
                                            if group_inv_cost > 0:
                                                roi_values = []
                                                for year in all_years:
                                                    year_stocking = stocking_items[stocking_items["invoice_date"].dt.year == year]
                                                    year_gp = year_stocking["gross_profit_usd"].sum()
                                                    roi = year_gp / group_inv_cost
                                                    roi_values.append(f"{year}: {roi:.2f}x")
                                                roi_text = " | ROI: " + ", ".join(roi_values)
                            
                            st.subheader(f"{group_name}: {group_value}{roi_text}")

                            # Aggregate sales by month - label by actual calendar year
                            def aggregate_monthly(df):
                                if df.empty:
                                    return pd.DataFrame(columns=["month", "sales", "period"])
                                df_copy = df.copy()
                                # Keep full date with year for proper separation
                                df_copy["month"] = df_copy["invoice_date"].dt.to_period("M").dt.to_timestamp()
                                df_copy["year"] = df_copy["invoice_date"].dt.year
                                # Group by both month and year to separate years properly
                                monthly = df_copy.groupby(["month", "year"])["extended_price_usd"].sum().reset_index()
                                monthly["period"] = monthly["year"].astype(str)
                                # Rename extended_price_usd to sales
                                monthly = monthly[["month", "extended_price_usd", "period"]].copy()
                                monthly.columns = ["month", "sales", "period"]
                                return monthly

                            # Use group_data for aggregation
                            group_monthly = aggregate_monthly(group_data)

                            # Create 3-column layout
                            col1, col2, col3 = st.columns(3)

                            # Column 1: Line chart
                            with col1:
                                if group_monthly.empty:
                                    st.info(f"No monthly data available")
                                else:
                                    # Create figure with secondary y-axis
                                    import plotly.graph_objects as go
                                    from plotly.subplots import make_subplots
                                    
                                    fig = make_subplots(specs=[[{"secondary_y": True}]])
                                    
                                    # Get unique years for group data (dynamically)
                                    years = sorted(group_monthly["period"].unique())
                                    
                                    # Dynamic color assignment for years
                                    year_colors = ["#1f77b4", "#ff7f0e", "#2ca02c", "#d62728", "#9467bd"]
                                    
                                    # Add group lines (primary y-axis)
                                    for idx, year in enumerate(years):
                                        year_data = group_monthly[group_monthly["period"] == year]
                                        fig.add_trace(
                                            go.Scatter(
                                                x=year_data["month"],
                                                y=year_data["sales"],
                                                name=year,
                                                mode="lines+markers",
                                                line=dict(color=year_colors[idx % len(year_colors)]),
                                                marker=dict(size=6)
                                            ),
                                            secondary_y=False
                                        )
                                    
                                    # Add single average line (secondary y-axis)
                                    if not avg_monthly.empty:
                                        fig.add_trace(
                                            go.Scatter(
                                                x=avg_monthly["month"],
                                                y=avg_monthly["sales"],
                                                name="Average",
                                                mode="lines+markers",
                                                line=dict(color="#8c564b", dash="dash"),
                                                marker=dict(size=4)
                                            ),
                                            secondary_y=True
                                        )
                                    
                                    # Update axes
                                    fig.update_xaxes(title_text="Month")
                                    fig.update_yaxes(title_text="Sales (USD)", tickformat="$,.0f", secondary_y=False)
                                    fig.update_yaxes(title_text="Avg Sales (USD)", tickformat="$,.0f", secondary_y=True)
                                    
                                    fig.update_layout(
                                        title="Sales Comparison",
                                        height=400,
                                        hovermode="x unified",
                                        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1)
                                    )
                                    st.plotly_chart(fig, use_container_width=True)

                            # Column 2: Top 10 Price Classes - Most Recent Year
                            with col2:
                                # Use group_data (already filtered to date range)
                                
                                if group_data.empty or "price_class_desc" not in group_data.columns:
                                    st.markdown(f"**Top 10 Price Classes**")
                                    st.info(f"No data available")
                                else:
                                    # Get all unique years and select the most recent
                                    all_years = sorted(group_data["invoice_date"].dt.year.unique())
                                    if len(all_years) > 0:
                                        current_year = all_years[-1]  # Most recent year
                                        st.markdown(f"**Top 10 Price Classes - {current_year}**")
                                        ytd_current_year = group_data[group_data["invoice_date"].dt.year == current_year].copy()
                                        
                                        if ytd_current_year.empty:
                                            st.info(f"No {current_year} data")
                                        else:
                                            # Aggregate by price class description
                                            ytd_top10 = ytd_current_year.groupby("price_class_desc")["extended_price_usd"].sum().reset_index()
                                            ytd_top10 = ytd_top10.sort_values("extended_price_usd", ascending=False).head(10)
                                            ytd_top10.columns = ["Price Class", "Sales (USD)"]
                                            
                                            # Calculate ROI or Growth % for each price class
                                            if group_field == "salesperson_desc":
                                                # For Salesperson: Calculate Growth %
                                                growth_list = []
                                                for _, row in ytd_top10.iterrows():
                                                    pc_desc = row["Price Class"]
                                                    if pc_desc and str(pc_desc).strip():
                                                        # Current year avg monthly sales
                                                        pc_current_data = ytd_current_year[ytd_current_year["price_class_desc"] == pc_desc].copy()
                                                        if not pc_current_data.empty:
                                                            pc_current_data["month"] = pc_current_data["invoice_date"].dt.to_period("M")
                                                            current_monthly = pc_current_data.groupby("month")["extended_price_usd"].sum()
                                                            current_avg = current_monthly.mean() if len(current_monthly) > 0 else 0
                                                        else:
                                                            current_avg = 0
                                                        
                                                        # Prior year avg monthly sales
                                                        if len(all_years) >= 2:
                                                            prior_year_val = all_years[-2]
                                                            prior_year_data = group_data[group_data["invoice_date"].dt.year == prior_year_val]
                                                            pc_prior_data = prior_year_data[prior_year_data["price_class_desc"] == pc_desc].copy()
                                                            if not pc_prior_data.empty:
                                                                pc_prior_data["month"] = pc_prior_data["invoice_date"].dt.to_period("M")
                                                                prior_monthly = pc_prior_data.groupby("month")["extended_price_usd"].sum()
                                                                prior_avg = prior_monthly.mean() if len(prior_monthly) > 0 else 0
                                                            else:
                                                                prior_avg = 0
                                                            
                                                            if prior_avg > 0:
                                                                growth_pct = (current_avg / prior_avg) * 100
                                                                growth_list.append(f"{growth_pct:.1f}%")
                                                            else:
                                                                growth_list.append("")
                                                        else:
                                                            growth_list.append("")
                                                    else:
                                                        growth_list.append("")
                                                
                                                ytd_top10["Growth %"] = growth_list
                                            
                                            elif "inventory_costs" in st.session_state and not st.session_state["inventory_costs"].empty:
                                                inventory_costs = st.session_state["inventory_costs"]
                                                if "gross_profit_usd" in ytd_current_year.columns:
                                                    roi_list = []
                                                    for _, row in ytd_top10.iterrows():
                                                        pc_desc = row["Price Class"]
                                                        # Get gross profit for this price class
                                                        if "inventory_flag" in ytd_current_year.columns and pc_desc and str(pc_desc).strip():
                                                            pc_stocking = ytd_current_year[
                                                                (ytd_current_year["price_class_desc"] == pc_desc) &
                                                                (ytd_current_year["inventory_flag"] == "Y")
                                                            ].copy()
                                                        else:
                                                            pc_stocking = pd.DataFrame()
                                                        
                                                        if not pc_stocking.empty:
                                                            pc_gp = pc_stocking["gross_profit_usd"].sum()
                                                            pc_inv_cost = inventory_costs[
                                                                (inventory_costs["price_class_desc"].astype(str).str.strip() == str(pc_desc).strip()) &
                                                                (inventory_costs["price_class_desc"].notna())
                                                            ]["total_cost"].sum()
                                                            
                                                            if pc_inv_cost > 0:
                                                                roi = pc_gp / pc_inv_cost
                                                                roi_list.append(f"{roi:.2f}x")
                                                            else:
                                                                roi_list.append("")
                                                        else:
                                                            roi_list.append("")
                                                    
                                                    ytd_top10["ROI"] = roi_list
                                            
                                            # Add total for top 10
                                            ytd_total = ytd_top10["Sales (USD)"].sum()
                                            # Calculate grand total for all sales in this year
                                            ytd_grand_total = ytd_current_year["extended_price_usd"].sum()
                                            ytd_top10_display = ytd_top10.copy()
                                            ytd_top10_display["Sales (USD)"] = ytd_top10_display["Sales (USD)"].apply(lambda x: f"${x:,.2f}")
                                            
                                            # Make price classes clickable
                                            for idx, row in ytd_top10.iterrows():
                                                price_class = row["Price Class"]
                                                sales_formatted = f"${row['Sales (USD)']:,.2f}"
                                                roi_val = ytd_top10.at[idx, "ROI"] if "ROI" in ytd_top10.columns else ""
                                                growth_val = ytd_top10.at[idx, "Growth %"] if "Growth %" in ytd_top10.columns else ""
                                                
                                                button_label = f"{price_class} - {sales_formatted}"
                                                if roi_val:
                                                    button_label += f" | {roi_val}"
                                                if growth_val:
                                                    button_label += f" | {growth_val}"
                                                
                                                if st.button(button_label, key=f"group_current_{group_value}_{price_class}", use_container_width=True):
                                                    st.session_state["selected_price_class"] = price_class
                                                    st.session_state["selected_group_value"] = group_value
                                                    st.session_state["show_price_class_detail"] = True
                                                    st.rerun()
                                            
                                            st.markdown(f"**Total:** ${ytd_total:,.2f}")
                                            st.markdown(f"**Grand Total:** ${ytd_grand_total:,.2f}")
                                    else:
                                        st.markdown(f"**Top 10 Price Classes**")
                                        st.info(f"No data available")

                            # Column 3: Top 10 Price Classes - Second Most Recent Year
                            with col3:
                                # Use group_data (already filtered to date range)
                                
                                if group_data.empty or "price_class_desc" not in group_data.columns:
                                    st.markdown(f"**Top 10 Price Classes**")
                                    st.info(f"No data available")
                                else:
                                    # Get all unique years in the data
                                    all_years = sorted(group_data["invoice_date"].dt.year.unique())
                                    
                                    # If we have 2+ years, use the second-to-last year
                                    if len(all_years) >= 2:
                                        prior_year_val = all_years[-2]
                                        st.markdown(f"**Top 10 Price Classes - {prior_year_val}**")
                                        prior_year_data = group_data[group_data["invoice_date"].dt.year == prior_year_val].copy()
                                        
                                        if prior_year_data.empty:
                                            st.info(f"No {prior_year_val} data")
                                        else:
                                            # Aggregate by price class description
                                            prior_top10 = prior_year_data.groupby("price_class_desc")["extended_price_usd"].sum().reset_index()
                                            prior_top10 = prior_top10.sort_values("extended_price_usd", ascending=False).head(10)
                                            prior_top10.columns = ["Price Class", "Sales (USD)"]
                                            
                                            # Only calculate ROI for non-salesperson groupings
                                            if group_field != "salesperson_desc" and "inventory_costs" in st.session_state and not st.session_state["inventory_costs"].empty:
                                                inventory_costs = st.session_state["inventory_costs"]
                                                if "gross_profit_usd" in prior_year_data.columns:
                                                    roi_list = []
                                                    for _, row in prior_top10.iterrows():
                                                        pc_desc = row["Price Class"]
                                                        # Get gross profit for this price class
                                                        if "inventory_flag" in prior_year_data.columns and pc_desc and str(pc_desc).strip():
                                                            pc_stocking = prior_year_data[
                                                                (prior_year_data["price_class_desc"] == pc_desc) &
                                                                (prior_year_data["inventory_flag"] == "Y")
                                                            ].copy()
                                                        else:
                                                            pc_stocking = pd.DataFrame()
                                                        
                                                        if not pc_stocking.empty:
                                                            pc_gp = pc_stocking["gross_profit_usd"].sum()
                                                            pc_inv_cost = inventory_costs[
                                                                (inventory_costs["price_class_desc"].astype(str).str.strip() == str(pc_desc).strip()) &
                                                                (inventory_costs["price_class_desc"].notna())
                                                            ]["total_cost"].sum()
                                                            
                                                            if pc_inv_cost > 0:
                                                                roi = pc_gp / pc_inv_cost
                                                                roi_list.append(f"{roi:.2f}x")
                                                            else:
                                                                roi_list.append("")
                                                        else:
                                                            roi_list.append("")
                                                    
                                                    prior_top10["ROI"] = roi_list
                                            
                                            # Add total for top 10
                                            prior_total = prior_top10["Sales (USD)"].sum()
                                            # Calculate grand total for all sales in this year
                                            prior_grand_total = prior_year_data["extended_price_usd"].sum()
                                            prior_top10_display = prior_top10.copy()
                                            prior_top10_display["Sales (USD)"] = prior_top10_display["Sales (USD)"].apply(lambda x: f"${x:,.2f}")
                                            
                                            # Make price classes clickable
                                            for idx, row in prior_top10.iterrows():
                                                price_class = row["Price Class"]
                                                sales_formatted = f"${row['Sales (USD)']:,.2f}"
                                                roi_val = prior_top10.at[idx, "ROI"] if "ROI" in prior_top10.columns else ""
                                                
                                                button_label = f"{price_class} - {sales_formatted}"
                                                if roi_val:
                                                    button_label += f" | {roi_val}"
                                                
                                                if st.button(button_label, key=f"group_prior_{group_value}_{price_class}", use_container_width=True):
                                                    st.session_state["selected_price_class"] = price_class
                                                    st.session_state["selected_group_value"] = group_value
                                                    st.session_state["show_price_class_detail"] = True
                                                    st.rerun()
                                            
                                            st.markdown(f"**Total:** ${prior_total:,.2f}")
                                            st.markdown(f"**Grand Total:** ${prior_grand_total:,.2f}")
                                    else:
                                        st.markdown(f"**Top 10 Price Classes**")
                                        st.info(f"No data available")
                            
                            # Add account growth analysis for Salesperson grouping
                            if group_field == "salesperson_desc" and len(all_years) >= 2:
                                st.markdown("---")
                                st.markdown("### Account Performance Analysis")
                                st.markdown("Top 5 and Bottom 5 accounts by growth percentage")
                                
                                current_year = all_years[-1]
                                prior_year = all_years[-2]
                                
                                # Create account identifier (bank_name - account_number)
                                if "bank_name" in group_data.columns and "account_number" in group_data.columns:
                                    group_data_copy = group_data.copy()
                                    group_data_copy["account_id"] = group_data_copy["bank_name"].astype(str) + " - " + group_data_copy["account_number"].astype(str)
                                    
                                    # Calculate average monthly sales for each account by year
                                    account_growth = []
                                    for account_id in group_data_copy["account_id"].unique():
                                        account_data = group_data_copy[group_data_copy["account_id"] == account_id]
                                        
                                        # Current year
                                        current_data = account_data[account_data["invoice_date"].dt.year == current_year].copy()
                                        if not current_data.empty:
                                            current_data["month"] = current_data["invoice_date"].dt.to_period("M")
                                            current_monthly = current_data.groupby("month")["extended_price_usd"].sum()
                                            current_avg = current_monthly.mean() if len(current_monthly) > 0 else 0
                                        else:
                                            current_avg = 0
                                        
                                        # Prior year
                                        prior_data = account_data[account_data["invoice_date"].dt.year == prior_year].copy()
                                        if not prior_data.empty:
                                            prior_data["month"] = prior_data["invoice_date"].dt.to_period("M")
                                            prior_monthly = prior_data.groupby("month")["extended_price_usd"].sum()
                                            prior_avg = prior_monthly.mean() if len(prior_monthly) > 0 else 0
                                        else:
                                            prior_avg = 0
                                        
                                        # Calculate growth %
                                        if prior_avg > 0:
                                            growth_pct = (current_avg / prior_avg) * 100
                                            account_growth.append({
                                                "Account": account_id,
                                                f"{prior_year} Avg Monthly": prior_avg,
                                                f"{current_year} Avg Monthly": current_avg,
                                                "Growth %": growth_pct
                                            })
                                        elif current_avg > 0:
                                            # New account with no prior year data
                                            account_growth.append({
                                                "Account": account_id,
                                                f"{prior_year} Avg Monthly": 0,
                                                f"{current_year} Avg Monthly": current_avg,
                                                "Growth %": 999.9  # Mark as new
                                                })
                                    
                                    if account_growth:
                                        account_df = pd.DataFrame(account_growth)
                                        account_df = account_df.sort_values("Growth %", ascending=False)
                                        
                                        # Get top 5 and bottom 5
                                        top5 = account_df.head(5).copy()
                                        bottom5 = account_df.tail(5).copy()
                                        
                                        col1, col2 = st.columns(2)
                                        
                                        with col1:
                                            st.markdown("**Top 5 Accounts**")
                                            top5_display = top5.copy()
                                            top5_display[f"{prior_year} Avg Monthly"] = top5_display[f"{prior_year} Avg Monthly"].apply(lambda x: f"${x:,.0f}")
                                            top5_display[f"{current_year} Avg Monthly"] = top5_display[f"{current_year} Avg Monthly"].apply(lambda x: f"${x:,.0f}")
                                            top5_display["Growth %"] = top5_display["Growth %"].apply(lambda x: "NEW" if x >= 999 else f"{x:.1f}%")
                                            st.dataframe(top5_display, use_container_width=True, hide_index=True)
                                        
                                        with col2:
                                            st.markdown("**Bottom 5 Accounts**")
                                            bottom5_display = bottom5.copy()
                                            bottom5_display[f"{prior_year} Avg Monthly"] = bottom5_display[f"{prior_year} Avg Monthly"].apply(lambda x: f"${x:,.0f}")
                                            bottom5_display[f"{current_year} Avg Monthly"] = bottom5_display[f"{current_year} Avg Monthly"].apply(lambda x: f"${x:,.0f}")
                                            bottom5_display["Growth %"] = bottom5_display["Growth %"].apply(lambda x: "NEW" if x >= 999 else f"{x:.1f}%")
                                            st.dataframe(bottom5_display, use_container_width=True, hide_index=True)
                                        
                                        # Add a combined chart
                                        st.markdown("**Growth Comparison Chart**")
                                        combined = pd.concat([top5, bottom5])
                                        combined = combined.sort_values("Growth %", ascending=True)
                                        combined["Growth % Clean"] = combined["Growth %"].apply(lambda x: min(x, 300))  # Cap at 300% for display
                                        
                                        import plotly.graph_objects as go
                                        fig = go.Figure()
                                        colors = ['#ef4444' if x < 0 else '#10b981' for x in combined["Growth % Clean"]]
                                        fig.add_trace(go.Bar(
                                            y=combined["Account"],
                                            x=combined["Growth % Clean"],
                                            orientation='h',
                                            marker=dict(color=colors),
                                            text=combined["Growth %"].apply(lambda x: "NEW" if x >= 999 else f"{x:.1f}%"),
                                            textposition='outside'
                                        ))
                                        fig.update_layout(
                                            title="Account Growth Percentage",
                                            xaxis_title="Growth %",
                                            yaxis_title="Account",
                                            height=400,
                                            showlegend=False
                                        )
                                        st.plotly_chart(fig, use_container_width=True)
                                    else:
                                        st.info("Not enough account data for growth analysis")
                                else:
                                    st.info("Account information not available")
                            
                            # Display rolling average chart for selected price class (group-specific)
                            if (st.session_state.get("show_price_class_detail") and 
                                st.session_state.get("selected_price_class") and 
                                st.session_state.get("selected_group_value") == group_value):
                                
                                selected_pc = st.session_state["selected_price_class"]
                                
                                st.markdown("---")
                                st.markdown(f"### 📈 Rolling Average Sales: {selected_pc}")
                                st.markdown(f"**{group_name}: {group_value}**")
                                
                                # Close button
                                if st.button("✖ Close Detail View", key=f"close_price_class_detail_{group_value}"):
                                    st.session_state["show_price_class_detail"] = False
                                    st.session_state["selected_price_class"] = None
                                    st.session_state["selected_group_value"] = None
                                    st.rerun()
                                
                                # Filter data for selected price class within this group
                                pc_data = group_data[group_data["price_class_desc"] == selected_pc].copy()
                                
                                if pc_data.empty:
                                    st.warning(f"No data available for price class: {selected_pc}")
                                else:
                                    # Ensure we have invoice_date as datetime
                                    pc_data["invoice_date"] = pd.to_datetime(pc_data["invoice_date"])
                                    
                                    # Get the end date (most recent date in data)
                                    max_date = pc_data["invoice_date"].max()
                                    
                                    # Define two periods:
                                    # Period 1: Last 12 months (from max_date going back 12 months)
                                    # Period 2: 12 months prior to period 1 (months 13-24 before max_date)
                                    period1_start = max_date - pd.DateOffset(months=12)
                                    period1_end = max_date
                                    period2_start = max_date - pd.DateOffset(months=24)
                                    period2_end = period1_start
                                    
                                    st.info(f"Period 1 (Last 12 months): {period1_start.strftime('%Y-%m-%d')} to {period1_end.strftime('%Y-%m-%d')}")
                                    st.info(f"Period 2 (Prior 12 months): {period2_start.strftime('%Y-%m-%d')} to {period2_end.strftime('%Y-%m-%d')}")
                                    
                                    # Function to calculate rolling 12-month average
                                    def calculate_rolling_12month_avg_group(data, start_date, end_date):
                                        # Filter to period
                                        period_data = data[(data["invoice_date"] >= start_date) & (data["invoice_date"] <= end_date)].copy()
                                        
                                        if period_data.empty:
                                            return pd.DataFrame(columns=["date", "rolling_avg"])
                                        
                                        # Aggregate by day
                                        daily_sales = period_data.groupby(period_data["invoice_date"].dt.date)["extended_price_usd"].sum().reset_index()
                                        daily_sales.columns = ["date", "sales"]
                                        daily_sales["date"] = pd.to_datetime(daily_sales["date"])
                                        daily_sales = daily_sales.sort_values("date")
                                        
                                        # Create a complete date range for the period
                                        date_range = pd.date_range(start=start_date, end=end_date, freq='D')
                                        full_df = pd.DataFrame({"date": date_range})
                                        full_df = full_df.merge(daily_sales, on="date", how="left")
                                        full_df["sales"] = full_df["sales"].fillna(0)
                                        
                                        # Calculate rolling 3-week (21-day) average
                                        full_df["rolling_avg"] = full_df["sales"].rolling(window=21, min_periods=1).mean()
                                        
                                        return full_df[["date", "rolling_avg"]]
                                    
                                    # Calculate rolling averages for both periods
                                    period1_rolling = calculate_rolling_12month_avg_group(pc_data, period1_start, period1_end)
                                    period2_rolling = calculate_rolling_12month_avg_group(pc_data, period2_start, period2_end)
                                    
                                    # Prepare data for plotting
                                    import plotly.graph_objects as go
                                    
                                    fig = go.Figure()
                                    
                                    if not period1_rolling.empty:
                                        fig.add_trace(go.Scatter(
                                            x=period1_rolling["date"],
                                            y=period1_rolling["rolling_avg"],
                                            name="Last 12 Months Rolling Avg",
                                            mode="lines",
                                            line=dict(color="#1f77b4", width=2)
                                        ))
                                    
                                    if not period2_rolling.empty:
                                        fig.add_trace(go.Scatter(
                                            x=period2_rolling["date"],
                                            y=period2_rolling["rolling_avg"],
                                            name="Prior 12 Months Rolling Avg",
                                            mode="lines",
                                            line=dict(color="#ff7f0e", width=2)
                                        ))
                                    
                                    fig.update_layout(
                                        title=f"Rolling 12-Month Average Sales: {selected_pc}",
                                        xaxis_title="Date",
                                        yaxis_title="Rolling Average Sales (USD)",
                                        yaxis_tickformat="$,.0f",
                                        height=500,
                                        hovermode="x unified",
                                        legend=dict(
                                            orientation="h",
                                            yanchor="bottom",
                                            y=1.02,
                                            xanchor="right",
                                            x=1
                                        )
                                    )
                                    
                                    st.plotly_chart(fig, use_container_width=True, key=f"group_rolling_avg_{group_value}_{selected_pc}")
                                    
                                    # Show summary statistics
                                    col1, col2 = st.columns(2)
                                    with col1:
                                        if not period1_rolling.empty:
                                            avg1 = period1_rolling["rolling_avg"].mean()
                                            max1 = period1_rolling["rolling_avg"].max()
                                            min1 = period1_rolling["rolling_avg"].min()
                                            st.markdown("**Last 12 Months Statistics:**")
                                            st.markdown(f"- Average: ${avg1:,.2f}")
                                            st.markdown(f"- Maximum: ${max1:,.2f}")
                                            st.markdown(f"- Minimum: ${min1:,.2f}")
                                    
                                    with col2:
                                        if not period2_rolling.empty:
                                            avg2 = period2_rolling["rolling_avg"].mean()
                                            max2 = period2_rolling["rolling_avg"].max()
                                            min2 = period2_rolling["rolling_avg"].min()
                                            st.markdown("**Prior 12 Months Statistics:**")
                                            st.markdown(f"- Average: ${avg2:,.2f}")
                                            st.markdown(f"- Maximum: ${max2:,.2f}")
                                            st.markdown(f"- Minimum: ${min2:,.2f}")
                        
                        except Exception as e:
                            st.error(f"Error processing {group_name} {group_value}: {str(e)}")
                            import traceback
                            st.code(traceback.format_exc())
    
    # Cost Center ROI Tab
    with tabs[8]:
        st.header("Cost Center ROI Analysis")
        st.markdown("Return on Investment (ROI) for each cost center based on gross profit and inventory costs.")
        
        if st.button("Load Cost Center ROI", type="primary", key="load_cc_roi"):
            with st.spinner("Loading cost center data..."):
                try:
                    # Load all supplier performance data (no filters)
                    all_data = loaders.load_supplier_performance(
                        config.connection_string,
                        cost_centers=None,
                        start_date=None,
                        end_date=None
                    )
                    
                    # Load inventory costs
                    inventory_costs = loaders.load_inventory_costs(config.connection_string)
                    
                    if all_data.empty:
                        st.warning("No data found.")
                    else:
                        st.session_state["cc_roi_data"] = all_data
                        st.session_state["cc_roi_inv_costs"] = inventory_costs
                        st.success(f"Loaded {len(all_data):,} order lines.")
                
                except Exception as e:
                    st.error(f"Failed to load data: {e}")
        
        # Display cost center ROI if data is loaded
        if "cc_roi_data" in st.session_state and not st.session_state["cc_roi_data"].empty:
            all_data = st.session_state["cc_roi_data"]
            inventory_costs = st.session_state.get("cc_roi_inv_costs", pd.DataFrame())
            
            # Ensure invoice_date is datetime
            if "invoice_date" in all_data.columns:
                all_data["invoice_date"] = pd.to_datetime(all_data["invoice_date"])
            
            if "cost_center" in all_data.columns and "gross_profit_usd" in all_data.columns and "inventory_flag" in all_data.columns:
                # Filter to stocking items only
                stocking_data = all_data[all_data["inventory_flag"] == "Y"].copy()
                
                if not stocking_data.empty and not inventory_costs.empty:
                    # Get unique years
                    unique_years = sorted(stocking_data["invoice_date"].dt.year.unique())
                    
                    if len(unique_years) >= 2:
                        current_year = unique_years[-1]
                        prior_year = unique_years[-2]
                        
                        # Get unique cost centers and sort, excluding those starting with 1
                        all_cost_centers = stocking_data["cost_center"].dropna().unique()
                        cost_centers = sorted([cc for cc in all_cost_centers if not str(cc).startswith("1")])
                        
                        # Calculate ROI for each cost center by year
                        roi_results_current = []
                        roi_results_prior = []
                        
                        for cc in cost_centers:
                            cc_data = stocking_data[stocking_data["cost_center"] == cc].copy()
                            
                            # Get cost center description
                            cc_desc = ""
                            if "cost_center_desc" in cc_data.columns:
                                cc_desc_values = cc_data["cost_center_desc"].dropna().unique()
                                if len(cc_desc_values) > 0:
                                    cc_desc = str(cc_desc_values[0])
                            
                            # Format display name
                            if cc_desc:
                                cc_display = f"{cc_desc} ({cc})"
                            else:
                                cc_display = str(cc)
                            
                            # Get inventory cost for items in this cost center
                            cc_skus = cc_data["sku"].unique()
                            cc_inv_cost = inventory_costs[
                                inventory_costs["sku"].isin(cc_skus)
                            ]["total_cost"].sum()
                            
                            if cc_inv_cost > 0:
                                # Current year
                                cc_current = cc_data[cc_data["invoice_date"].dt.year == current_year]
                                if not cc_current.empty:
                                    cc_gp_current = cc_current["gross_profit_usd"].sum()
                                    roi_current = cc_gp_current / cc_inv_cost
                                    cc_sales_current = cc_current["extended_price_usd"].sum()
                                    
                                    roi_results_current.append({
                                        "Cost Center": cc_display,
                                        "ROI": roi_current,
                                        "Gross Profit": cc_gp_current,
                                        "Inventory Cost": cc_inv_cost,
                                        "Total Sales": cc_sales_current
                                    })
                                
                                # Prior year
                                cc_prior = cc_data[cc_data["invoice_date"].dt.year == prior_year]
                                if not cc_prior.empty:
                                    cc_gp_prior = cc_prior["gross_profit_usd"].sum()
                                    roi_prior = cc_gp_prior / cc_inv_cost
                                    cc_sales_prior = cc_prior["extended_price_usd"].sum()
                                    
                                    roi_results_prior.append({
                                        "Cost Center": cc_display,
                                        "ROI": roi_prior,
                                        "Gross Profit": cc_gp_prior,
                                        "Inventory Cost": cc_inv_cost,
                                        "Total Sales": cc_sales_prior
                                    })
                        
                        if roi_results_current or roi_results_prior:
                            # Create side-by-side columns
                            col1, col2 = st.columns(2)
                            
                            # Current year
                            with col1:
                                st.markdown(f"### {current_year} Cost Center ROI")
                                if roi_results_current:
                                    roi_df_current = pd.DataFrame(roi_results_current)
                                    roi_df_current = roi_df_current.sort_values("ROI", ascending=False)
                                    
                                    # Format for display
                                    roi_display_current = roi_df_current.copy()
                                    roi_display_current["ROI"] = roi_display_current["ROI"].apply(lambda x: f"{x:.2f}x")
                                    roi_display_current["Gross Profit"] = roi_display_current["Gross Profit"].apply(lambda x: f"${x:,.0f}")
                                    roi_display_current["Inventory Cost"] = roi_display_current["Inventory Cost"].apply(lambda x: f"${x:,.0f}")
                                    roi_display_current["Total Sales"] = roi_display_current["Total Sales"].apply(lambda x: f"${x:,.0f}")
                                    
                                    st.dataframe(roi_display_current, use_container_width=True, hide_index=True, height=400)
                                    
                                    # Download button with professional CSV formatting
                                    csv_export = roi_df_current[["Cost Center", "ROI", "Gross Profit", "Inventory Cost", "Total Sales"]].copy()
                                    csv_export.columns = ["Cost Center", "Return on Investment (ROI)", "Gross Profit ($)", "Inventory Cost ($)", "Total Sales ($)"]
                                    csv_export["Return on Investment (ROI)"] = csv_export["Return on Investment (ROI)"].apply(lambda x: f"{x:.2f}")
                                    csv_export["Gross Profit ($)"] = csv_export["Gross Profit ($)"].apply(lambda x: f"{x:.2f}")
                                    csv_export["Inventory Cost ($)"] = csv_export["Inventory Cost ($)"].apply(lambda x: f"{x:.2f}")
                                    csv_export["Total Sales ($)"] = csv_export["Total Sales ($)"].apply(lambda x: f"{x:.2f}")
                                    
                                    csv_current = csv_export.to_csv(index=False)
                                    st.download_button(
                                        label=f"📥 Download {current_year} Data",
                                        data=csv_current,
                                        file_name=f"cost_center_roi_{current_year}.csv",
                                        mime="text/csv",
                                        key=f"download_cc_roi_{current_year}"
                                    )
                                    
                                    # Summary metrics
                                    avg_roi = roi_df_current["ROI"].mean()
                                    st.metric("Average ROI", f"{avg_roi:.2f}x")
                                else:
                                    st.info(f"No data for {current_year}")
                            
                            # Prior year
                            with col2:
                                st.markdown(f"### {prior_year} Cost Center ROI")
                                if roi_results_prior:
                                    roi_df_prior = pd.DataFrame(roi_results_prior)
                                    roi_df_prior = roi_df_prior.sort_values("ROI", ascending=False)
                                    
                                    # Format for display
                                    roi_display_prior = roi_df_prior.copy()
                                    roi_display_prior["ROI"] = roi_display_prior["ROI"].apply(lambda x: f"{x:.2f}x")
                                    roi_display_prior["Gross Profit"] = roi_display_prior["Gross Profit"].apply(lambda x: f"${x:,.0f}")
                                    roi_display_prior["Inventory Cost"] = roi_display_prior["Inventory Cost"].apply(lambda x: f"${x:,.0f}")
                                    roi_display_prior["Total Sales"] = roi_display_prior["Total Sales"].apply(lambda x: f"${x:,.0f}")
                                    
                                    st.dataframe(roi_display_prior, use_container_width=True, hide_index=True, height=400)
                                    
                                    # Download button with professional CSV formatting
                                    csv_export = roi_df_prior[["Cost Center", "ROI", "Gross Profit", "Inventory Cost", "Total Sales"]].copy()
                                    csv_export.columns = ["Cost Center", "Return on Investment (ROI)", "Gross Profit ($)", "Inventory Cost ($)", "Total Sales ($)"]
                                    csv_export["Return on Investment (ROI)"] = csv_export["Return on Investment (ROI)"].apply(lambda x: f"{x:.2f}")
                                    csv_export["Gross Profit ($)"] = csv_export["Gross Profit ($)"].apply(lambda x: f"{x:.2f}")
                                    csv_export["Inventory Cost ($)"] = csv_export["Inventory Cost ($)"].apply(lambda x: f"{x:.2f}")
                                    csv_export["Total Sales ($)"] = csv_export["Total Sales ($)"].apply(lambda x: f"{x:.2f}")
                                    
                                    csv_prior = csv_export.to_csv(index=False)
                                    st.download_button(
                                        label=f"📥 Download {prior_year} Data",
                                        data=csv_prior,
                                        file_name=f"cost_center_roi_{prior_year}.csv",
                                        mime="text/csv",
                                        key=f"download_cc_roi_{prior_year}"
                                    )
                                    
                                    # Summary metrics
                                    avg_roi = roi_df_prior["ROI"].mean()
                                    st.metric("Average ROI", f"{avg_roi:.2f}x")
                                else:
                                    st.info(f"No data for {prior_year}")
                        else:
                            st.info("No cost centers with inventory costs found.")
                    else:
                        st.warning("Need at least 2 years of data for comparison.")
                else:
                    st.warning("No stocking items or inventory costs available.")
            else:
                st.error("Required columns not found in data.")
    
    # Sales Rep Performance Tab
    with tabs[9]:
        _build_sales_rep_performance_tab()

    # TAB 11: Executive Intelligence
    # ========================================================
    with tabs[10]:
        st.header("🎯 Executive Intelligence Hub")
        st.markdown("*Actionable insights, predictive analytics, and strategic opportunities*")
        
        # Load data button
        if st.button("🚀 Generate Intelligence Report", type="primary", key="load_exec_intel"):
            with st.spinner("Analyzing business data and generating insights..."):
                try:
                    # Load comprehensive dataset
                    exec_data = loaders.load_supplier_performance(
                        config.connection_string,
                        cost_centers=selected_cost_centers if selected_cost_centers else None
                    )
                    
                    inventory_costs = loaders.load_inventory_costs(config.connection_string)
                    
                    if exec_data.empty:
                        st.warning("No data available for analysis")
                    else:
                        # Store in session state
                        st.session_state["exec_intel_data"] = exec_data
                        st.session_state["exec_inventory_costs"] = inventory_costs
                        st.success(f"✅ Analyzed {len(exec_data):,} transactions")
                
                except Exception as e:
                    st.error(f"Failed to load data: {e}")
        
        # Display intelligence if data is loaded
        if "exec_intel_data" in st.session_state and not st.session_state["exec_intel_data"].empty:
            exec_data = st.session_state["exec_intel_data"].copy()
            inventory_costs = st.session_state.get("exec_inventory_costs", pd.DataFrame())
            
            # Ensure dates are parsed
            exec_data["invoice_date"] = pd.to_datetime(exec_data["invoice_date"])
            exec_data = exec_data[exec_data["invoice_date"].notna()]
            
            # Get date range
            max_date = exec_data["invoice_date"].max()
            min_date = exec_data["invoice_date"].min()
            
            # Define time periods
            last_12mo_start = max_date - pd.DateOffset(months=12)
            last_6mo_start = max_date - pd.DateOffset(months=6)
            last_3mo_start = max_date - pd.DateOffset(months=3)
            last_30d_start = max_date - pd.DateOffset(days=30)
            prior_12mo_start = max_date - pd.DateOffset(months=24)
            prior_12mo_end = last_12mo_start
            
            # Filter datasets
            last_12mo = exec_data[exec_data["invoice_date"] >= last_12mo_start].copy()
            last_6mo = exec_data[exec_data["invoice_date"] >= last_6mo_start].copy()
            last_3mo = exec_data[exec_data["invoice_date"] >= last_3mo_start].copy()
            last_30d = exec_data[exec_data["invoice_date"] >= last_30d_start].copy()
            prior_12mo = exec_data[(exec_data["invoice_date"] >= prior_12mo_start) & 
                                   (exec_data["invoice_date"] < prior_12mo_end)].copy()
            
            st.markdown("---")
            
            # ==============================================================
            # SECTION 1: KEY PERFORMANCE INDICATORS
            # ==============================================================
            st.markdown("## 📊 Key Performance Indicators")
            
            # Calculate KPIs
            revenue_12mo = last_12mo["extended_price_usd"].sum()
            revenue_prior = prior_12mo["extended_price_usd"].sum()
            revenue_growth = ((revenue_12mo - revenue_prior) / revenue_prior * 100) if revenue_prior > 0 else 0
            
            gp_12mo = last_12mo["gross_profit_usd"].sum() if "gross_profit_usd" in last_12mo.columns else 0
            gp_prior = prior_12mo["gross_profit_usd"].sum() if "gross_profit_usd" in prior_12mo.columns else 0
            margin_12mo = (gp_12mo / revenue_12mo * 100) if revenue_12mo > 0 else 0
            margin_prior = (gp_prior / revenue_prior * 100) if revenue_prior > 0 else 0
            
            # ROI calculation
            roi_12mo = 0
            if not inventory_costs.empty and "inventory_flag" in last_12mo.columns:
                stocking_12mo = last_12mo[last_12mo["inventory_flag"] == "Y"]
                inv_cost = inventory_costs["total_cost"].sum()
                if inv_cost > 0:
                    roi_12mo = gp_12mo / inv_cost
            
            # Transaction velocity
            avg_daily_transactions_12mo = len(last_12mo) / 365
            avg_daily_transactions_30d = len(last_30d) / 30
            velocity_change = ((avg_daily_transactions_30d - avg_daily_transactions_12mo) / 
                              avg_daily_transactions_12mo * 100) if avg_daily_transactions_12mo > 0 else 0
            
            # Display KPI cards
            kpi_col1, kpi_col2, kpi_col3, kpi_col4 = st.columns(4)
            
            with kpi_col1:
                st.metric(
                    label="💰 Revenue (12mo)",
                    value=f"${revenue_12mo:,.0f}",
                    delta=f"{revenue_growth:+.1f}% YoY"
                )
            
            with kpi_col2:
                st.metric(
                    label="📈 Gross Margin",
                    value=f"{margin_12mo:.1f}%",
                    delta=f"{margin_12mo - margin_prior:+.1f}pp YoY"
                )
            
            with kpi_col3:
                st.metric(
                    label="🎯 ROI",
                    value=f"{roi_12mo:.2f}x",
                    delta="On Inventory" if roi_12mo > 0 else "N/A"
                )
            
            with kpi_col4:
                st.metric(
                    label="⚡ Transaction Velocity",
                    value=f"{avg_daily_transactions_30d:.0f}/day",
                    delta=f"{velocity_change:+.1f}% vs avg"
                )
            
            # ==============================================================
            # SECTION 2: STRATEGIC INSIGHTS & ALERTS
            # ==============================================================
            st.markdown("---")
            st.markdown("## 🚨 Strategic Alerts & Opportunities")
            
            alerts = []
            opportunities = []
            
            # Alert 1: Declining revenue categories
            if not last_6mo.empty and not prior_12mo.empty:
                cat_revenue_current = last_6mo.groupby("item_class_1_desc")["extended_price_usd"].sum()
                cat_revenue_prior = prior_12mo.groupby("item_class_1_desc")["extended_price_usd"].sum()
                
                for cat in cat_revenue_current.index:
                    current_val = cat_revenue_current.get(cat, 0)
                    prior_val = cat_revenue_prior.get(cat, 0)
                    if prior_val > 0:
                        change_pct = (current_val - prior_val) / prior_val * 100
                        if change_pct < -20 and current_val > 10000:
                            alerts.append({
                                "type": "⚠️ Revenue Decline",
                                "message": f"{cat}: Revenue down {abs(change_pct):.1f}% in last 6 months",
                                "severity": "High" if change_pct < -30 else "Medium"
                            })
                        elif change_pct > 30 and current_val > 10000:
                            opportunities.append({
                                "type": "🚀 Growth Opportunity",
                                "message": f"{cat}: Strong growth at {change_pct:.1f}% - consider expanding",
                                "impact": "High"
                            })
            
            # Alert 2: Margin compression by supplier
            if not last_3mo.empty and "supplier_number" in last_3mo.columns:
                for supplier in last_3mo["supplier_number"].unique()[:20]:
                    supp_data = last_3mo[last_3mo["supplier_number"] == supplier]
                    supp_revenue = supp_data["extended_price_usd"].sum()
                    if supp_revenue > 50000:
                        supp_gp = supp_data["gross_profit_usd"].sum() if "gross_profit_usd" in supp_data.columns else 0
                        supp_margin = (supp_gp / supp_revenue * 100) if supp_revenue > 0 else 0
                        
                        if supp_margin < 15:
                            alerts.append({
                                "type": "💸 Low Margin Alert",
                                "message": f"Supplier {supplier}: Margin only {supp_margin:.1f}% on ${supp_revenue:,.0f}",
                                "severity": "High" if supp_margin < 10 else "Medium"
                            })
            
            # Alert 3: High velocity items with low stock
            if not inventory_costs.empty and not last_30d.empty:
                sku_velocity = last_30d.groupby("sku")["extended_price_usd"].agg(["sum", "count"])
                sku_velocity.columns = ["revenue_30d", "order_count_30d"]
                sku_velocity["avg_order_value"] = sku_velocity["revenue_30d"] / sku_velocity["order_count_30d"]
                
                high_velocity_skus = sku_velocity[
                    (sku_velocity["order_count_30d"] >= 5) & 
                    (sku_velocity["revenue_30d"] > 5000)
                ].head(10)
                
                for sku in high_velocity_skus.index:
                    opportunities.append({
                        "type": "⭐ High-Value SKU",
                        "message": f"SKU {sku}: ${high_velocity_skus.loc[sku, 'revenue_30d']:,.0f} in 30 days - ensure adequate stock",
                        "impact": "High"
                    })
            
            # Alert 4: Seasonality patterns
            if not last_12mo.empty:
                last_12mo_copy = last_12mo.copy()
                last_12mo_copy["month"] = last_12mo_copy["invoice_date"].dt.month
                monthly_revenue = last_12mo_copy.groupby("month")["extended_price_usd"].sum()
                
                if len(monthly_revenue) >= 6:
                    avg_monthly = monthly_revenue.mean()
                    current_month = max_date.month
                    current_month_revenue = monthly_revenue.get(current_month, 0)
                    
                    if current_month_revenue > avg_monthly * 1.3:
                        opportunities.append({
                            "type": "📅 Seasonal Peak",
                            "message": f"Current month tracking {(current_month_revenue/avg_monthly - 1)*100:.0f}% above average - capitalize on demand",
                            "impact": "High"
                        })
            
            # Display alerts and opportunities
            col1, col2 = st.columns(2)
            
            with col1:
                st.markdown("### 🚨 Priority Alerts")
                if alerts:
                    for alert in sorted(alerts, key=lambda x: x["severity"], reverse=True)[:5]:
                        severity_color = "🔴" if alert["severity"] == "High" else "🟡"
                        st.warning(f"{severity_color} **{alert['type']}**\n\n{alert['message']}")
                else:
                    st.success("✅ No critical alerts at this time")
            
            with col2:
                st.markdown("### 💡 Growth Opportunities")
                if opportunities:
                    for opp in opportunities[:5]:
                        st.info(f"**{opp['type']}**\n\n{opp['message']}")
                else:
                    st.info("Continue monitoring for emerging opportunities")
            
            # ==============================================================
            # SECTION 3: PROFITABILITY HEATMAP
            # ==============================================================
            st.markdown("---")
            st.markdown("## 🔥 Profitability Heatmap")
            st.markdown("*Identify your most and least profitable business segments*")
            
            if not last_12mo.empty and "supplier_number" in last_12mo.columns:
                # Create profitability matrix: Supplier vs Item Class
                top_suppliers = last_12mo.groupby("supplier_number")["extended_price_usd"].sum().nlargest(15).index
                
                heatmap_data = []
                for supplier in top_suppliers:
                    supp_data = last_12mo[last_12mo["supplier_number"] == supplier]
                    for item_class in supp_data["item_class_1_desc"].unique():
                        class_data = supp_data[supp_data["item_class_1_desc"] == item_class]
                        revenue = class_data["extended_price_usd"].sum()
                        gp = class_data["gross_profit_usd"].sum() if "gross_profit_usd" in class_data.columns else 0
                        margin = (gp / revenue * 100) if revenue > 0 else 0
                        
                        if revenue > 5000:  # Filter small values
                            heatmap_data.append({
                                "Supplier": supplier,
                                "Category": item_class,
                                "Revenue": revenue,
                                "Margin %": margin,
                                "Gross Profit": gp
                            })
                
                if heatmap_data:
                    heatmap_df = pd.DataFrame(heatmap_data)
                    
                    # Create pivot for heatmap
                    pivot = heatmap_df.pivot_table(
                        values="Margin %",
                        index="Supplier",
                        columns="Category",
                        aggfunc="mean"
                    )
                    
                    # Create heatmap
                    import plotly.graph_objects as go
                    
                    fig = go.Figure(data=go.Heatmap(
                        z=pivot.values,
                        x=pivot.columns,
                        y=pivot.index,
                        colorscale="RdYlGn",
                        text=np.round(pivot.values, 1),
                        texttemplate="%{text}%",
                        textfont={"size": 10},
                        colorbar=dict(title="Margin %")
                    ))
                    
                    fig.update_layout(
                        title="Margin % by Supplier × Category",
                        xaxis_title="Item Category",
                        yaxis_title="Supplier",
                        height=600
                    )
                    
                    st.plotly_chart(fig, use_container_width=True, key="exec_profitability_heatmap")
                    
                    # Show top and bottom performers
                    col1, col2 = st.columns(2)
                    
                    with col1:
                        st.markdown("#### 🏆 Top 5 Most Profitable Segments")
                        top5 = heatmap_df.nlargest(5, "Gross Profit")[["Supplier", "Category", "Revenue", "Margin %", "Gross Profit"]]
                        top5["Revenue"] = top5["Revenue"].apply(lambda x: f"${x:,.0f}")
                        top5["Margin %"] = top5["Margin %"].apply(lambda x: f"{x:.1f}%")
                        top5["Gross Profit"] = top5["Gross Profit"].apply(lambda x: f"${x:,.0f}")
                        st.dataframe(top5, use_container_width=True, hide_index=True)
                    
                    with col2:
                        st.markdown("#### ⚠️ Bottom 5 Segments (Needs Attention)")
                        bottom5 = heatmap_df.nsmallest(5, "Margin %")[["Supplier", "Category", "Revenue", "Margin %", "Gross Profit"]]
                        bottom5["Revenue"] = bottom5["Revenue"].apply(lambda x: f"${x:,.0f}")
                        bottom5["Margin %"] = bottom5["Margin %"].apply(lambda x: f"{x:.1f}%")
                        bottom5["Gross Profit"] = bottom5["Gross Profit"].apply(lambda x: f"${x:,.0f}")
                        st.dataframe(bottom5, use_container_width=True, hide_index=True)
            
            # ==============================================================
            # SECTION 4: TREND ANALYSIS & FORECASTING
            # ==============================================================
            st.markdown("---")
            st.markdown("## 📈 Trend Analysis & Forecasting")
            
            if not last_12mo.empty:
                # Monthly revenue trend with forecast
                monthly_data = last_12mo.copy()
                monthly_data["month"] = monthly_data["invoice_date"].dt.to_period("M").dt.to_timestamp()
                monthly_revenue = monthly_data.groupby("month")["extended_price_usd"].sum().reset_index()
                monthly_revenue.columns = ["month", "revenue"]
                monthly_revenue = monthly_revenue.sort_values("month")
                
                if len(monthly_revenue) >= 6:
                    # Try to import scipy for forecasting
                    try:
                        from scipy import stats
                        scipy_available = True
                    except ImportError:
                        scipy_available = False
                    
                    if scipy_available:
                        # Simple linear forecast for next 3 months
                        x = np.arange(len(monthly_revenue))
                        y = monthly_revenue["revenue"].values
                        slope, intercept, r_value, p_value, std_err = stats.linregress(x, y)
                        
                        # Forecast next 3 months
                        forecast_months = pd.date_range(
                            start=monthly_revenue["month"].max() + pd.DateOffset(months=1),
                            periods=3,
                            freq="MS"
                        )
                        forecast_x = np.arange(len(monthly_revenue), len(monthly_revenue) + 3)
                        forecast_y = slope * forecast_x + intercept
                        
                        # Create visualization
                        import plotly.graph_objects as go
                        
                        fig = go.Figure()
                        
                        # Historical data
                        fig.add_trace(go.Scatter(
                            x=monthly_revenue["month"],
                            y=monthly_revenue["revenue"],
                            name="Actual Revenue",
                            mode="lines+markers",
                            line=dict(color="#1f77b4", width=3),
                            marker=dict(size=8)
                        ))
                        
                        # Trend line
                        trend_y = slope * x + intercept
                        fig.add_trace(go.Scatter(
                            x=monthly_revenue["month"],
                            y=trend_y,
                            name="Trend",
                            mode="lines",
                            line=dict(color="#ff7f0e", width=2, dash="dash")
                        ))
                        
                        # Forecast
                        fig.add_trace(go.Scatter(
                            x=forecast_months,
                            y=forecast_y,
                            name="Forecast",
                            mode="lines+markers",
                            line=dict(color="#2ca02c", width=2, dash="dot"),
                            marker=dict(size=8, symbol="diamond")
                        ))
                        
                        fig.update_layout(
                            title="Revenue Trend & 3-Month Forecast",
                            xaxis_title="Month",
                            yaxis_title="Revenue (USD)",
                            yaxis_tickformat="$,.0f",
                            height=500,
                            hovermode="x unified"
                        )
                        
                        st.plotly_chart(fig, use_container_width=True, key="exec_forecast_chart")
                        
                        # Show forecast values
                        st.markdown("#### 🔮 Revenue Forecast")
                        forecast_df = pd.DataFrame({
                            "Month": forecast_months.strftime("%B %Y"),
                            "Forecasted Revenue": [f"${v:,.0f}" for v in forecast_y],
                            "Trend": "Growing" if slope > 0 else "Declining"
                        })
                        st.dataframe(forecast_df, use_container_width=True, hide_index=True)
                        
                        # Trend interpretation
                        trend_direction = "📈 Upward" if slope > 0 else "📉 Downward"
                        confidence = "High" if abs(r_value) > 0.7 else "Medium" if abs(r_value) > 0.4 else "Low"
                        st.info(f"**Trend Direction:** {trend_direction} | **Confidence:** {confidence} (R² = {r_value**2:.2f})")
                    else:
                        # Fallback: Simple trend visualization without forecast
                        st.info("📊 Advanced forecasting requires scipy library. Showing historical trend only.")
                        
                        import plotly.graph_objects as go
                        
                        fig = go.Figure()
                        
                        # Historical data
                        fig.add_trace(go.Scatter(
                            x=monthly_revenue["month"],
                            y=monthly_revenue["revenue"],
                            name="Monthly Revenue",
                            mode="lines+markers",
                            line=dict(color="#1f77b4", width=3),
                            marker=dict(size=8)
                        ))
                        
                        fig.update_layout(
                            title="Monthly Revenue Trend (Last 12 Months)",
                            xaxis_title="Month",
                            yaxis_title="Revenue (USD)",
                            yaxis_tickformat="$,.0f",
                            height=500,
                            hovermode="x unified"
                        )
                        
                        st.plotly_chart(fig, use_container_width=True, key="exec_trend_chart")
                        
                        # Simple trend analysis
                        first_half = monthly_revenue.head(6)["revenue"].mean()
                        second_half = monthly_revenue.tail(6)["revenue"].mean()
                        trend_pct = ((second_half - first_half) / first_half * 100) if first_half > 0 else 0
                        
                        if trend_pct > 5:
                            st.success(f"📈 **Positive Trend:** Revenue increased {trend_pct:.1f}% in recent 6 months vs prior 6 months")
                        elif trend_pct < -5:
                            st.warning(f"📉 **Negative Trend:** Revenue decreased {abs(trend_pct):.1f}% in recent 6 months vs prior 6 months")
                        else:
                            st.info(f"➡️ **Stable Trend:** Revenue relatively flat ({trend_pct:+.1f}% change)")
            
            # ==============================================================
            # SECTION 5: EFFICIENCY METRICS
            # ==============================================================
            st.markdown("---")
            st.markdown("## ⚙️ Operational Efficiency Metrics")
            
            col1, col2, col3 = st.columns(3)
            
            with col1:
                # Average order value
                aov_12mo = last_12mo["extended_price_usd"].mean() if not last_12mo.empty else 0
                aov_prior = prior_12mo["extended_price_usd"].mean() if not prior_12mo.empty else 0
                aov_change = ((aov_12mo - aov_prior) / aov_prior * 100) if aov_prior > 0 else 0
                
                st.metric(
                    label="💵 Avg Order Value",
                    value=f"${aov_12mo:,.2f}",
                    delta=f"{aov_change:+.1f}% YoY"
                )
                
                # SKU diversity
                unique_skus_12mo = last_12mo["sku"].nunique() if not last_12mo.empty else 0
                unique_skus_prior = prior_12mo["sku"].nunique() if not prior_12mo.empty else 0
                st.metric(
                    label="📦 Active SKUs",
                    value=f"{unique_skus_12mo:,}",
                    delta=f"{unique_skus_12mo - unique_skus_prior:+,} YoY"
                )
            
            with col2:
                # Revenue concentration (top 20% of SKUs)
                if not last_12mo.empty:
                    sku_revenue = last_12mo.groupby("sku")["extended_price_usd"].sum().sort_values(ascending=False)
                    top_20pct_count = int(len(sku_revenue) * 0.2)
                    top_20pct_revenue = sku_revenue.head(top_20pct_count).sum()
                    concentration = (top_20pct_revenue / sku_revenue.sum() * 100)
                    
                    st.metric(
                        label="🎯 Revenue Concentration",
                        value=f"{concentration:.1f}%",
                        delta="Top 20% of SKUs",
                        delta_color="off"
                    )
                
                # Supplier diversity
                unique_suppliers = last_12mo["supplier_number"].nunique() if not last_12mo.empty else 0
                st.metric(
                    label="🤝 Active Suppliers",
                    value=f"{unique_suppliers:,}",
                    delta="Last 12 months",
                    delta_color="off"
                )
            
            with col3:
                # Orders per day
                orders_per_day = len(last_12mo) / 365 if not last_12mo.empty else 0
                orders_per_day_30d = len(last_30d) / 30 if not last_30d.empty else 0
                
                st.metric(
                    label="📋 Orders Per Day",
                    value=f"{orders_per_day_30d:.1f}",
                    delta=f"{((orders_per_day_30d - orders_per_day) / orders_per_day * 100):+.1f}% vs avg" if orders_per_day > 0 else "N/A"
                )
                
                # Revenue per supplier
                revenue_per_supplier = (last_12mo["extended_price_usd"].sum() / unique_suppliers) if unique_suppliers > 0 else 0
                st.metric(
                    label="💰 Revenue per Supplier",
                    value=f"${revenue_per_supplier:,.0f}",
                    delta="Avg across all",
                    delta_color="off"
                )
            
            # ==============================================================
            # SECTION 6: CATEGORY PERFORMANCE MATRIX
            # ==============================================================
            st.markdown("---")
            st.markdown("## 📊 Category Performance Matrix")
            st.markdown("*Stars, Cash Cows, Question Marks, and Dogs*")
            
            if not last_12mo.empty and not prior_12mo.empty and "item_class_1_desc" in last_12mo.columns:
                # Calculate category metrics
                category_metrics = []
                
                # Get unique categories, filtering out nulls
                unique_categories = last_12mo["item_class_1_desc"].dropna().unique()
                
                for category in unique_categories:
                    if pd.isna(category) or str(category).strip() == "":
                        continue
                    
                    cat_current = last_12mo[last_12mo["item_class_1_desc"] == category]
                    cat_prior = prior_12mo[prior_12mo["item_class_1_desc"] == category]
                    
                    revenue_current = cat_current["extended_price_usd"].sum()
                    revenue_prior = cat_prior["extended_price_usd"].sum()
                    
                    if revenue_current > 1000:  # Lower threshold to 1000
                        growth_rate = ((revenue_current - revenue_prior) / revenue_prior * 100) if revenue_prior > 0 else 100
                        market_share = (revenue_current / last_12mo["extended_price_usd"].sum() * 100)
                        
                        # Classify using BCG matrix logic
                        if growth_rate > 10 and market_share > 5:
                            quadrant = "⭐ Star"
                        elif growth_rate <= 10 and market_share > 5:
                            quadrant = "🐄 Cash Cow"
                        elif growth_rate > 10 and market_share <= 5:
                            quadrant = "❓ Question Mark"
                        else:
                            quadrant = "🐕 Dog"
                        
                        category_metrics.append({
                            "Category": category,
                            "Revenue": revenue_current,
                            "Growth %": growth_rate,
                            "Market Share %": market_share,
                            "Classification": quadrant
                        })
                
                if category_metrics:
                    cat_df = pd.DataFrame(category_metrics)
                    
                    # Create scatter plot
                    fig = px.scatter(
                        cat_df,
                        x="Market Share %",
                        y="Growth %",
                        size="Revenue",
                        color="Classification",
                        hover_name="Category",
                        hover_data={"Revenue": ":$,.0f", "Growth %": ":.1f%", "Market Share %": ":.1f%"},
                        text="Category",
                        title="Category Performance Matrix (BCG-Style)",
                        color_discrete_map={
                            "⭐ Star": "#FFD700",
                            "🐄 Cash Cow": "#32CD32",
                            "❓ Question Mark": "#FFA500",
                            "🐕 Dog": "#DC143C"
                        }
                    )
                    
                    fig.update_traces(textposition="top center", textfont_size=8)
                    fig.add_hline(y=10, line_dash="dash", line_color="gray", annotation_text="Growth Threshold")
                    fig.add_vline(x=5, line_dash="dash", line_color="gray", annotation_text="Share Threshold")
                    
                    fig.update_layout(height=600)
                    
                    st.plotly_chart(fig, use_container_width=True, key="exec_bcg_matrix")
                    
                    # Show recommendations by quadrant
                    st.markdown("#### 💡 Strategic Recommendations")
                    
                    col1, col2, col3, col4 = st.columns(4)
                    
                    with col1:
                        stars = cat_df[cat_df["Classification"] == "⭐ Star"]
                        st.markdown("**⭐ Stars**")
                        st.success(f"{len(stars)} categories\n\n*Invest aggressively*")
                    
                    with col2:
                        cows = cat_df[cat_df["Classification"] == "🐄 Cash Cow"]
                        st.markdown("**🐄 Cash Cows**")
                        st.info(f"{len(cows)} categories\n\n*Maintain & harvest*")
                    
                    with col3:
                        questions = cat_df[cat_df["Classification"] == "❓ Question Mark"]
                        st.markdown("**❓ Question Marks**")
                        st.warning(f"{len(questions)} categories\n\n*Invest or divest*")
                    
                    with col4:
                        dogs = cat_df[cat_df["Classification"] == "🐕 Dog"]
                        st.markdown("**🐕 Dogs**")
                        st.error(f"{len(dogs)} categories\n\n*Consider exiting*")
                else:
                    st.warning("📊 No category data available for BCG matrix analysis. Categories may not have sufficient revenue history.")
            else:
                st.warning("📊 Insufficient data for category performance analysis. Need both current and prior year data with item classifications.")
            
            # ==============================================================
            # SECTION 7: EXECUTIVE SUMMARY & ACTION ITEMS
            # ==============================================================
            st.markdown("---")
            st.markdown("## 📋 Executive Summary & Action Items")
            
            # Generate automated insights
            summary_insights = []
            
            if revenue_growth > 15:
                summary_insights.append("✅ Strong revenue growth momentum - capitalize on market conditions")
            elif revenue_growth < -5:
                summary_insights.append("⚠️ Revenue declining - implement growth initiatives urgently")
            
            if margin_12mo > margin_prior:
                summary_insights.append("✅ Margin improvement achieved - maintain pricing discipline")
            else:
                summary_insights.append("⚠️ Margin compression detected - review pricing and costs")
            
            if roi_12mo > 1.5:
                summary_insights.append("✅ Excellent ROI on inventory - efficient capital deployment")
            elif roi_12mo < 0.8 and roi_12mo > 0:
                summary_insights.append("⚠️ Low ROI on inventory - optimize stock levels")
            
            if velocity_change > 10:
                summary_insights.append("✅ Transaction velocity increasing - scale operations accordingly")
            elif velocity_change < -10:
                summary_insights.append("⚠️ Transaction velocity declining - investigate root causes")
            
            st.markdown("### 🎯 Key Takeaways")
            for insight in summary_insights:
                st.markdown(f"- {insight}")
            
            # Recommended actions
            st.markdown("### 🚀 Recommended Actions")
            
            action_col1, action_col2 = st.columns(2)
            
            with action_col1:
                st.markdown("**Immediate (0-30 days)**")
                st.markdown("""
                1. Address low-margin supplier segments
                2. Review and optimize inventory for high-velocity SKUs
                3. Follow up on declining revenue categories
                4. Implement pricing adjustments where margin is compressed
                """)
            
            with action_col2:
                st.markdown("**Strategic (30-90 days)**")
                st.markdown("""
                1. Develop growth plan for "Star" categories
                2. Evaluate "Dog" categories for potential exit
                3. Invest in "Question Mark" categories with highest potential
                4. Harvest cash from "Cash Cow" categories efficiently
                """)
            
            # ==============================================================
            # SECTION 8: PRICE CLASS DEEP DIVE
            # ==============================================================
            st.markdown("---")
            st.markdown("## 💎 Price Class Performance Deep Dive")
            
            if not last_12mo.empty and "price_class_desc" in last_12mo.columns:
                # Calculate detailed metrics by price class
                pc_metrics = []
                
                for pc in last_12mo["price_class_desc"].unique():
                    if pd.isna(pc) or str(pc).strip() == "":
                        continue
                    
                    pc_current = last_12mo[last_12mo["price_class_desc"] == pc]
                    pc_prior = prior_12mo[prior_12mo["price_class_desc"] == pc] if not prior_12mo.empty else pd.DataFrame()
                    
                    revenue_current = pc_current["extended_price_usd"].sum()
                    revenue_prior = pc_prior["extended_price_usd"].sum() if not pc_prior.empty else 0
                    
                    if revenue_current > 5000:  # Minimum threshold
                        gp_current = pc_current["gross_profit_usd"].sum() if "gross_profit_usd" in pc_current.columns else 0
                        margin_current = (gp_current / revenue_current * 100) if revenue_current > 0 else 0
                        
                        order_count = len(pc_current)
                        unique_customers = pc_current["account_number"].nunique() if "account_number" in pc_current.columns else 0
                        avg_order_value = revenue_current / order_count if order_count > 0 else 0
                        
                        revenue_growth = ((revenue_current - revenue_prior) / revenue_prior * 100) if revenue_prior > 0 else 0
                        
                        # Calculate velocity (orders per day)
                        velocity = order_count / 365
                        
                        pc_metrics.append({
                            "Price Class": pc,
                            "Revenue": revenue_current,
                            "Growth %": revenue_growth,
                            "Margin %": margin_current,
                            "Orders": order_count,
                            "Avg Order $": avg_order_value,
                            "Customers": unique_customers,
                            "Velocity": velocity
                        })
                
                if pc_metrics:
                    pc_df = pd.DataFrame(pc_metrics)
                    pc_df = pc_df.sort_values("Revenue", ascending=False)
                    
                    # Top performers table
                    col1, col2 = st.columns(2)
                    
                    with col1:
                        st.markdown("### 🏆 Top 10 Price Classes by Revenue")
                        top10_pc = pc_df.head(10).copy()
                        top10_pc["Revenue"] = top10_pc["Revenue"].apply(lambda x: f"${x:,.0f}")
                        top10_pc["Growth %"] = top10_pc["Growth %"].apply(lambda x: f"{x:+.1f}%")
                        top10_pc["Margin %"] = top10_pc["Margin %"].apply(lambda x: f"{x:.1f}%")
                        top10_pc["Avg Order $"] = top10_pc["Avg Order $"].apply(lambda x: f"${x:,.0f}")
                        top10_pc["Velocity"] = top10_pc["Velocity"].apply(lambda x: f"{x:.1f}/day")
                        st.dataframe(top10_pc, use_container_width=True, hide_index=True)
                    
                    with col2:
                        st.markdown("### 📊 Price Class Metrics Distribution")
                        # Create bubble chart: Margin vs Growth, size = Revenue
                        fig = px.scatter(
                            pc_df.head(20),
                            x="Margin %",
                            y="Growth %",
                            size="Revenue",
                            color="Velocity",
                            hover_name="Price Class",
                            hover_data={
                                "Revenue": ":$,.0f",
                                "Orders": True,
                                "Customers": True
                            },
                            title="Margin vs Growth (Size = Revenue)",
                            color_continuous_scale="Viridis"
                        )
                        fig.update_layout(height=400)
                        st.plotly_chart(fig, use_container_width=True, key="pc_bubble_chart")
                    
                    # Price class trends over time
                    st.markdown("### 📈 Top 5 Price Classes - Revenue Trend")
                    top5_classes = pc_df.head(5)["Price Class"].tolist()
                    
                    if top5_classes:
                        trend_data = []
                        for pc in top5_classes:
                            pc_data = last_12mo[last_12mo["price_class_desc"] == pc].copy()
                            pc_data["month"] = pc_data["invoice_date"].dt.to_period("M").dt.to_timestamp()
                            monthly = pc_data.groupby("month")["extended_price_usd"].sum().reset_index()
                            monthly["price_class"] = pc
                            trend_data.append(monthly)
                        
                        if trend_data:
                            trend_df = pd.concat(trend_data, ignore_index=True)
                            
                            fig = px.line(
                                trend_df,
                                x="month",
                                y="extended_price_usd",
                                color="price_class",
                                title="Monthly Revenue Trend - Top 5 Price Classes",
                                labels={
                                    "month": "Month",
                                    "extended_price_usd": "Revenue (USD)",
                                    "price_class": "Price Class"
                                }
                            )
                            fig.update_layout(height=400, hovermode="x unified")
                            fig.update_traces(mode="lines+markers")
                            st.plotly_chart(fig, use_container_width=True, key="pc_trend_chart")
                    
                    # Alert: Low margin price classes with high volume
                    st.markdown("### ⚠️ Attention Required")
                    low_margin_high_volume = pc_df[
                        (pc_df["Margin %"] < 15) & 
                        (pc_df["Orders"] > 50)
                    ]
                    
                    if not low_margin_high_volume.empty:
                        st.warning(f"**{len(low_margin_high_volume)} price classes** have low margin (<15%) but high order volume")
                        display_concern = low_margin_high_volume.copy()
                        display_concern["Revenue"] = display_concern["Revenue"].apply(lambda x: f"${x:,.0f}")
                        display_concern["Margin %"] = display_concern["Margin %"].apply(lambda x: f"{x:.1f}%")
                        display_concern["Orders"] = display_concern["Orders"].apply(lambda x: f"{x:,}")
                        st.dataframe(
                            display_concern[["Price Class", "Revenue", "Margin %", "Orders"]],
                            use_container_width=True,
                            hide_index=True
                        )
                    else:
                        st.success("✅ All high-volume price classes maintain healthy margins")
            
            # ==============================================================
            # SECTION 9: SALES REP PERFORMANCE ANALYSIS
            # ==============================================================
            st.markdown("---")
            st.markdown("## 👥 Sales Rep Performance Analysis")
            
            if not last_12mo.empty and "salesperson_desc" in last_12mo.columns:
                # Calculate rep metrics
                rep_metrics = []
                
                for rep in last_12mo["salesperson_desc"].unique():
                    if pd.isna(rep) or str(rep).strip() == "":
                        continue
                    
                    rep_current = last_12mo[last_12mo["salesperson_desc"] == rep]
                    rep_prior = prior_12mo[prior_12mo["salesperson_desc"] == rep] if not prior_12mo.empty else pd.DataFrame()
                    
                    revenue_current = rep_current["extended_price_usd"].sum()
                    revenue_prior = rep_prior["extended_price_usd"].sum() if not rep_prior.empty else 0
                    
                    if revenue_current > 10000:  # Minimum threshold
                        gp_current = rep_current["gross_profit_usd"].sum() if "gross_profit_usd" in rep_current.columns else 0
                        margin_current = (gp_current / revenue_current * 100) if revenue_current > 0 else 0
                        
                        # Calculate margin consistency (std dev)
                        if "gross_profit_usd" in rep_current.columns and len(rep_current) > 0:
                            rep_current_copy = rep_current.copy()
                            rep_current_copy["margin"] = (rep_current_copy["gross_profit_usd"] / 
                                                         rep_current_copy["extended_price_usd"] * 100)
                            margin_std = rep_current_copy["margin"].std()
                        else:
                            margin_std = 0
                        
                        order_count = len(rep_current)
                        unique_customers = rep_current["account_number"].nunique() if "account_number" in rep_current.columns else 0
                        avg_order_value = revenue_current / order_count if order_count > 0 else 0
                        
                        revenue_growth = ((revenue_current - revenue_prior) / revenue_prior * 100) if revenue_prior > 0 else 0
                        
                        # Calculate monthly average
                        rep_current_copy = rep_current.copy()
                        rep_current_copy["month"] = rep_current_copy["invoice_date"].dt.to_period("M")
                        monthly_avg = rep_current_copy.groupby("month")["extended_price_usd"].sum().mean()
                        
                        rep_metrics.append({
                            "Sales Rep": rep,
                            "Revenue": revenue_current,
                            "Growth %": revenue_growth,
                            "Margin %": margin_current,
                            "Margin Consistency": margin_std,
                            "Orders": order_count,
                            "Customers": unique_customers,
                            "Avg Order $": avg_order_value,
                            "Monthly Avg": monthly_avg
                        })
                
                if rep_metrics:
                    rep_df = pd.DataFrame(rep_metrics)
                    rep_df = rep_df.sort_values("Revenue", ascending=False)
                    
                    # Summary metrics
                    col1, col2, col3, col4 = st.columns(4)
                    
                    with col1:
                        top_rep_revenue = rep_df.iloc[0]["Revenue"]
                        top_rep_name = rep_df.iloc[0]["Sales Rep"]
                        st.metric(
                            label="🏆 Top Performer",
                            value=top_rep_name,
                            delta=f"${top_rep_revenue:,.0f}"
                        )
                    
                    with col2:
                        avg_revenue = rep_df["Revenue"].mean()
                        st.metric(
                            label="📊 Average Revenue",
                            value=f"${avg_revenue:,.0f}",
                            delta=f"{len(rep_df)} reps"
                        )
                    
                    with col3:
                        avg_margin = rep_df["Margin %"].mean()
                        st.metric(
                            label="💰 Average Margin",
                            value=f"{avg_margin:.1f}%",
                            delta="Team average"
                        )
                    
                    with col4:
                        growing_reps = len(rep_df[rep_df["Growth %"] > 0])
                        growth_pct = (growing_reps / len(rep_df) * 100) if len(rep_df) > 0 else 0
                        st.metric(
                            label="📈 Growing Reps",
                            value=f"{growing_reps}/{len(rep_df)}",
                            delta=f"{growth_pct:.0f}%"
                        )
                    
                    # Performance comparison
                    st.markdown("### 📊 Sales Rep Performance Comparison")
                    
                    col1, col2 = st.columns(2)
                    
                    with col1:
                        # Top 10 by revenue
                        st.markdown("**Top 10 by Revenue**")
                        top10_reps = rep_df.head(10).copy()
                        top10_reps["Revenue"] = top10_reps["Revenue"].apply(lambda x: f"${x:,.0f}")
                        top10_reps["Growth %"] = top10_reps["Growth %"].apply(lambda x: f"{x:+.1f}%")
                        top10_reps["Margin %"] = top10_reps["Margin %"].apply(lambda x: f"{x:.1f}%")
                        st.dataframe(
                            top10_reps[["Sales Rep", "Revenue", "Growth %", "Margin %", "Customers"]],
                            use_container_width=True,
                            hide_index=True,
                            height=350
                        )
                    
                    with col2:
                        # Top 10 by growth
                        st.markdown("**Top 10 by Growth %**")
                        growth_leaders = rep_df.sort_values("Growth %", ascending=False).head(10).copy()
                        growth_leaders["Revenue"] = growth_leaders["Revenue"].apply(lambda x: f"${x:,.0f}")
                        growth_leaders["Growth %"] = growth_leaders["Growth %"].apply(lambda x: f"{x:+.1f}%")
                        growth_leaders["Monthly Avg"] = growth_leaders["Monthly Avg"].apply(lambda x: f"${x:,.0f}")
                        st.dataframe(
                            growth_leaders[["Sales Rep", "Growth %", "Revenue", "Monthly Avg"]],
                            use_container_width=True,
                            hide_index=True,
                            height=350
                        )
                    
                    # Visual analysis
                    st.markdown("### 📈 Rep Performance Matrix")
                    
                    # Create scatter plot: Revenue vs Margin, color by Growth
                    fig = px.scatter(
                        rep_df,
                        x="Revenue",
                        y="Margin %",
                        size="Orders",
                        color="Growth %",
                        hover_name="Sales Rep",
                        hover_data={
                            "Revenue": ":$,.0f",
                            "Customers": True,
                            "Margin Consistency": ":.1f"
                        },
                        title="Revenue vs Margin (Size = Order Count, Color = Growth %)",
                        color_continuous_scale="RdYlGn",
                        color_continuous_midpoint=0
                    )
                    fig.update_layout(height=500)
                    st.plotly_chart(fig, use_container_width=True, key="rep_scatter")
                    
                    # Margin discipline alerts
                    st.markdown("### 🎯 Margin Discipline Report")
                    
                    # Find reps with inconsistent margins
                    inconsistent_reps = rep_df[rep_df["Margin Consistency"] > 10].sort_values("Margin Consistency", ascending=False)
                    low_margin_reps = rep_df[rep_df["Margin %"] < 15]
                    
                    col1, col2 = st.columns(2)
                    
                    with col1:
                        if not inconsistent_reps.empty:
                            st.warning(f"⚠️ **{len(inconsistent_reps)} reps** with inconsistent margins (std dev > 10)")
                            display_incon = inconsistent_reps.head(5).copy()
                            display_incon["Margin %"] = display_incon["Margin %"].apply(lambda x: f"{x:.1f}%")
                            display_incon["Margin Consistency"] = display_incon["Margin Consistency"].apply(lambda x: f"{x:.1f}")
                            st.dataframe(
                                display_incon[["Sales Rep", "Margin %", "Margin Consistency"]],
                                use_container_width=True,
                                hide_index=True
                            )
                        else:
                            st.success("✅ All reps maintain consistent margins")
                    
                    with col2:
                        if not low_margin_reps.empty:
                            st.warning(f"⚠️ **{len(low_margin_reps)} reps** with margins below 15%")
                            display_low = low_margin_reps.head(5).copy()
                            display_low["Revenue"] = display_low["Revenue"].apply(lambda x: f"${x:,.0f}")
                            display_low["Margin %"] = display_low["Margin %"].apply(lambda x: f"{x:.1f}%")
                            st.dataframe(
                                display_low[["Sales Rep", "Revenue", "Margin %"]],
                                use_container_width=True,
                                hide_index=True
                            )
                        else:
                            st.success("✅ All reps maintain healthy margins")
                    
                    # Rep ranking by composite score
                    st.markdown("### 🏅 Composite Performance Score")
                    st.markdown("*Weighted score based on Revenue (40%), Growth (30%), Margin (20%), Customer Count (10%)*")
                    
                    # Normalize and calculate scores
                    rep_df_score = rep_df.copy()
                    rep_df_score["revenue_score"] = (rep_df_score["Revenue"] - rep_df_score["Revenue"].min()) / (rep_df_score["Revenue"].max() - rep_df_score["Revenue"].min()) * 40
                    
                    # Growth score (normalize to 0-30)
                    growth_normalized = (rep_df_score["Growth %"] + 50) / 100  # Shift and normalize
                    rep_df_score["growth_score"] = growth_normalized.clip(0, 1) * 30
                    
                    # Margin score (normalize to 0-20)
                    rep_df_score["margin_score"] = (rep_df_score["Margin %"] / rep_df_score["Margin %"].max()) * 20
                    
                    # Customer score (normalize to 0-10)
                    rep_df_score["customer_score"] = (rep_df_score["Customers"] - rep_df_score["Customers"].min()) / (rep_df_score["Customers"].max() - rep_df_score["Customers"].min()) * 10
                    
                    rep_df_score["Composite Score"] = (
                        rep_df_score["revenue_score"] + 
                        rep_df_score["growth_score"] + 
                        rep_df_score["margin_score"] + 
                        rep_df_score["customer_score"]
                    )
                    
                    rep_df_score = rep_df_score.sort_values("Composite Score", ascending=False)
                    
                    # Display top 15
                    top15_composite = rep_df_score.head(15).copy()
                    top15_composite["Rank"] = range(1, len(top15_composite) + 1)
                    top15_composite["Revenue"] = top15_composite["Revenue"].apply(lambda x: f"${x:,.0f}")
                    top15_composite["Growth %"] = top15_composite["Growth %"].apply(lambda x: f"{x:+.1f}%")
                    top15_composite["Margin %"] = top15_composite["Margin %"].apply(lambda x: f"{x:.1f}%")
                    top15_composite["Composite Score"] = top15_composite["Composite Score"].apply(lambda x: f"{x:.1f}")
                    
                    st.dataframe(
                        top15_composite[["Rank", "Sales Rep", "Composite Score", "Revenue", "Growth %", "Margin %", "Customers"]],
                        use_container_width=True,
                        hide_index=True,
                        height=400
                    )
            
            # Export option
            st.markdown("---")
            if st.button("📄 Export Intelligence Report", key="export_exec_intel"):
                st.info("Export functionality ready - integrate with existing PDF export framework")
        
        else:
            st.info("👆 Click 'Generate Intelligence Report' to begin analysis")

    # TAB 12: Returns
    # ========================================================
    with tabs[11]:
        st.header("↩️ Returns")
        st.markdown("*Returns volume, restock fees, and return rate insights by month, account, supplier, and cost center.*")

        # Date range + scoped filters (persist start date; end date defaults to today)
        pref_returns_start = user_prefs.get("returns_start_date")
        pref_returns_start_date = None
        if pref_returns_start:
            try:
                pref_returns_start_date = pd.to_datetime(pref_returns_start).date()
            except Exception:
                pref_returns_start_date = None
        default_returns_start = pref_returns_start_date or (date.today() - timedelta(days=365))

        fcol1, fcol2, fcol3, fcol4, fcol5, fcol6, fcol7, fcol8 = st.columns([1, 1, 1.2, 1.4, 1.2, 1.0, 1.4, 1.4])
        with fcol1:
            returns_start_date = st.date_input("Start date", value=default_returns_start, key="returns_start_date_input")
        with fcol2:
            returns_end_date = st.date_input("End date", value=date.today(), key="returns_end_date_input")
        with fcol3:
            returns_cost_centers = st.multiselect(
                "Cost centers",
                options=cost_center_options,
                default=[],
                key="returns_cost_centers",
            )
        with fcol4:
            returns_supplier_options = loaders.load_suppliers(config.connection_string, returns_cost_centers)
            returns_suppliers = st.multiselect(
                "Suppliers",
                options=returns_supplier_options,
                default=[s for s in selected_suppliers if s in returns_supplier_options],
                key="returns_suppliers",
            )
        with fcol5:
            account_filter_container = st.empty()
        with fcol6:
            samples_filter = st.radio(
                "Samples filter",
                options=["No samples", "Just samples", "All returns"],
                index=0,
                horizontal=True,
                key="returns_samples_filter",
            )
        with fcol7:
            return_reason_container = st.empty()
        with fcol8:
            reason_code_category_container = st.empty()

        # Persist start date preference when it changes
        try:
            if (pref_returns_start_date is None) or (returns_start_date != pref_returns_start_date):
                prefs_update = _load_prefs()
                prefs_update["returns_start_date"] = returns_start_date.isoformat()
                _save_prefs(prefs_update)
        except Exception:
            pass

        @st.cache_data(ttl=config.cache_ttl_seconds, show_spinner=False)
        def _load_returns_sources(cost_centers_tuple: tuple[str, ...], start: date, end: date):
            orders_df = loaders.load_orders(
                config.connection_string,
                cost_centers=list(cost_centers_tuple) if cost_centers_tuple else None,
                start_date=start,
                end_date=end,
            )
            openpo_m_df = loaders.load_openpo_m_lines(config.connection_string)
            return orders_df, openpo_m_df

        with st.spinner("Loading returns data..."):
            orders_src, openpo_m_src = _load_returns_sources(tuple(returns_cost_centers), returns_start_date, returns_end_date)

        if orders_src.empty:
            st.warning("No orders found for the selected date range and cost centers.")
        else:
            orders = orders_src.copy()

            # Exclude account # 1
            acct_series = pd.to_numeric(orders.get("account_number"), errors="coerce")
            orders = orders[acct_series.ne(1)]

            # Samples filter (cost centers starting with '1')
            samples_filter = st.session_state.get("returns_samples_filter", "No samples")
            cc_series = orders.get("cost_center", pd.Series(index=orders.index, dtype=object)).fillna("").astype(str).str.strip()
            if samples_filter == "Just samples":
                orders = orders[cc_series.str.startswith("1")]
            elif samples_filter == "No samples":
                orders = orders[~cc_series.str.startswith("1")]

            # Supplier filter
            if returns_suppliers and "supplier_number" in orders.columns:
                orders = orders[orders["supplier_number"].astype(str).isin([str(s) for s in returns_suppliers])]

            # Normalize account fields on orders for grouping
            orders["account_number"] = orders.get("account_number").fillna("").astype(str).str.strip()
            orders["bank_name2"] = orders.get("bank_name2").fillna("").astype(str).str.strip()
            orders["account_id"] = orders["bank_name2"].where(orders["bank_name2"].str.len() > 0, "Unknown") + " - " + orders["account_number"]
            if "cost_center_desc" in orders.columns:
                orders["cost_center_desc"] = orders.get("cost_center_desc").fillna("").astype(str).str.strip()
            else:
                orders["cost_center_desc"] = ""

            # Account filter (populate options dynamically from filtered orders)
            account_options = (
                orders.get("account_number")
                .dropna()
                .astype(str)
                .str.strip()
            )
            account_options = sorted(account_options[account_options.str.len() > 0].unique().tolist())
            st.session_state["returns_account_filter"] = [
                a for a in st.session_state.get("returns_account_filter", []) if a in account_options
            ]
            with account_filter_container:
                st.multiselect(
                    "Account #",
                    options=account_options,
                    default=st.session_state["returns_account_filter"],
                    key="returns_account_filter",
                )
            returns_account_filter = st.session_state.get("returns_account_filter", [])
            if returns_account_filter:
                orders = orders[orders.get("account_number").astype(str).isin([str(a) for a in returns_account_filter])]

            # Return reason filter (populate options dynamically from filtered orders)
            reason_col = "credit_type_desc" if "credit_type_desc" in orders.columns else "credit_type_code"
            reason_options = (
                orders.get(reason_col, pd.Series(dtype=str))
                .fillna("Unknown")
                .astype(str)
                .str.strip()
            )
            reason_options = sorted(reason_options[reason_options.str.len() > 0].unique().tolist())
            st.session_state["returns_reason_filter"] = [
                r for r in st.session_state.get("returns_reason_filter", []) if r in reason_options
            ]
            with return_reason_container:
                st.multiselect(
                    "Return Reason",
                    options=reason_options,
                    default=st.session_state["returns_reason_filter"],
                    key="returns_reason_filter",
                )
            selected_reasons = st.session_state.get("returns_reason_filter", [])
            if selected_reasons:
                orders = orders[orders.get(reason_col).fillna("Unknown").astype(str).str.strip().isin(selected_reasons)]

            # Reason code category filter (based on first character of reason_code)
            # Map: R=Customer, N=NRF, H=Special Deal, M=Mill
            reason_code_category_map = {
                "R": "Customer",
                "N": "NRF",
                "H": "Special Deal",
                "M": "Mill",
            }
            
            if "reason_code" in orders.columns:
                reason_code_series = orders.get("reason_code", pd.Series(dtype=str)).fillna("").astype(str).str.strip()
                orders["reason_code_category"] = reason_code_series.str[0].str.upper()
                orders["reason_code_category_label"] = orders["reason_code_category"].map(
                    lambda x: reason_code_category_map.get(x, "Other") if pd.notna(x) and x else "Unknown"
                )
                
                category_options = sorted([v for v in reason_code_category_map.values()])
                st.session_state["returns_reason_code_category_filter"] = [
                    c for c in st.session_state.get("returns_reason_code_category_filter", []) if c in category_options
                ]
                with reason_code_category_container:
                    st.multiselect(
                        "Return Category",
                        options=category_options,
                        default=st.session_state["returns_reason_code_category_filter"],
                        key="returns_reason_code_category_filter",
                    )
                selected_reason_categories = st.session_state.get("returns_reason_code_category_filter", [])
                if selected_reason_categories:
                    orders = orders[orders.get("reason_code_category_label").isin(selected_reason_categories)]
            else:
                orders["reason_code_category"] = "Unknown"
                orders["reason_code_category_label"] = "Unknown"

            # Build return date
            return_date = orders.get("order_entry_date")
            if "order_ship_date" in orders.columns:
                return_date = return_date.combine_first(orders["order_ship_date"])
            if "invoice_ship_date" in orders.columns:
                return_date = return_date.combine_first(orders["invoice_ship_date"])
            orders["return_date"] = pd.to_datetime(return_date, errors="coerce")

            # Value column (use ENTENDED_PRICE_NO_FUNDS)
            returned_value_series = pd.to_numeric(orders.get("extended_price_no_funds"), errors="coerce")
            orders["returned_value"] = returned_value_series.fillna(0.0).abs()

            # Identify returns from _ORDERS
            restock_charge = pd.to_numeric(orders.get("restocking_charge_p"), errors="coerce").fillna(0.0)
            fee_amount = pd.to_numeric(orders.get("discount_handling_charged"), errors="coerce").fillna(0.0)
            order_type = orders.get("order_type", pd.Series(index=orders.index, dtype=object)).fillna("").astype(str).str.strip().str.upper()
            customer_po = orders.get("customer_po", pd.Series(index=orders.index, dtype=object)).fillna("").astype(str).str.strip()
            detail_status = orders.get("detail_line_status", pd.Series(index=orders.index, dtype=object)).fillna("").astype(str).str.strip()
            detail_has_number = detail_status.str.contains(r"\d", regex=True)

            returns_with_fee = orders[restock_charge.gt(0)].copy()
            returns_with_fee["return_source"] = "orders_fee"
            returns_with_fee["invoice_number"] = returns_with_fee.get("invoice_number").fillna("").astype(str).str.strip()
            fee_series = pd.to_numeric(returns_with_fee.get("discount_handling_charged"), errors="coerce").fillna(0.0)
            inv_key = returns_with_fee["invoice_number"]
            returns_with_fee["invoice_fee_total"] = fee_series.groupby(inv_key, dropna=False).transform("max")
            returns_with_fee["invoice_return_value"] = returns_with_fee.groupby(inv_key, dropna=False)["returned_value"].transform("sum")
            returns_with_fee["invoice_return_lines"] = returns_with_fee.groupby(inv_key, dropna=False)["order_number"].transform("count")
            returns_with_fee["restock_fee"] = returns_with_fee.apply(
                lambda r: (r.get("invoice_fee_total", 0.0) / r.get("invoice_return_lines", 1)) if r.get("invoice_return_lines", 0) > 0 else 0.0,
                axis=1,
            )

            returns_no_fee = orders[
                fee_amount.le(0)
                & restock_charge.eq(0)   # exclude rows already captured in returns_with_fee
                & order_type.eq("C")
                & customer_po.str.startswith("RET")
                & ((detail_status == "") | detail_has_number)
            ].copy()
            returns_no_fee["invoice_number"] = returns_no_fee.get("invoice_number").fillna("").astype(str).str.strip()
            returns_no_fee["restock_fee"] = 0.0
            returns_no_fee["invoice_fee_total"] = 0.0
            returns_no_fee["invoice_return_value"] = returns_no_fee.get("returned_value").fillna(0.0)
            returns_no_fee["return_source"] = "orders_no_fee"

            # OPENPO_M fee returns (GL 9140) - merge fees into existing returns instead of creating duplicates
            openpo_m_fees = openpo_m_src.copy()
            if not openpo_m_fees.empty and "order_number" in openpo_m_fees.columns and "line_number" in openpo_m_fees.columns:
                openpo_m_fees = openpo_m_fees[openpo_m_fees.get("gl_number").fillna(0).astype(float).eq(9140)]
                openpo_m_fees["base_line"] = (openpo_m_fees["line_number"].fillna(0).astype(int) // 10) * 10
                openpo_m_fees["openpo_fee"] = pd.to_numeric(openpo_m_fees.get("fee_amount"), errors="coerce").fillna(0.0)
                # Create lookup: order_number + base_line -> openpo_fee
                openpo_fee_map = openpo_m_fees.groupby(["order_number", "base_line"])["openpo_fee"].sum().to_dict()
            else:
                openpo_fee_map = {}

            # Consolidate returns (without openpo_m as separate records)
            returns_frames = [returns_with_fee, returns_no_fee]
            returns_df = pd.concat(returns_frames, ignore_index=True) if returns_frames else pd.DataFrame()

            # Safety-net dedup: if the same order line somehow appears in both frames, keep the
            # returns_with_fee version (which appears first) since it carries the actual fee data.
            if not returns_df.empty and "order_number" in returns_df.columns and "line_number" in returns_df.columns:
                returns_df = returns_df.drop_duplicates(subset=["order_number", "line_number"], keep="first").reset_index(drop=True)
            
            # Merge OPENPO_M fees into existing return lines
            if not returns_df.empty and openpo_fee_map:
                returns_df["order_number_num"] = pd.to_numeric(returns_df.get("order_number"), errors="coerce")
                returns_df["line_number_num"] = pd.to_numeric(returns_df.get("line_number"), errors="coerce")
                returns_df["base_line_calc"] = (returns_df["line_number_num"].fillna(0).astype(int) // 10) * 10
                returns_df["openpo_additional_fee"] = returns_df.apply(
                    lambda r: openpo_fee_map.get((r.get("order_number_num"), r.get("base_line_calc")), 0.0),
                    axis=1,
                )
                # Add OPENPO_M fee to existing restock_fee
                returns_df["restock_fee"] = pd.to_numeric(returns_df.get("restock_fee"), errors="coerce").fillna(0.0) + returns_df["openpo_additional_fee"]
                # Recalculate invoice-level totals after adding openpo fees
                if "invoice_number" in returns_df.columns:
                    inv_key = returns_df["invoice_number"].fillna("").astype(str)
                    returns_df["invoice_fee_total"] = returns_df.groupby(inv_key, dropna=False)["restock_fee"].transform("sum")
                    returns_df["invoice_return_value"] = returns_df.groupby(inv_key, dropna=False)["returned_value"].transform("sum")

            if returns_df.empty:
                st.warning("No returns found for the selected filters and date range.")
            else:
                # Ensure invoice-level fields exist for all return rows
                if "invoice_number" in returns_df.columns:
                    returns_df["invoice_number"] = returns_df.get("invoice_number").fillna("").astype(str).str.strip()
                else:
                    returns_df["invoice_number"] = ""
                if "invoice_fee_total" not in returns_df.columns:
                    returns_df["invoice_fee_total"] = pd.to_numeric(returns_df.get("restock_fee"), errors="coerce").fillna(0.0)
                if "invoice_return_value" not in returns_df.columns:
                    returns_df["invoice_return_value"] = pd.to_numeric(returns_df.get("returned_value"), errors="coerce").fillna(0.0)

                # Ensure cost_center_desc is available on returns
                if "cost_center_desc" not in returns_df.columns:
                    try:
                        orders_cc = orders[["order_number", "line_number", "cost_center_desc"]].drop_duplicates()
                        returns_df = returns_df.merge(
                            orders_cc,
                            on=["order_number", "line_number"],
                            how="left",
                        )
                    except Exception:
                        returns_df["cost_center_desc"] = ""
                else:
                    returns_df["cost_center_desc"] = returns_df.get("cost_center_desc").fillna("").astype(str).str.strip()

                # Attach OPENPO_M message lines (M@MSG) grouped by base line (e.g., 010 -> 011-019, 020 -> 021-029)
                # Also include order-wide messages where M@LINE < 10 or M@LINE > 9000 for all lines on that order
                msg_map: dict[tuple[int, int], str] = {}
                msg_global_map: dict[int, str] = {}
                msg_src = openpo_m_src.copy()
                if isinstance(msg_src, pd.DataFrame) and not msg_src.empty and "message_text" in msg_src.columns:
                    msg_src = msg_src.copy()
                    msg_src["order_number_num"] = pd.to_numeric(msg_src.get("order_number"), errors="coerce")
                    msg_src["line_number_num"] = pd.to_numeric(msg_src.get("line_number"), errors="coerce")
                    msg_src = msg_src[msg_src["message_text"].fillna("").astype(str).str.strip().ne("")]
                    if not msg_src.empty:
                        msg_src["base_line"] = (msg_src["line_number_num"].fillna(0).astype(int) // 10) * 10

                        # Order-wide messages (apply to all lines)
                        global_mask = (msg_src["line_number_num"] < 10) | (msg_src["line_number_num"] > 9000)
                        global_msgs = msg_src[global_mask]
                        if not global_msgs.empty:
                            grouped_global = global_msgs.groupby(["order_number_num"], dropna=False)["message_text"].apply(
                                lambda s: " | ".join([t for t in s.astype(str).str.strip().tolist() if t])
                            )
                            msg_global_map = {(int(k) if pd.notna(k) else -1): v for k, v in grouped_global.items()}

                        # Base-line messages (e.g., 011-019 for base 010)
                        msg_detail = msg_src[
                            (msg_src["line_number_num"] > msg_src["base_line"])
                            & (msg_src["line_number_num"] < (msg_src["base_line"] + 10))
                        ]
                        msg_detail = msg_detail.sort_values(["order_number_num", "base_line", "line_number_num"])
                        grouped = msg_detail.groupby(["order_number_num", "base_line"], dropna=False)["message_text"].apply(
                            lambda s: " | ".join([t for t in s.astype(str).str.strip().tolist() if t])
                        )
                        msg_map = {(int(k[0]) if pd.notna(k[0]) else -1, int(k[1]) if pd.notna(k[1]) else -1): v for k, v in grouped.items()}

                returns_df["order_number_num"] = pd.to_numeric(returns_df.get("order_number"), errors="coerce")
                returns_df["line_number_num"] = pd.to_numeric(returns_df.get("line_number"), errors="coerce")
                returns_df["base_line"] = (returns_df["line_number_num"].fillna(0).astype(int) // 10) * 10
                returns_df["return_messages"] = returns_df.apply(
                    lambda r: " | ".join([
                        msg_global_map.get(int(r.get("order_number_num")) if pd.notna(r.get("order_number_num")) else -1, ""),
                        msg_map.get((int(r.get("order_number_num")) if pd.notna(r.get("order_number_num")) else -1,
                                     int(r.get("base_line")) if pd.notna(r.get("base_line")) else -1), ""),
                    ]).strip(" |"),
                    axis=1,
                )

                # Mark fee collection status BEFORE backfilling estimated fees.
                # fee_collected = True  → actual billed dollar fee already in system (item returned, fee charged)
                # fee_collected = False → outstanding return — fee not yet collected; amount estimated from %
                returns_df["fee_collected"] = pd.to_numeric(returns_df.get("restock_fee"), errors="coerce").fillna(0.0).gt(0)

                # Calculate missing fees from RESTOCKING_CHARGE_P percentage
                # If restock_fee is 0 or missing, use restocking_charge_p * returned_value
                # Note: RESTOCKING_CHARGE_P is already stored as a decimal (0.1 = 10%, not 10 = 10%)
                if "restocking_charge_p" in returns_df.columns and "returned_value" in returns_df.columns and "restock_fee" in returns_df.columns:
                    restock_fee = pd.to_numeric(returns_df.get("restock_fee"), errors="coerce").fillna(0.0)
                    restocking_pct = pd.to_numeric(returns_df.get("restocking_charge_p"), errors="coerce").fillna(0.0)
                    returned_value = pd.to_numeric(returns_df.get("returned_value"), errors="coerce").fillna(0.0)
                    
                    # For zero fees, calculate from percentage (already in decimal form)
                    zero_fee_mask = restock_fee.eq(0.0) | restock_fee.isna()
                    calculated_fee = restocking_pct * returned_value
                    returns_df["restock_fee"] = restock_fee.where(~zero_fee_mask, calculated_fee)
                    
                    # Recalculate invoice-level fee totals after fee calculation
                    if "invoice_number" in returns_df.columns:
                        inv_key = returns_df["invoice_number"].fillna("").astype(str)
                        returns_df["invoice_fee_total"] = returns_df.groupby(inv_key, dropna=False)["restock_fee"].transform("sum")

                returns_df_all = returns_df.copy()

                # Returns fee filter
                fee_filter = st.radio(
                    "Return fee filter",
                    options=["All returns", "Returns with fee", "Returns without fee"],
                    horizontal=True,
                    key="returns_fee_filter",
                )

                if fee_filter == "Returns with fee":
                    returns_df = returns_df_all[returns_df_all["restock_fee"].fillna(0) > 0].copy()
                elif fee_filter == "Returns without fee":
                    returns_df = returns_df_all[returns_df_all["restock_fee"].fillna(0) <= 0].copy()
                else:
                    returns_df = returns_df_all.copy()

                # Fee simulation controls
                sim_pct_enabled = False
                sim_min_enabled = False
                sim_from_pct = 30.0
                sim_to_pct = 20.0
                sim_min_fee = 0.0
                with st.expander("Simulate Return Fees", expanded=False):
                    st.caption("Simulations apply only to the currently filtered returns.")
                    sim_pct_enabled = st.checkbox("Simulate percent change", value=False, key="returns_sim_pct_enabled")
                    if sim_pct_enabled:
                        c1, c2 = st.columns(2)
                        with c1:
                            sim_from_pct = float(st.number_input("Replace %", min_value=0.0, max_value=100.0, value=30.0, step=0.5, key="returns_sim_from_pct"))
                        with c2:
                            sim_to_pct = float(st.number_input("With %", min_value=0.0, max_value=100.0, value=20.0, step=0.5, key="returns_sim_to_pct"))
                        st.caption("Matches restock % within +/- 0.5 of the selected value.")

                    sim_min_enabled = st.checkbox("Apply minimum fee ($)", value=False, key="returns_sim_min_enabled")
                    if sim_min_enabled:
                        sim_min_fee = float(st.number_input("Minimum fee ($)", min_value=0.0, value=0.0, step=1.0, key="returns_sim_min_fee"))

                # Apply fee simulations (percent change then minimum fee)
                sim_active = bool(sim_pct_enabled or sim_min_enabled)
                if sim_active and not returns_df.empty:
                    returns_df["restock_fee_original"] = pd.to_numeric(returns_df.get("restock_fee"), errors="coerce").fillna(0.0)
                    restock_pct_raw = pd.to_numeric(returns_df.get("restocking_charge_p"), errors="coerce").fillna(0.0)
                    returned_value_sim = pd.to_numeric(returns_df.get("returned_value"), errors="coerce").fillna(0.0)

                    simulated_fee = returns_df["restock_fee_original"].copy()
                    if sim_pct_enabled:
                        tol = 0.5
                        pct_mask = restock_pct_raw.between(sim_from_pct - tol, sim_from_pct + tol, inclusive="both")
                        simulated_fee = simulated_fee.where(~pct_mask, returned_value_sim * (sim_to_pct / 100.0))

                    if sim_min_enabled:
                        min_mask = returned_value_sim > 0
                        simulated_fee = simulated_fee.where(~min_mask, np.maximum(simulated_fee, sim_min_fee))

                    returns_df["restock_fee"] = simulated_fee

                # Normalize account identifiers
                returns_df["account_number"] = returns_df.get("account_number").fillna("").astype(str).str.strip()
                returns_df["bank_name2"] = returns_df.get("bank_name2").fillna("").astype(str).str.strip()
                if "bank_name2" in returns_df.columns:
                    returns_df["account_id"] = (
                        returns_df["bank_name2"].fillna("").astype(str).str.strip()
                        + " - "
                        + returns_df.get("account_number").fillna("").astype(str).str.strip()
                    )
                else:
                    returns_df["account_id"] = returns_df.get("account_number").fillna("").astype(str).str.strip()

                # Build order key for return-rate calculations
                returns_df["_return_key"] = (
                    returns_df.get("order_number").astype(str) + "-" + returns_df.get("line_number").astype(str)
                )
                if "_return_key" not in returns_df_all.columns:
                    returns_df_all["_return_key"] = (
                        returns_df_all.get("order_number").astype(str) + "-" + returns_df_all.get("line_number").astype(str)
                    )
                orders["_order_key"] = orders.get("order_number").astype(str) + "-" + orders.get("line_number").astype(str)

                # Monthly aggregates (all returns — fees include estimated outstanding amounts)
                returns_df["return_month"] = returns_df["return_date"].dt.to_period("M").dt.to_timestamp()
                monthly = returns_df.groupby("return_month", as_index=False).agg(
                    returns_count=("_return_key", "nunique"),
                    restock_fees=("restock_fee", "sum"),
                    returned_value=("returned_value", "sum"),
                )
                monthly["avg_restock_pct"] = monthly.apply(
                    lambda r: (r["restock_fees"] / r["returned_value"]) if r["returned_value"] > 0 else 0.0,
                    axis=1,
                )

                # Monthly aggregates — collected fees only (actual billed/received fees)
                _fee_col_mask = returns_df.get("fee_collected", pd.Series(index=returns_df.index, dtype=bool)).fillna(False)
                _returns_coll = returns_df[_fee_col_mask]
                if not _returns_coll.empty:
                    _monthly_c = _returns_coll.groupby("return_month", as_index=False).agg(
                        restock_fees_collected=("restock_fee", "sum"),
                        returned_value_collected=("returned_value", "sum"),
                    )
                    _monthly_c["avg_restock_pct_collected"] = _monthly_c.apply(
                        lambda r: (r["restock_fees_collected"] / r["returned_value_collected"]) if r["returned_value_collected"] > 0 else 0.0,
                        axis=1,
                    )
                    monthly = monthly.merge(
                        _monthly_c[["return_month", "restock_fees_collected", "avg_restock_pct_collected"]],
                        on="return_month",
                        how="left",
                    )
                else:
                    monthly["restock_fees_collected"] = 0.0
                    monthly["avg_restock_pct_collected"] = 0.0
                monthly["restock_fees_collected"] = monthly["restock_fees_collected"].fillna(0.0)
                monthly["avg_restock_pct_collected"] = monthly["avg_restock_pct_collected"].fillna(0.0)

                # Overall metrics
                total_returns = int(returns_df["_return_key"].nunique())
                total_value = float(returns_df["returned_value"].sum())
                # Fee metrics — collected only (actual billed fees; excludes estimated outstanding)
                _all_collected_mask = returns_df.get("fee_collected", pd.Series(index=returns_df.index, dtype=bool)).fillna(False)
                total_fees_collected = float(returns_df.loc[_all_collected_mask, "restock_fee"].sum())
                avg_fee_pct_collected = (total_fees_collected / total_value) if total_value > 0 else 0.0

                returns_df["return_date"] = pd.to_datetime(returns_df.get("return_date"), errors="coerce")
                mtd_start = pd.Timestamp(returns_end_date).replace(day=1)
                mtd_mask = returns_df["return_date"].between(mtd_start, pd.Timestamp(returns_end_date), inclusive="both")
                mtd_df = returns_df.loc[mtd_mask]
                mtd_returns = int(mtd_df.get("_return_key", pd.Series(dtype=str)).nunique())
                mtd_value = float(mtd_df.get("returned_value", pd.Series(dtype=float)).sum())
                # MTD collected fees only
                _mtd_collected_mask = mtd_df.get("fee_collected", pd.Series(index=mtd_df.index, dtype=bool)).fillna(False)
                mtd_fees = float(mtd_df.loc[_mtd_collected_mask, "restock_fee"].sum())
                mtd_avg_pct = (mtd_fees / mtd_value) if mtd_value > 0 else 0.0

                avg_monthly_returns = float(monthly.get("returns_count", pd.Series(dtype=float)).mean()) if not monthly.empty else 0.0
                avg_monthly_fees = float(monthly.get("restock_fees_collected", pd.Series(dtype=float)).mean()) if not monthly.empty else 0.0
                avg_monthly_value = float(monthly.get("returned_value", pd.Series(dtype=float)).mean()) if not monthly.empty else 0.0
                avg_monthly_pct = float(monthly.get("avg_restock_pct_collected", pd.Series(dtype=float)).mean()) if not monthly.empty else 0.0

                msg_series = returns_df.get("return_messages", pd.Series(index=returns_df.index, dtype=str)).fillna("").astype(str)
                error_mask = msg_series.str.lower().str.contains("error", regex=False)
                total_errors = int(returns_df.loc[error_mask, "_return_key"].nunique())

                swap_mask = msg_series.str.lower().str.contains("swap", regex=False)
                swap_returns = returns_df.loc[swap_mask].copy()
                total_swap_fees = float(swap_returns.get("restock_fee", pd.Series(dtype=float)).sum())
                swap_return_count = int(swap_returns.get("_return_key", pd.Series(dtype=str)).nunique())
                avg_swap_fee = (total_swap_fees / swap_return_count) if swap_return_count > 0 else 0.0

                mcol1, mcol2, mcol3, mcol4 = st.columns(4)
                mcol1.metric("Total Returns", f"{total_returns:,}")
                mcol2.metric("Total Fees Collected", f"${total_fees_collected:,.0f}", help="Collected only — excludes outstanding returns where fee not yet billed")
                mcol3.metric("Total Value Returned", f"${total_value:,.0f}")
                mcol4.metric("Avg Restock % (Collected)", f"{avg_fee_pct_collected:.1%}", help="Based on collected fees only")

                mtd1, mtd2, mtd3, mtd4 = st.columns(4)
                mtd1.metric("MTD Returns", f"{mtd_returns:,}")
                mtd2.metric("MTD Fees Collected", f"${mtd_fees:,.0f}", help="Collected only — excludes outstanding returns where fee not yet billed")
                mtd3.metric("MTD Value Returned", f"${mtd_value:,.0f}")
                mtd4.metric("MTD Avg Restock % (Collected)", f"{mtd_avg_pct:.1%}", help="Based on collected fees only")

                avg1, avg2, avg3, avg4 = st.columns(4)
                avg1.metric("Avg Monthly Returns", f"{avg_monthly_returns:,.0f}")
                avg2.metric("Avg Monthly Fees Collected", f"${avg_monthly_fees:,.0f}", help="Collected only — excludes outstanding returns where fee not yet billed")
                avg3.metric("Avg Monthly Value Returned", f"${avg_monthly_value:,.0f}")
                avg4.metric("Avg Monthly Restock % (Collected)", f"{avg_monthly_pct:.1%}", help="Based on collected fees only")

                mcol5, mcol6, mcol7, mcol8 = st.columns(4)
                mcol5.metric("Total Errors", f"{total_errors:,}")
                mcol6.metric("Total Swap Out Fees", f"${total_swap_fees:,.0f}")
                mcol7.metric("Avg Swap Out Fee", f"${avg_swap_fee:,.0f}")
                mcol8.metric("Swap Out Orders", f"{swap_return_count:,}")

                st.markdown("---")
                st.subheader("Monthly Trends")
                tcol1, tcol2 = st.columns(2)
                with tcol1:
                    fig_returns = px.line(
                        monthly,
                        x="return_month",
                        y="returns_count",
                        markers=True,
                        title="Returns Count (Monthly)",
                    )
                    st.plotly_chart(fig_returns, use_container_width=True, key="returns_count_monthly")
                with tcol2:
                    _fees_melted = monthly[["return_month", "restock_fees", "restock_fees_collected"]].rename(columns={
                        "restock_fees": "All Returns (incl. estimated)",
                        "restock_fees_collected": "Collected Only",
                    }).melt(id_vars="return_month", var_name="Type", value_name="Restock Fees")
                    fig_fees = px.line(
                        _fees_melted,
                        x="return_month",
                        y="Restock Fees",
                        color="Type",
                        markers=True,
                        title="Restock Fees by Month",
                        color_discrete_map={"All Returns (incl. estimated)": "#636efa", "Collected Only": "#00cc96"},
                    )
                    st.plotly_chart(fig_fees, use_container_width=True, key="returns_fees_monthly")

                tcol3, tcol4 = st.columns(2)
                with tcol3:
                    fig_value = px.line(
                        monthly,
                        x="return_month",
                        y="returned_value",
                        markers=True,
                        title="Total Value Returned (Monthly)",
                    )
                    st.plotly_chart(fig_value, use_container_width=True, key="returns_value_monthly")
                with tcol4:
                    _pct_melted = monthly[["return_month", "avg_restock_pct", "avg_restock_pct_collected"]].rename(columns={
                        "avg_restock_pct": "All Returns (incl. estimated)",
                        "avg_restock_pct_collected": "Collected Only",
                    }).melt(id_vars="return_month", var_name="Type", value_name="Avg Restock %")
                    fig_pct = px.line(
                        _pct_melted,
                        x="return_month",
                        y="Avg Restock %",
                        color="Type",
                        markers=True,
                        title="Average Restock % by Month",
                        color_discrete_map={"All Returns (incl. estimated)": "#636efa", "Collected Only": "#00cc96"},
                    )
                    fig_pct.update_yaxes(tickformat=".0%")
                    st.plotly_chart(fig_pct, use_container_width=True, key="returns_avg_pct_monthly")

                st.markdown("---")
                st.subheader("Return Value Distribution")
                return_values = returns_df.get("returned_value", pd.Series(dtype=float)).fillna(0.0).abs()
                avg_return_value = float(return_values.mean()) if not return_values.empty else 0.0
                if return_values.empty:
                    st.info("No return values available for distribution.")
                else:
                    bin_size = 200
                    max_value = float(return_values.max()) if return_values.max() > 0 else float(bin_size)
                    max_edge = int((max_value // bin_size) + 1) * bin_size
                    bins = list(range(0, max_edge + bin_size, bin_size))
                    labels = [f"${b:,.0f}-${b + bin_size:,.0f}" for b in bins[:-1]]
                    value_bins = pd.cut(return_values, bins=bins, labels=labels, right=False, include_lowest=True)
                    bin_counts = value_bins.value_counts().reindex(labels, fill_value=0).reset_index()
                    bin_counts.columns = ["Value Range", "Returns"]
                    bin_counts = bin_counts.sort_values("Returns", ascending=False).head(10)
                    fig_value_dist = px.bar(
                        bin_counts,
                        x="Value Range",
                        y="Returns",
                        title=f"Returns by Value ($200 bins) | Avg Return ${avg_return_value:,.0f}",
                    )
                    fig_value_dist.update_xaxes(tickangle=45)
                    st.plotly_chart(fig_value_dist, use_container_width=True, key="returns_value_distribution")

                # Return reason breakdown
                st.markdown("---")
                st.subheader("Return Reasons")
                reason_col = "credit_type_desc" if "credit_type_desc" in returns_df.columns else "credit_type_code"
                if reason_col in returns_df.columns:
                    reason_series = returns_df[reason_col].fillna("Unknown").astype(str).str.strip()
                    reason_counts = reason_series.value_counts().reset_index()
                    reason_counts.columns = ["Reason", "Count"]
                    fig_reasons = px.pie(
                        reason_counts,
                        names="Reason",
                        values="Count",
                        title="Returns by Reason",
                        hole=0.4,
                    )
                    st.plotly_chart(fig_reasons, use_container_width=True, key="returns_reason_pie")
                else:
                    st.info("Return reason data not available.")

                # Return reason code description breakdown
                if "order_reason_code_desc" in returns_df.columns:
                    reason_code_desc_series = returns_df["order_reason_code_desc"].fillna("Unknown").astype(str).str.strip()
                    reason_code_desc_counts = reason_code_desc_series.value_counts().reset_index()
                    reason_code_desc_counts.columns = ["Reason Code Description", "Count"]
                    fig_reason_codes = px.pie(
                        reason_code_desc_counts,
                        names="Reason Code Description",
                        values="Count",
                        title="Returns by Reason Code Description",
                        hole=0.4,
                    )
                    st.plotly_chart(fig_reason_codes, use_container_width=True, key="returns_reason_code_desc_pie")
                else:
                    st.info("Return reason code description data not available.")

                # Helper to build group summary tables
                def _group_returns(group_col: str | list[str]) -> pd.DataFrame:
                    orders_totals = orders.groupby(group_col, dropna=False)["_order_key"].nunique().rename("total_orders")
                    
                    # Count returns with "error" in messages
                    if "return_messages" in returns_df.columns:
                        error_mask = returns_df["return_messages"].fillna("").astype(str).str.lower().str.contains("error", regex=False)
                        returns_df_temp = returns_df.copy()
                        returns_df_temp["has_error"] = error_mask.astype(int)
                    else:
                        returns_df_temp = returns_df.copy()
                        returns_df_temp["has_error"] = 0
                    
                    returns_totals = returns_df_temp.groupby(group_col, dropna=False).agg(
                        returns_count=("_return_key", "nunique"),
                        restock_fees=("restock_fee", "sum"),
                        returned_value=("returned_value", "sum"),
                        error_count=("has_error", "sum"),
                    )
                    combined = returns_totals.join(orders_totals, how="left").fillna({"total_orders": 0})
                    combined["return_rate"] = combined.apply(
                        lambda r: (r["returns_count"] / r["total_orders"]) if r["total_orders"] > 0 else 0.0,
                        axis=1,
                    )
                    combined["avg_restock_fee"] = combined.apply(
                        lambda r: (r["restock_fees"] / r["returns_count"]) if r["returns_count"] > 0 else 0.0,
                        axis=1,
                    )
                    combined["avg_restock_pct"] = combined.apply(
                        lambda r: (r["restock_fees"] / r["returned_value"]) if r["returned_value"] > 0 else 0.0,
                        axis=1,
                    )
                    combined = combined.reset_index()
                    return combined

                # Account-level analysis
                account_summary = _group_returns(["account_number", "bank_name2"])
                account_summary = account_summary.sort_values("returns_count", ascending=False)
                account_summary["account_label"] = account_summary["bank_name2"].fillna("").astype(str).str.strip()
                account_summary["account_label"] = account_summary["account_label"].where(
                    account_summary["account_label"].str.len() > 0,
                    "Unknown"
                ) + " - " + account_summary["account_number"].fillna("").astype(str).str.strip()

                st.markdown("---")
                st.subheader("Account Return Drivers")
                acol1, acol2 = st.columns(2)
                with acol1:
                    top_accounts_count = account_summary.head(10)
                    fig_acct_count = px.bar(
                        top_accounts_count,
                        x="returns_count",
                        y="account_label",
                        orientation="h",
                        title="Top Accounts by Return Count",
                        hover_data={
                            "return_rate": ":.1%",
                            "avg_restock_pct": ":.1%",
                            "returned_value": ":$,.0f",
                            "restock_fees": ":$,.0f",
                        },
                    )
                    st.plotly_chart(fig_acct_count, use_container_width=True, key="returns_top_accounts_count")
                with acol2:
                    top_accounts_rate = account_summary.sort_values("return_rate", ascending=False).head(10)
                    fig_acct_rate = px.bar(
                        top_accounts_rate,
                        x="return_rate",
                        y="account_label",
                        orientation="h",
                        title="Top Accounts by Return Rate",
                        hover_data={
                            "returns_count": True,
                            "avg_restock_pct": ":.1%",
                            "returned_value": ":$,.0f",
                            "restock_fees": ":$,.0f",
                        },
                    )
                    fig_acct_rate.update_xaxes(tickformat=".0%")
                    st.plotly_chart(fig_acct_rate, use_container_width=True, key="returns_top_accounts_rate")

                # Supplier and cost center summaries
                if "salesperson_desc" not in returns_df.columns:
                    returns_df["salesperson_desc"] = "Unknown"
                supplier_label_col = "usual_supplier" if "usual_supplier" in returns_df.columns else "supplier_number"
                supplier_summary = _group_returns(supplier_label_col)
                supplier_summary = supplier_summary.sort_values("returns_count", ascending=False)
                salesperson_summary = _group_returns("salesperson_desc")
                salesperson_summary = salesperson_summary.sort_values("returns_count", ascending=False)
                cost_center_summary = _group_returns(["cost_center", "cost_center_desc"] if "cost_center_desc" in returns_df.columns else "cost_center")
                cost_center_summary = cost_center_summary.sort_values("return_rate", ascending=False)

                st.markdown("---")
                st.subheader("Summary Tables")
                t1, t2, t3, t4, t5 = st.tabs(["Accounts", "Suppliers", "Cost Centers", "Salespersons", "MTD Returns"])

                def _coerce_numeric(df: pd.DataFrame, cols: list[str]) -> pd.DataFrame:
                    df_out = df.copy()
                    for c in cols:
                        if c in df_out.columns:
                            df_out[c] = pd.to_numeric(df_out[c], errors="coerce")
                    return df_out

                def _style_summary(df: pd.DataFrame) -> "pd.io.formats.style.Styler":
                    format_map = {
                        "Returns": "{:,.0f}",
                        "Return Rate (%)": "{:,.1f}%",
                        "Returned Value": "${:,.2f}",
                        "Restock Fees": "${:,.2f}",
                        "Avg Restock Fee": "${:,.2f}",
                        "Avg Restock %": "{:,.1f}%",
                        "Total Orders": "{:,.0f}",
                    }
                    return df.style.format(format_map, na_rep="")

                def _render_selectable_table(df: pd.DataFrame, key: str, sort_cols: dict[str, str]) -> dict | None:
                    sort_col = st.selectbox(
                        "Sort by",
                        options=list(sort_cols.keys()),
                        index=0,
                        key=f"{key}_sort",
                    )
                    ascending = st.checkbox("Ascending", value=False, key=f"{key}_asc")
                    sort_key = sort_cols.get(sort_col)
                    df_sorted = df.sort_values(sort_key, ascending=ascending).reset_index(drop=True)

                    df_display = df_sorted.copy()
                    df_display["Return Rate (%)"] = df_display["Return Rate (%)"].apply(lambda x: f"{x:.1f}%")
                    df_display["Avg Restock %"] = df_display["Avg Restock %"].apply(lambda x: f"{x:.1f}%")
                    df_display["Returned Value"] = df_display["Returned Value"].apply(lambda x: f"${x:,.2f}")
                    df_display["Restock Fees"] = df_display["Restock Fees"].apply(lambda x: f"${x:,.2f}")
                    df_display["Avg Restock Fee"] = df_display["Avg Restock Fee"].apply(lambda x: f"${x:,.2f}")
                    df_display["Returns"] = df_display["Returns"].apply(lambda x: f"{x:,.0f}")
                    if "Error Count" in df_display.columns:
                        df_display["Error Count"] = df_display["Error Count"].apply(lambda x: f"{int(x):,.0f}")
                    df_display["Total Orders"] = df_display["Total Orders"].apply(lambda x: f"{x:,.0f}")

                    event = st.dataframe(
                        df_display,
                        use_container_width=True,
                        hide_index=True,
                        height=400,
                        selection_mode="single-row",
                        on_select="rerun",
                        key=key,
                    )
                    try:
                        if event and event.get("selection") and event["selection"].get("rows"):
                            idx = event["selection"]["rows"][0]
                            return df_sorted.iloc[idx].to_dict()
                    except Exception:
                        return None
                    return None

                def _render_return_details(title: str, filtered_returns: pd.DataFrame) -> None:
                    if filtered_returns.empty:
                        st.info("No return detail rows found for the selection.")
                        return
                    detail_df = filtered_returns.copy()
                    detail_df["returned_value"] = pd.to_numeric(detail_df.get("returned_value"), errors="coerce").fillna(0.0).abs()
                    detail_df["restock_fee"] = pd.to_numeric(detail_df.get("restock_fee"), errors="coerce").fillna(0.0).abs()
                    detail_df["invoice_fee_total"] = pd.to_numeric(detail_df.get("invoice_fee_total"), errors="coerce").fillna(0.0)
                    detail_df["invoice_return_value"] = pd.to_numeric(detail_df.get("invoice_return_value"), errors="coerce").fillna(0.0)
                    detail_df["return_pct"] = detail_df.apply(
                        lambda r: (r.get("invoice_fee_total", 0) / r.get("invoice_return_value", 0)) if r.get("invoice_return_value", 0) > 0 else 0.0,
                        axis=1,
                    )
                    reason_col_detail = "credit_type_desc" if "credit_type_desc" in detail_df.columns else "credit_type_code"
                    if reason_col_detail in detail_df.columns:
                        detail_df["reason_desc"] = detail_df.get(reason_col_detail).fillna("").astype(str).str.strip()
                    else:
                        detail_df["reason_desc"] = ""
                    detail_df["return_messages"] = detail_df.get("return_messages").fillna("").astype(str)
                    detail_df["invoice_number"] = detail_df.get("invoice_number").fillna("").astype(str).str.strip()
                    detail_df = detail_df[[
                        "sku",
                        "order_number",
                        "invoice_number",
                        "reason_desc",
                        "returned_value",
                        "restock_fee",
                        "return_pct",
                        "return_messages",
                    ]]
                    detail_df.columns = [
                        "Item Number",
                        "Order #",
                        "Invoice #",
                        "Reason",
                        "Return Value",
                        "Return Fee",
                        "Return %",
                        "Messages",
                    ]
                    st.markdown("#### " + title)
                    detail_display = detail_df.copy()
                    detail_display["Order #"] = detail_display["Order #"].apply(lambda x: f"{int(float(x))}" if pd.notna(x) and str(x).strip() != "" else "")
                    detail_display["Invoice #"] = detail_display["Invoice #"].apply(lambda x: f"{int(float(x))}" if pd.notna(x) and str(x).strip() != "" else "")
                    detail_display["Return Value"] = detail_display["Return Value"].apply(lambda x: f"${x:,.2f}")
                    detail_display["Return Fee"] = detail_display["Return Fee"].apply(lambda x: f"${x:,.2f}")
                    detail_display["Return %"] = detail_display["Return %"].apply(lambda x: f"{x:.1%}")
                    st.markdown(
                        """
                        <style>
                        div[data-testid="stDataFrame"] td {
                            white-space: normal !important;
                            line-height: 1.4 !important;
                            vertical-align: top !important;
                            height: auto !important;
                            min-height: 75px !important;
                            padding: 8px !important;
                        }
                        div[data-testid="stDataFrame"] td div {
                            white-space: normal !important;
                            word-wrap: break-word !important;
                            overflow-wrap: break-word !important;
                        }
                        div[data-testid="stDataFrame"] tr {
                            height: auto !important;
                        }
                        </style>
                        """,
                        unsafe_allow_html=True,
                    )
                    inv_series = detail_display["Invoice #"].fillna("").astype(str)
                    inv_unique = [v for v in inv_series.unique().tolist()]
                    inv_colors = {v: ("#f7f9fc" if i % 2 == 0 else "#ffffff") for i, v in enumerate(inv_unique)}
                    def _row_style(row):
                        color = inv_colors.get(str(row.get("Invoice #", "")), "#ffffff")
                        return [f"background-color: {color};"] * len(row)
                    st.dataframe(
                        detail_display.style.apply(_row_style, axis=1),
                        width="stretch",
                        hide_index=True,
                        height=600,
                        column_config={
                            "Item Number": st.column_config.TextColumn("Item Number", width="medium"),
                            "Order #": st.column_config.TextColumn("Order #", width="small"),
                            "Invoice #": st.column_config.TextColumn("Invoice #", width="small"),
                            "Reason": st.column_config.TextColumn("Reason", width="medium"),
                            "Return Value": st.column_config.TextColumn("Return Value", width="small"),
                            "Return Fee": st.column_config.TextColumn("Return Fee", width="small"),
                            "Return %": st.column_config.TextColumn("Return %", width="small"),
                            "Messages": st.column_config.TextColumn(
                                "Messages",
                                width="large",
                                help="Return messages from OPENPO_M",
                            ),
                        },
                    )

                def _format_summary(df_in: pd.DataFrame, label_col: str) -> pd.DataFrame:
                    df_out = df_in.copy()
                    df_out = df_out[[label_col, "returns_count", "return_rate", "returned_value", "restock_fees", "avg_restock_fee", "avg_restock_pct", "error_count", "total_orders"]]
                    df_out.columns = [
                        label_col,
                        "Returns",
                        "Return Rate (%)",
                        "Returned Value",
                        "Restock Fees",
                        "Avg Restock Fee",
                        "Avg Restock %",
                        "Error Count",
                        "Total Orders",
                    ]
                    df_out["Return Rate (%)"] = df_out["Return Rate (%)"] * 100
                    df_out["Avg Restock %"] = df_out["Avg Restock %"] * 100
                    df_out = _coerce_numeric(
                        df_out,
                        [
                            "Returns",
                            "Return Rate (%)",
                            "Returned Value",
                            "Restock Fees",
                            "Avg Restock Fee",
                            "Avg Restock %",
                            "Error Count",
                            "Total Orders",
                        ],
                    )
                    return df_out

                with t1:
                    df_accounts = account_summary.copy()
                    df_accounts = df_accounts[[
                        "account_number",
                        "bank_name2",
                        "returns_count",
                        "return_rate",
                        "returned_value",
                        "restock_fees",
                        "avg_restock_fee",
                        "avg_restock_pct",
                        "error_count",
                        "total_orders",
                    ]]
                    df_accounts.columns = [
                        "Account #",
                        "Bank Name 2",
                        "Returns",
                        "Return Rate (%)",
                        "Returned Value",
                        "Restock Fees",
                        "Avg Restock Fee",
                        "Avg Restock %",
                        "Error Count",
                        "Total Orders",
                    ]
                    df_accounts["Return Rate (%)"] = df_accounts["Return Rate (%)"] * 100
                    df_accounts["Avg Restock %"] = df_accounts["Avg Restock %"] * 100
                    df_accounts = _coerce_numeric(
                        df_accounts,
                        [
                            "Returns",
                            "Return Rate (%)",
                            "Returned Value",
                            "Restock Fees",
                            "Avg Restock Fee",
                            "Avg Restock %",
                            "Error Count",
                            "Total Orders",
                        ],
                    )
                    selected_row = _render_selectable_table(
                        df_accounts,
                        "returns_accounts_table",
                        {
                            "Return Rate (%)": "Return Rate (%)",
                            "Returns": "Returns",
                            "Returned Value": "Returned Value",
                            "Restock Fees": "Restock Fees",
                            "Avg Restock Fee": "Avg Restock Fee",
                            "Avg Restock %": "Avg Restock %",
                            "Error Count": "Error Count",
                            "Total Orders": "Total Orders",
                        },
                    )
                    if selected_row:
                        acct_num = str(selected_row.get("Account #", "")).strip()
                        bank_name = str(selected_row.get("Bank Name 2", "")).strip()
                        detail_returns = returns_df[
                            (returns_df.get("account_number").astype(str).str.strip() == acct_num)
                            & (returns_df.get("bank_name2").astype(str).str.strip() == bank_name)
                        ]
                        _render_return_details(f"Account {acct_num} - {bank_name}", detail_returns)

                with t2:
                    df_suppliers = _format_summary(supplier_summary, supplier_label_col)
                    selected_row = _render_selectable_table(
                        df_suppliers,
                        "returns_suppliers_table",
                        {
                            "Return Rate (%)": "Return Rate (%)",
                            "Returns": "Returns",
                            "Returned Value": "Returned Value",
                            "Restock Fees": "Restock Fees",
                            "Avg Restock Fee": "Avg Restock Fee",
                            "Avg Restock %": "Avg Restock %",
                            "Error Count": "Error Count",
                            "Total Orders": "Total Orders",
                        },
                    )
                    if selected_row:
                        supplier_label = str(selected_row.get(supplier_label_col, "")).strip()
                        detail_returns = returns_df[returns_df.get(supplier_label_col).astype(str).str.strip() == supplier_label]
                        _render_return_details(f"Supplier {supplier_label}", detail_returns)

                with t3:
                    if "cost_center_desc" in cost_center_summary.columns:
                        df_cc = cost_center_summary.copy()
                        df_cc = df_cc[[
                            "cost_center",
                            "cost_center_desc",
                            "returns_count",
                            "return_rate",
                            "returned_value",
                            "restock_fees",
                            "avg_restock_fee",
                            "avg_restock_pct",
                            "error_count",
                            "total_orders",
                        ]]
                        df_cc.columns = [
                            "Cost Center",
                            "Description",
                            "Returns",
                            "Return Rate (%)",
                            "Returned Value",
                            "Restock Fees",
                            "Avg Restock Fee",
                            "Avg Restock %",
                            "Error Count",
                            "Total Orders",
                        ]
                        df_cc["Return Rate (%)"] = df_cc["Return Rate (%)"] * 100
                        df_cc["Avg Restock %"] = df_cc["Avg Restock %"] * 100
                        df_cc = _coerce_numeric(
                            df_cc,
                            [
                                "Returns",
                                "Return Rate (%)",
                                "Returned Value",
                                "Restock Fees",
                                "Avg Restock Fee",
                                "Avg Restock %",
                                "Error Count",
                                "Total Orders",
                            ],
                        )
                    else:
                        df_cc = _format_summary(cost_center_summary, "cost_center")
                    selected_row = _render_selectable_table(
                        df_cc,
                        "returns_cost_centers_table",
                        {
                            "Return Rate (%)": "Return Rate (%)",
                            "Returns": "Returns",
                            "Returned Value": "Returned Value",
                            "Restock Fees": "Restock Fees",
                            "Avg Restock Fee": "Avg Restock Fee",
                            "Avg Restock %": "Avg Restock %",
                            "Error Count": "Error Count",
                            "Total Orders": "Total Orders",
                        },
                    )
                    if selected_row:
                        cc = str(selected_row.get("Cost Center", selected_row.get("cost_center", ""))).strip()
                        detail_returns = returns_df[returns_df.get("cost_center").astype(str).str.strip() == cc]
                        _render_return_details(f"Cost Center {cc}", detail_returns)

                with t4:
                    df_sales = _format_summary(salesperson_summary, "salesperson_desc")
                    df_sales = df_sales.rename(columns={"salesperson_desc": "Salesperson"})
                    selected_row = _render_selectable_table(
                        df_sales,
                        "returns_salespersons_table",
                        {
                            "Return Rate (%)": "Return Rate (%)",
                            "Returns": "Returns",
                            "Returned Value": "Returned Value",
                            "Restock Fees": "Restock Fees",
                            "Avg Restock Fee": "Avg Restock Fee",
                            "Avg Restock %": "Avg Restock %",
                            "Error Count": "Error Count",
                            "Total Orders": "Total Orders",
                        },
                    )
                    if selected_row:
                        salesperson_label = str(selected_row.get("Salesperson", "")).strip()
                        detail_returns = returns_df[returns_df.get("salesperson_desc").astype(str).str.strip() == salesperson_label]
                        _render_return_details(f"Salesperson {salesperson_label}", detail_returns)

                with t5:
                    st.markdown("#### Month-To-Date Returns")
                    mtd_start = pd.Timestamp(returns_end_date).replace(day=1)
                    mtd_end   = pd.Timestamp(returns_end_date)

                    reason_col_mtd = "credit_type_desc" if "credit_type_desc" in returns_df_all.columns else "credit_type_code"

                    # Helper to build a display-ready DataFrame for both sub-tables
                    def _build_mtd_display(df_in: pd.DataFrame, date_col: str, date_label: str) -> pd.DataFrame:
                        cols_src = [
                            date_col,
                            "account_number",
                            "bank_name2",
                            "cost_center",
                            "returned_value",
                            "restock_fee",
                            "reason_code_category_label",
                            reason_col_mtd,
                            "order_number",
                        ]
                        # Only include columns that exist
                        cols_src = [c for c in cols_src if c in df_in.columns]
                        out = df_in[cols_src].copy()
                        # Rename date column to the display label
                        out = out.rename(columns={date_col: date_label})
                        # Build full column list (in case some were missing)
                        col_map = {
                            "account_number":            "Account #",
                            "bank_name2":                "Bank Name 2",
                            "cost_center":               "Cost Center",
                            "returned_value":            "Return Value",
                            "restock_fee":               "Return Fee",
                            "reason_code_category_label": "Return Category",
                            reason_col_mtd:              "Return Reason",
                            "order_number":              "Order #",
                        }
                        out = out.rename(columns=col_map)
                        # Format
                        out[date_label] = pd.to_datetime(out[date_label], errors="coerce").dt.strftime("%Y-%m-%d")
                        if "Return Value" in out.columns:
                            out["Return Value"] = pd.to_numeric(out["Return Value"], errors="coerce").fillna(0.0).apply(lambda x: f"${x:,.2f}")
                        if "Return Fee" in out.columns:
                            out["Return Fee"] = pd.to_numeric(out["Return Fee"], errors="coerce").fillna(0.0).apply(lambda x: f"${x:,.2f}")
                        if "Order #" in out.columns:
                            out["Order #"] = out["Order #"].astype(str).apply(lambda x: f"{int(float(x))}" if x.strip() and x != "nan" else "")
                        # Sort newest first
                        out["_sort"] = pd.to_datetime(out[date_label], errors="coerce")
                        out = out.sort_values("_sort", ascending=False).drop(columns=["_sort"])
                        return out

                    # -- invoice_number numeric helper applied to returns_df_all --
                    _inv_num_all = pd.to_numeric(returns_df_all.get("invoice_number", pd.Series(index=returns_df_all.index, dtype=object)), errors="coerce").fillna(0)

                    # ── Tab A: Returns Received MTD ────────────────────────────────────
                    # return_date in current month AND invoice_number > 0 (fully closed)
                    _came_back_mask = (
                        returns_df_all["return_date"].between(mtd_start, mtd_end, inclusive="both")
                        & _inv_num_all.gt(0)
                    )
                    df_came_back = returns_df_all.loc[_came_back_mask].copy()

                    # ── Tab B: Returns Issued MTD ──────────────────────────────────────
                    # order_date in current month AND invoice_number = 0 (not yet closed)
                    if "order_date" in returns_df_all.columns:
                        _order_date_series = pd.to_datetime(returns_df_all["order_date"], errors="coerce")
                        _issued_mask = (
                            _order_date_series.between(mtd_start, mtd_end, inclusive="both")
                            & _inv_num_all.eq(0)
                        )
                        df_issued = returns_df_all.loc[_issued_mask].copy()
                    else:
                        df_issued = pd.DataFrame()

                    mtd_tab_a, mtd_tab_b = st.tabs(["Returns Received MTD", "Returns Issued MTD"])

                    with mtd_tab_a:
                        if df_came_back.empty:
                            st.info("No received returns found for the current month.")
                        else:
                            n_cb = len(df_came_back)
                            val_cb = float(df_came_back["returned_value"].sum())
                            fee_cb = float(df_came_back["restock_fee"].sum())
                            ca1, ca2, ca3 = st.columns(3)
                            ca1.metric("Returns Received", f"{n_cb:,}")
                            ca2.metric("Return Value", f"${val_cb:,.2f}")
                            ca3.metric("Fees Collected", f"${fee_cb:,.2f}")
                            st.dataframe(
                                _build_mtd_display(df_came_back, "return_date", "Return Date"),
                                use_container_width=True,
                                hide_index=True,
                                height=600,
                            )

                    with mtd_tab_b:
                        if df_issued.empty:
                            st.info("No issued returns found for the current month.")
                        else:
                            n_is = len(df_issued)
                            val_is = float(df_issued["returned_value"].sum())
                            fee_is = float(df_issued["restock_fee"].sum())
                            cb1, cb2, cb3 = st.columns(3)
                            cb1.metric("Returns Issued", f"{n_is:,}")
                            cb2.metric("Return Value", f"${val_is:,.2f}")
                            cb3.metric("Estimated Fees", f"${fee_is:,.2f}")
                            st.dataframe(
                                _build_mtd_display(df_issued, "order_date", "Order Date"),
                                use_container_width=True,
                                hide_index=True,
                                height=600,
                            )

                # PDF export (premium summary)
                st.markdown("---")
                st.subheader("PDF Export")

                def _fmt_currency(val: float) -> str:
                    try:
                        return f"${float(val):,.2f}"
                    except Exception:
                        return ""

                def _fmt_pct(val: float) -> str:
                    try:
                        return f"{float(val):,.1f}%"
                    except Exception:
                        return ""

                def _fmt_int(val) -> str:
                    try:
                        if val is None or pd.isna(val):
                            return "0"
                        return f"{int(val):,}"
                    except Exception:
                        return "0"

                def _build_returns_pdf() -> bytes:
                    if not _HAS_REPORTLAB:
                        raise RuntimeError("PDF export requires 'reportlab' package. Please install it.")
                    from reportlab.lib import colors
                    from reportlab.lib.pagesizes import letter, landscape
                    from matplotlib.ticker import FuncFormatter
                    import sys
                    if True:
                        buffer = BytesIO()
                        page_width = landscape(letter)[0] - 36 - 36
                        doc = SimpleDocTemplate(
                            buffer,
                            pagesize=landscape(letter),
                            leftMargin=36,
                            rightMargin=36,
                            topMargin=36,
                            bottomMargin=36,
                        )
                        styles = getSampleStyleSheet()
                        title_style = styles["Heading1"]
                        subtitle_style = styles["Normal"]
                        section_style = styles["Heading2"]

                        story: list = []
                        story.append(Paragraph("Returns Report", title_style))
                        story.append(Paragraph(
                            f"Date Range: {returns_start_date.isoformat()} to {returns_end_date.isoformat()}",
                            subtitle_style,
                        ))
                        story.append(Paragraph(
                            f"Cost Centers: {', '.join(returns_cost_centers) if returns_cost_centers else 'All'}", subtitle_style
                        ))
                        story.append(Paragraph(
                            f"Suppliers: {', '.join([str(s) for s in returns_suppliers]) if returns_suppliers else 'All'}", subtitle_style
                        ))
                        story.append(Spacer(1, 12))

                    def _monthly_stats(df_returns: pd.DataFrame) -> pd.DataFrame:
                        df_m = df_returns.copy()
                        df_m["restock_fee"] = pd.to_numeric(df_m.get("restock_fee"), errors="coerce")
                        df_m["returned_value"] = pd.to_numeric(df_m.get("returned_value"), errors="coerce")
                        df_m["return_month"] = df_m["return_date"].dt.to_period("M").dt.to_timestamp()
                        m = df_m.groupby("return_month", as_index=False).agg(
                            returns_count=("_return_key", "nunique"),
                            restock_fees=("restock_fee", "sum"),
                            returned_value=("returned_value", "sum"),
                        )
                        m["avg_restock_pct"] = m.apply(
                            lambda r: (r["restock_fees"] / r["returned_value"]) if r["returned_value"] > 0 else 0.0,
                            axis=1,
                        )
                        return m.sort_values("return_month")

                    def _metrics(df_returns: pd.DataFrame) -> dict:
                        total_r = int(df_returns.get("_return_key").nunique()) if "_return_key" in df_returns.columns else 0
                        total_f = float(pd.to_numeric(df_returns.get("restock_fee"), errors="coerce").sum())
                        total_v = float(pd.to_numeric(df_returns.get("returned_value"), errors="coerce").sum())
                        avg_pct = (total_f / total_v) if total_v > 0 else 0.0
                        return {
                            "total_returns": total_r,
                            "total_fees": total_f,
                            "total_value": total_v,
                            "avg_pct": avg_pct,
                        }

                    def _plot_line(df_m: pd.DataFrame, y_col: str, title: str, fmt: str | None = None) -> BytesIO:
                        fig, ax = plt.subplots(figsize=(4.0, 2.2), dpi=150)
                        ax.plot(df_m["return_month"], df_m[y_col], marker="o", linewidth=1.5, color="#1f77b4")
                        ax.set_title(title, fontsize=9)
                        ax.grid(True, alpha=0.2)
                        if fmt == "currency":
                            ax.yaxis.set_major_formatter(FuncFormatter(lambda x, _: f"${x:,.0f}"))
                        elif fmt == "percent":
                            ax.yaxis.set_major_formatter(FuncFormatter(lambda x, _: f"{x*100:.0f}%"))
                        fig.autofmt_xdate()
                        buf = BytesIO()
                        fig.savefig(buf, format="png", bbox_inches="tight")
                        plt.close(fig)
                        buf.seek(0)
                        return buf

                    def _plot_pie(series: pd.Series, title: str) -> BytesIO:
                        fig, ax = plt.subplots(figsize=(7.2, 4.2), dpi=150)
                        data = series.value_counts().head(8)
                        if data.empty:
                            data = pd.Series({"Unknown": 1})
                        ax.pie(data.values, labels=data.index, autopct="%1.0f%%", textprops={"fontsize": 9})
                        ax.set_title(title, fontsize=11)
                        buf = BytesIO()
                        fig.savefig(buf, format="png", bbox_inches="tight")
                        plt.close(fig)
                        buf.seek(0)
                        return buf

                    def _add_perspective_section(title: str, df_returns: pd.DataFrame) -> None:
                        story.append(Paragraph(title, section_style))
                        m = _metrics(df_returns)
                        metrics_data = [
                            ["Total Returns", f"{m['total_returns']:,}"],
                            ["Total Restock Fees", _fmt_currency(m["total_fees"])],
                            ["Total Value Returned", _fmt_currency(m["total_value"])],
                            ["Avg Restock %", _fmt_pct(m["avg_pct"] * 100)],
                        ]
                        metrics_tbl = Table(metrics_data, colWidths=[170, 160])
                        metrics_tbl.setStyle(TableStyle([
                            ("GRID", (0,0), (-1,-1), 0.25, colors.black),
                            ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#f0f0f0")),
                            ("FONTNAME", (0,0), (-1,-1), "Helvetica"),
                            ("FONTSIZE", (0,0), (-1,-1), 8),
                        ]))
                        story.append(metrics_tbl)
                        story.append(Spacer(1, 8))

                        monthly_local = _monthly_stats(df_returns)
                        story.append(Paragraph("Monthly Trends", styles["Heading3"]))
                        monthly_rows = [[
                            "Month",
                            "Returns",
                            "Returned Value",
                            "Restock Fees",
                            "Avg Restock %",
                        ]]
                        for _, r in monthly_local.iterrows():
                            monthly_rows.append([
                                r.get("return_month").strftime("%Y-%m"),
                                _fmt_int(r.get("returns_count", 0)),
                                _fmt_currency(r.get("returned_value", 0)),
                                _fmt_currency(r.get("restock_fees", 0)),
                                _fmt_pct((r.get("avg_restock_pct", 0)) * 100),
                            ])
                        m_row_heights = [16] + [14] * (len(monthly_rows) - 1)
                        m_tbl = Table(monthly_rows, repeatRows=1, rowHeights=m_row_heights)
                        m_tbl.setStyle(TableStyle([
                            ("GRID", (0,0), (-1,-1), 0.25, colors.black),
                            ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#f0f0f0")),
                            ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
                            ("FONTSIZE", (0,0), (-1,-1), 8),
                        ]))
                        m_tbl.splitByRow = 0
                        story.append(m_tbl)
                        story.append(Spacer(1, 10))

                        img1 = _plot_line(monthly_local, "returns_count", "Returns (Monthly)")
                        img2 = _plot_line(monthly_local, "restock_fees", "Restock Fees (Monthly)", fmt="currency")
                        img3 = _plot_line(monthly_local, "returned_value", "Returned Value (Monthly)", fmt="currency")
                        img4 = _plot_line(monthly_local, "avg_restock_pct", "Avg Restock % (Monthly)", fmt="percent")
                        reason_col_pdf = "credit_type_desc" if "credit_type_desc" in df_returns.columns else "credit_type_code"
                        reason_series = df_returns.get(reason_col_pdf).fillna("Unknown").astype(str).str.strip()
                        img5 = _plot_pie(reason_series, "Return Reasons")

                        pie_w = page_width * 0.9
                        pie_h = 300
                        charts_rows = [
                            [Image(img1, width=240, height=135), Image(img2, width=240, height=135)],
                            [Image(img3, width=240, height=135), Image(img4, width=240, height=135)],
                            [Image(img5, width=pie_w, height=pie_h), ""],
                        ]
                        charts = Table(
                            charts_rows,
                            colWidths=[page_width * 0.5, page_width * 0.5],
                            rowHeights=[140, 140, pie_h + 10],
                        )
                        charts.setStyle(TableStyle([
                            ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
                            ("SPAN", (0,2), (1,2)),
                        ]))
                        story.append(charts)
                        story.append(Spacer(1, 16))

                    # Build perspectives
                    returns_with_fee = returns_df_all[returns_df_all["restock_fee"].fillna(0) > 0].copy()
                    returns_no_fee = returns_df_all[returns_df_all["restock_fee"].fillna(0) <= 0].copy()

                    _add_perspective_section("All Returns", returns_df_all)
                    story.append(PageBreak())
                    _add_perspective_section("Returns Without Fee", returns_no_fee)
                    story.append(PageBreak())
                    _add_perspective_section("Returns With Fee", returns_with_fee)
                    story.append(PageBreak())

                    def _add_table(title: str, df: pd.DataFrame) -> None:
                        print(f"DEBUG: _add_table called with title='{title}', df shape={df.shape}, dtypes={dict(df.dtypes)}", file=sys.stderr)
                        story.append(Paragraph(title, section_style))
                        story.append(Paragraph("Format per column: With Fee / No Fee / All", styles["Normal"]))
                        story.append(Spacer(1, 6))
                        cols = [
                            "Label",
                            "Total Orders",
                            "Returns",
                            "Return Rate %",
                            "Returned Value",
                            "Restock Fees",
                            "Avg Restock Fee",
                            "Avg Restock %",
                        ]
                        rows = [cols]
                        def _clip_text(val: str, max_len: int = 26) -> str:
                            txt = "" if val is None else str(val)
                            return txt if len(txt) <= max_len else txt[:max_len].rstrip()

                        for _, r in df.iterrows():
                            rows.append([
                                _clip_text(r.get("label", "")),
                                _fmt_int(r.get("total_orders", 0)),
                                f"{_fmt_int(r.get('returns_with', 0))} / {_fmt_int(r.get('returns_no', 0))} / {_fmt_int(r.get('returns_all', 0))}",
                                f"{_fmt_pct(r.get('return_rate_with', 0) * 100)} / {_fmt_pct(r.get('return_rate_no', 0) * 100)} / {_fmt_pct(r.get('return_rate_all', 0) * 100)}",
                                f"{_fmt_currency(r.get('returned_value_with', 0))} / {_fmt_currency(r.get('returned_value_no', 0))} / {_fmt_currency(r.get('returned_value_all', 0))}",
                                f"{_fmt_currency(r.get('restock_fees_with', 0))} / {_fmt_currency(r.get('restock_fees_no', 0))} / {_fmt_currency(r.get('restock_fees_all', 0))}",
                                f"{_fmt_currency(r.get('avg_restock_fee_with', 0))} / {_fmt_currency(r.get('avg_restock_fee_no', 0))} / {_fmt_currency(r.get('avg_restock_fee_all', 0))}",
                                f"{_fmt_pct(r.get('avg_restock_pct_with', 0) * 100)} / {_fmt_pct(r.get('avg_restock_pct_no', 0) * 100)} / {_fmt_pct(r.get('avg_restock_pct_all', 0) * 100)}",
                            ])

                        # Normalize cell values to strings
                        for i in range(len(rows)):
                            rows[i] = ["" if v is None else str(v) for v in rows[i]]

                        # Fixed proportional column widths (stable and avoids None comparisons)
                        col_widths = [
                            page_width * 0.18,  # Label
                            page_width * 0.08,  # Total Orders
                            page_width * 0.11,  # Returns
                            page_width * 0.11,  # Return Rate %
                            page_width * 0.17,  # Returned Value
                            page_width * 0.15,  # Restock Fees
                            page_width * 0.10,  # Avg Restock Fee
                            page_width * 0.10,  # Avg Restock %
                        ]

                        def _clip_cell(val: str, col_w: float) -> str:
                            safe_w = col_w if col_w is not None else 0.0
                            max_len = max(4, int(safe_w / 4.1) + 2)
                            txt = "" if val is None else str(val)
                            return txt if len(txt) <= max_len else txt[:max_len].rstrip()

                        # Clamp all cells to their column width to prevent overflow
                        for i in range(1, len(rows)):
                            for j in range(len(rows[i])):
                                if j == 0:
                                    base_w = col_widths[j] if col_widths[j] is not None else 0.0
                                    adj_w = max(0.0, base_w - 8.2)  # ~2 chars
                                    rows[i][j] = _clip_cell(rows[i][j], adj_w)
                                else:
                                    rows[i][j] = _clip_cell(rows[i][j], col_widths[j])
                        row_heights = [18] + [14] * (len(rows) - 1)
                        tbl = Table(rows, repeatRows=1, colWidths=col_widths, rowHeights=row_heights)
                        tbl.setStyle(TableStyle([
                            ("GRID", (0,0), (-1,-1), 0.25, colors.black),
                            ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#f0f0f0")),
                            ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
                            ("FONTSIZE", (0,0), (-1,-1), 7),
                            ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
                        ]))
                        tbl.splitByRow = 1
                        story.append(tbl)
                        story.append(Spacer(1, 12))

                    def _group_returns_pdf(df_returns: pd.DataFrame, group_col: str | list[str]) -> pd.DataFrame:
                        print(f"DEBUG: _group_returns_pdf called with group_col={group_col}, df_returns shape={df_returns.shape}", file=sys.stderr)
                        orders_totals = orders.groupby(group_col, dropna=False)["_order_key"].nunique().rename("total_orders")
                        returns_totals = df_returns.groupby(group_col, dropna=False).agg(
                            returns_count=("_return_key", "nunique"),
                            restock_fees=("restock_fee", "sum"),
                            returned_value=("returned_value", "sum"),
                        )
                        combined = returns_totals.join(orders_totals, how="left").fillna({"total_orders": 0})
                        combined["return_rate"] = combined.apply(
                            lambda r: (r["returns_count"] / r["total_orders"]) if r["total_orders"] > 0 else 0.0,
                            axis=1,
                        )
                        combined["avg_restock_fee"] = combined.apply(
                            lambda r: (r["restock_fees"] / r["returns_count"]) if r["returns_count"] > 0 else 0.0,
                            axis=1,
                        )
                        combined["avg_restock_pct"] = combined.apply(
                            lambda r: (r["restock_fees"] / r["returned_value"]) if r["returned_value"] > 0 else 0.0,
                            axis=1,
                        )
                        combined = combined.reset_index()
                        return combined

                    def _merge_perspectives(group_col: str | list[str], label_series: pd.Series | None = None) -> pd.DataFrame:
                        print(f"DEBUG: _merge_perspectives called with group_col={group_col}", file=sys.stderr)
                        all_df = _group_returns_pdf(returns_df_all, group_col)
                        with_df = _group_returns_pdf(returns_with_fee, group_col)
                        no_df = _group_returns_pdf(returns_no_fee, group_col)

                        def _rename(df: pd.DataFrame, suffix: str) -> pd.DataFrame:
                            return df.rename(columns={
                                "returns_count": f"returns_{suffix}",
                                "return_rate": f"return_rate_{suffix}",
                                "returned_value": f"returned_value_{suffix}",
                                "restock_fees": f"restock_fees_{suffix}",
                                "avg_restock_fee": f"avg_restock_fee_{suffix}",
                                "avg_restock_pct": f"avg_restock_pct_{suffix}",
                            })

                        all_df = _rename(all_df, "all")
                        with_df = _rename(with_df, "with")
                        no_df = _rename(no_df, "no")

                        merged = all_df.merge(with_df, on=group_col, how="left").merge(no_df, on=group_col, how="left")
                        # Coerce numeric columns to avoid NoneType comparisons
                        numeric_cols = [
                            "total_orders",
                            "returns_all", "returns_with", "returns_no",
                            "return_rate_all", "return_rate_with", "return_rate_no",
                            "returned_value_all", "returned_value_with", "returned_value_no",
                            "restock_fees_all", "restock_fees_with", "restock_fees_no",
                            "avg_restock_fee_all", "avg_restock_fee_with", "avg_restock_fee_no",
                            "avg_restock_pct_all", "avg_restock_pct_with", "avg_restock_pct_no",
                        ]
                        for col in numeric_cols:
                            if col in merged.columns:
                                merged[col] = pd.to_numeric(merged[col], errors="coerce").fillna(0.0)
                        if label_series is not None:
                            merged["label"] = label_series
                        return merged

                    account_pdf = _merge_perspectives(["account_number", "bank_name2"]) 
                    account_pdf["label"] = account_pdf["bank_name2"].fillna("").astype(str).str.strip().where(
                        account_pdf["bank_name2"].fillna("").astype(str).str.strip().str.len() > 0,
                        "Unknown"
                    ) + " - " + account_pdf["account_number"].fillna("").astype(str).str.strip()
                    account_pdf = account_pdf.sort_values("return_rate_all", ascending=False)

                    supplier_pdf = _merge_perspectives(supplier_label_col)
                    supplier_pdf["label"] = supplier_pdf[supplier_label_col].astype(str)
                    supplier_pdf = supplier_pdf.sort_values("return_rate_all", ascending=False)

                    cost_center_cols = ["cost_center", "cost_center_desc"] if "cost_center_desc" in returns_df_all.columns else "cost_center"
                    cost_center_pdf = _merge_perspectives(cost_center_cols)
                    if "cost_center_desc" in cost_center_pdf.columns:
                        cost_center_pdf["label"] = cost_center_pdf["cost_center"].astype(str) + " - " + cost_center_pdf["cost_center_desc"].astype(str)
                    else:
                        cost_center_pdf["label"] = cost_center_pdf["cost_center"].astype(str)
                    cost_center_pdf = cost_center_pdf.sort_values("return_rate_all", ascending=False)

                    _add_table("Cost Centers (With Fee/No Fee/All)", cost_center_pdf)
                    _add_table("Suppliers (With Fee/No Fee/All)", supplier_pdf)
                    _add_table("Accounts (With Fee/No Fee/All)", account_pdf)

                    doc.build(story)
                    buffer.seek(0)
                    print(f"DEBUG: PDF built successfully, size={len(buffer.getvalue())} bytes", file=sys.stderr)
                    return buffer.getvalue()

                try:
                    pdf_bytes = _build_returns_pdf()
                    st.download_button(
                        label="📥 Download Returns PDF",
                        data=pdf_bytes,
                        file_name=f"returns_report_{returns_start_date.isoformat()}_{returns_end_date.isoformat()}.pdf",
                        mime="application/pdf",
                        key="returns_pdf_download",
                    )
                except Exception as e:
                    import traceback
                    import sys
                    print(f"\n{'='*80}", file=sys.stderr)
                    print(f"ERROR BUILDING RETURNS PDF: {type(e).__name__}", file=sys.stderr)
                    print(f"Message: {str(e)}", file=sys.stderr)
                    traceback.print_exc(file=sys.stderr)
                    print(f"{'='*80}\n", file=sys.stderr)
                    st.warning(f"Unable to build PDF: {e}")

    # TAB 13: CCA
    # ========================================================
    with tabs[12]:
        st.header("CCA")
        st.markdown(
            "*Customer group performance for Flooring America (ACA), ProSource (ACP), and Carpet One (AC1), "
            "classified from BILL_CD with one account per hierarchy (AC1 > ACP > ACA).*"
        )

        # Persistent CCA filters
        try:
            _cca_pref_start = pd.to_datetime(user_prefs.get("cca_start_date"), errors="coerce").date() if user_prefs.get("cca_start_date") else (date.today() - timedelta(days=365))
        except Exception:
            _cca_pref_start = date.today() - timedelta(days=365)
        try:
            _cca_pref_end = pd.to_datetime(user_prefs.get("cca_end_date"), errors="coerce").date() if user_prefs.get("cca_end_date") else date.today()
        except Exception:
            _cca_pref_end = date.today()
        _cca_pref_cc = user_prefs.get("cca_cost_centers", selected_cost_centers if selected_cost_centers else config.default_cost_centers)

        cca_col1, cca_col2, cca_col3 = st.columns([1, 1, 2])
        with cca_col1:
            cca_start_date = st.date_input("Start date", value=_cca_pref_start, key="cca_start_date_input")
        with cca_col2:
            cca_end_date = st.date_input("End date", value=_cca_pref_end, key="cca_end_date_input")
        with cca_col3:
            cca_cost_centers = st.multiselect(
                "Cost centers",
                options=cost_center_options,
                default=_ensure_selection(cost_center_options, _cca_pref_cc),
                key="cca_cost_centers_input",
                help="Filters CCA sales to selected cost centers.",
            )

        # Persist CCA filter preferences
        try:
            _needs_save = (
                str(cca_start_date) != str(user_prefs.get("cca_start_date", ""))
                or str(cca_end_date) != str(user_prefs.get("cca_end_date", ""))
                or list(cca_cost_centers) != list(user_prefs.get("cca_cost_centers", []))
            )
            if _needs_save:
                _prefs = _load_prefs()
                _prefs["cca_start_date"] = cca_start_date.isoformat()
                _prefs["cca_end_date"] = cca_end_date.isoformat()
                _prefs["cca_cost_centers"] = list(cca_cost_centers)
                _save_prefs(_prefs)
        except Exception:
            pass

        @st.cache_data(ttl=config.cache_ttl_seconds, show_spinner=False)
        def _load_cca_sources(cost_centers_tuple: tuple[str, ...], start: date, end: date):
            # Use dedicated CCA sales source to avoid inventory-only filtering side effects.
            orders_df = loaders.load_cca_sales_orders(
                config.connection_string,
                start_date=start,
                end_date=end,
            )
            bill_df = loaders.load_cca_account_groups(config.connection_string)
            return orders_df, bill_df

        with st.spinner("Loading CCA data..."):
            cca_orders_raw, cca_bill_raw = _load_cca_sources(tuple(cca_cost_centers), cca_start_date, cca_end_date)

        if cca_bill_raw.empty:
            st.warning("No CCA account definitions were found in BILL_CD for BCCAT='MP' and BCCODE in ACA/ACP/AC1.")
        elif cca_orders_raw.empty:
            st.warning("No sales orders were found for the selected CCA date range and cost centers.")
        else:
            group_name_map = {
                "ACA": "Flooring America",
                "ACP": "ProSource",
                "AC1": "Carpet One",
            }
            group_priority = {"AC1": 1, "ACP": 2, "ACA": 3}

            def _norm_acct(series: pd.Series) -> pd.Series:
                s = series.fillna("").astype(str).str.strip()
                n = pd.to_numeric(s, errors="coerce")
                out = s.copy()
                mask = n.notna()
                out.loc[mask] = n.loc[mask].astype("Int64").astype(str)
                return out

            cca_bill = cca_bill_raw.copy()
            cca_bill["group_code"] = cca_bill.get("group_code").fillna("").astype(str).str.strip().str.upper()
            cca_bill["account_number_norm"] = _norm_acct(cca_bill.get("account_number", pd.Series(dtype=str)))
            cca_bill = cca_bill[cca_bill["group_code"].isin(list(group_name_map.keys()))]
            cca_bill = cca_bill[cca_bill["account_number_norm"] != ""]
            cca_bill["priority"] = cca_bill["group_code"].map(group_priority).fillna(999)
            cca_bill.sort_values(["account_number_norm", "priority"], inplace=True)
            cca_members = cca_bill.drop_duplicates(subset=["account_number_norm"], keep="first").copy()
            cca_members["group_name"] = cca_members["group_code"].map(group_name_map)

            cca_orders = cca_orders_raw.copy()
            cca_orders["order_entry_date"] = pd.to_datetime(cca_orders.get("order_entry_date"), errors="coerce")
            cca_orders["order_ship_date"] = pd.to_datetime(cca_orders.get("order_ship_date"), errors="coerce")
            cca_orders["invoice_ship_date"] = pd.to_datetime(cca_orders.get("invoice_ship_date"), errors="coerce")

            # Build effective cost center and apply selected CC filter in-memory.
            _cc_item = cca_orders.get("cost_center", pd.Series(index=cca_orders.index, dtype=object)).fillna("").astype(str).str.strip().str.zfill(3)
            _cc_desc = cca_orders.get("cost_center_desc", pd.Series(index=cca_orders.index, dtype=object)).fillna("").astype(str)
            _cc_from_desc = _cc_desc.str.extract(r"(\d{3})", expand=False).fillna("").astype(str).str.strip().str.zfill(3)
            cca_orders["effective_cost_center"] = np.where(_cc_from_desc.str.len() == 3, _cc_from_desc, _cc_item)
            if cca_cost_centers:
                _cc_sel = {str(c).strip().zfill(3) for c in cca_cost_centers}
                _cc_mask = cca_orders["effective_cost_center"].astype(str).isin(_cc_sel) | _cc_item.astype(str).isin(_cc_sel)
                cca_orders = cca_orders.loc[_cc_mask].copy()

            # Use a sales-centric date: invoiced ship date when invoiced, else order ship date, else entry date.
            _inv_num = pd.to_numeric(cca_orders.get("invoice_number"), errors="coerce").fillna(0)
            _sales_date = cca_orders.get("order_ship_date")
            _sales_date = _sales_date.combine_first(cca_orders.get("order_entry_date"))
            _inv_ship = cca_orders.get("invoice_ship_date")
            _sales_date = _sales_date.where(_inv_num <= 0, _inv_ship.combine_first(_sales_date))
            cca_orders["sales_date"] = pd.to_datetime(_sales_date, errors="coerce")

            # Normalize order-line key and deduplicate (same strategy as the metrics path).
            def _norm_id_local(series: pd.Series) -> pd.Series:
                s = series.fillna("").astype(str).str.strip()
                n = pd.to_numeric(s, errors="coerce")
                out = s.copy()
                mask = n.notna()
                out.loc[mask] = n.loc[mask].astype("Int64").astype(str)
                return out

            cca_orders["order_number_norm"] = _norm_id_local(cca_orders.get("order_number", pd.Series(dtype=str)))
            cca_orders["line_number_norm"] = _norm_id_local(cca_orders.get("line_number", pd.Series(dtype=str)))
            cca_orders["order_line_id"] = cca_orders["order_number_norm"] + "-" + cca_orders["line_number_norm"]
            cca_orders["invoice_number_norm"] = _norm_id_local(cca_orders.get("invoice_number", pd.Series(dtype=str)))

            # Stage diagnostics to explain differences vs external totals.
            _raw_rows = len(cca_orders)
            cca_orders["sales_amount"] = pd.to_numeric(cca_orders.get("extended_price_no_funds"), errors="coerce").fillna(0.0)
            _raw_total = float(cca_orders["sales_amount"].sum())

            cca_orders = cca_orders[
                cca_orders["sales_date"].between(pd.Timestamp(cca_start_date), pd.Timestamp(cca_end_date), inclusive="both")
            ].copy()
            _dated_rows = len(cca_orders)
            _dated_total = float(cca_orders["sales_amount"].sum())

            cca_orders.sort_values(["sales_date", "order_entry_date"], inplace=True)
            # Deduplicate only truly identical sales rows; preserve legitimate multi-invoice
            # activity for the same order/line.
            _dedup_subset = [
                c for c in [
                    "order_number_norm",
                    "line_number_norm",
                    "invoice_number_norm",
                    "sku",
                    "sales_amount",
                    "sales_date",
                ] if c in cca_orders.columns
            ]
            if _dedup_subset:
                cca_orders = cca_orders.drop_duplicates(subset=_dedup_subset, keep="first")
            _dedup_rows = len(cca_orders)
            _dedup_total = float(cca_orders["sales_amount"].sum())

            # Keep signed sales values (including credits) so totals reconcile to source systems.
            _pos_rows = int((cca_orders["sales_amount"] > 0).sum())
            _pos_total = float(cca_orders.loc[cca_orders["sales_amount"] > 0, "sales_amount"].sum())

            cca_orders["account_number_norm"] = _norm_acct(cca_orders.get("account_number", pd.Series(dtype=str)))
            cca_orders["salesperson"] = cca_orders.get("salesperson_desc").fillna("Unknown").astype(str).str.strip().replace("", "Unknown")
            cca_orders["bank_name2"] = cca_orders.get("bank_name2").fillna("").astype(str).str.strip()
            cca_orders["order_month"] = cca_orders["sales_date"].dt.to_period("M").dt.to_timestamp()

            cca_sales = cca_orders.merge(
                cca_members[["account_number_norm", "group_code", "group_name"]],
                on="account_number_norm",
                how="inner",
            )

            if cca_sales.empty:
                st.info("No matched CCA sales were found for the selected filters.")
            else:
                with st.expander("CCA diagnostics (why totals can differ)", expanded=False):
                    d1, d2, d3, d4 = st.columns(4)
                    d1.metric("Raw Rows / Sales", f"{_raw_rows:,} / ${_raw_total:,.0f}")
                    d2.metric("Date-Filtered", f"{_dated_rows:,} / ${_dated_total:,.0f}")
                    d3.metric("After Dedup", f"{_dedup_rows:,} / ${_dedup_total:,.0f}")
                    d4.metric("Positive Portion", f"{_pos_rows:,} / ${_pos_total:,.0f}")

                    _acct_probe = st.text_input(
                        "Probe account #",
                        value="51149",
                        key="cca_probe_account",
                        help="Enter account with or without leading zeros (e.g., 051149 or 51149).",
                    )
                    _acct_probe_norm = _norm_acct(pd.Series([_acct_probe])).iloc[0] if str(_acct_probe).strip() else ""
                    if _acct_probe_norm:
                        _probe_df = cca_orders.copy()
                        _probe_df = _probe_df[_probe_df.get("account_number_norm") == _acct_probe_norm]
                        _probe_total = float(_probe_df.get("sales_amount", pd.Series(dtype=float)).sum()) if not _probe_df.empty else 0.0
                        st.caption(f"Current CCA total for account {_acct_probe_norm}: ${_probe_total:,.0f}")
                        if not _probe_df.empty:
                            _probe_detail = (
                                _probe_df.groupby("order_month", as_index=False)["sales_amount"].sum()
                                .sort_values("order_month")
                            )
                            st.dataframe(
                                _probe_detail.rename(columns={"order_month": "Month", "sales_amount": "Sales"}).style.format({"Sales": "${:,.0f}"}),
                                use_container_width=True,
                                hide_index=True,
                            )

                cca_sales["account_label"] = (
                    cca_sales["bank_name2"].where(cca_sales["bank_name2"].str.len() > 0, "Unknown")
                    + " - "
                    + cca_sales["account_number_norm"]
                )

                # Top-level KPI cards
                total_sales = float(cca_sales["sales_amount"].sum())
                active_accounts = int(cca_sales["account_number_norm"].nunique())
                active_reps = int(cca_sales["salesperson"].nunique())
                avg_monthly_sales = float(
                    cca_sales.groupby("order_month")["sales_amount"].sum().mean()
                ) if not cca_sales.empty else 0.0

                monthly_all = cca_sales.groupby("order_month", as_index=False)["sales_amount"].sum().sort_values("order_month")
                if len(monthly_all) >= 2:
                    latest = float(monthly_all.iloc[-1]["sales_amount"])
                    prev = float(monthly_all.iloc[-2]["sales_amount"])
                    mom_pct = ((latest - prev) / prev) if prev != 0 else 0.0
                else:
                    mom_pct = 0.0

                k1, k2, k3, k4, k5 = st.columns(5)
                k1.metric("Total Sales", f"${total_sales:,.0f}")
                k2.metric("Active Accounts", f"{active_accounts:,}")
                k3.metric("Active Reps", f"{active_reps:,}")
                k4.metric("Avg Monthly Sales", f"${avg_monthly_sales:,.0f}")
                k5.metric("Latest MoM", f"{mom_pct:.1%}")

                # Group summary
                group_summary = (
                    cca_sales.groupby(["group_name", "group_code"], as_index=False)
                    .agg(
                        sales=("sales_amount", "sum"),
                        accounts=("account_number_norm", "nunique"),
                        reps=("salesperson", "nunique"),
                    )
                    .sort_values("sales", ascending=False)
                )
                if total_sales > 0:
                    group_summary["sales_share"] = group_summary["sales"] / total_sales
                else:
                    group_summary["sales_share"] = 0.0
                st.markdown("#### CCA Group Snapshot")
                st.dataframe(
                    group_summary.rename(
                        columns={
                            "group_name": "Group",
                            "group_code": "Code",
                            "sales": "Sales",
                            "accounts": "Accounts",
                            "reps": "Reps",
                            "sales_share": "Sales Share",
                        }
                    ).style.format({"Sales": "${:,.0f}", "Sales Share": "{:.1%}"}),
                    use_container_width=True,
                    hide_index=True,
                )

                # Trends: all groups together + rolling average
                monthly_group = (
                    cca_sales.groupby(["order_month", "group_name"], as_index=False)["sales_amount"].sum()
                    .sort_values("order_month")
                )
                st.markdown("#### Sales Trends")
                tr1, tr2 = st.columns(2)
                with tr1:
                    fig_groups = px.line(
                        monthly_group,
                        x="order_month",
                        y="sales_amount",
                        color="group_name",
                        markers=True,
                        title="Monthly Sales by CCA Group",
                    )
                    fig_groups.update_layout(legend_title_text="Group")
                    st.plotly_chart(fig_groups, use_container_width=True)

                with tr2:
                    rolling = (
                        monthly_group.pivot_table(index="order_month", columns="group_name", values="sales_amount", aggfunc="sum")
                        .fillna(0.0)
                        .sort_index()
                    )
                    rolling_3m = rolling.rolling(3, min_periods=1).mean().reset_index().melt(
                        id_vars="order_month", var_name="group_name", value_name="rolling_sales"
                    )
                    fig_roll = px.line(
                        rolling_3m,
                        x="order_month",
                        y="rolling_sales",
                        color="group_name",
                        markers=True,
                        title="3-Month Rolling Sales by Group",
                    )
                    fig_roll.update_layout(legend_title_text="Group")
                    st.plotly_chart(fig_roll, use_container_width=True)

                # Separate per-group trend charts
                st.markdown("#### Group-Specific Trends")
                gcol_a, gcol_b, gcol_c = st.columns(3)
                _group_plot_order = ["Carpet One", "ProSource", "Flooring America"]
                _group_cols = [gcol_a, gcol_b, gcol_c]
                for _idx, _gname in enumerate(_group_plot_order):
                    with _group_cols[_idx]:
                        _gdf = monthly_group[monthly_group["group_name"] == _gname].copy()
                        if _gdf.empty:
                            st.info(f"No data for {_gname} in selected range.")
                        else:
                            _gdf = _gdf.sort_values("order_month")
                            _gdf["rolling_3m"] = _gdf["sales_amount"].rolling(3, min_periods=1).mean()
                            fig_g = px.line(
                                _gdf,
                                x="order_month",
                                y=["sales_amount", "rolling_3m"],
                                markers=True,
                                title=f"{_gname}: Monthly vs 3M Rolling",
                            )
                            fig_g.update_layout(legend_title_text="Series")
                            st.plotly_chart(fig_g, use_container_width=True)

                # Sales by rep across CCA groups (columns = groups)
                st.markdown("#### Sales by Salesperson and Group")
                rep_group = (
                    cca_sales.pivot_table(
                        index="salesperson",
                        columns="group_name",
                        values="sales_amount",
                        aggfunc="sum",
                        fill_value=0.0,
                    )
                    .reindex(columns=_group_plot_order, fill_value=0.0)
                )
                rep_group["Total"] = rep_group.sum(axis=1)
                rep_group = rep_group.sort_values("Total", ascending=False)
                st.dataframe(
                    rep_group.reset_index().style.format("${:,.0f}", subset=[c for c in rep_group.columns if c != "salesperson"]).background_gradient(
                        cmap="Blues", subset=[c for c in rep_group.columns if c != "salesperson"]
                    ),
                    use_container_width=True,
                    hide_index=True,
                )

                # Sales by rep and month with totals (row + bottom total)
                st.markdown("#### Monthly Sales by Salesperson")
                rep_month = cca_sales.pivot_table(
                    index="salesperson",
                    columns="order_month",
                    values="sales_amount",
                    aggfunc="sum",
                    fill_value=0.0,
                ).sort_index(axis=1)

                if not rep_month.empty:
                    rep_month["Total"] = rep_month.sum(axis=1)
                    rep_month = rep_month.sort_values("Total", ascending=False)

                    _month_cols = [c for c in rep_month.columns if isinstance(c, pd.Timestamp)]
                    if len(_month_cols) >= 2:
                        rep_month["MoM Delta"] = rep_month[_month_cols[-1]] - rep_month[_month_cols[-2]]
                    else:
                        rep_month["MoM Delta"] = 0.0

                    total_row = rep_month.sum(axis=0)
                    total_row.name = "TOTAL"
                    rep_month = pd.concat([rep_month, total_row.to_frame().T], axis=0)

                    rep_month_display = rep_month.copy()
                    rep_month_display.columns = [c.strftime("%Y-%m") if isinstance(c, pd.Timestamp) else c for c in rep_month_display.columns]
                    rep_month_display = rep_month_display.reset_index().rename(columns={"index": "Salesperson"})

                    def _mom_color(v):
                        try:
                            x = float(v)
                            if x > 0:
                                return "background-color: #e8f5e9; color: #1b5e20;"
                            if x < 0:
                                return "background-color: #ffebee; color: #b71c1c;"
                            return "background-color: #f5f5f5;"
                        except Exception:
                            return ""

                    _num_cols_rep = [c for c in rep_month_display.columns if c != "Salesperson"]
                    st.dataframe(
                        rep_month_display.style.format({c: "${:,.0f}" for c in _num_cols_rep}).applymap(_mom_color, subset=["MoM Delta"]),
                        use_container_width=True,
                        hide_index=True,
                    )
                else:
                    st.info("No salesperson-month data available for selected filters.")

                # Accounts by group (tabbed), month-to-month with totals and trend color
                st.markdown("#### Accounts by CCA Group")
                acc_tabs = st.tabs(_group_plot_order)
                for _idx, _gname in enumerate(_group_plot_order):
                    with acc_tabs[_idx]:
                        _gacc = cca_sales[cca_sales["group_name"] == _gname].copy()
                        if _gacc.empty:
                            st.info(f"No accounts found for {_gname} with current filters.")
                            continue

                        acc_month = _gacc.pivot_table(
                            index="account_label",
                            columns="order_month",
                            values="sales_amount",
                            aggfunc="sum",
                            fill_value=0.0,
                        ).sort_index(axis=1)
                        acc_month["Total"] = acc_month.sum(axis=1)

                        _acct_month_cols = [c for c in acc_month.columns if isinstance(c, pd.Timestamp)]
                        if len(_acct_month_cols) >= 2:
                            acc_month["MoM Delta"] = acc_month[_acct_month_cols[-1]] - acc_month[_acct_month_cols[-2]]
                        else:
                            acc_month["MoM Delta"] = 0.0

                        acc_month = acc_month.sort_values("Total", ascending=False)
                        total_row_acc = acc_month.sum(axis=0)
                        total_row_acc.name = "TOTAL"
                        acc_month = pd.concat([acc_month, total_row_acc.to_frame().T], axis=0)

                        acc_disp = acc_month.copy()
                        acc_disp.columns = [c.strftime("%Y-%m") if isinstance(c, pd.Timestamp) else c for c in acc_disp.columns]
                        acc_disp = acc_disp.reset_index().rename(columns={"index": "Account"})

                        def _trend_color(v):
                            try:
                                x = float(v)
                                if x > 0:
                                    return "background-color: #e8f5e9; color: #1b5e20;"
                                if x < 0:
                                    return "background-color: #ffebee; color: #b71c1c;"
                                return "background-color: #f5f5f5;"
                            except Exception:
                                return ""

                        _acct_num_cols = [c for c in acc_disp.columns if c != "Account"]
                        st.dataframe(
                            acc_disp.style.format({c: "${:,.0f}" for c in _acct_num_cols}).applymap(_trend_color, subset=["MoM Delta"]),
                            use_container_width=True,
                            hide_index=True,
                        )

                        # Small account trend chart for this group
                        _acc_monthly = _gacc.groupby("order_month", as_index=False)["sales_amount"].sum().sort_values("order_month")
                        _acc_monthly["rolling_3m"] = _acc_monthly["sales_amount"].rolling(3, min_periods=1).mean()
                        fig_acc = px.line(
                            _acc_monthly,
                            x="order_month",
                            y=["sales_amount", "rolling_3m"],
                            markers=True,
                            title=f"{_gname} Total Sales Trend",
                        )
                        st.plotly_chart(fig_acc, use_container_width=True)


# ============================================================================
# SALES REP PERFORMANCE PDF GENERATION
# ============================================================================

def _calculate_driver_insights(rep_metrics: dict, rolling_metrics: dict, 
                               all_reps_metrics: list, df: pd.DataFrame,
                               rep_name: str, peer_group: list) -> dict:
    """Calculate detailed driver insights for PDF explanations.
    
    Returns dict with insights for each component.
    """
    insights = {}
    
    # Revenue Momentum Drivers
    weekly_revenue = rolling_metrics.get("weekly_revenue_list", [])
    if len(weekly_revenue) >= 2:
        # Find best and worst weeks
        sorted_weeks = sorted(enumerate(weekly_revenue), key=lambda x: x[1], reverse=True)
        best_weeks = sorted_weeks[:2] if len(sorted_weeks) >= 2 else sorted_weeks
        worst_weeks = sorted_weeks[-2:] if len(sorted_weeks) >= 2 else []
        
        # Growth rate
        growth_pct = rolling_metrics.get("revenue_growth_pct", 0)
        trend_slope = rolling_metrics.get("weekly_trend_slope", 0)
        
        insights["revenue_momentum"] = {
            "weekly_revenue": weekly_revenue,
            "growth_pct": growth_pct,
            "trend_direction": "Up" if trend_slope > 0 else "Down" if trend_slope < 0 else "Flat",
            "best_weeks": [(i+1, v) for i, v in best_weeks],
            "worst_weeks": [(i+1, v) for i, v in worst_weeks],
        }
    else:
        insights["revenue_momentum"] = {"weekly_revenue": [], "growth_pct": 0, "trend_direction": "Unknown"}
    
    # Margin Discipline Drivers
    avg_margin = rep_metrics.get("avg_margin_pct", 0)
    margin_std = rep_metrics.get("margin_std_dev", 0)
    
    # Find lowest margin customers/categories
    rep_df = df[df["salesperson"] == rep_name].copy()
    if not rep_df.empty:
        cust_margin = rep_df.groupby("account_number").apply(
            lambda g: (g["line_gross_profit"].sum() / g["line_revenue"].sum() * 100) if g["line_revenue"].sum() > 0 else 0
        ).sort_values()
        
        cat_margin = {}
        for cat in rep_metrics.get("margin_by_category", {}).keys():
            cat_margin[cat] = rep_metrics["margin_by_category"][cat]
        
        # Peer margin comparison (category-aware)
        peer_metrics_list = [m for m in all_reps_metrics if m.get("salesperson") in peer_group]
        if peer_metrics_list and len(peer_metrics_list) >= 3:
            peer_margins = [m.get("avg_margin_pct", 0) for m in peer_metrics_list if m.get("avg_margin_pct", 0) > 0]
            peer_avg_margin = sum(peer_margins) / len(peer_margins) if peer_margins else 0
            
            peer_stds = [m.get("margin_std_dev", 0) for m in peer_metrics_list if m.get("margin_std_dev", 0) > 0]
            peer_avg_std = sum(peer_stds) / len(peer_stds) if peer_stds else 10
        else:
            peer_avg_margin = 0
            peer_avg_std = 10
        
        insights["margin_discipline"] = {
            "avg_margin": avg_margin,
            "std_dev": margin_std,
            "consistency": "Good" if margin_std < peer_avg_std else "Watch" if margin_std < peer_avg_std * 1.5 else "High Risk",
            "lowest_margin_customers": list(cust_margin.head(3).items()),
            "category_margins": cat_margin,
            "peer_avg_margin": peer_avg_margin,
            "peer_avg_std": peer_avg_std,
            "vs_peer_margin": avg_margin - peer_avg_margin
        }
    else:
        insights["margin_discipline"] = {"avg_margin": avg_margin, "std_dev": margin_std, "consistency": "Unknown"}
    
    # Customer Health Drivers
    active_customers = rep_metrics.get("active_customers", 0)
    top_cust_conc = rep_metrics.get("top_customer_concentration", 0)
    top3_cust_conc = rep_metrics.get("top3_customer_concentration", 0)
    rev_per_cust = rep_metrics.get("revenue_per_customer", 0)
    
    # Peer comparison
    peer_metrics = [m for m in all_reps_metrics if m.get("salesperson") in peer_group]
    if peer_metrics:
        peer_active_custs = [m.get("active_customers", 0) for m in peer_metrics]
        avg_peer_customers = sum(peer_active_custs) / len(peer_active_custs) if peer_active_custs else 0
    else:
        avg_peer_customers = 0
    
    insights["customer_health"] = {
        "active_customers": active_customers,
        "top1_concentration": top_cust_conc,
        "top3_concentration": top3_cust_conc,
        "revenue_per_customer": rev_per_cust,
        "vs_peers": active_customers - avg_peer_customers,
        "concentration_risk": "High" if top3_cust_conc > 60 else "Medium" if top3_cust_conc > 40 else "Low"
    }
    
    # Sales Efficiency Drivers
    gp_per_cust = rep_metrics.get("gp_per_customer", 0)
    gp_per_order = rep_metrics.get("avg_gp_per_order", 0)    
    # Peer comparison
    if peer_metrics:
        peer_gp_per_cust = [m.get("gp_per_customer", 0) for m in peer_metrics if m.get("gp_per_customer", 0) > 0]
        avg_peer_gp_per_cust = sum(peer_gp_per_cust) / len(peer_gp_per_cust) if peer_gp_per_cust else 0
    else:
        avg_peer_gp_per_cust = 0
    
    insights["sales_efficiency"] = {
        "gp_per_customer": gp_per_cust,
        "gp_per_order": gp_per_order,
        "vs_peer_gp_per_cust": gp_per_cust - avg_peer_gp_per_cust,
        "peer_avg_gp_per_customer": avg_peer_gp_per_cust
    }
    
    # Product Mix Drivers
    category_margins = rep_metrics.get("margin_by_category", {})
    revenue_by_cat = rep_metrics.get("revenue_by_category", {})
    
    # Sort by revenue
    top_categories = sorted(revenue_by_cat.items(), key=lambda x: x[1], reverse=True)[:5]
    
    insights["product_mix"] = {
        "top_categories": [(cat, rev, category_margins.get(cat, 0)) for cat, rev in top_categories],
        "category_count": len(revenue_by_cat)
    }
    
    return insights


def _generate_actionable_recommendations(insights: dict, component_scores: dict, 
                                        rep_metrics: dict, max_recommendations: int = 4) -> list:
    """Generate specific, actionable recommendations based on driver insights.
    
    Args:
        max_recommendations: Maximum number of recommendations (default 4 for space)
    
    Returns list of recommendation strings.
    """
    recommendations = []
    
    # Sort components by score to prioritize improvements
    sorted_components = sorted(component_scores.items(), key=lambda x: x[1] if isinstance(x[1], (int, float)) else 50)
    
    # Revenue Momentum recommendations
    if component_scores.get("revenue_momentum", 50) < 60:
        growth_pct = insights["revenue_momentum"].get("growth_pct", 0)
        if growth_pct < 0:
            recommendations.append(
                f"📈 URGENT: Your revenue dropped {abs(growth_pct):.1f}% over 12 weeks. "
                "Focus on reactivating dormant customers and cross-selling to existing accounts."
            )
        else:
            recommendations.append(
                "📈 Your revenue growth is below average. Try setting weekly sales targets "
                "and tracking progress to build momentum."
            )
    
    # Customer Health recommendations
    cust_insights = insights.get("customer_health", {})
    if cust_insights.get("concentration_risk") == "High":
        top3_pct = cust_insights.get("top3_concentration", 0)
        recommendations.append(
            f"⚠️ HIGH RISK: {top3_pct:.0f}% of your revenue comes from just 3 customers. "
            "Grow 3-5 mid-sized customers to reduce dependency risk."
        )
    
    if cust_insights.get("active_customers", 0) > 0:
        vs_peers = cust_insights.get("vs_peers", 0)
        if vs_peers < -5:
            recommendations.append(
                f"👥 You have {abs(vs_peers):.0f} fewer customers than the average for reps selling similar product categories. "
                "Focus on prospecting and customer acquisition to diversify your base."
            )
    
    # Margin Discipline recommendations (category-aware)
    margin_insights = insights.get("margin_discipline", {})
    vs_peer_margin = margin_insights.get("vs_peer_margin", 0)
    peer_avg_margin = margin_insights.get("peer_avg_margin", 0)
    
    if vs_peer_margin < -3 and peer_avg_margin > 0:
        recommendations.append(
            f"💰 Your margin ({margin_insights.get('avg_margin', 0):.1f}%) is below the average for reps selling similar product categories "
            f"({peer_avg_margin:.1f}%). Review pricing strategy within your categories."
        )
    
    if margin_insights.get("consistency") in ["Watch", "High Risk"]:
        std_dev = margin_insights.get("std_dev", 0)
        peer_avg_std = margin_insights.get("peer_avg_std", 10)
        recommendations.append(
            f"💰 Your margin volatility ({std_dev:.1f}%) is higher than the average for reps selling similar categories ({peer_avg_std:.1f}%). "
            "Review pricing consistency within your product categories."
        )
    
    # Check for low-margin customers
    low_margin_custs = margin_insights.get("lowest_margin_customers", [])
    if low_margin_custs and len(low_margin_custs) > 0:
        worst_cust, worst_margin = low_margin_custs[0]
        if worst_margin < 10:
            recommendations.append(
                f"💡 Customer #{worst_cust} has very low margin ({worst_margin:.1f}%). "
                "Consider repricing or focus energy on higher-margin accounts."
            )
    
    # Sales Efficiency recommendations
    eff_insights = insights.get("sales_efficiency", {})
    if eff_insights.get("vs_peer_gp_per_cust", 0) < 0:
        recommendations.append(
            "💵 Your profit per customer is below average for reps selling similar product categories. Try upselling complementary products "
            "or focusing on higher-margin items with existing customers."
        )
    
    # Product Mix recommendations
    mix_insights = insights.get("product_mix", {})
    top_cats = mix_insights.get("top_categories", [])
    if top_cats:
        # Find highest margin category
        best_cat = max(top_cats, key=lambda x: x[2]) if top_cats else None
        if best_cat and best_cat[2] > 20:
            cat_name, cat_rev, cat_margin = best_cat
            recommendations.append(
                f"✨ Your best category is '{cat_name}' with {cat_margin:.1f}% margin. "
                "Try to grow sales in this category to boost overall profitability."
            )
    
    # If no specific recommendations, add general encouragement
    if not recommendations:
        recommendations.append(
            "✅ Keep up the good work! Continue focusing on customer relationships "
            "and maintaining your current performance level."
        )
    
    # Limit to max_recommendations (default 4 for PDF space)
    return recommendations[:max_recommendations]


def _create_matplotlib_sparkline(weekly_revenue: list) -> BytesIO:
    """Create a compact line chart for weekly revenue (12 weeks)."""
    fig, ax = plt.subplots(figsize=(3, 0.9))  # Reduced size
    
    if weekly_revenue:
        weeks = list(range(1, len(weekly_revenue) + 1))
        ax.plot(weeks, weekly_revenue, color='#2E86AB', linewidth=1.5, marker='o', markersize=2)
        ax.fill_between(weeks, weekly_revenue, alpha=0.3, color='#A23B72')
        ax.set_xlabel("Week", fontsize=6)
        ax.set_ylabel("Revenue", fontsize=6)
        ax.tick_params(labelsize=5)
        ax.grid(True, alpha=0.2, linestyle='--')
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
    
    plt.tight_layout(pad=0.2)
    
    # Save to BytesIO
    img_buffer = BytesIO()
    plt.savefig(img_buffer, format='png', dpi=120, bbox_inches='tight')  # Reduced DPI
    plt.close(fig)
    img_buffer.seek(0)
    
    return img_buffer


def _create_matplotlib_category_bars(top_categories: list, peer_avg_margins: dict = None) -> BytesIO:
    """Create compact bar chart comparing category margins (rep vs peers)."""
    fig, ax = plt.subplots(figsize=(3.5, 1.5))  # Significantly reduced
    
    if top_categories:
        # Limit to top 3 categories for space
        top_categories = top_categories[:3]
        cat_names = [cat[0][:12] + '...' if len(cat[0]) > 12 else cat[0] for cat in top_categories]
        cat_margins = [cat[2] for cat in top_categories]
        
        x_pos = np.arange(len(cat_names))
        bars = ax.bar(x_pos, cat_margins, color='#2E86AB', alpha=0.8, width=0.6)
        
        # Color bars by margin level
        for i, bar in enumerate(bars):
            if cat_margins[i] > 25:
                bar.set_color('#27AE60')  # Green for good
            elif cat_margins[i] < 15:
                bar.set_color('#E74C3C')  # Red for low
            else:
                bar.set_color('#F39C12')  # Orange for medium
        
        ax.set_xticks(x_pos)
        ax.set_xticklabels(cat_names, rotation=30, ha='right', fontsize=6)
        ax.set_ylabel("Margin %", fontsize=7)
        ax.tick_params(labelsize=6)
        ax.grid(True, alpha=0.2, axis='y', linestyle='--')
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        
        # Add value labels on bars
        for i, v in enumerate(cat_margins):
            ax.text(i, v + 0.5, f'{v:.0f}%', ha='center', va='bottom', fontsize=5)
    
    plt.tight_layout(pad=0.2)
    
    img_buffer = BytesIO()
    plt.savefig(img_buffer, format='png', dpi=120, bbox_inches='tight')
    plt.close(fig)
    img_buffer.seek(0)
    
    return img_buffer


def _generate_executive_summary_page(all_reps_data: list, start_date: date, end_date: date, 
                                     story: list, styles: dict) -> None:
    """Generate 1-page executive summary showing all reps and winners/losers.
    
    Args:
        all_reps_data: List of rep data dictionaries
        start_date: Analysis start date
        end_date: Analysis end date
        story: ReportLab story list to append to
        styles: Dictionary of paragraph styles
    """
    from reportlab.lib.colors import HexColor
    
    # Title
    story.append(Paragraph("Sales Rep Performance - Executive Summary", styles['title']))
    story.append(Paragraph(
        f"Analysis Period: Last 12 closed ISO weeks ending {end_date.strftime('%B %d, %Y')}<br/>"
        f"<i>Note: Scores compare reps only to similar reps selling similar product categories.</i>",
        styles['caption']
    ))
    story.append(Spacer(1, 0.15*inch))
    
    # Build all reps summary table
    table_data = [[
        "Rep", "Score", "Band", "Revenue\nMomentum", "Margin\nDiscipline", 
        "Customer\nHealth", "Sales\nEfficiency", "Product\nMix"
    ]]
    
    # Sort by health score descending
    sorted_reps = sorted(all_reps_data, key=lambda x: x['health_score'], reverse=True)
    
    for rep_data in sorted_reps:
        rep_name = rep_data['rep_name']
        health_score = rep_data['health_score']
        components = rep_data['components']
        
        # Determine band
        if health_score >= 80:
            band = "Healthy"
        elif health_score >= 60:
            band = "Watch"
        elif health_score >= 40:
            band = "At Risk"
        else:
            band = "Critical"
        
        table_data.append([
            rep_name[:15],  # Truncate long names
            f"{health_score:.0f}",
            band,
            f"{components.get('revenue_momentum', 0):.0f}",
            f"{components.get('margin_discipline', 0):.0f}",
            f"{components.get('customer_health', 0):.0f}",
            f"{components.get('sales_efficiency', 0):.0f}",
            f"{components.get('product_mix_quality', 0):.0f}"
        ])
    
    # Create table with conditional formatting
    col_widths = [1.2*inch, 0.5*inch, 0.8*inch, 0.7*inch, 0.7*inch, 0.7*inch, 0.7*inch, 0.7*inch]
    summary_table = Table(table_data, colWidths=col_widths, repeatRows=1)
    
    # Build table style with heatmap coloring
    table_style = [
        ('BACKGROUND', (0, 0), (-1, 0), HexColor('#1F4788')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 7),
        ('FONTSIZE', (0, 1), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, HexColor('#F9F9F9')]),
    ]
    
    # Add conditional formatting for bands
    for i, rep_data in enumerate(sorted_reps, start=1):
        health_score = rep_data['health_score']
        if health_score >= 80:
            table_style.append(('BACKGROUND', (2, i), (2, i), HexColor('#C8E6C9')))  # Light green
        elif health_score < 40:
            table_style.append(('BACKGROUND', (2, i), (2, i), HexColor('#FFCDD2')))  # Light red
    
    summary_table.setStyle(TableStyle(table_style))
    story.append(summary_table)
    story.append(Spacer(1, 0.15*inch))
    
    # Winners & Losers section
    story.append(Paragraph("🏆 Top Performers & 🚨 Needs Attention", styles['subheader']))
    
    # Calculate winners/losers for each component
    component_labels = {
        'revenue_momentum': 'Revenue Growth',
        'margin_discipline': 'Margin Consistency',
        'customer_health': 'Customer Base',
        'sales_efficiency': 'Sales Efficiency',
        'product_mix_quality': 'Product Mix'
    }
    
    winners_data = []
    for comp_key, comp_label in component_labels.items():
        # Filter out invalid scores
        valid_reps = [
            (r['rep_name'], r['components'].get(comp_key, 0))
            for r in all_reps_data
            if isinstance(r['components'].get(comp_key), (int, float)) and r['components'].get(comp_key) > 0
        ]
        
        if len(valid_reps) >= 2:
            valid_reps.sort(key=lambda x: x[1], reverse=True)
            winner_name, winner_score = valid_reps[0]
            loser_name, loser_score = valid_reps[-1]
            
            winners_data.append([
                f"<b>{comp_label}</b>",
                f"🏆 {winner_name[:12]}: {winner_score:.0f}",
                f"🚨 {loser_name[:12]}: {loser_score:.0f}"
            ])
    
    # Add overall health score winners/losers
    health_sorted = sorted(all_reps_data, key=lambda x: x['health_score'], reverse=True)
    if len(health_sorted) >= 2:
        winners_data.insert(0, [
            "<b>Overall Health</b>",
            f"🏆 {health_sorted[0]['rep_name'][:12]}: {health_sorted[0]['health_score']:.0f}",
            f"🚨 {health_sorted[-1]['rep_name'][:12]}: {health_sorted[-1]['health_score']:.0f}"
        ])
    
    winners_table = Table(winners_data, colWidths=[1.5*inch, 2.5*inch, 2.5*inch])
    winners_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), HexColor('#E8F4F8')),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (-1, -1), 'LEFT'),
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]))
    story.append(winners_table)
    story.append(Spacer(1, 0.15*inch))
    
    # Executive Insights (data-driven)
    story.append(Paragraph("📊 Key Insights", styles['subheader']))
    
    insights = []
    
    # Insight 1: Weakest component overall
    component_avgs = {}
    for comp_key, comp_label in component_labels.items():
        scores = [r['components'].get(comp_key, 0) for r in all_reps_data 
                 if isinstance(r['components'].get(comp_key), (int, float))]
        if scores:
            component_avgs[comp_label] = sum(scores) / len(scores)
    
    if component_avgs:
        weakest_comp = min(component_avgs.items(), key=lambda x: x[1])
        insights.append(f"• <b>Common weakness:</b> Most reps struggle with {weakest_comp[0]} (avg: {weakest_comp[1]:.0f}/100)")
    
    # Insight 2: Customer concentration risk
    high_concentration = sum(1 for r in all_reps_data 
                            if r['metrics'].get('top3_customer_concentration', 0) > 60)
    if high_concentration > 0:
        insights.append(f"• <b>Risk alert:</b> {high_concentration} reps have >60% revenue from top 3 customers")
    
    # Insight 3: Top performer characteristic
    if len(health_sorted) >= 3:
        top_3_avg_customers = sum(r['metrics'].get('active_customers', 0) for r in health_sorted[:3]) / 3
        bottom_3_avg_customers = sum(r['metrics'].get('active_customers', 0) for r in health_sorted[-3:]) / 3
        if top_3_avg_customers > bottom_3_avg_customers * 1.2:
            insights.append(f"• <b>Success pattern:</b> Top performers have {top_3_avg_customers:.0f} customers on average vs {bottom_3_avg_customers:.0f} for bottom performers")
    
    for insight in insights[:3]:  # Limit to 3
        story.append(Paragraph(insight, styles['body']))
    
    story.append(PageBreak())


def _add_rep_detail_pages(rep_name: str, df: pd.DataFrame, story: list, style_dict: dict) -> None:
    """Add detailed performance analysis pages for a sales rep.
    
    Implements progressive drill-down with proper monthly aggregation and peer comparisons.
    
    Args:
        rep_name: Name of the sales rep
        df: Full orders dataframe
        story: ReportLab story list to append to
        style_dict: Dict of ParagraphStyle objects
    """
    # Extract styles
    header_style = style_dict['header']
    subheader_style = style_dict['subheader']
    body_style = style_dict['body']
    caption_style = style_dict['caption']
    
    # ========================================
    # STEP 0: VALIDATE AND PREPARE DATA
    # ========================================
    
    # Determine which date field to use
    date_field = None
    for field in ['order_date', 'invoice_date', 'order_entry_date']:
        if field in df.columns:
            date_field = field
            break
    
    if date_field is None:
        # No date field - skip silently
        return
    
    # Validate required fields
    required = ['salesperson', 'line_revenue']
    if not all(f in df.columns for f in required):
        return
    
    # Get rep's data
    rep_df = df[df['salesperson'] == rep_name].copy()
    if rep_df.empty:
        return
    
    # Parse dates
    rep_df[date_field] = pd.to_datetime(rep_df[date_field], errors='coerce')
    rep_df = rep_df.dropna(subset=[date_field])
    
    if rep_df.empty:
        return
    
    # ========================================
    # STEP 1: BUILD MONTHLY REVENUE FACT TABLE
    # ========================================
    
    # Add month_start column (first day of month)
    df_working = df.copy()
    df_working[date_field] = pd.to_datetime(df_working[date_field], errors='coerce')
    df_working = df_working.dropna(subset=[date_field])
    
    if df_working.empty:
        return
    
    df_working['month_start'] = df_working[date_field].dt.to_period('M').dt.to_timestamp()
    
    # Get last complete month and prior month
    today = pd.Timestamp.today()
    current_month_start = pd.Timestamp(today.year, today.month, 1)
    last_complete_month = current_month_start - pd.DateOffset(months=1)
    prior_month = last_complete_month - pd.DateOffset(months=1)
    
    # Filter to only closed months (exclude current partial month)
    df_monthly = df_working[df_working['month_start'] < current_month_start].copy()
    
    if df_monthly.empty:
        return
    
    # Determine grouping columns
    has_cost_center = 'cost_center' in df_monthly.columns
    has_account = 'account_number' in df_monthly.columns
    
    if not has_account:
        # Can't do account-level analysis without accounts
        return
    
    # Build monthly aggregation
    group_cols = ['salesperson', 'month_start']
    if has_account:
        group_cols.append('account_number')
    if has_cost_center:
        group_cols.append('cost_center')
    
    # Aggregate to monthly level
    monthly_facts = df_monthly.groupby(group_cols, as_index=False).agg({
        'line_revenue': 'sum'
    }).rename(columns={'line_revenue': 'revenue'})
    
    if monthly_facts.empty:
        return
    
    # Filter to rep's data for last 2 months
    rep_monthly = monthly_facts[
        (monthly_facts['salesperson'] == rep_name) &
        (monthly_facts['month_start'].isin([last_complete_month, prior_month]))
    ].copy()
    
    if rep_monthly.empty:
        return
    
    # Check if we have data in both months
    months_present = rep_monthly['month_start'].nunique()
    if months_present < 2:
        # Only one month of data - can't do MoM
        return
    
    # ========================================
    # FORMAT MONTH LABELS FOR ALL PAGES
    # ========================================
    prior_month_label = prior_month.strftime('%b %Y')  # "Dec 2025"
    current_month_label = last_complete_month.strftime('%b %Y')  # "Jan 2026"
    month_comparison = f"{prior_month_label} → {current_month_label}"
    
    # ========================================
    # STEP 2: COST CENTER BASELINE COMPARISON (MoM)
    # ========================================
    
    if has_cost_center:
        story.append(PageBreak())
        story.append(Paragraph(f"Cost Center Baseline Comparison: {rep_name}", header_style))
        story.append(Paragraph(
            f"Month-over-month change: {month_comparison}",
            subheader_style
        ))
        story.append(Spacer(1, 0.1*inch))
        
        cc_summary = _compute_cc_mom_comparison(
            rep_name, monthly_facts, last_complete_month, prior_month
        )
        
        if not cc_summary.empty:
            story.append(Paragraph(
                f"How you're trending vs the overall cost center across ALL reps.",
                body_style
            ))
            story.append(Spacer(1, 0.1*inch))
            
            # Build table with clear month labels
            table_data = [[
                "Cost Center", 
                f"Prior ({prior_month_label})", 
                f"Current ({current_month_label})", 
                f"Rep MoM%", 
                "CC Total MoM%", 
                "Indicator"
            ]]
            
            for _, row in cc_summary.head(10).iterrows():
                cc = row['cost_center']
                prior = row['rep_prior_rev']
                curr = row['rep_curr_rev']
                rep_mom = row['rep_mom_pct']
                cc_mom = row['cc_baseline_mom_pct']
                status = row['status']
                
                # Format
                prior_str = f"${prior:,.0f}" if pd.notna(prior) else "$0"
                curr_str = f"${curr:,.0f}" if pd.notna(curr) else "$0"
                rep_str = f"{rep_mom:+.1f}%" if pd.notna(rep_mom) and np.isfinite(rep_mom) else "—"
                cc_str = f"{cc_mom:+.1f}%" if pd.notna(cc_mom) and np.isfinite(cc_mom) else "—"
                
                if status == "Above":
                    status_str = "↑ Above"
                elif status == "Below":
                    status_str = "↓ Below"
                else:
                    status_str = "— In Line"
                
                table_data.append([str(cc), prior_str, curr_str, rep_str, cc_str, status_str])
            
            cc_table = Table(table_data, colWidths=[1*inch, 1.0*inch, 1.0*inch, 0.9*inch, 1.0*inch, 0.9*inch])
            cc_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E8F4F8')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#1F4788')),
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 7),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9F9F9')]),
                ('TOPPADDING', (0, 0), (-1, -1), 3),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ]))
            
            story.append(cc_table)
            story.append(Spacer(1, 0.15*inch))
            
            # Add narrative summary
            above = cc_summary[cc_summary['status'] == 'Above']
            below = cc_summary[cc_summary['status'] == 'Below']
            total_ccs = len(cc_summary)
            above_count = len(above)
            
            story.append(Paragraph("📊 Summary:", body_style))
            story.append(Spacer(1, 0.05*inch))
            
            if not above.empty:
                best = above.iloc[0]
                story.append(Paragraph(
                    f"• <b>Biggest above-baseline cost center:</b> {best['cost_center']} "
                    f"(you: {best['rep_mom_pct']:+.1f}% vs CC total: {best['cc_baseline_mom_pct']:+.1f}%)",
                    body_style
                ))
            
            if not below.empty:
                worst = below.iloc[0]
                story.append(Paragraph(
                    f"• <b>Biggest below-baseline cost center:</b> {worst['cost_center']} "
                    f"(you: {worst['rep_mom_pct']:+.1f}% vs CC total: {worst['cc_baseline_mom_pct']:+.1f}%)",
                    body_style
                ))
            
            story.append(Paragraph(
                f"• <b>Overall:</b> You are outperforming baseline in {above_count} of {total_ccs} cost centers.",
                body_style
            ))
            
            story.append(Spacer(1, 0.1*inch))
            story.append(Paragraph(
                "<i>Note: CC Total MoM% = overall cost center change across ALL reps (not peer-grouped).</i>",
                caption_style
            ))
    
    # ========================================
    # STEP 3: SAME-ACCOUNT REP COMPARISON (MoM)
    # ========================================
    
    if has_account:
        acct_detail = _compute_account_mom_comparison(
            rep_name, monthly_facts, last_complete_month, prior_month
        )
        
        if not acct_detail.empty:
            # Take top 15 accounts by current revenue
            top_accounts = acct_detail.nlargest(15, 'rep_curr_rev')
            
            if not top_accounts.empty:
                story.append(PageBreak())
                story.append(Paragraph(f"Same-Account Rep Comparison: {rep_name}", header_style))
                story.append(Paragraph(
                    f"Month-over-month change: {month_comparison}",
                    subheader_style
                ))
                story.append(Spacer(1, 0.1*inch))
                
                story.append(Paragraph(
                    "How your MoM% compares to OTHER REPS selling on the SAME accounts.",
                    body_style
                ))
                story.append(Spacer(1, 0.1*inch))
                
                # Build table
                acct_table_data = [[
                    "Account",
                    "CC",
                    f"Prior ({prior_month_label})",
                    f"Current ({current_month_label})",
                    "Δ Rev",
                    "Rep MoM%",
                    "Others Avg MoM%",
                    "Indicator"
                ]]
                
                for _, row in top_accounts.iterrows():
                    acct = row['account']
                    cc = row.get('cost_center', '')
                    prior = row['rep_prior_rev']
                    curr = row['rep_curr_rev']
                    delta = curr - prior
                    rep_mom = row['rep_mom_pct']
                    others_mom = row['others_avg_mom_pct']
                    status = row['status']
                    
                    prior_str = f"${prior:,.0f}" if pd.notna(prior) else "$0"
                    curr_str = f"${curr:,.0f}" if pd.notna(curr) else "$0"
                    delta_str = f"${delta:,.0f}"
                    rep_str = f"{rep_mom:+.1f}%" if pd.notna(rep_mom) and np.isfinite(rep_mom) else "—"
                    others_str = f"{others_mom:+.1f}%" if pd.notna(others_mom) and np.isfinite(others_mom) else "No other reps"
                    
                    if status == "Above":
                        status_str = "↑"
                    elif status == "Below":
                        status_str = "↓"
                    else:
                        status_str = "—"
                    
                    acct_table_data.append([
                        str(acct)[:22],
                        str(cc)[:8] if cc else "",
                        prior_str,
                        curr_str,
                        delta_str,
                        rep_str,
                        others_str,
                        status_str
                    ])
                
                acct_table = Table(acct_table_data, colWidths=[1.3*inch, 0.5*inch, 0.7*inch, 0.7*inch, 0.65*inch, 0.65*inch, 0.9*inch, 0.4*inch])
                acct_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E8F4F8')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#1F4788')),
                    ('ALIGN', (0, 0), (1, -1), 'LEFT'),
                    ('ALIGN', (2, 0), (-1, -1), 'RIGHT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 6),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9F9F9')]),
                    ('TOPPADDING', (0, 0), (-1, -1), 3),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
                ]))
                
                story.append(acct_table)
                story.append(Spacer(1, 0.15*inch))
                
                # Callouts
                grew_faster = top_accounts[
                    (top_accounts['status'] == 'Above') & 
                    (top_accounts['rep_mom_pct'] > 0)
                ]
                declined_while_others_grew = top_accounts[
                    (top_accounts['status'] == 'Below') &
                    (top_accounts['rep_mom_pct'] < 0) &
                    (top_accounts['others_avg_mom_pct'] > 0)
                ]
                
                story.append(Paragraph("💡 Key Insights:", body_style))
                story.append(Spacer(1, 0.05*inch))
                
                if not grew_faster.empty:
                    best = grew_faster.iloc[0]
                    story.append(Paragraph(
                        f"• <b>Accounts where you grew faster than others:</b> {best['account']} "
                        f"(you: {best['rep_mom_pct']:+.1f}% vs others: {best['others_avg_mom_pct']:+.1f}%)",
                        body_style
                    ))
                
                if not declined_while_others_grew.empty:
                    worst = declined_while_others_grew.iloc[0]
                    story.append(Paragraph(
                        f"• <b>Accounts where you declined while others grew:</b> {worst['account']} "
                        f"(you: {worst['rep_mom_pct']:+.1f}% vs others: {worst['others_avg_mom_pct']:+.1f}%)",
                        body_style
                    ))
                
                # Share loss risk
                big_declines_with_peers = top_accounts[
                    (top_accounts['rep_mom_pct'] < -10) &
                    (top_accounts['others_avg_mom_pct'].notna())
                ]
                if not big_declines_with_peers.empty:
                    risky = big_declines_with_peers.iloc[0]
                    story.append(Paragraph(
                        f"• <b>Accounts with biggest share loss risk:</b> {risky['account']} "
                        f"(you declined {risky['rep_mom_pct']:.1f}% while competing reps are active)",
                        body_style
                    ))
                
                story.append(Spacer(1, 0.1*inch))
                story.append(Paragraph(
                    "<i>Note: 'Others Avg MoM%' = average MoM for OTHER REPS on that specific account.</i>",
                    caption_style
                ))
    
    # ========================================
    # STEP 4: TOP/BOTTOM ACCOUNT MOVERS
    # ========================================
    
    if has_account:
        movers = _compute_top_bottom_movers(
            rep_name, monthly_facts, last_complete_month, prior_month
        )
        
        if not movers.empty and len(movers) >= 2:
            story.append(PageBreak())
            story.append(Paragraph(f"Top & Bottom Account Movers: {rep_name}", header_style))
            story.append(Paragraph(
                f"Month-over-month change: {month_comparison}",
                subheader_style
            ))
            story.append(Spacer(1, 0.1*inch))
            
            # Top 5
            top_5 = movers.nlargest(5, 'revenue_delta')
            if not top_5.empty and (top_5['revenue_delta'] > 0).any():
                story.append(Paragraph("🔼 Top 5 Growing Accounts", subheader_style))
                story.append(Spacer(1, 0.05*inch))
                
                top_data = [[
                    "Rank", 
                    "Account", 
                    "CC", 
                    f"Prior ({prior_month_label})", 
                    f"Current ({current_month_label})", 
                    "Δ Revenue", 
                    f"MoM% ({prior_month_label}→{current_month_label})"
                ]]
                for i, (_, row) in enumerate(top_5.iterrows(), 1):
                    acct = str(row['account'])[:22]
                    cc = str(row.get('cost_center', ''))[:8] if has_cost_center else ""
                    prior = row['prior_rev']
                    curr = row['curr_rev']
                    delta = row['revenue_delta']
                    pct = row['mom_pct']
                    
                    prior_str = f"${prior:,.0f}" if pd.notna(prior) else "$0"
                    curr_str = f"${curr:,.0f}" if pd.notna(curr) else "$0"
                    delta_str = f"${delta:,.0f}"
                    pct_str = f"{pct:+.1f}%" if pd.notna(pct) and np.isfinite(pct) and prior > 0 else "New"
                    
                    top_data.append([str(i), acct, cc, prior_str, curr_str, delta_str, pct_str])
                
                top_table = Table(top_data, colWidths=[0.35*inch, 1.5*inch, 0.6*inch, 0.8*inch, 0.8*inch, 0.8*inch, 0.95*inch])
                top_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#D5F4E6')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#27AE60')),
                    ('ALIGN', (0, 0), (0, -1), 'CENTER'),
                    ('ALIGN', (1, 0), (2, -1), 'LEFT'),
                    ('ALIGN', (3, 0), (-1, -1), 'RIGHT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 7),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('TOPPADDING', (0, 0), (-1, -1), 3),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
                ]))
                
                story.append(top_table)
                story.append(Spacer(1, 0.2*inch))
            
            # Bottom 5
            bottom_5 = movers.nsmallest(5, 'revenue_delta')
            if not bottom_5.empty and (bottom_5['revenue_delta'] < 0).any():
                story.append(Paragraph("🔽 Top 5 Declining Accounts", subheader_style))
                story.append(Spacer(1, 0.05*inch))
                
                bottom_data = [[
                    "Rank", 
                    "Account", 
                    "CC", 
                    f"Prior ({prior_month_label})", 
                    f"Current ({current_month_label})", 
                    "Δ Revenue", 
                    f"MoM% ({prior_month_label}→{current_month_label})"
                ]]
                for i, (_, row) in enumerate(bottom_5.iterrows(), 1):
                    acct = str(row['account'])[:22]
                    cc = str(row.get('cost_center', ''))[:8] if has_cost_center else ""
                    prior = row['prior_rev']
                    curr = row['curr_rev']
                    delta = row['revenue_delta']
                    pct = row['mom_pct']
                    
                    prior_str = f"${prior:,.0f}" if pd.notna(prior) else "$0"
                    curr_str = f"${curr:,.0f}" if pd.notna(curr) else "$0"
                    delta_str = f"${delta:,.0f}"
                    pct_str = f"{pct:+.1f}%" if pd.notna(pct) and np.isfinite(pct) and prior > 0 else "Lost"
                    
                    bottom_data.append([str(i), acct, cc, prior_str, curr_str, delta_str, pct_str])
                
                bottom_table = Table(bottom_data, colWidths=[0.35*inch, 1.5*inch, 0.6*inch, 0.8*inch, 0.8*inch, 0.8*inch, 0.95*inch])
                bottom_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#FADBD8')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#E74C3C')),
                    ('ALIGN', (0, 0), (0, -1), 'CENTER'),
                    ('ALIGN', (1, 0), (2, -1), 'LEFT'),
                    ('ALIGN', (3, 0), (-1, -1), 'RIGHT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 7),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('TOPPADDING', (0, 0), (-1, -1), 3),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
                ]))
                
                story.append(bottom_table)
                story.append(Spacer(1, 0.15*inch))
            
            # Narrative box
            if len(movers) >= 10:
                story.append(Paragraph("📊 Impact Analysis:", body_style))
                story.append(Spacer(1, 0.05*inch))
                
                top_gain = top_5['revenue_delta'].sum()
                bottom_loss = bottom_5['revenue_delta'].sum()
                net = top_gain + bottom_loss
                
                # Identify drivers
                if not top_5.empty:
                    top_driver = top_5.iloc[0]
                    story.append(Paragraph(
                        f"• <b>Biggest driver of increase:</b> {top_driver['account']} "
                        f"added ${top_driver['revenue_delta']:,.0f} "
                        f"({top_driver['mom_pct']:+.1f}% growth)",
                        body_style
                    ))
                
                if not bottom_5.empty:
                    bottom_driver = bottom_5.iloc[0]
                    story.append(Paragraph(
                        f"• <b>Biggest driver of decline:</b> {bottom_driver['account']} "
                        f"lost ${abs(bottom_driver['revenue_delta']):,.0f} "
                        f"({bottom_driver['mom_pct']:+.1f}% decline)",
                        body_style
                    ))
                
                story.append(Paragraph(
                    f"• <b>Net impact from top/bottom 10:</b> ${net:,.0f}",
                    body_style
                ))
            
            story.append(Spacer(1, 0.1*inch))
            story.append(Paragraph(
                "<i>MoM% = (Current - Prior) / Prior. If Prior = 0 and Current > 0, labeled as 'New'.</i>",
                caption_style
            ))
    
    # ========================================
    # NEW SECTION 1: COST CENTER MONTHLY SALES TABLES (24 MONTHS)
    # ========================================
    
    if has_cost_center:
        cc_monthly_data = _compute_cc_monthly_trends(
            rep_name, monthly_facts, last_complete_month
        )
        
        if not cc_monthly_data.empty:
            _render_cc_monthly_tables(
                rep_name, cc_monthly_data, story, style_dict
            )
    
    # ========================================
    # NEW SECTION 2: SHARED-ACCOUNT REP VS REP MONTHLY CHANGE TABLES (24 MONTHS)
    # ========================================
    
    if has_account:
        shared_rep_data = _compute_shared_account_trends(
            rep_name, monthly_facts, last_complete_month
        )
        
        if shared_rep_data:
            _render_shared_account_tables(
                rep_name, shared_rep_data, story, style_dict
            )


def _compute_cc_mom_comparison(
    rep_name: str,
    monthly_facts: pd.DataFrame,
    current_month: pd.Timestamp,
    prior_month: pd.Timestamp
) -> pd.DataFrame:
    """Compute cost center MoM comparison for rep vs baseline.
    
    Returns DataFrame with:
    - cost_center
    - rep_prior_rev, rep_curr_rev, rep_mom_pct
    - cc_baseline_mom_pct (all reps in CC)
    - status (Above/In Line/Below)
    """
    if 'cost_center' not in monthly_facts.columns:
        return pd.DataFrame()
    
    results = []
    
    # Get rep's cost centers
    rep_data = monthly_facts[monthly_facts['salesperson'] == rep_name].copy()
    cost_centers = rep_data['cost_center'].dropna().unique()
    
    for cc in cost_centers:
        # Rep's revenue in this CC
        rep_cc_prior = rep_data[
            (rep_data['cost_center'] == cc) &
            (rep_data['month_start'] == prior_month)
        ]['revenue'].sum()
        
        rep_cc_curr = rep_data[
            (rep_data['cost_center'] == cc) &
            (rep_data['month_start'] == current_month)
        ]['revenue'].sum()
        
        # Skip if no activity
        if rep_cc_prior == 0 and rep_cc_curr == 0:
            continue
        
        rep_mom_pct = ((rep_cc_curr - rep_cc_prior) / rep_cc_prior * 100) if rep_cc_prior > 0 else (100 if rep_cc_curr > 0 else 0)
        
        # Cost center baseline (all reps)
        cc_data = monthly_facts[monthly_facts['cost_center'] == cc]
        
        cc_prior = cc_data[cc_data['month_start'] == prior_month]['revenue'].sum()
        cc_curr = cc_data[cc_data['month_start'] == current_month]['revenue'].sum()
        
        cc_mom_pct = ((cc_curr - cc_prior) / cc_prior * 100) if cc_prior > 0 else None
        
        # Determine status
        if cc_mom_pct is not None and np.isfinite(cc_mom_pct):
            if rep_mom_pct > cc_mom_pct + 5:
                status = "Above"
            elif rep_mom_pct < cc_mom_pct - 5:
                status = "Below"
            else:
                status = "In Line"
        else:
            status = "N/A"
        
        results.append({
            'cost_center': str(cc),
            'rep_prior_rev': rep_cc_prior,
            'rep_curr_rev': rep_cc_curr,
            'rep_mom_pct': rep_mom_pct,
            'cc_baseline_mom_pct': cc_mom_pct,
            'status': status
        })
    
    if not results:
        return pd.DataFrame()
    
    return pd.DataFrame(results).sort_values('rep_curr_rev', ascending=False)


def _compute_account_mom_comparison(
    rep_name: str,
    monthly_facts: pd.DataFrame,
    current_month: pd.Timestamp,
    prior_month: pd.Timestamp
) -> pd.DataFrame:
    """Compute account-level MoM comparison for rep vs other reps on same account.
    
    Returns DataFrame with:
    - account, cost_center
    - rep_prior_rev, rep_curr_rev, rep_mom_pct
    - others_avg_mom_pct (other reps on THIS account)
    - status
    """
    if 'account_number' not in monthly_facts.columns:
        return pd.DataFrame()
    
    results = []
    
    # Get rep's accounts
    rep_data = monthly_facts[monthly_facts['salesperson'] == rep_name].copy()
    accounts = rep_data['account_number'].dropna().unique()
    
    for acct in accounts:
        # Rep's revenue on this account
        rep_acct_data = rep_data[rep_data['account_number'] == acct]
        
        rep_prior = rep_acct_data[rep_acct_data['month_start'] == prior_month]['revenue'].sum()
        rep_curr = rep_acct_data[rep_acct_data['month_start'] == current_month]['revenue'].sum()
        
        # Skip if no activity
        if rep_prior == 0 and rep_curr == 0:
            continue
        
        rep_mom_pct = ((rep_curr - rep_prior) / rep_prior * 100) if rep_prior > 0 else (100 if rep_curr > 0 else 0)
        
        # Get cost center
        cc = rep_acct_data['cost_center'].mode()[0] if 'cost_center' in rep_acct_data.columns and not rep_acct_data['cost_center'].empty else None
        
        # Other reps on this account
        others_data = monthly_facts[
            (monthly_facts['account_number'] == acct) &
            (monthly_facts['salesperson'] != rep_name)
        ]
        
        if not others_data.empty:
            # Compute each other rep's MoM on this account
            other_reps = others_data['salesperson'].unique()
            other_mom_pcts = []
            
            for other_rep in other_reps:
                other_rep_data = others_data[others_data['salesperson'] == other_rep]
                other_prior = other_rep_data[other_rep_data['month_start'] == prior_month]['revenue'].sum()
                other_curr = other_rep_data[other_rep_data['month_start'] == current_month]['revenue'].sum()
                
                if other_prior > 0:
                    other_mom = ((other_curr - other_prior) / other_prior * 100)
                    other_mom_pcts.append(other_mom)
            
            others_avg_mom = np.mean(other_mom_pcts) if other_mom_pcts else None
        else:
            others_avg_mom = None
        
        # Determine status
        if others_avg_mom is not None and np.isfinite(others_avg_mom):
            if rep_mom_pct > others_avg_mom + 5:
                status = "Above"
            elif rep_mom_pct < others_avg_mom - 5:
                status = "Below"
            else:
                status = "In Line"
        else:
            status = "N/A"
        
        results.append({
            'account': str(acct),
            'cost_center': str(cc) if cc is not None else "Unknown",
            'rep_prior_rev': rep_prior,
            'rep_curr_rev': rep_curr,
            'rep_mom_pct': rep_mom_pct,
            'others_avg_mom_pct': others_avg_mom,
            'status': status
        })
    
    if not results:
        return pd.DataFrame()
    
    return pd.DataFrame(results)


def _compute_top_bottom_movers(
    rep_name: str,
    monthly_facts: pd.DataFrame,
    current_month: pd.Timestamp,
    prior_month: pd.Timestamp
) -> pd.DataFrame:
    """Compute top and bottom account movers by absolute revenue change.
    
    Returns DataFrame with:
    - account, cost_center
    - prior_rev, curr_rev
    - revenue_delta, mom_pct
    """
    if 'account_number' not in monthly_facts.columns:
        return pd.DataFrame()
    
    # Get rep's data
    rep_data = monthly_facts[monthly_facts['salesperson'] == rep_name].copy()
    accounts = rep_data['account_number'].dropna().unique()
    
    results = []
    
    for acct in accounts:
        acct_data = rep_data[rep_data['account_number'] == acct]
        
        prior_rev = acct_data[acct_data['month_start'] == prior_month]['revenue'].sum()
        curr_rev = acct_data[acct_data['month_start'] == current_month]['revenue'].sum()
        
        delta = curr_rev - prior_rev
        mom_pct = ((curr_rev - prior_rev) / prior_rev * 100) if prior_rev > 0 else None
        
        # Get cost center
        cc = acct_data['cost_center'].mode()[0] if 'cost_center' in acct_data.columns and not acct_data['cost_center'].empty else None
        
        results.append({
            'account': str(acct),
            'cost_center': str(cc) if cc is not None else "Unknown",
            'prior_rev': prior_rev,
            'curr_rev': curr_rev,
            'revenue_delta': delta,
            'mom_pct': mom_pct
        })
    
    if not results:
        return pd.DataFrame()
    
    return pd.DataFrame(results).sort_values('revenue_delta', ascending=False)
    # If no cost center or account data, skip detail pages
    if not has_cost_center or not has_accounts:
        return
    
    # ========================================
    # STEP 4: DETERMINE WHICH PAGES CAN BE RENDERED
    # ========================================
    can_show_mom_comparison = has_current_month and has_prior_month
    can_show_current_only = has_current_month
    
    # If we can't show anything meaningful, exit
    if not can_show_current_only:
        return
    
    # ========================================
    # PAGE TYPE A: COST CENTER PERFORMANCE OVERVIEW
    # (Only if we have MoM data)
    # ========================================
    
    if can_show_mom_comparison:
        story.append(PageBreak())
        story.append(Paragraph(f"Detail Analysis: {rep_name}", header_style))
        story.append(Spacer(1, 0.1*inch))
        story.append(Paragraph("Cost Center Performance Overview", subheader_style))
        story.append(Paragraph(
            f"Month-over-month comparison: {last_month_start.strftime('%B %Y')} vs {prior_month_start.strftime('%B %Y')}",
            caption_style
        ))
        story.append(Spacer(1, 0.1*inch))
        
        # Compute cost center performance
        cc_performance = _compute_cost_center_performance(
            rep_name, df, last_month_start, last_month_end, 
            prior_month_start, prior_month_end, date_field
        )
        
        if not cc_performance.empty:
            # Build table
            table_data = [["Cost Center", "Current Month", "Prior Month", "Rep %Δ", "Same-Acct Peers %Δ", "All CC Reps %Δ", "Relative"]]
            
            for _, row in cc_performance.iterrows():
                cc = row['cost_center']
                curr_rev = row['current_revenue']
                prior_rev = row['prior_revenue']
                rep_chg = row['rep_pct_change']
                peer_chg = row['same_account_peers_pct_change']
                cc_chg = row['all_cc_reps_pct_change']
                relative = row['relative_performance']
                
                # Format values
                curr_str = f"${curr_rev:,.0f}" if pd.notna(curr_rev) else "$0"
                prior_str = f"${prior_rev:,.0f}" if pd.notna(prior_rev) else "$0"
                rep_str = f"{rep_chg:+.1f}%" if pd.notna(rep_chg) and np.isfinite(rep_chg) else "N/A"
                peer_str = f"{peer_chg:+.1f}%" if pd.notna(peer_chg) and np.isfinite(peer_chg) else "N/A"
                cc_str = f"{cc_chg:+.1f}%" if pd.notna(cc_chg) and np.isfinite(cc_chg) else "N/A"
                
                # Relative indicator with symbol
                if relative == "Above":
                    rel_str = "↑ Above"
                elif relative == "Below":
                    rel_str = "↓ Below"
                else:
                    rel_str = "— In Line"
                
                table_data.append([cc, curr_str, prior_str, rep_str, peer_str, cc_str, rel_str])
            
            cc_table = Table(table_data, colWidths=[0.8*inch, 1*inch, 1*inch, 0.8*inch, 1.1*inch, 1*inch, 0.9*inch])
            cc_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E8F4F8')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#1F4788')),
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9F9F9')]),
                ('TOPPADDING', (0, 0), (-1, -1), 3),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ]))
            
            story.append(cc_table)
            story.append(Spacer(1, 0.15*inch))
            
            # Add narrative insights
            story.append(Paragraph("📊 Key Insights:", body_style))
            story.append(Spacer(1, 0.05*inch))
            
            # Find best and worst performing cost centers
            above_cc = cc_performance[cc_performance['relative_performance'] == 'Above']
            below_cc = cc_performance[cc_performance['relative_performance'] == 'Below']
            
            if not above_cc.empty:
                best_cc = above_cc.iloc[0]['cost_center']
                best_chg = above_cc.iloc[0]['rep_pct_change']
                if pd.notna(best_chg) and np.isfinite(best_chg):
                    story.append(Paragraph(
                        f"• In Cost Center <b>{best_cc}</b>, you outperformed other reps on the same accounts with {best_chg:+.1f}% growth.",
                        body_style
                    ))
            
            if not below_cc.empty:
                worst_cc = below_cc.iloc[0]['cost_center']
                worst_chg = below_cc.iloc[0]['rep_pct_change']
                if pd.notna(worst_chg) and np.isfinite(worst_chg):
                    story.append(Paragraph(
                        f"• Cost Center <b>{worst_cc}</b> showed {worst_chg:+.1f}% change, below peer performance on those accounts.",
                        body_style
                    ))
            
            story.append(Spacer(1, 0.1*inch))
            story.append(Paragraph(
                "<i>Note: 'Same-Acct Peers' = other reps selling to YOUR accounts. 'All CC Reps' = all reps selling in that cost center.</i>",
                caption_style
            ))
        else:
            # No cost center data available
            story.append(Paragraph(
                "Cost center performance comparison unavailable due to insufficient historical data.",
                body_style
            ))
    
    # ========================================
    # PAGE TYPE C: ACCOUNT-LEVEL PERFORMANCE DETAIL
    # (Only if we have MoM data and accounts)
    # ========================================
    
    if can_show_mom_comparison and has_accounts:
        # Compute account-level performance
        account_perf = _compute_account_performance(
            rep_name, df, last_month_start, last_month_end, 
            prior_month_start, prior_month_end, date_field
        )
        
        if not account_perf.empty:
            story.append(PageBreak())
            story.append(Paragraph(f"Account-Level Detail: {rep_name}", header_style))
            story.append(Spacer(1, 0.1*inch))
            
            # Group by cost center
            cost_centers = account_perf['cost_center'].unique()
            
            for cc in cost_centers:
                cc_accounts = account_perf[account_perf['cost_center'] == cc].copy()
                cc_accounts = cc_accounts.sort_values('rep_pct_change', ascending=False)
                
                story.append(Paragraph(f"Cost Center: {cc}", subheader_style))
                story.append(Spacer(1, 0.05*inch))
                
                # Build table
                acct_table_data = [["Account", "Current", "Prior", "Rep %Δ", "Peers %Δ", "Relative"]]
                
                for _, row in cc_accounts.head(10).iterrows():  # Limit to top 10 per CC
                    acct = row['account']
                    curr = row['current_revenue']
                    prior = row['prior_revenue']
                    rep_chg = row['rep_pct_change']
                    peer_chg = row['peers_pct_change']
                    relative = row['relative_performance']
                    
                    curr_str = f"${curr:,.0f}" if pd.notna(curr) else "$0"
                    prior_str = f"${prior:,.0f}" if pd.notna(prior) else "$0"
                    rep_str = f"{rep_chg:+.1f}%" if pd.notna(rep_chg) and np.isfinite(rep_chg) else "N/A"
                    peer_str = f"{peer_chg:+.1f}%" if pd.notna(peer_chg) and np.isfinite(peer_chg) else "N/A"
                    
                    if relative == "Above":
                        rel_str = "↑"
                    elif relative == "Below":
                        rel_str = "↓"
                    else:
                        rel_str = "—"
                    
                    acct_table_data.append([acct, curr_str, prior_str, rep_str, peer_str, rel_str])
                
                acct_table = Table(acct_table_data, colWidths=[2*inch, 1*inch, 1*inch, 0.8*inch, 0.8*inch, 0.6*inch])
                acct_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E8F4F8')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#1F4788')),
                    ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                    ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9F9F9')]),
                    ('TOPPADDING', (0, 0), (-1, -1), 3),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
                ]))
                
                story.append(acct_table)
                
                if len(cc_accounts) > 10:
                    story.append(Paragraph(f"<i>Showing top 10 of {len(cc_accounts)} accounts</i>", caption_style))
                
                story.append(Spacer(1, 0.15*inch))
    
    # ========================================
    # TOP / BOTTOM ACCOUNT MOVERS
    # (Only if we have MoM data)
    # ========================================
    
    if can_show_mom_comparison and has_accounts:
        account_perf = _compute_account_performance(
            rep_name, df, last_month_start, last_month_end, 
            prior_month_start, prior_month_end, date_field
        )
        
        if not account_perf.empty and len(account_perf) >= 2:
            story.append(PageBreak())
            story.append(Paragraph(f"Top & Bottom Account Movers: {rep_name}", header_style))
            story.append(Spacer(1, 0.1*inch))
            
            # Top 5 Accounts by revenue change
            account_perf['revenue_change'] = account_perf['current_revenue'] - account_perf['prior_revenue']
            account_perf['revenue_change'] = account_perf['revenue_change'].fillna(0)
            
            # Only show if we have meaningful data
            has_positive = (account_perf['revenue_change'] > 0).any()
            has_negative = (account_perf['revenue_change'] < 0).any()
            
            if has_positive:
                top_5 = account_perf.nlargest(5, 'revenue_change')
                
                # Top 5
                story.append(Paragraph("🔼 Top 5 Growing Accounts", subheader_style))
                story.append(Spacer(1, 0.05*inch))
                
                top_table_data = [["Rank", "Account", "Cost Center", "Revenue Δ", "%Δ"]]
                for i, (_, row) in enumerate(top_5.iterrows(), 1):
                    acct = row['account']
                    cc = row['cost_center']
                    rev_chg = row['revenue_change']
                    pct_chg = row['rep_pct_change']
                    
                    top_table_data.append([
                        str(i),
                        acct,
                        str(cc),
                        f"${rev_chg:,.0f}",
                        f"{pct_chg:+.1f}%" if pd.notna(pct_chg) and np.isfinite(pct_chg) else "N/A"
                    ])
                
                top_table = Table(top_table_data, colWidths=[0.5*inch, 2.5*inch, 1*inch, 1*inch, 0.8*inch])
                top_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#D5F4E6')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#27AE60')),
                    ('ALIGN', (0, 0), (0, -1), 'CENTER'),
                    ('ALIGN', (1, 0), (1, -1), 'LEFT'),
                    ('ALIGN', (2, 0), (-1, -1), 'RIGHT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 9),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('TOPPADDING', (0, 0), (-1, -1), 4),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ]))
                
                story.append(top_table)
                story.append(Spacer(1, 0.2*inch))
            
            if has_negative:
                bottom_5 = account_perf.nsmallest(5, 'revenue_change')
                
                # Bottom 5
                story.append(Paragraph("🔽 Top 5 Declining Accounts", subheader_style))
                story.append(Spacer(1, 0.05*inch))
                
                bottom_table_data = [["Rank", "Account", "Cost Center", "Revenue Δ", "%Δ"]]
                for i, (_, row) in enumerate(bottom_5.iterrows(), 1):
                    acct = row['account']
                    cc = row['cost_center']
                    rev_chg = row['revenue_change']
                    pct_chg = row['rep_pct_change']
                    
                    bottom_table_data.append([
                        str(i),
                        acct,
                        str(cc),
                        f"${rev_chg:,.0f}",
                        f"{pct_chg:+.1f}%" if pd.notna(pct_chg) and np.isfinite(pct_chg) else "N/A"
                    ])
                
                bottom_table = Table(bottom_table_data, colWidths=[0.5*inch, 2.5*inch, 1*inch, 1*inch, 0.8*inch])
                bottom_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#FADBD8')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#E74C3C')),
                    ('ALIGN', (0, 0), (0, -1), 'CENTER'),
                    ('ALIGN', (1, 0), (1, -1), 'LEFT'),
                    ('ALIGN', (2, 0), (-1, -1), 'RIGHT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 9),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('TOPPADDING', (0, 0), (-1, -1), 4),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ]))
                
                story.append(bottom_table)
                story.append(Spacer(1, 0.15*inch))
            
            # Add insights only if we have both positive and negative
            if has_positive and has_negative:
                story.append(Paragraph("📖 What This Means:", body_style))
                story.append(Spacer(1, 0.05*inch))
                
                top_5 = account_perf.nlargest(5, 'revenue_change')
                bottom_5 = account_perf.nsmallest(5, 'revenue_change')
                
                top_total_gain = top_5['revenue_change'].sum()
                bottom_total_loss = bottom_5['revenue_change'].sum()
                
                story.append(Paragraph(
                    f"• Your top 5 growing accounts added <b>${top_total_gain:,.0f}</b> in revenue month-over-month.",
                    body_style
                ))
                story.append(Paragraph(
                    f"• Your top 5 declining accounts lost <b>${abs(bottom_total_loss):,.0f}</b> in revenue.",
                    body_style
                ))
                story.append(Paragraph(
                    f"• Net impact from these 10 accounts: <b>${(top_total_gain + bottom_total_loss):,.0f}</b>",
                    body_style
                ))
                
                story.append(Spacer(1, 0.1*inch))
                story.append(Paragraph(
                    "<i>Focus: Replicate success from top growers and investigate root causes of declines.</i>",
                    caption_style
                ))
    
    # If no pages were rendered, don't leave the rep hanging
    # (This shouldn't happen due to early returns, but just in case)
    if not can_show_mom_comparison:
        story.append(Spacer(1, 0.1*inch))
        story.append(Paragraph(
            f"<i>Note: Detailed month-over-month analysis for {rep_name} is unavailable due to insufficient transaction history in the prior month. At least two complete months of data are required for trend analysis.</i>",
            caption_style
        ))


def _compute_cc_monthly_trends(
    rep_name: str,
    monthly_facts: pd.DataFrame,
    last_complete_month: pd.Timestamp
) -> pd.DataFrame:
    """Compute cost center monthly sales trends for up to 24 months.
    
    Returns DataFrame with monthly data for cost centers that are >= 5% of rep's total revenue.
    Columns: cost_center, month_start, rep_sales, rep_mom_pct, cc_baseline_mom_pct
    """
    if 'cost_center' not in monthly_facts.columns:
        return pd.DataFrame()
    
    # Get 24 months of history (or what's available)
    start_month = last_complete_month - pd.DateOffset(months=23)
    
    # Filter to rep's data in the window
    rep_data = monthly_facts[
        (monthly_facts['salesperson'] == rep_name) &
        (monthly_facts['month_start'] >= start_month) &
        (monthly_facts['month_start'] <= last_complete_month)
    ].copy()
    
    if rep_data.empty:
        return pd.DataFrame()
    
    # Calculate rep's total revenue across all cost centers in the window
    rep_total_revenue = rep_data['revenue'].sum()
    
    if rep_total_revenue == 0:
        return pd.DataFrame()
    
    # Identify cost centers >= 5% of rep total
    cc_totals = rep_data.groupby('cost_center')['revenue'].sum()
    significant_ccs = cc_totals[cc_totals >= rep_total_revenue * 0.05].index.tolist()
    
    if not significant_ccs:
        return pd.DataFrame()
    
    # Get all months in the window
    all_months_data = monthly_facts[
        (monthly_facts['month_start'] >= start_month) &
        (monthly_facts['month_start'] <= last_complete_month)
    ].copy()
    
    results = []
    
    for cc in significant_ccs:
        # Get rep's monthly sales for this CC
        rep_cc_monthly = rep_data[rep_data['cost_center'] == cc].groupby('month_start')['revenue'].sum()
        
        # Get all reps' monthly sales for this CC (baseline)
        cc_monthly = all_months_data[all_months_data['cost_center'] == cc].groupby('month_start')['revenue'].sum()
        
        # Get all unique months
        months = sorted(rep_cc_monthly.index.union(cc_monthly.index))
        
        prev_rep_sales = None
        prev_cc_sales = None
        
        for month in months:
            rep_sales = rep_cc_monthly.get(month, 0)
            cc_sales = cc_monthly.get(month, 0)
            
            # Calculate MoM%
            if prev_rep_sales is not None and prev_rep_sales > 0:
                rep_mom_pct = ((rep_sales - prev_rep_sales) / prev_rep_sales * 100)
            elif prev_rep_sales == 0 and rep_sales > 0:
                rep_mom_pct = None  # Will be marked as "New"
            else:
                rep_mom_pct = 0.0
            
            if prev_cc_sales is not None and prev_cc_sales > 0:
                cc_mom_pct = ((cc_sales - prev_cc_sales) / prev_cc_sales * 100)
            elif prev_cc_sales == 0 and cc_sales > 0:
                cc_mom_pct = None  # Will be marked as "New"
            else:
                cc_mom_pct = 0.0
            
            results.append({
                'cost_center': str(cc),
                'month_start': month,
                'rep_sales': rep_sales,
                'rep_mom_pct': rep_mom_pct,
                'cc_baseline_mom_pct': cc_mom_pct
            })
            
            prev_rep_sales = rep_sales
            prev_cc_sales = cc_sales
    
    if not results:
        return pd.DataFrame()
    
    return pd.DataFrame(results)


def _compute_shared_account_trends(
    rep_name: str,
    monthly_facts: pd.DataFrame,
    last_complete_month: pd.Timestamp
) -> dict:
    """Compute shared-account trends between rep and other reps.
    
    Returns dict mapping shared_rep_name -> DataFrame with monthly MoM% for shared accounts.
    DataFrame columns: month_start, rep_a_mom_pct, rep_b_mom_pct
    """
    if 'account_number' not in monthly_facts.columns:
        return {}
    
    # Get 24 months of history
    start_month = last_complete_month - pd.DateOffset(months=23)
    
    # Filter to window
    window_data = monthly_facts[
        (monthly_facts['month_start'] >= start_month) &
        (monthly_facts['month_start'] <= last_complete_month)
    ].copy()
    
    if window_data.empty:
        return {}
    
    # Get rep's accounts
    rep_accounts = set(window_data[window_data['salesperson'] == rep_name]['account_number'].dropna().unique())
    
    if not rep_accounts:
        return {}
    
    # Find other reps who share accounts
    other_reps_data = window_data[
        (window_data['salesperson'] != rep_name) &
        (window_data['account_number'].isin(rep_accounts))
    ]
    
    if other_reps_data.empty:
        return {}
    
    # Calculate shared revenue per rep
    shared_revenue_by_rep = other_reps_data.groupby('salesperson')['revenue'].sum().sort_values(ascending=False)
    
    # Take top 5 shared reps
    top_shared_reps = shared_revenue_by_rep.head(5).index.tolist()
    
    results = {}
    
    for shared_rep in top_shared_reps:
        # Find accounts shared between rep_name and shared_rep
        shared_rep_accounts = set(window_data[window_data['salesperson'] == shared_rep]['account_number'].dropna().unique())
        shared_accounts = rep_accounts.intersection(shared_rep_accounts)
        
        if not shared_accounts:
            continue
        
        # Get monthly revenue on shared accounts only
        rep_a_monthly = window_data[
            (window_data['salesperson'] == rep_name) &
            (window_data['account_number'].isin(shared_accounts))
        ].groupby('month_start')['revenue'].sum()
        
        rep_b_monthly = window_data[
            (window_data['salesperson'] == shared_rep) &
            (window_data['account_number'].isin(shared_accounts))
        ].groupby('month_start')['revenue'].sum()
        
        # Get all months
        months = sorted(rep_a_monthly.index.union(rep_b_monthly.index))
        
        monthly_results = []
        prev_a = None
        prev_b = None
        
        for month in months:
            a_rev = rep_a_monthly.get(month, 0)
            b_rev = rep_b_monthly.get(month, 0)
            
            # Calculate MoM%
            if prev_a is not None and prev_a > 0:
                a_mom = ((a_rev - prev_a) / prev_a * 100)
            elif prev_a == 0 and a_rev > 0:
                a_mom = None
            else:
                a_mom = 0.0
            
            if prev_b is not None and prev_b > 0:
                b_mom = ((b_rev - prev_b) / prev_b * 100)
            elif prev_b == 0 and b_rev > 0:
                b_mom = None
            else:
                b_mom = 0.0
            
            monthly_results.append({
                'month_start': month,
                'rep_a_mom_pct': a_mom,
                'rep_b_mom_pct': b_mom,
                'rep_a_rev': a_rev,
                'rep_b_rev': b_rev
            })
            
            prev_a = a_rev
            prev_b = b_rev
        
        if monthly_results:
            results[shared_rep] = pd.DataFrame(monthly_results)
    
    return results


def _render_cc_monthly_tables(
    rep_name: str,
    cc_monthly_data: pd.DataFrame,
    story: list,
    style_dict: dict
) -> None:
    """Render cost center monthly trend comparison tables (matching shared-account style)."""
    from reportlab.lib.colors import HexColor
    
    header_style = style_dict['header']
    subheader_style = style_dict['subheader']
    body_style = style_dict['body']
    caption_style = style_dict['caption']
    
    story.append(PageBreak())
    story.append(Paragraph(f"Cost Center Trend Comparison: {rep_name}", header_style))
    story.append(Paragraph(
        "24-month MoM% comparison: Your performance vs overall cost center baseline (all reps).",
        body_style
    ))
    story.append(Spacer(1, 0.1*inch))
    
    # Get unique months and cost centers
    months = sorted(cc_monthly_data['month_start'].unique())
    cost_centers = sorted(cc_monthly_data['cost_center'].unique())
    
    # Process each cost center separately (like shared-account comparison)
    for cc in cost_centers:
        cc_data = cc_monthly_data[cc_monthly_data['cost_center'] == cc].copy()
        
        if cc_data.empty:
            continue
        
        cc_months = sorted(cc_data['month_start'].unique())
        
        # Split into 12-month chunks
        for chunk_idx, month_chunk_start in enumerate(range(0, len(cc_months), 12)):
            month_chunk = cc_months[month_chunk_start:month_chunk_start + 12]
            
            if chunk_idx > 0:
                story.append(PageBreak())
                story.append(Paragraph(f"Cost Center Trend Comparison: {rep_name} (continued)", header_style))
                story.append(Spacer(1, 0.1*inch))
            
            start_label = month_chunk[0].strftime('%b %Y')
            end_label = month_chunk[-1].strftime('%b %Y')
            
            story.append(Paragraph(f"<b>Cost Center: {cc}</b>", subheader_style))
            story.append(Paragraph(f"Period: {start_label} – {end_label}", caption_style))
            story.append(Spacer(1, 0.05*inch))
            
            # Build table with 2 rows: Rep MoM% and CC Total MoM%
            table_data = []
            
            # Header row
            header = ["Metric"] + [m.strftime('%b %y') for m in month_chunk]
            table_data.append(header)
            
            # Two rows: Rep MoM%, CC Total MoM%
            rep_row = [f"{rep_name} MoM%"]
            cc_row = [f"{cc} Total MoM%"]
            
            for month in month_chunk:
                month_data = cc_data[cc_data['month_start'] == month]
                
                if not month_data.empty:
                    row = month_data.iloc[0]
                    
                    # Rep MoM%
                    if pd.isna(row['rep_mom_pct']):
                        rep_row.append("New")
                    elif np.isfinite(row['rep_mom_pct']):
                        rep_row.append(f"{row['rep_mom_pct']:+.1f}%")
                    else:
                        rep_row.append("—")
                    
                    # CC Baseline MoM%
                    if pd.isna(row['cc_baseline_mom_pct']):
                        cc_row.append("New")
                    elif np.isfinite(row['cc_baseline_mom_pct']):
                        cc_row.append(f"{row['cc_baseline_mom_pct']:+.1f}%")
                    else:
                        cc_row.append("—")
                else:
                    rep_row.append("—")
                    cc_row.append("—")
            
            table_data.append(rep_row)
            table_data.append(cc_row)
            
            # Calculate column widths
            col_width = 0.5*inch
            metric_col_width = 1.2*inch
            col_widths = [metric_col_width] + [col_width] * len(month_chunk)
            
            table = Table(table_data, colWidths=col_widths)
            
            # Apply styling with color coding
            table_style = [
                ('BACKGROUND', (0, 0), (-1, 0), HexColor('#E8F4F8')),
                ('TEXTCOLOR', (0, 0), (-1, 0), HexColor('#1F4788')),
                ('BACKGROUND', (0, 1), (0, -1), HexColor('#F5F5F5')),
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 6),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 0), (-1, -1), 2),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
            ]
            
            # Add color coding for MoM% rows
            for col_idx, month in enumerate(month_chunk, start=1):
                month_data = cc_data[cc_data['month_start'] == month]
                if not month_data.empty:
                    row = month_data.iloc[0]
                    
                    # Rep MoM% color (row 1)
                    if pd.notna(row['rep_mom_pct']) and np.isfinite(row['rep_mom_pct']):
                        if row['rep_mom_pct'] > 2:
                            table_style.append(('BACKGROUND', (col_idx, 1), (col_idx, 1), HexColor('#E8F5E9')))
                        elif row['rep_mom_pct'] < -2:
                            table_style.append(('BACKGROUND', (col_idx, 1), (col_idx, 1), HexColor('#FFEBEE')))
                    
                    # CC Total MoM% color (row 2)
                    if pd.notna(row['cc_baseline_mom_pct']) and np.isfinite(row['cc_baseline_mom_pct']):
                        if row['cc_baseline_mom_pct'] > 2:
                            table_style.append(('BACKGROUND', (col_idx, 2), (col_idx, 2), HexColor('#E8F5E9')))
                        elif row['cc_baseline_mom_pct'] < -2:
                            table_style.append(('BACKGROUND', (col_idx, 2), (col_idx, 2), HexColor('#FFEBEE')))
            
            table.setStyle(TableStyle(table_style))
            story.append(table)
            story.append(Spacer(1, 0.1*inch))
            
            # Add summary bullets for this chunk
            chunk_data = cc_data[cc_data['month_start'].isin(month_chunk)]
            
            you_outperformed = []
            cc_outperformed = []
            
            for _, row in chunk_data.iterrows():
                if (pd.notna(row['rep_mom_pct']) and pd.notna(row['cc_baseline_mom_pct']) and
                    np.isfinite(row['rep_mom_pct']) and np.isfinite(row['cc_baseline_mom_pct'])):
                    month_label = row['month_start'].strftime('%b %Y')
                    if row['rep_mom_pct'] > row['cc_baseline_mom_pct'] + 2:
                        you_outperformed.append(month_label)
                    elif row['cc_baseline_mom_pct'] > row['rep_mom_pct'] + 2:
                        cc_outperformed.append(month_label)
            
            if you_outperformed or cc_outperformed:
                story.append(Paragraph("📊 Performance Summary:", body_style))
                story.append(Spacer(1, 0.03*inch))
                
                if you_outperformed:
                    story.append(Paragraph(
                        f"• <b>Months you outperformed the cost center:</b> {', '.join(you_outperformed[:6])}",
                        body_style
                    ))
                
                if cc_outperformed:
                    story.append(Paragraph(
                        f"• <b>Months the cost center outperformed you:</b> {', '.join(cc_outperformed[:6])}",
                        body_style
                    ))
            
            story.append(Spacer(1, 0.15*inch))
            
            # Add MoM definition note on first table
            if chunk_idx == 0 and cc == cost_centers[0]:
                story.append(Paragraph(
                    "<i>Note: MoM% compares each month to the previous closed month. "
                    "Cost Center Total includes ALL reps (baseline). "
                    "Light green indicates growth (>2%), light red indicates decline (<-2%).</i>",
                    caption_style
                ))
                story.append(Spacer(1, 0.1*inch))


def _render_shared_account_tables(
    rep_name: str,
    shared_rep_data: dict,
    story: list,
    style_dict: dict
) -> None:
    """Render shared-account rep vs rep comparison tables."""
    from reportlab.lib.colors import HexColor
    
    header_style = style_dict['header']
    subheader_style = style_dict['subheader']
    body_style = style_dict['body']
    caption_style = style_dict['caption']
    
    story.append(PageBreak())
    story.append(Paragraph(f"Shared-Account Rep Comparison: {rep_name}", header_style))
    story.append(Paragraph(
        "24-month MoM% trends on SHARED ACCOUNTS ONLY (you vs other reps who sell to your accounts).",
        body_style
    ))
    story.append(Spacer(1, 0.1*inch))
    
    for shared_rep, df in shared_rep_data.items():
        months = sorted(df['month_start'].unique())
        
        # Split into 12-month chunks
        for chunk_idx, month_chunk_start in enumerate(range(0, len(months), 12)):
            month_chunk = months[month_chunk_start:month_chunk_start + 12]
            
            if chunk_idx > 0:
                story.append(PageBreak())
                story.append(Paragraph(f"Shared-Account Rep Comparison: {rep_name} (continued)", header_style))
                story.append(Spacer(1, 0.1*inch))
            
            start_label = month_chunk[0].strftime('%b %Y')
            end_label = month_chunk[-1].strftime('%b %Y')
            
            story.append(Paragraph(f"<b>You vs {shared_rep}</b> (Shared Accounts)", subheader_style))
            story.append(Paragraph(f"Period: {start_label} – {end_label}", caption_style))
            story.append(Spacer(1, 0.05*inch))
            
            # Build table
            table_data = []
            
            # Header row with month labels
            header = ["Metric"] + [m.strftime('%b %y') for m in month_chunk]
            table_data.append(header)
            
            # Two rows: You MoM%, Other Rep MoM%
            you_row = [f"{rep_name} MoM%"]
            other_row = [f"{shared_rep} MoM%"]
            
            for month in month_chunk:
                month_data = df[df['month_start'] == month]
                
                if not month_data.empty:
                    row = month_data.iloc[0]
                    
                    if pd.isna(row['rep_a_mom_pct']):
                        you_row.append("New")
                    elif np.isfinite(row['rep_a_mom_pct']):
                        you_row.append(f"{row['rep_a_mom_pct']:+.1f}%")
                    else:
                        you_row.append("—")
                    
                    if pd.isna(row['rep_b_mom_pct']):
                        other_row.append("New")
                    elif np.isfinite(row['rep_b_mom_pct']):
                        other_row.append(f"{row['rep_b_mom_pct']:+.1f}%")
                    else:
                        other_row.append("—")
                else:
                    you_row.append("—")
                    other_row.append("—")
            
            table_data.append(you_row)
            table_data.append(other_row)
            
            # Calculate column widths
            col_width = 0.5*inch
            metric_col_width = 1.2*inch
            col_widths = [metric_col_width] + [col_width] * len(month_chunk)
            
            table = Table(table_data, colWidths=col_widths)
            
            # Apply styling
            table_style = [
                ('BACKGROUND', (0, 0), (-1, 0), HexColor('#E8F4F8')),
                ('TEXTCOLOR', (0, 0), (-1, 0), HexColor('#1F4788')),
                ('BACKGROUND', (0, 1), (0, -1), HexColor('#F5F5F5')),
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 6),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 0), (-1, -1), 2),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
            ]
            
            # Add color coding
            for col_idx, month in enumerate(month_chunk, start=1):
                month_data = df[df['month_start'] == month]
                if not month_data.empty:
                    row = month_data.iloc[0]
                    
                    # Your MoM% color (row 1)
                    if pd.notna(row['rep_a_mom_pct']) and np.isfinite(row['rep_a_mom_pct']):
                        if row['rep_a_mom_pct'] > 2:
                            table_style.append(('BACKGROUND', (col_idx, 1), (col_idx, 1), HexColor('#E8F5E9')))
                        elif row['rep_a_mom_pct'] < -2:
                            table_style.append(('BACKGROUND', (col_idx, 1), (col_idx, 1), HexColor('#FFEBEE')))
                    
                    # Other rep MoM% color (row 2)
                    if pd.notna(row['rep_b_mom_pct']) and np.isfinite(row['rep_b_mom_pct']):
                        if row['rep_b_mom_pct'] > 2:
                            table_style.append(('BACKGROUND', (col_idx, 2), (col_idx, 2), HexColor('#E8F5E9')))
                        elif row['rep_b_mom_pct'] < -2:
                            table_style.append(('BACKGROUND', (col_idx, 2), (col_idx, 2), HexColor('#FFEBEE')))
            
            table.setStyle(TableStyle(table_style))
            story.append(table)
            story.append(Spacer(1, 0.1*inch))
            
            # Add callouts for this chunk
            chunk_df = df[df['month_start'].isin(month_chunk)]
            
            you_outgrew = []
            other_outgrew = []
            
            for _, row in chunk_df.iterrows():
                if (pd.notna(row['rep_a_mom_pct']) and pd.notna(row['rep_b_mom_pct']) and
                    np.isfinite(row['rep_a_mom_pct']) and np.isfinite(row['rep_b_mom_pct'])):
                    month_label = row['month_start'].strftime('%b %Y')
                    if row['rep_a_mom_pct'] > row['rep_b_mom_pct'] + 5:
                        you_outgrew.append(month_label)
                    elif row['rep_b_mom_pct'] > row['rep_a_mom_pct'] + 5:
                        other_outgrew.append(month_label)
            
            if you_outgrew or other_outgrew:
                story.append(Paragraph("📊 Performance Summary:", body_style))
                story.append(Spacer(1, 0.03*inch))
                
                if you_outgrew:
                    story.append(Paragraph(
                        f"• <b>Months you outgrew {shared_rep} on shared accounts:</b> {', '.join(you_outgrew[:6])}",
                        body_style
                    ))
                
                if other_outgrew:
                    story.append(Paragraph(
                        f"• <b>Months {shared_rep} outgrew you on shared accounts:</b> {', '.join(other_outgrew[:6])}",
                        body_style
                    ))
            
            story.append(Spacer(1, 0.15*inch))
    
    story.append(Paragraph(
        "<i>Note: MoM% calculated on shared accounts only (accounts both reps sold to during 24-month window). "
        "Light green indicates growth (>2%), light red indicates decline (<-2%).</i>",
        caption_style
    ))


def _generate_sales_rep_pdf(
    all_reps_data: list,
    df: pd.DataFrame,
    all_reps_metrics: list,
    all_reps_rolling: list,
    company_avg_margin: float,
    start_date: date,
    end_date: date
) -> BytesIO:
    """Generate comprehensive PDF for all sales reps.
    
    Args:
        all_reps_data: List of dicts with rep_name, metrics, rolling, health_score, components, insights
        df: Full sales rep orders dataframe
        all_reps_metrics: All reps metrics for peer comparison
        all_reps_rolling: All reps rolling metrics
        company_avg_margin: Company-wide average margin
        start_date: Analysis start date
        end_date: Analysis end date
    
    Returns:
        BytesIO buffer containing PDF
    """
    from app.services import sales_rep_service
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=0.5*inch,
        leftMargin=0.5*inch,
        topMargin=0.5*inch,
        bottomMargin=0.5*inch
    )
    
    # Styles
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#1F4788'),
        spaceAfter=12,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    header_style = ParagraphStyle(
        'CustomHeader',
        parent=styles['Heading2'],
        fontSize=16,
        textColor=colors.HexColor('#2E86AB'),
        spaceAfter=10,
        spaceBefore=12,
        fontName='Helvetica-Bold'
    )
    
    subheader_style = ParagraphStyle(
        'CustomSubHeader',
        parent=styles['Heading3'],
        fontSize=12,
        textColor=colors.HexColor('#333333'),
        spaceAfter=6,
        spaceBefore=8,
        fontName='Helvetica-Bold'
    )
    
    body_style = ParagraphStyle(
        'CustomBody',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.HexColor('#333333'),
        spaceAfter=6,
        leading=14
    )
    
    caption_style = ParagraphStyle(
        'CustomCaption',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.HexColor('#666666'),
        spaceAfter=4,
        fontStyle='italic'
    )
    
    # Build PDF content
    story = []
    
    # Package styles for passing to helper functions
    style_dict = {
        'title': title_style,
        'header': header_style,
        'subheader': subheader_style,
        'body': body_style,
        'caption': caption_style
    }
    
    # Cover page
    story.append(Spacer(1, 1*inch))
    story.append(Paragraph("Sales Rep Performance Report", title_style))
    story.append(Spacer(1, 0.2*inch))
    story.append(Paragraph(
        f"Analysis Period: {start_date.strftime('%B %d, %Y')} to {end_date.strftime('%B %d, %Y')}",
        body_style
    ))
    story.append(Paragraph(
        f"Generated: {date.today().strftime('%B %d, %Y')}",
        caption_style
    ))
    story.append(Spacer(1, 0.3*inch))
    story.append(Paragraph(
        "This report uses category-aware comparisons to ensure fair evaluation of reps with different product specializations.",
        body_style
    ))
    story.append(PageBreak())
    
    # Executive Summary (new)
    _generate_executive_summary_page(all_reps_data, start_date, end_date, story, style_dict)
    
    # Build peer groups once
    peer_groups = sales_rep_service.build_category_peer_groups(all_reps_metrics, min_peers=3)
    
    # Generate professional 1-2 page report per rep with rich insights
    # DESIGN PHILOSOPHY: Page 1 = Executive Snapshot (visual, scannable)
    #                    Page 2 = Performance Story (narrative, deep insights) - added when valuable
    for idx, rep_data in enumerate(all_reps_data):
        rep_name = rep_data["rep_name"]
        rep_metrics = rep_data["metrics"]
        rolling_metrics = rep_data["rolling"]
        health_score = rep_data["health_score"]
        component_scores = rep_data["components"]
        
        # Get peer group
        peer_names = peer_groups.get(rep_name, [m.get("salesperson") for m in all_reps_metrics])
        
        # Calculate insights
        insights = _calculate_driver_insights(
            rep_metrics, rolling_metrics, all_reps_metrics, df, rep_name, peer_names
        )
        
        # Generate recommendations (full set, we'll show 5 on page 1, rest on page 2 if needed)
        recommendations = _generate_actionable_recommendations(insights, component_scores, rep_metrics, max_recommendations=7)
        
        # Health band
        if health_score >= 80:
            band, band_color = "Healthy", colors.HexColor('#27AE60')
        elif health_score >= 60:
            band, band_color = "Watch", colors.HexColor('#F39C12')
        elif health_score >= 40:
            band, band_color = "At Risk", colors.HexColor('#E67E22')
        else:
            band, band_color = "Critical", colors.HexColor('#E74C3C')
        
        # ========================================
        # PAGE 1: EXECUTIVE SNAPSHOT
        # ========================================
        
        # A) HEADER - Professional, informative
        story.append(Paragraph(f"Sales Rep: {rep_name}", header_style))
        
        header_data = [
            ["Health Score", "Status Band", "Analysis Period", "Comparison Group"],
            [
                f"{health_score:.1f} / 100",
                band,
                f"{rolling_metrics.get('weeks_analyzed', 12)} closed ISO weeks",
                f"{len(peer_names)} reps selling similar categories"
            ]
        ]
        
        header_table = Table(header_data, colWidths=[1.5*inch, 1.5*inch, 2*inch, 2.3*inch])
        header_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E8F4F8')),
            ('BACKGROUND', (1, 1), (1, 1), band_color),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#1F4788')),
            ('TEXTCOLOR', (1, 1), (1, 1), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (1, 1), (1, 1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        story.append(header_table)
        story.append(Spacer(1, 0.15*inch))
        
        # B) OVERALL SUMMARY - Insightful, plain language (2-3 sentences)
        story.append(Paragraph("📊 Overall Performance Summary", subheader_style))
        
        # Build intelligent summary from actual metrics
        summary_parts = []
        
        # Overall status
        if health_score >= 75:
            summary_parts.append("Performance is <b>strong</b> across most dimensions.")
        elif health_score >= 60:
            summary_parts.append("Performance is <b>solid overall</b> with some areas for improvement.")
        elif health_score >= 40:
            summary_parts.append("Performance is <b>below expectations</b> and requires focused attention.")
        else:
            summary_parts.append("Performance shows <b>significant challenges</b> requiring immediate action.")
        
        # Identify key strengths and weaknesses (exclude metadata keys starting with _)
        sorted_components = sorted(
            [(k, v) for k, v in component_scores.items() 
             if isinstance(v, (int, float)) and not k.startswith('_')],
            key=lambda x: x[1], reverse=True
        )
        
        if sorted_components:
            strongest = sorted_components[0]
            weakest = sorted_components[-1]
            
            strength_names = {
                "revenue_momentum": "revenue growth",
                "margin_discipline": "margin consistency",
                "customer_health": "customer base strength",
                "sales_efficiency": "profit efficiency",
                "product_mix_quality": "product mix margins"
            }
            
            if strongest[1] >= 70:
                summary_parts.append(f"Key strength is <b>{strength_names.get(strongest[0], strongest[0])}</b>.")
            
            if weakest[1] < 60:
                summary_parts.append(f"Primary challenge is <b>{strength_names.get(weakest[0], weakest[0])}</b>, which is impacting the overall score.")
        
        # Add concentration risk if relevant
        cust_insights = insights.get("customer_health", {})
        top3_conc = cust_insights.get('top3_concentration', 0)
        if top3_conc > 60:
            summary_parts.append(f"Revenue is heavily concentrated in top 3 customers ({top3_conc:.0f}%), creating risk.")
        
        summary_text = " ".join(summary_parts)
        story.append(Paragraph(summary_text, body_style))
        story.append(Spacer(1, 0.15*inch))
        
        # C) SCORE BREAKDOWN - Full, clear, explanatory
        story.append(Paragraph("🎯 Health Score Breakdown", subheader_style))
        
        comp_data = [["Component", "Score", "What This Measures"]]
        
        comp_labels = {
            "revenue_momentum": "Revenue Momentum",
            "margin_discipline": "Margin Discipline",
            "customer_health": "Customer Health",
            "sales_efficiency": "Sales Efficiency",
            "product_mix_quality": "Product Mix Quality"
        }
        
        comp_explanations = {
            "revenue_momentum": "12-week growth trajectory (seasonally adjusted)",
            "margin_discipline": "Pricing consistency compared to reps selling similar categories",
            "customer_health": "Customer count, quality, and diversification",
            "sales_efficiency": "Gross profit generated per customer",
            "product_mix_quality": "Margin performance by product category"
        }
        
        for comp_key in ["revenue_momentum", "margin_discipline", "customer_health", "sales_efficiency", "product_mix_quality"]:
            if comp_key in component_scores:
                score = component_scores[comp_key]
                if isinstance(score, (int, float)):
                    comp_data.append([
                        comp_labels.get(comp_key, comp_key),
                        f"{score:.1f}",
                        comp_explanations.get(comp_key, "")
                    ])
        
        comp_table = Table(comp_data, colWidths=[1.8*inch, 0.8*inch, 4*inch])
        comp_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E8F4F8')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#1F4788')),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'CENTER'),
            ('ALIGN', (2, 0), (2, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9F9F9')]),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        story.append(comp_table)
        story.append(Spacer(1, 0.15*inch))
        
        # D) WHAT'S DRIVING THE SCORES - Data-backed, specific
        story.append(Paragraph("🔍 What's Driving Your Scores", subheader_style))
        
        # Revenue momentum with context
        rev_insights = insights.get("revenue_momentum", {})
        growth_pct = rev_insights.get("growth_pct", 0)
        trend = rev_insights.get("trend_direction", "Unknown")
        percentile = rev_insights.get("percentile_rank", 50)
        
        if growth_pct > 10:
            rev_context = "showing strong growth"
        elif growth_pct > 0:
            rev_context = "showing modest growth"
        elif growth_pct > -10:
            rev_context = "relatively flat"
        else:
            rev_context = "declining"
        
        story.append(Paragraph(
            f"<b>Revenue Trend:</b> {growth_pct:+.1f}% over 12 weeks ({rev_context}), "
            f"ranking at {percentile:.0f}th percentile compared to all reps. "
            f"Trend direction: {trend}.",
            body_style
        ))
        
        # Customer health with specific metrics
        active_custs = cust_insights.get('active_customers', 0)
        rev_per_cust = cust_insights.get('revenue_per_customer', 0)
        conc_risk = cust_insights.get('concentration_risk', 'Unknown')
        
        story.append(Paragraph(
            f"<b>Customer Base:</b> {active_custs} active customers generating "
            f"${rev_per_cust:,.0f} revenue per customer on average. "
            f"Top 3 customers represent {top3_conc:.0f}% of revenue ({conc_risk} concentration risk).",
            body_style
        ))
        
        # Margin discipline with peer comparison (category-aware)
        margin_insights = insights.get("margin_discipline", {})
        avg_margin = margin_insights.get('avg_margin', 0)
        peer_avg_margin = margin_insights.get("peer_avg_margin", 0)
        vs_peer_margin = margin_insights.get("vs_peer_margin", 0)
        std_dev = margin_insights.get('std_dev', 0)
        peer_avg_std = margin_insights.get('peer_avg_std', 10)
        consistency = margin_insights.get('consistency', 'Unknown')
        
        margin_comparison = ""
        if peer_avg_margin > 0:
            if vs_peer_margin > 2:
                margin_comparison = f"— <b>above</b> the {peer_avg_margin:.1f}% avg for similar-category reps ({vs_peer_margin:+.1f}%)"
            elif vs_peer_margin < -2:
                margin_comparison = f"— <b>below</b> the {peer_avg_margin:.1f}% avg for similar-category reps ({vs_peer_margin:+.1f}%)"
            else:
                margin_comparison = f"— <b>in line with</b> the {peer_avg_margin:.1f}% avg for similar-category reps"
        
        story.append(Paragraph(
            f"<b>Margin Performance:</b> Your margin: {avg_margin:.1f}% {margin_comparison}. "
            f"Consistency is {consistency.lower()} (std dev: {std_dev:.1f}% vs similar-category rep avg: {peer_avg_std:.1f}%). "
            f"<i>Compared to reps selling similar product categories.</i>",
            body_style
        ))
        
        # Sales efficiency
        eff_insights = insights.get("sales_efficiency", {})
        gp_per_cust = eff_insights.get('gp_per_customer', 0)
        gp_per_order = eff_insights.get('gp_per_order', 0)
        peer_gp_per_cust = eff_insights.get('peer_avg_gp_per_customer', 0)
        
        eff_comparison = ""
        if peer_gp_per_cust > 0:
            diff_pct = ((gp_per_cust - peer_gp_per_cust) / peer_gp_per_cust) * 100
            if abs(diff_pct) > 10:
                eff_comparison = f" ({diff_pct:+.0f}% vs avg for reps selling similar categories: ${peer_gp_per_cust:,.0f})"
            else:
                eff_comparison = f" (similar-category rep avg: ${peer_gp_per_cust:,.0f})"
        
        story.append(Paragraph(
            f"<b>Profit Efficiency:</b> Your profit per customer: ${gp_per_cust:,.0f}{eff_comparison}, "
            f"with ${gp_per_order:,.0f} gross profit per order.",
            body_style
        ))
        
        story.append(Spacer(1, 0.15*inch))
        
        # E) VISUALS - Readable charts
        weekly_rev = rev_insights.get("weekly_revenue", [])
        if weekly_rev and len(weekly_rev) >= 4:
            try:
                sparkline_img = _create_matplotlib_sparkline(weekly_rev)
                story.append(Paragraph("<b>12-Week Revenue Trend:</b>", body_style))
                story.append(Image(sparkline_img, width=4*inch, height=1.2*inch))
                story.append(Spacer(1, 0.1*inch))
            except Exception:
                pass
        
        mix_insights = insights.get("product_mix", {})
        top_cats = mix_insights.get("top_categories", [])
        if top_cats and len(top_cats) >= 2:
            try:
                cat_chart = _create_matplotlib_category_bars(top_cats)
                story.append(Paragraph("<b>Top Product Categories (Your Margin vs Similar-Category Rep Avg):</b>", body_style))
                story.append(Image(cat_chart, width=4.5*inch, height=1.8*inch))
                story.append(Spacer(1, 0.1*inch))
            except Exception:
                pass
        
        # F) ACTIONABLE NEXT STEPS - First 5 on page 1
        story.append(Paragraph("💡 Recommended Actions (Priority Order)", subheader_style))
        
        for i, rec in enumerate(recommendations[:5], 1):
            story.append(Paragraph(f"{i}. {rec}", body_style))
            story.append(Spacer(1, 0.05*inch))
        
        # ========================================
        # DECISION: ADD PAGE 2 FOR NARRATIVE DEPTH?
        # Add page 2 if: 
        # - Health score needs explanation (< 70)
        # - High concentration risk (> 60%)
        # - More than 5 recommendations
        # - Significant performance variance
        # ========================================
        
        needs_page2 = (
            health_score < 70 or
            top3_conc > 60 or
            len(recommendations) > 5 or
            abs(vs_peer_margin) > 5 or
            (sorted_components and (sorted_components[0][1] - sorted_components[-1][1]) > 40)
        )
        
        if needs_page2:
            story.append(PageBreak())
            
            # ========================================
            # PAGE 2: PERFORMANCE STORY
            # ========================================
            
            story.append(Paragraph(f"Performance Story: {rep_name} (continued)", header_style))
            story.append(Spacer(1, 0.15*inch))
            
            # A) NARRATIVE: What's Really Happening
            story.append(Paragraph("📖 What's Really Happening Behind the Numbers", subheader_style))
            
            narrative_parts = []
            
            # Revenue narrative
            if growth_pct < -5:
                narrative_parts.append(
                    f"Revenue has declined {abs(growth_pct):.1f}% over the past 12 weeks, "
                    f"placing performance in the bottom tier compared to other reps. "
                    f"This decline suggests either market challenges, customer churn, or reduced order frequency."
                )
            elif growth_pct > 15:
                narrative_parts.append(
                    f"Revenue growth of {growth_pct:.1f}% is exceptional and significantly outpaces other reps. "
                    f"This momentum should be analyzed to identify replicable success factors."
                )
            elif growth_pct < 5:
                narrative_parts.append(
                    f"Revenue is relatively flat with {growth_pct:+.1f}% change. "
                    f"While stability has value, lack of growth limits potential and suggests opportunity for expansion."
                )
            
            # Customer concentration narrative
            if top3_conc > 65:
                narrative_parts.append(
                    f"A critical concern is revenue concentration: the top 3 customers drive {top3_conc:.0f}% of sales. "
                    f"This creates significant risk—losing even one major account could substantially impact revenue. "
                    f"Diversifying the customer base should be an immediate priority."
                )
            elif active_custs < 10:
                narrative_parts.append(
                    f"With only {active_custs} active customers, the business base is narrow. "
                    f"Expanding the customer portfolio would provide stability and growth opportunity."
                )
            
            # Margin narrative
            if vs_peer_margin < -3:
                narrative_parts.append(
                    f"Margins are {abs(vs_peer_margin):.1f}% below the average for reps selling similar product categories. "
                    f"This suggests either aggressive discounting, unfavorable product mix within categories, "
                    f"or missed opportunities to maintain pricing discipline."
                )
            elif std_dev > peer_avg_std * 1.5:
                narrative_parts.append(
                    f"Margin volatility is notably high (std dev: {std_dev:.1f}% vs similar-category rep avg: {peer_avg_std:.1f}%). "
                    f"This inconsistency may indicate reactive pricing, one-off deals, or lack of systematic pricing strategy."
                )
            
            # Efficiency narrative
            if peer_gp_per_cust > 0 and gp_per_cust < peer_gp_per_cust * 0.8:
                narrative_parts.append(
                    f"Profit per customer is {((gp_per_cust/peer_gp_per_cust - 1) * 100):.0f}% below the average for reps selling similar product categories. "
                    f"This could stem from smaller order sizes, lower-margin product selection, or serving price-sensitive customers."
                )
            
            for para in narrative_parts:
                story.append(Paragraph(para, body_style))
                story.append(Spacer(1, 0.1*inch))
            
            story.append(Spacer(1, 0.1*inch))
            
            # B) DEEP-DIVE HIGHLIGHTS
            story.append(Paragraph("🔬 Deep-Dive Highlights", subheader_style))
            
            # Show component score distribution
            if sorted_components and len(sorted_components) >= 3:
                high_scores = [c for c in sorted_components if c[1] >= 70]
                low_scores = [c for c in sorted_components if c[1] < 50]
                
                if high_scores:
                    high_names = ", ".join([comp_labels.get(c[0], c[0]) for c in high_scores[:2]])
                    story.append(Paragraph(
                        f"<b>Strengths to maintain:</b> {high_names} are performing well above average. "
                        f"These areas provide a foundation to build upon.",
                        body_style
                    ))
                
                if low_scores:
                    low_names = ", ".join([comp_labels.get(c[0], c[0]) for c in low_scores[:2]])
                    story.append(Paragraph(
                        f"<b>Critical gaps:</b> {low_names} need immediate attention. "
                        f"Improvement in these areas would have the largest impact on overall health score.",
                        body_style
                    ))
            
            # Category-specific insights
            # Note: top_cats is list of tuples: (category_name, revenue, margin)
            if top_cats and len(top_cats) >= 2:
                best_cat = max(top_cats, key=lambda x: x[2])  # x[2] is margin
                story.append(Paragraph(
                    f"<b>Category performance:</b> Strongest margin performance is in "
                    f"{best_cat[0]} at {best_cat[2]:.1f}%. "
                    f"Consider whether this category can be expanded.",
                    body_style
                ))
            
            story.append(Spacer(1, 0.15*inch))
            
            # C) FORWARD-LOOKING GUIDANCE
            story.append(Paragraph("🎯 Forward-Looking Guidance", subheader_style))
            
            # Determine what would move score most
            if sorted_components:
                weakest_comp = sorted_components[-1]
                impact_estimate = (100 - weakest_comp[1]) * 0.2  # Rough estimate of score impact
                
                story.append(Paragraph(
                    f"<b>Highest-impact opportunity:</b> Improving {comp_labels.get(weakest_comp[0], weakest_comp[0])} "
                    f"from {weakest_comp[1]:.0f} to even 65 could increase overall health score by ~{impact_estimate:.0f} points.",
                    body_style
                ))
            
            # Risk management
            if top3_conc > 60:
                story.append(Paragraph(
                    f"<b>Risk management priority:</b> With {top3_conc:.0f}% concentration, actively cultivate "
                    f"additional relationships. Target: reduce top-3 concentration to below 50% within 6 months.",
                    body_style
                ))
            
            # Additional recommendations if available
            if len(recommendations) > 5:
                story.append(Spacer(1, 0.1*inch))
                story.append(Paragraph("<b>Additional Recommended Actions:</b>", body_style))
                for i, rec in enumerate(recommendations[5:], 6):
                    story.append(Paragraph(f"{i}. {rec}", body_style))
                    story.append(Spacer(1, 0.05*inch))
        
        # ========================================
        # NEW: DETAILED PERFORMANCE PAGES
        # Add progressive drill-down: Cost Center → Account Level → Top/Bottom Movers
        # ========================================
        try:
            _add_rep_detail_pages(
                rep_name=rep_name,
                df=df,
                story=story,
                style_dict=style_dict
            )
        except Exception as e:
            # Don't fail entire PDF if detail pages have issues
            story.append(PageBreak())
            story.append(Paragraph(f"Detail pages unavailable for {rep_name}: {str(e)}", caption_style))
        
        # Page break between reps
        story.append(PageBreak())
    
    # Definitions page
    story.append(PageBreak())
    story.append(Paragraph("📚 Metric Definitions", header_style))
    story.append(Spacer(1, 0.1*inch))
    
    definitions = [
        ("<b>Health Score:</b>", "A 0-100 score combining five key metrics. Higher is better."),
        ("<b>Category-Aware Comparison:</b>", "Reps are compared only to others selling similar product categories (30%+ overlap). This ensures fair evaluation."),
        ("<b>12 Closed Weeks:</b>", "We use complete Monday-Sunday weeks, excluding the current in-progress week, to avoid seasonal distortion."),
        ("<b>Revenue Momentum:</b>", "Growth rate over 12 weeks compared to all reps. Accounts for seasonality."),
        ("<b>Margin Discipline:</b>", "Average margin % and consistency compared to reps selling the same product categories. Volatility penalty based on peer group standards."),
        ("<b>Customer Health:</b>", "Number of customers and revenue per customer. High concentration in top 3 customers = risk."),
        ("<b>Sales Efficiency:</b>", "Gross profit per customer and per order, compared to reps selling similar categories."),
        ("<b>Product Mix:</b>", "Margin performance within each category you sell, compared to peers in those categories."),
    ]
    
    for term, definition in definitions:
        story.append(Paragraph(f"{term} {definition}", body_style))
        story.append(Spacer(1, 0.08*inch))
    
    story.append(Spacer(1, 0.2*inch))
    story.append(Paragraph(
        "<i>Note: This report uses industry-standard distributor metrics with category-aware fairness adjustments. "
        "Scores are relative to peer groups, not absolute benchmarks.</i>",
        caption_style
    ))
    
    # Build PDF
    doc.build(story)
    buffer.seek(0)
    
    return buffer


def _build_sales_rep_performance_tab() -> None:
    """Build the Sales Rep Performance dashboard tab."""
    from app.services import sales_rep_service
    
    st.header("Sales Rep Performance Dashboard")
    st.caption("Analyze sales rep effectiveness with industry-standard distributor metrics and health scoring")
    
    # PDF Download Button (placeholder - will be enabled after data loads)
    pdf_button_placeholder = st.empty()
    
    # Health Score Explanation
    with st.expander("ℹ️ What is the Health Score and How is it Calculated?", expanded=False):
        st.markdown("""
        ### Rep Health Score Overview
        
        The **Rep Health Score** is a comprehensive 0-100 metric that evaluates sales representative performance 
        across five key dimensions. This score uses **category-aware comparisons** to ensure fair evaluation 
        of reps with different product specializations.
        
        ---
        
        ### 🎯 Category-Aware Fairness
        
        **Why it matters:** Reps sell different product portfolios - some specialize in narrow categories while 
        others sell broadly. Global averages are misleading when comparing specialists to generalists.
        
        **How it works:**
        - Reps are grouped into **peer groups** based on product category overlap (30%+ similarity)
        - Metrics like GP/Order and Product Mix are compared only within relevant peer groups
        - Dynamic weight adjustment: If insufficient peers for comparison, that component's weight is redistributed
        - This ensures specialists aren't penalized for high-value/low-volume product focus
        
        ---
        
        ### Score Components (Weighted)
        
        #### 1. Revenue Momentum (15%) — **SEASONALLY ADJUSTED**
        **What it measures:** Growth trajectory using closed ISO weeks to handle seasonal patterns
        
        **Comparison scope:** All reps (seasonality affects everyone equally)
        
        **Calculation:**
        - Uses **12 closed ISO weeks** (Monday–Sunday) excluding current in-progress week
        - Three scoring signals to avoid binary outcomes:
          - **Percentile Rank (40%):** Compare 12-week growth % to all other reps
          - **Capped Growth Rate (30%):** Growth % capped at ±30% to prevent extremes
          - **Trend Direction (30%):** Linear slope over 12 weeks (upward/flat/downward)
        - **Scoring:** 
          - Top 25% of reps = 75-100 points (percentile)
          - Growth capped: +30% = 100, 0% = 50, -30% = 0
          - Upward trend adds 10 points, downward subtracts 10
        
        **Why it matters:** Seasonally robust - compares to peers during same period instead of absolute month-over-month which penalizes during slow seasons.
        
        ---
        
        #### 2. Margin Discipline (25%) — ⚙️ **CATEGORY-AWARE**
        **What it measures:** Ability to maintain healthy margins with consistency within your product categories
        
        **Comparison scope:** Category-similar peer group (30%+ product overlap)
        
        **Calculation:**
        - **Base Score:** Compare rep's average margin % to peer group average (reps selling same categories)
          - At or above peer avg = 100 base points
          - Each 1% below peer avg = -5 points
          - Each 1% above peer avg = +5 points (capped at 100)
        - **Volatility Penalty:** Standard deviation of order-level margins compared to peer group
          - If std dev higher than peer average, penalty applied
          - Ensures fair comparison: high-value categories naturally have different margin patterns
        
        **Why it matters:** High margins drive profitability; consistency indicates pricing discipline. Category-aware comparison ensures reps aren't penalized for selling categories with inherently different margin characteristics.
        
        ---
        
        #### 3. Customer Health (25%)
        **What it measures:** Breadth of customer base and revenue quality
        
        **Comparison scope:** All reps (customer diversification is universally important)
        
        **Calculation:**
        - **Customer Count Score (60%):** Percentile rank among all reps
          - More customers = higher score
          - Formula: (# of reps with fewer customers / total reps) × 100
        - **Revenue per Customer Score (40%):** Percentile rank
          - Higher revenue/customer = higher score
        - **Concentration Penalty:** If top 3 customers > 50% of revenue
          - Subtract 0.5 points for each % over 50%
          - Example: 65% concentration = -7.5 points
        
        **Why it matters:** Diversified customer base reduces risk; high revenue/customer indicates strong relationships.
        
        ---
        
        #### 4. Sales Efficiency (20%) — ⚙️ **CATEGORY-AWARE**
        **What it measures:** How effectively the rep converts activities into revenue and profit
        **Comparison scope:** Category-similar peer group (30%+ product overlap)
        
        **Calculation:**
        - **GP per Customer (50%):** Percentile rank within peer group
          - Higher GP/customer = better customer value extraction
        - **Gross Profit per Order (50%):** Percentile rank within peer group
          - Higher GP/order = better profitability per transaction
        - Combined score: (GP/Customer percentile × 0.5) + (GP/Order percentile × 0.5)
        - **Note:** Average Order Value removed - not meaningful when comparing different product types
        
        **Why it matters:** Customer-based efficiency (GP/customer) is more fair than order-based when product portfolios differ.
        
        ---
        
        #### 5. Product Mix Quality (15%) — ⚙️ **CATEGORY-AWARE**
        **What it measures:** Margin performance within the categories you actually sell
        
        **Comparison scope:** Category-specific peer groups (evaluated per category)
        
        **Calculation:**
        - For each product category you sell:
          - Compare your margin % to peers who also sell that category
          - Calculate percentile rank within category
        - **Weighted Average:** Combine category scores weighted by your revenue in each category
        - **Fallback:** If insufficient peers (<3), uses margin discipline as proxy
        
        **Why it matters:** Ensures you're not penalized for selling high-value/low-margin categories vs low-value/high-margin ones. Evaluates margin discipline within context.
        
        ---
        
        ### Health Bands
        
        | Score Range | Band | Color | Interpretation |
        |-------------|------|-------|----------------|
        | 80-100 | **Healthy** | 🟢 Green | Top performer - share best practices |
        | 60-79 | **Watch** | 🟠 Orange | Solid performer - minor improvements needed |
        | 40-59 | **At Risk** | 🟠 Orange-Red | Needs coaching - specific areas require attention |
        | 0-39 | **Critical** | 🔴 Red | Immediate intervention required |
        
        ---
        
        ### How to Use the Health Score
        
        **For Managers:**
        1. Monitor weekly for downward trends
        2. Review component scores to identify specific improvement areas
        3. Compare reps within their peer groups (similar product portfolios)
        4. Celebrate high scores and share success strategies
        
        **For Sales Reps:**
        1. Track your score over time
        2. Focus on lowest-scoring components first
        3. Understand your peer group - you're compared to reps selling similar products
        4. Review customer concentration and margin consistency regularly
        
        **Red Flags:**
        - Score dropping > 10 points month-over-month
        - Margin discipline score < 40
        - Customer concentration > 60% in top 3
        - Negative revenue momentum for 2+ consecutive periods
        """)
    
    # Load data
    with st.spinner("Loading sales rep data..."):
        try:
            rep_orders_df = loaders.load_sales_rep_orders(config.connection_string)
            assignments_df = loaders.load_account_assignments(config.connection_string)
        except Exception as e:
            st.error(f"Failed to load sales rep data: {e}")
            return
    
    if rep_orders_df.empty:
        st.warning("No sales rep order data available.")
        return
    
    # Build salesman number to name mapping
    salesman_map = {}
    if not rep_orders_df.empty and "salesman_number" in rep_orders_df.columns and "salesperson" in rep_orders_df.columns:
        temp_map = rep_orders_df[["salesman_number", "salesperson"]].drop_duplicates()
        salesman_map = dict(zip(temp_map["salesman_number"], temp_map["salesperson"]))
    
    # Filter to only include dates from August 4, 2025 onwards
    cutoff_date = pd.Timestamp(date(2025, 8, 4))
    rep_orders_df = rep_orders_df[rep_orders_df["order_date"] >= cutoff_date].copy()
    
    if rep_orders_df.empty:
        st.warning("No sales rep order data available from August 4, 2025 onwards.")
        return
    
    # Sidebar filters
    st.sidebar.header("Sales Rep Filters")
    
    # Get unique sales reps
    available_reps = sorted(rep_orders_df["salesperson"].dropna().unique().tolist())
    
    if not available_reps:
        st.warning("No sales reps found in data.")
        return
    
    # Rep selection
    selected_rep = st.sidebar.selectbox(
        "Select Sales Rep",
        options=available_reps,
        help="Choose a sales rep to analyze"
    )
    
    # Date range selection
    min_date = rep_orders_df["order_date"].min().date()
    max_date = rep_orders_df["order_date"].max().date()
    
    default_start = max_date - timedelta(days=180)  # Last 6 months
    default_end = max_date
    
    date_range = st.sidebar.date_input(
        "Date Range",
        value=(default_start, default_end),
        min_value=min_date,
        max_value=max_date,
        help="Select the analysis period"
    )
    
    if isinstance(date_range, tuple) and len(date_range) == 2:
        start_date, end_date = date_range
    else:
        start_date = default_start
        end_date = default_end
    
    # Product category filter (optional)
    available_categories = sorted(rep_orders_df["product_category"].dropna().unique().tolist())
    selected_categories = st.sidebar.multiselect(
        "Product Categories",
        options=available_categories,
        default=[],
        help="Filter by product categories (leave empty for all)"
    )
    
    # Filter data
    filtered_df = rep_orders_df.copy()
    if selected_categories:
        filtered_df = filtered_df[filtered_df["product_category"].isin(selected_categories)]
    
    # Calculate metrics for selected rep
    rep_metrics = sales_rep_service.calculate_rep_metrics(
        filtered_df, selected_rep, start_date, end_date
    )
    
    rolling_metrics = sales_rep_service.calculate_rolling_metrics(
        filtered_df, selected_rep, weeks=12  # CHANGED: Now uses 12 closed ISO weeks instead of months
    )
    
    new_vs_existing = sales_rep_service.calculate_new_vs_existing_customers(
        filtered_df, selected_rep, lookback_days=90
    )
    
    # Calculate company-wide average margin for benchmarking
    company_df = filtered_df[
        (filtered_df["order_date"] >= pd.Timestamp(start_date)) &
        (filtered_df["order_date"] <= pd.Timestamp(end_date))
    ]
    company_total_gp = company_df["line_gross_profit"].sum()
    company_total_rev = company_df["line_revenue"].sum()
    company_avg_margin = (company_total_gp / company_total_rev * 100) if company_total_rev > 0 else 0
    
    # Calculate metrics for all reps (for normalization in health score)
    all_reps_metrics = []
    all_reps_rolling = []  # NEW: Collect rolling metrics for all reps
    for rep in available_reps:
        rep_m = sales_rep_service.calculate_rep_metrics(
            filtered_df, rep, start_date, end_date
        )
        rep_roll = sales_rep_service.calculate_rolling_metrics(
            filtered_df, rep, weeks=12  # Changed from months to weeks
        )
        all_reps_metrics.append(rep_m)
        all_reps_rolling.append(rep_roll)  # NEW: Store rolling metrics
    
    # Calculate health score (UPDATED: Now passes BSCODE-based peer grouping data)
    health_score, component_scores = sales_rep_service.calculate_health_score(
        rep_metrics, rolling_metrics, all_reps_metrics, all_reps_rolling, company_avg_margin,
        assignments_df=assignments_df, salesman_map=salesman_map
    )
    
    health_band, health_color = sales_rep_service.get_health_band(health_score)
    
    # PDF Download Button
    st.markdown("---")
    
    if _HAS_REPORTLAB:
        col_pdf1, col_pdf2, col_pdf3 = st.columns([2, 1, 2])
        
        with col_pdf2:
            if st.button("📄 Download All Reps PDF", type="primary", use_container_width=True):
                with st.spinner("Generating PDF report for all reps... This may take a moment."):
                    try:
                        # Prepare data for all reps
                        all_reps_data = []
                        for rep in available_reps:
                            rep_m = sales_rep_service.calculate_rep_metrics(
                                filtered_df, rep, start_date, end_date
                            )
                            rep_roll = sales_rep_service.calculate_rolling_metrics(
                                filtered_df, rep, weeks=12
                            )
                            rep_health, rep_comp = sales_rep_service.calculate_health_score(
                                rep_m, rep_roll, all_reps_metrics, all_reps_rolling, company_avg_margin,
                                assignments_df=assignments_df, salesman_map=salesman_map
                            )
                            
                            all_reps_data.append({
                                "rep_name": rep,
                                "metrics": rep_m,
                                "rolling": rep_roll,
                                "health_score": rep_health,
                                "components": rep_comp
                            })
                        
                        # Generate PDF
                        pdf_buffer = _generate_sales_rep_pdf(
                            all_reps_data=all_reps_data,
                            df=filtered_df,
                            all_reps_metrics=all_reps_metrics,
                            all_reps_rolling=all_reps_rolling,
                            company_avg_margin=company_avg_margin,
                            start_date=start_date,
                            end_date=end_date
                        )
                        
                        # Download button
                        st.download_button(
                            label="⬇️ Download PDF Report",
                            data=pdf_buffer,
                            file_name=f"sales_rep_performance_{date.today().strftime('%Y%m%d')}.pdf",
                            mime="application/pdf",
                            use_container_width=True
                        )
                        
                        st.success(f"✅ PDF generated successfully for {len(all_reps_data)} reps!")
                        
                    except Exception as e:
                        st.error(f"Error generating PDF: {e}")
                        st.exception(e)
    else:
        st.warning("⚠️ PDF generation requires ReportLab library. Please install: pip install reportlab")
    
    st.markdown("---")
    
    # === SECTION 1: Rep Health Overview ===
    st.subheader("🎯 Rep Health Overview")
    
    col1, col2, col3, col4, col5 = st.columns(5)
    
    with col1:
        st.metric(
            "Health Score",
            f"{health_score}/100",
            delta=health_band,
            delta_color="normal" if health_score >= 60 else "inverse",
            help="Composite 0-100 score across 5 performance dimensions. Higher is better. Click ℹ️ above for detailed calculation."
        )
    
    with col2:
        st.metric(
            "Total Revenue",
            f"${rep_metrics['total_revenue']:,.0f}",
            help="Sum of all line-level revenue (extended price without funds) for selected date range"
        )
    
    with col3:
        st.metric(
            "Gross Profit",
            f"${rep_metrics['total_gross_profit']:,.0f}",
            help="Sum of all line-level gross profit (revenue minus cost) for selected date range"
        )
    
    with col4:
        st.metric(
            "Avg Margin %",
            f"{rep_metrics['avg_margin_pct']:.1f}%",
            delta=f"{rep_metrics['avg_margin_pct'] - company_avg_margin:+.1f}% vs company",
            help="Average margin % calculated as (Total GP / Total Revenue) × 100. Delta shows difference from company average."
        )
    
    with col5:
        st.metric(
            "Orders",
            f"{rep_metrics['orders_count']:,}",
            help="Total count of unique orders placed by this rep during selected date range"
        )
    
    # Health Score Component Breakdown
    st.markdown("#### Health Score Components")
    st.caption(f"⚙️ Category-Aware Scoring: Comparing to {component_scores.get('_peer_count', len(all_reps_metrics))} reps with similar product portfolios (30%+ overlap)")
    
    comp_col1, comp_col2, comp_col3, comp_col4, comp_col5 = st.columns(5)
    
    # Get adjusted weights if available
    adjusted_weights = component_scores.get("_weights", {
        "revenue_momentum": 0.15,
        "margin_discipline": 0.25,
        "customer_health": 0.25,
        "sales_efficiency": 0.20,
        "product_mix_quality": 0.15
    })
    
    with comp_col1:
        weight_pct = adjusted_weights.get("revenue_momentum", 0.15) * 100
        st.metric(
            f"Revenue Momentum ({weight_pct:.0f}%)",
            f"{component_scores['revenue_momentum']:.1f}",
            help="Seasonally-adjusted momentum using 12 closed ISO weeks. Combines percentile rank among ALL reps, capped growth rate (±30%), and trend direction."
        )
    with comp_col2:
        weight_pct = adjusted_weights.get("margin_discipline", 0.25) * 100
        st.metric(
            f"Margin Discipline ({weight_pct:.0f}%)",
            f"{component_scores['margin_discipline']:.1f}",
            help="Margin % and consistency compared to CATEGORY-SIMILAR peer group. Compares reps selling the same product categories. Penalties for high volatility vs peers."
        )
    with comp_col3:
        weight_pct = adjusted_weights.get("customer_health", 0.25) * 100
        st.metric(
            f"Customer Health ({weight_pct:.0f}%)",
            f"{component_scores['customer_health']:.1f}",
            help="Customer count + revenue/customer percentile rank (among ALL reps), minus concentration penalty. Higher = more diversified customer base."
        )
    with comp_col4:
        weight_pct = adjusted_weights.get("sales_efficiency", 0.20) * 100
        st.metric(
            f"Sales Efficiency ({weight_pct:.0f}%)",
            f"{component_scores['sales_efficiency']:.1f}",
            help="GP/customer + GP/order percentiles within CATEGORY-SIMILAR peer group. Ensures fair comparison between specialists and generalists."
        )
    with comp_col5:
        weight_pct = adjusted_weights.get("product_mix_quality", 0.15) * 100
        st.metric(
            f"Product Mix ({weight_pct:.0f}%)",
            f"{component_scores['product_mix_quality']:.1f}",
            help="Category-scoped margin performance. Compares margin within each product category you sell vs peers who sell the same categories."
        )
    
    st.divider()
    
    # === SECTION 2: Trend Diagnostics ===
    st.subheader("📈 Trend Diagnostics")
    
    trend_col1, trend_col2 = st.columns(2)
    
    with trend_col1:
        st.markdown("#### 12-Week Rolling Revenue Trend")
        if rolling_metrics["rolling_revenue"]:
            rolling_df = pd.DataFrame(rolling_metrics["rolling_revenue"])
            fig_revenue = px.line(
                rolling_df,
                x="order_date",
                y="rolling_revenue",
                title="Rolling 12-Week Average Revenue (Closed Weeks Only)",
                labels={"order_date": "Week Start", "rolling_revenue": "Revenue ($)"}
            )
            fig_revenue.update_traces(line_color="#3498DB", line_width=3)
            st.plotly_chart(fig_revenue, use_container_width=True)
        else:
            st.info("Not enough data for rolling trend")
    
    with trend_col2:
        st.markdown("#### 12-Week Rolling Margin Trend")
        if rolling_metrics["rolling_margin"]:
            rolling_margin_df = pd.DataFrame(rolling_metrics["rolling_margin"])
            fig_margin = px.line(
                rolling_margin_df,
                x="order_date",
                y="rolling_margin",
                title="Rolling 12-Week Average Margin % (Closed Weeks Only)",
                labels={"order_date": "Week Start", "rolling_margin": "Margin %"}
            )
            fig_margin.update_traces(line_color="#27AE60", line_width=3)
            st.plotly_chart(fig_margin, use_container_width=True)
        else:
            st.info("Not enough data for rolling trend")
    
    # Growth metrics
    growth_col1, growth_col2, growth_col3 = st.columns(3)
    
    with growth_col1:
        st.metric(
            "12-Week Revenue Growth",
            f"{rolling_metrics['revenue_growth_pct']:+.1f}%",
            delta="vs prior 12 weeks",
            help="Percentage change comparing most recent 12 closed weeks to the 12 weeks before that. Uses ISO weeks (Mon-Sun) to handle seasonality."
        )
    
    with growth_col2:
        st.metric(
            "New Customer Revenue %",
            f"{new_vs_existing['new_pct']:.1f}%",
            delta=f"${new_vs_existing['new_customer_revenue']:,.0f} from new customers",
            help="Percentage of revenue from customers whose first order was within last 90 days. Higher = strong acquisition"
        )
    
    with growth_col3:
        st.metric(
            "Existing Customer Revenue",
            f"${new_vs_existing['existing_customer_revenue']:,.0f}",
            delta=f"{100 - new_vs_existing['new_pct']:.1f}% of total",
            help="Revenue from customers with orders older than 90 days. Shows retention and relationship strength"
        )
    
    st.divider()
    
    # === SECTION 3: Customer Risk & Concentration ===
    st.subheader("⚠️ Customer Risk & Concentration")
    
    risk_col1, risk_col2, risk_col3, risk_col4 = st.columns(4)
    
    with risk_col1:
        st.metric(
            "Active Customers",
            f"{rep_metrics['active_customers']:,}",
            help="Count of unique customers (account numbers) with orders during selected date range"
        )
    
    with risk_col2:
        st.metric(
            "Revenue/Customer",
            f"${rep_metrics['revenue_per_customer']:,.0f}",
            help="Average revenue per customer (Total Revenue / Active Customers). Higher indicates stronger relationships"
        )
    
    with risk_col3:
        concentration_color = "🔴" if rep_metrics['top_customer_concentration'] > 40 else "🟢"
        st.metric(
            "Top Customer %",
            f"{concentration_color} {rep_metrics['top_customer_concentration']:.1f}%",
            delta="High risk if >40%",
            help="Percentage of total revenue from single largest customer. >40% = high concentration risk"
        )
    
    with risk_col4:
        concentration3_color = "🔴" if rep_metrics['top3_customer_concentration'] > 60 else "🟢"
        st.metric(
            "Top 3 Customers %",
            f"{concentration3_color} {rep_metrics['top3_customer_concentration']:.1f}%",
            delta="High risk if >60%",
            help="Percentage of total revenue from top 3 customers combined. >60% = high concentration risk"
        )
    
    # Margin consistency
    st.markdown("#### Margin Consistency")
    st.metric(
        "Order Margin Std Dev",
        f"{rep_metrics['margin_std_dev']:.2f}%",
        delta="Lower is more consistent",
        help="Standard deviation of order-level margin percentages. <5% = excellent, 5-15% = moderate, >15% = high volatility"
    )
    
    if rep_metrics['margin_std_dev'] > 15:
        st.warning("⚠️ High margin volatility detected. Review pricing consistency.")
    elif rep_metrics['margin_std_dev'] < 5:
        st.success("✅ Excellent margin consistency")
    else:
        st.info("ℹ️ Moderate margin consistency")
    
    st.divider()
    
    # === SECTION 4: Product Mix & Margin Quality ===
    st.subheader("📦 Product Mix & Margin Quality")
    
    # Revenue by category
    if rep_metrics['revenue_by_category']:
        category_df = pd.DataFrame([
            {"Category": cat, "Revenue": rev, "Margin %": rep_metrics['margin_by_category'].get(cat, 0)}
            for cat, rev in rep_metrics['revenue_by_category'].items()
        ])
        category_df = category_df.sort_values("Revenue", ascending=False)
        
        cat_col1, cat_col2 = st.columns(2)
        
        with cat_col1:
            st.markdown("#### Revenue by Product Category")
            fig_cat = px.bar(
                category_df,
                x="Revenue",
                y="Category",
                orientation="h",
                title="Revenue Distribution by Category",
                color="Margin %",
                color_continuous_scale="RdYlGn"
            )
            st.plotly_chart(fig_cat, use_container_width=True)
        
        with cat_col2:
            st.markdown("#### Category Performance Table")
            display_df = category_df.copy()
            display_df["Revenue"] = display_df["Revenue"].apply(lambda x: f"${x:,.0f}")
            display_df["Margin %"] = display_df["Margin %"].apply(lambda x: f"{x:.1f}%")
            st.dataframe(display_df, use_container_width=True, hide_index=True)
    else:
        st.info("No product category data available")
    
    st.divider()
    
    # === SECTION 5: Sales Efficiency Metrics ===
    st.subheader("⚡ Sales Efficiency")
    
    eff_col1, eff_col2, eff_col3 = st.columns(3)
    
    with eff_col1:
        st.metric(
            "GP per Customer",
            f"${rep_metrics['gp_per_customer']:,.0f}",
            delta="Higher is better",
            help="Gross profit per active customer (Total GP / Active Customers). Compares within category-similar peer group to ensure fair evaluation"
        )
    
    with eff_col2:
        st.metric(
            "Revenue per Customer",
            f"${rep_metrics['revenue_per_customer']:,.0f}",
            delta="Higher is better",
            help="Revenue per active customer (Total Revenue / Active Customers). Shows average customer value"
        )
    
    with eff_col3:
        st.metric(
            "Avg GP per Order",
            f"${rep_metrics['avg_gp_per_order']:,.0f}",
            delta="Higher is better",
            help="Average gross profit per order (Total GP / Orders Count). Compared within category-similar peer group for fairness"
        )
    
    st.divider()
    
    # === SECTION 6: All Reps Comparison ===
    st.subheader("📊 All Reps Comparison")
    
    # Build comparison table
    comparison_data = []
    for rep in available_reps:
        rep_m = sales_rep_service.calculate_rep_metrics(
            filtered_df, rep, start_date, end_date
        )
        rep_rolling = sales_rep_service.calculate_rolling_metrics(
            filtered_df, rep, weeks=12  # CHANGED: Now uses 12 closed ISO weeks
        )
        rep_health, rep_components = sales_rep_service.calculate_health_score(
            rep_m, rep_rolling, all_reps_metrics, all_reps_rolling, company_avg_margin,
            assignments_df=assignments_df, salesman_map=salesman_map
        )
        rep_band, _ = sales_rep_service.get_health_band(rep_health)
        
        comparison_data.append({
            "Sales Rep": rep,
            "Health Score": rep_health,
            "Health Band": rep_band,
            "Revenue Momentum": rep_components["revenue_momentum"],
            "Margin Discipline": rep_components["margin_discipline"],
            "Customer Health": rep_components["customer_health"],
            "Sales Efficiency": rep_components["sales_efficiency"],
            "Product Mix": rep_components["product_mix_quality"],
            "Revenue": rep_m["total_revenue"],
            "Gross Profit": rep_m["total_gross_profit"],
            "Margin %": rep_m["avg_margin_pct"],
            "Orders": rep_m["orders_count"],
            "Customers": rep_m["active_customers"],
            "Rev/Customer": rep_m["revenue_per_customer"],
            "GP/Customer": rep_m["gp_per_customer"],
            "GP/Order": rep_m["avg_gp_per_order"]
        })
    
    comparison_df = pd.DataFrame(comparison_data)
    comparison_df = comparison_df.sort_values("Health Score", ascending=False)
    
    # Health Score Component Breakdown Table
    st.markdown("#### Health Score Component Breakdown")
    st.caption("Shows the component scores (0-100) that make up each rep's overall health score")
    
    component_breakdown_df = comparison_df[[
        "Sales Rep", "Health Score", "Health Band",
        "Revenue Momentum", "Margin Discipline", "Customer Health",
        "Sales Efficiency", "Product Mix"
    ]].copy()
    
    # Add min/max summary rows
    min_row = {
        "Sales Rep": "MIN",
        "Health Score": component_breakdown_df["Health Score"].min(),
        "Health Band": "-",
        "Revenue Momentum": component_breakdown_df["Revenue Momentum"].min(),
        "Margin Discipline": component_breakdown_df["Margin Discipline"].min(),
        "Customer Health": component_breakdown_df["Customer Health"].min(),
        "Sales Efficiency": component_breakdown_df["Sales Efficiency"].min(),
        "Product Mix": component_breakdown_df["Product Mix"].min()
    }
    
    max_row = {
        "Sales Rep": "MAX",
        "Health Score": component_breakdown_df["Health Score"].max(),
        "Health Band": "-",
        "Revenue Momentum": component_breakdown_df["Revenue Momentum"].max(),
        "Margin Discipline": component_breakdown_df["Margin Discipline"].max(),
        "Customer Health": component_breakdown_df["Customer Health"].max(),
        "Sales Efficiency": component_breakdown_df["Sales Efficiency"].max(),
        "Product Mix": component_breakdown_df["Product Mix"].max()
    }
    
    avg_row = {
        "Sales Rep": "AVERAGE",
        "Health Score": component_breakdown_df["Health Score"].mean(),
        "Health Band": "-",
        "Revenue Momentum": component_breakdown_df["Revenue Momentum"].mean(),
        "Margin Discipline": component_breakdown_df["Margin Discipline"].mean(),
        "Customer Health": component_breakdown_df["Customer Health"].mean(),
        "Sales Efficiency": component_breakdown_df["Sales Efficiency"].mean(),
        "Product Mix": component_breakdown_df["Product Mix"].mean()
    }
    
    # Add summary rows to dataframe
    component_breakdown_df = pd.concat([
        component_breakdown_df,
        pd.DataFrame([min_row, max_row, avg_row])
    ], ignore_index=True)
    
    # Format for display
    display_breakdown = component_breakdown_df.copy()
    numeric_cols = ["Health Score", "Revenue Momentum", "Margin Discipline", 
                    "Customer Health", "Sales Efficiency", "Product Mix"]
    for col in numeric_cols:
        display_breakdown[col] = display_breakdown[col].apply(lambda x: f"{x:.1f}" if pd.notna(x) else "-")
    
    st.dataframe(display_breakdown, use_container_width=True, hide_index=True, height=500)
    
    # Download button for component breakdown
    csv_breakdown = component_breakdown_df.copy()
    csv_breakdown_data = csv_breakdown.to_csv(index=False)
    st.download_button(
        label="📥 Download Component Breakdown",
        data=csv_breakdown_data,
        file_name=f"sales_rep_component_breakdown_{start_date}_{end_date}.csv",
        mime="text/csv",
        key="download_component_breakdown"
    )
    
    st.divider()
    
    # Full Metrics Comparison Table
    st.markdown("#### Full Metrics Comparison")
    
    # Format for display
    display_comparison = comparison_df.copy()
    display_comparison["Revenue"] = display_comparison["Revenue"].apply(lambda x: f"${x:,.0f}")
    display_comparison["Gross Profit"] = display_comparison["Gross Profit"].apply(lambda x: f"${x:,.0f}")
    display_comparison["Margin %"] = display_comparison["Margin %"].apply(lambda x: f"{x:.1f}%")
    display_comparison["Rev/Customer"] = display_comparison["Rev/Customer"].apply(lambda x: f"${x:,.0f}")
    display_comparison["GP/Customer"] = display_comparison["GP/Customer"].apply(lambda x: f"${x:,.0f}")
    display_comparison["GP/Order"] = display_comparison["GP/Order"].apply(lambda x: f"${x:,.0f}")
    
    # Format component scores
    for col in ["Revenue Momentum", "Margin Discipline", "Customer Health", "Sales Efficiency", "Product Mix"]:
        display_comparison[col] = display_comparison[col].apply(lambda x: f"{x:.1f}")
    
    st.dataframe(display_comparison, use_container_width=True, hide_index=True, height=400)
    
    # Download button
    csv_export = comparison_df.copy()
    csv_data = csv_export.to_csv(index=False)
    st.download_button(
        label="📥 Download Full Comparison Data",
        data=csv_data,
        file_name=f"sales_rep_comparison_{start_date}_{end_date}.csv",
        mime="text/csv"
    )
    
    # Highlight at-risk reps
    at_risk_reps = comparison_df[comparison_df["Health Score"] < 60]
    if not at_risk_reps.empty:
        st.warning(f"⚠️ {len(at_risk_reps)} rep(s) need attention (Health Score < 60)")
        st.dataframe(
            at_risk_reps[["Sales Rep", "Health Score", "Health Band"]],
            use_container_width=True,
            hide_index=True
        )
